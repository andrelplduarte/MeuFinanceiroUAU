#!/usr/bin/env python
# -*- coding: utf-8 -*-
from __future__ import annotations

import hashlib
import json
import os
import subprocess
import sys
import time
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BASE_DIR))
OUT_DIR = BASE_DIR / "outputs"
BEFORE_GERAL = OUT_DIR / "CARTEIRAS GERAL.xlsx"
BEFORE_BASE = OUT_DIR / "CARTEIRAS BANCO DE DADOS.xlsx"

TEST_ROOT = OUT_DIR / "validacao_run" / "before_after_real"
INPUT_DIR = TEST_ROOT / "inputs_git_head"
AFTER_OUT = TEST_ROOT / "after_outputs"
REPORT_JSON = TEST_ROOT / "relatorio_validacao_before_after.json"


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def workbook_sheetnames(path: Path) -> list[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def read_df(path: Path, sheet: str, header_row_0based: int = 0) -> pd.DataFrame:
    try:
        return pd.read_excel(path, sheet_name=sheet, header=header_row_0based)
    except Exception:
        return pd.DataFrame()


def get_base_metrics(path_base: Path) -> dict:
    sheets = workbook_sheetnames(path_base)
    dr = read_df(path_base, "DADOS_RECEBER", 0) if "DADOS_RECEBER" in sheets else pd.DataFrame()
    cols_dr = [str(c) for c in dr.columns.tolist()]
    obrigatorias = [
        "DIA_VENCIMENTO_BOLETO",
        "CLASSIFICACAO_ADIMPLENCIA",
        "MES_VENCIMENTO",
        "ANO_VENCIMENTO",
    ]
    return {
        "path": str(path_base),
        "exists": path_base.is_file(),
        "size_bytes": path_base.stat().st_size if path_base.is_file() else None,
        "sheets": sheets,
        "dados_receber_columns": cols_dr,
        "required_columns_present": {c: (c in cols_dr) for c in obrigatorias},
    }


def get_geral_metrics(path_geral: Path) -> dict:
    sheets = workbook_sheetnames(path_geral)
    resumo_df = read_df(path_geral, "RESUMO GERAL", 7) if "RESUMO GERAL" in sheets else pd.DataFrame()
    return {
        "path": str(path_geral),
        "exists": path_geral.is_file(),
        "size_bytes": path_geral.stat().st_size if path_geral.is_file() else None,
        "sha256": sha256_file(path_geral) if path_geral.is_file() else None,
        "sheets": sheets,
        "resumo_rows": int(len(resumo_df)) if resumo_df is not None else 0,
        "resumo_columns": [str(c) for c in resumo_df.columns.tolist()] if resumo_df is not None else [],
    }


def git_show_to_file(repo: Path, git_path: str, dest: Path) -> None:
    raw = subprocess.check_output(
        ["git", "show", f"HEAD:{git_path}"],
        cwd=str(repo),
        stderr=subprocess.STDOUT,
    )
    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_bytes(raw)


def recover_real_inputs_from_git() -> tuple[list[str], list[str]]:
    ls = subprocess.check_output(
        ["git", "ls-tree", "-r", "--name-only", "HEAD"],
        cwd=str(BASE_DIR),
        text=True,
        encoding="utf-8",
        errors="replace",
    ).splitlines()

    receber = []
    recebidos = []
    for p in ls:
        if not p.startswith("uploads/rec_"):
            continue
        up = p.upper()
        if "_-_" not in up:
            # Ignora amostras genéricas rec_00_receber.txt que não representam lote real por empreendimento.
            continue
        if up.endswith("_RECEBER.TXT"):
            receber.append(p)
        elif up.endswith("_RECEBIDOS.TXT"):
            recebidos.append(p)

    # Pareamento por prefixo sem sufixo final.
    def key_of(git_path: str) -> str:
        s = git_path[:-4]  # .txt
        s = s.rsplit("_", 1)[0]  # remove RECEBER/RECEBIDOS
        return s

    map_r = {key_of(p): p for p in receber}
    map_p = {key_of(p): p for p in recebidos}
    keys = sorted(set(map_r) & set(map_p))
    if not keys:
        raise RuntimeError("Nenhum par real rec_* RECEBER/RECEBIDOS encontrado no HEAD.")

    receber_out: list[str] = []
    recebidos_out: list[str] = []
    max_pairs = int(os.environ.get("VALIDACAO_MAX_PARES", "1") or "1")
    keys = keys[: max(1, max_pairs)]

    for k in keys:
        p_r = map_r[k]
        p_p = map_p[k]
        dest_r = INPUT_DIR / Path(p_r).name
        dest_p = INPUT_DIR / Path(p_p).name
        git_show_to_file(BASE_DIR, p_r, dest_r)
        git_show_to_file(BASE_DIR, p_p, dest_p)
        receber_out.append(str(dest_r))
        recebidos_out.append(str(dest_p))
    return receber_out, recebidos_out


def main() -> int:
    from services.orquestrador_lote_uau import processar_lote_uau

    TEST_ROOT.mkdir(parents=True, exist_ok=True)
    INPUT_DIR.mkdir(parents=True, exist_ok=True)
    AFTER_OUT.mkdir(parents=True, exist_ok=True)

    if not BEFORE_BASE.is_file() or not BEFORE_GERAL.is_file():
        print("Arquivos 'antes' não encontrados em outputs/.", file=sys.stderr)
        return 2

    before_base = get_base_metrics(BEFORE_BASE)
    before_geral = get_geral_metrics(BEFORE_GERAL)

    caminhos_receber, caminhos_recebidos = recover_real_inputs_from_git()
    saida_base = AFTER_OUT / "base_placeholder.xlsx"

    t0 = time.perf_counter()
    (path_geral_after, path_base_after), reported_seconds = processar_lote_uau(
        caminhos_receber=caminhos_receber,
        caminhos_recebidos=caminhos_recebidos,
        caminho_saida_base=str(saida_base),
        modo_geracao="POR_EMPREENDIMENTO",
        caminhos_estoque=None,
    )
    elapsed = time.perf_counter() - t0

    path_geral_after = Path(path_geral_after)
    path_base_after = Path(path_base_after)

    after_base = get_base_metrics(path_base_after)
    after_geral = get_geral_metrics(path_geral_after)

    # Integridade de negócio (proxy objetivo): mesma estrutura de abas da geral e mesmo número de linhas do RESUMO.
    integridade = {
        "sheets_equal": before_geral["sheets"] == after_geral["sheets"],
        "resumo_rows_equal": before_geral["resumo_rows"] == after_geral["resumo_rows"],
        "resumo_columns_equal": before_geral["resumo_columns"] == after_geral["resumo_columns"],
        "sha_equal": before_geral["sha256"] == after_geral["sha256"],
    }

    report = {
        "before": {
            "base": before_base,
            "geral": before_geral,
        },
        "after": {
            "base": after_base,
            "geral": after_geral,
            "tempo_total_segundos_elapsed": round(elapsed, 2),
            "tempo_total_segundos_reported": round(float(reported_seconds), 2),
            "pares_processados": len(caminhos_receber),
        },
        "checks": {
            "abas_base_somente_duas": after_base["sheets"] == ["DADOS_RECEBER", "DADOS_RECEBIDOS"],
            "colunas_obrigatorias_dados_receber": after_base["required_columns_present"],
            "integridade_consolidada_proxy": integridade,
        },
    }

    REPORT_JSON.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Relatório salvo em: {REPORT_JSON}")
    print(json.dumps(report["checks"], ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
