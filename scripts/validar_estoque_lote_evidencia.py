# -*- coding: utf-8 -*-
"""Gera evidência objetiva do fluxo vigente POR_EMPREENDIMENTO."""
from __future__ import annotations

import os
import sys
import tempfile

from openpyxl import load_workbook

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, BASE)

from services.orquestrador_lote_uau import processar_lote_uau  # noqa: E402

CR = [
    os.path.join(BASE, "uploads", "ALVLT_-_LOT.SPE_RESIDENCIAL_OURILANDIA_-_RECEBER.txt"),
    os.path.join(BASE, "uploads", "CIDAN_-_LOT.RES.CIDADE_NOVA_-_RECEBER.txt"),
]
CP = [
    os.path.join(BASE, "uploads", "ALVLT_-_LOT.SPE_RESIDENCIAL_OURILANDIA_-_RECEBIDOS.txt"),
    os.path.join(BASE, "uploads", "CIDAN_-_LOT.RES.CIDADE_NOVA_-_RECEBIDOS.txt"),
]
EST = [
    os.path.join(BASE, "validacao_lote_estoque", "00_ALVLT_ESTOQUE.txt"),
    os.path.join(BASE, "validacao_lote_estoque", "01_CIDAN_ESTOQUE.txt"),
    os.path.join(BASE, "validacao_lote_estoque", "99_LTMAG_ORFAO_ESTOQUE.txt"),
]

ABAS_BASE_ESPERADAS = [
    "DADOS RECEBER",
    "DADOS RECEBIDOS",
    "DADOS GERAL",
    "PEND.PARCELAS",
    "CONSOLIDADO ESTOQUE",
    "CRITERIOS ANALISES",
]


def _sigla_da_aba(nome_aba: str) -> str:
    for sep in (" – ", " - "):
        if sep in nome_aba:
            return nome_aba.split(sep, 1)[0].strip()
    return nome_aba.strip()


def _validar_arquivo_geral(caminho: str) -> bool:
    print(f"\n{'='*70}\nA. VALIDAÇÃO — CARTEIRAS GERAL.xlsx\n{'='*70}")
    print(f"Arquivo: {caminho}")
    existe = os.path.isfile(caminho)
    print(f"Existe: {existe}")
    if not existe:
        return False
    wb = load_workbook(caminho, read_only=True, data_only=True)
    abas = wb.sheetnames
    wb.close()
    print(f"Total de abas: {len(abas)}")
    print(f"Abas: {abas}")
    ok_primeira = bool(abas) and abas[0] == "RESUMO GERAL"
    ok_demais = all("CONSOLIDADO" in a.upper() for a in abas[1:])
    print(f"Primeira aba = RESUMO GERAL: {ok_primeira}")
    print(f"Demais abas são consolidados: {ok_demais}")
    print(f"Siglas nos consolidados: {[_sigla_da_aba(a) for a in abas[1:]]}")
    return ok_primeira and ok_demais


def _validar_arquivo_base(caminho: str) -> bool:
    print(f"\n{'='*70}\nB. VALIDAÇÃO — CARTEIRAS BANCO DE DADOS.xlsx\n{'='*70}")
    print(f"Arquivo: {caminho}")
    existe = os.path.isfile(caminho)
    print(f"Existe: {existe}")
    if not existe:
        return False
    wb = load_workbook(caminho, read_only=True, data_only=True)
    abas = wb.sheetnames
    wb.close()
    print(f"Total de abas: {len(abas)}")
    print(f"Abas: {abas}")
    ok_exatas = abas == ABAS_BASE_ESPERADAS
    print(f"Abas exatas esperadas: {ok_exatas}")
    return ok_exatas


def main() -> None:
    out_dir = os.path.join(BASE, "outputs")
    os.makedirs(out_dir, exist_ok=True)

    for p in CR + CP + EST:
        if not os.path.isfile(p):
            print(f"FALTA ARQUIVO: {p}")
            sys.exit(1)

    fd, placeholder = tempfile.mkstemp(prefix="uau_val_", suffix=".xlsx", dir=out_dir)
    os.close(fd)
    ok = True
    try:
        (path_geral, path_base), tempo = processar_lote_uau(
            CR, CP, placeholder, "POR_EMPREENDIMENTO", EST
        )
        print(f"\nTempo POR_EMPREENDIMENTO: {tempo:.2f}s")
        print("\nArquivos gerados:")
        print(f"- {path_geral}")
        print(f"- {path_base}")
        ok = _validar_arquivo_geral(path_geral) and ok
        ok = _validar_arquivo_base(path_base) and ok

    finally:
        try:
            os.unlink(placeholder)
        except OSError:
            pass

    print("\n" + "=" * 70)
    print("C. RESULTADO FINAL")
    print(f"Validação fluxo POR_EMPREENDIMENTO: {'OK' if ok else 'FALHA'}")
    print("=" * 70)
    if not ok:
        sys.exit(1)


if __name__ == "__main__":
    main()
