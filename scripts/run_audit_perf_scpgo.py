#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Auditoria de performance SCPGO: baseline (HEAD commit) vs otimizado (working tree).
Gera relatório JSON e roda auditar_carteiras_geral_final nos dois .xlsx finais.
"""
from __future__ import annotations

import json
import os
import re
import shutil
import subprocess
import sys
import time
from pathlib import Path

REPO = Path(__file__).resolve().parents[1]
PROC = REPO / "services" / "processador_uau.py"
OUT_ROOT = REPO / "outputs" / "_audit_perf"
BASELINE_DIR = OUT_ROOT / "baseline_run"
OPT_DIR = OUT_ROOT / "optimized_run"
BACKUP_PROC = OUT_ROOT / "processador_uau_otimizado_worktree.py"

REC = REPO / "uploads" / "rec_00_SCPGO_-LOT.SCP_RESIDENCIAL_GOIANIA_-_RECEBER.txt"
REB = REPO / "uploads" / "reb_00_SCPGO_-LOT.SCP_RESIDENCIAL_GOIANIA_-_RECEBIDOS.txt"
EST = REPO / "uploads" / "est_00_ESTOQUE_ATUALIZADO.txt"

RE_TEMPO = re.compile(
    r"\[TEMPO\]\s+(?P<nome>[^\s:]+):\s+(?P<seg>[0-9]+\.?[0-9]*)s"
)


def _run_git(args: list[str]) -> None:
    subprocess.run(["git", *args], cwd=str(REPO), check=True)


def _parse_tempos(text: str) -> dict[str, float]:
    out: dict[str, float] = {}
    for m in RE_TEMPO.finditer(text):
        out[m.group("nome")] = float(m.group("seg"))
    return out


def _run_processador(
    label: str,
    out_dir: Path,
    nome_override: str,
) -> tuple[str, float, dict[str, float], str]:
    out_dir.mkdir(parents=True, exist_ok=True)
    placeholder = out_dir / "_placeholder_out.xlsx"
    code = f"""
import sys
sys.path.insert(0, r"{REPO}")
from services.processador_uau import processar_e_gerar_excel
import time
t0 = time.perf_counter()
path, tt = processar_e_gerar_excel(
    r"{REC}",
    r"{REB}",
    r"{placeholder}",
    caminho_estoque=r"{EST}",
    gerar_aba_consolidado_estoque=True,
    nome_arquivo_xlsx_override=r"{nome_override}",
)
print("__PATH_FINAL__", path)
print("__TEMPO_TOTAL__", tt)
print("__WALL__", time.perf_counter() - t0)
"""
    env = os.environ.copy()
    env["PYTHONUTF8"] = "1"
    p = subprocess.run(
        [str(REPO / "venv" / "Scripts" / "python.exe"), "-c", code],
        cwd=str(REPO),
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        env=env,
    )
    full = (p.stdout or "") + "\n" + (p.stderr or "")
    log_path = out_dir / f"console_{label}.log"
    log_path.write_text(full, encoding="utf-8")

    path_final = ""
    tempo_total = 0.0
    for line in full.splitlines():
        if line.startswith("__PATH_FINAL__"):
            parts = line.split(None, 2)
            if len(parts) >= 2:
                path_final = parts[1].strip()
        if line.startswith("__TEMPO_TOTAL__"):
            try:
                tempo_total = float(line.split()[1])
            except (IndexError, ValueError):
                pass

    tempos = _parse_tempos(full)
    if p.returncode != 0:
        raise RuntimeError(
            f"Falha run {label} rc={p.returncode}. Ver {log_path}"
        )
    return path_final, tempo_total, tempos, full


def _pct_antes_depois(antes: float, depois: float) -> float | None:
    if antes <= 0:
        return None
    return round(100.0 * (antes - depois) / antes, 2)


def _checar_estrutura_openpyxl(path: str) -> dict:
    from openpyxl import load_workbook
    from openpyxl.utils import column_index_from_string

    wb = load_workbook(path, data_only=False)
    try:
        names = wb.sheetnames
        resumo = "RESUMO GERAL" in names
        crit = []
        for sn in names:
            if sn == "RESUMO GERAL":
                continue
            ws = wb[sn]
            af = ws.auto_filter.ref if ws.auto_filter else None
            fp = ws.freeze_panes
            ok_af = bool(af) and str(af).upper().startswith("A8")
            ok_fp = fp in ("A9", "a9") or (fp and str(fp).upper().replace("$", "") == "A9")
            merges7 = [
                str(mr)
                for mr in ws.merged_cells.ranges
                if mr.min_row == mr.max_row == 7
            ]
            crit.append(
                {
                    "aba": sn,
                    "auto_filter_ref": str(af or ""),
                    "auto_filter_ok": ok_af,
                    "freeze_panes": str(fp or ""),
                    "freeze_ok": bool(ok_fp),
                    "max_column": ws.max_column,
                    "max_column_ge_aa": (ws.max_column or 0) >= column_index_from_string("AA"),
                    "merges_linha7_count": len(merges7),
                }
            )
        return {"ok_resumo": resumo, "abas": names, "emp_checks": crit}
    finally:
        wb.close()


def main() -> int:
    if not REC.is_file() or not REB.is_file():
        print("Arquivos SCPGO em uploads/ não encontrados.", file=sys.stderr)
        return 2
    OUT_ROOT.mkdir(parents=True, exist_ok=True)

    # 1) Backup do processador otimizado (working tree)
    shutil.copy2(PROC, BACKUP_PROC)
    print(f"[audit] backup otimizado -> {BACKUP_PROC}")

    report: dict = {"repo": str(REPO), "baseline_commit": None, "runs": {}}

    # 2) Baseline = versão commitada (HEAD)
    rev = subprocess.check_output(
        ["git", "rev-parse", "HEAD"], cwd=str(REPO), text=True
    ).strip()
    report["baseline_commit"] = rev

    try:
        _run_git(["checkout", "HEAD", "--", "services/processador_uau.py"])
        print("[audit] processador_uau.py restaurado para HEAD (baseline)")
        path_b, tt_b, tempos_b, _ = _run_processador(
            "baseline", BASELINE_DIR, "audit_baseline_scpgo.xlsx"
        )
        report["runs"]["baseline"] = {
            "path": path_b,
            "tempo_total": tt_b,
            "tempos": tempos_b,
        }
    finally:
        shutil.copy2(BACKUP_PROC, PROC)
        print("[audit] processador_uau.py restaurado para cópia otimizada (working tree)")

    path_o, tt_o, tempos_o, _ = _run_processador(
        "optimized", OPT_DIR, "audit_optimized_scpgo.xlsx"
    )
    report["runs"]["optimized"] = {
        "path": path_o,
        "tempo_total": tt_o,
        "tempos": tempos_o,
    }

    chaves = [
        "excel_escrever_openpyxl",
        "excel_pos_formatacao_openpyxl",
        "_validar_pre_exportacao",
        "montar_consolidado_total",
    ]
    comparativo = []
    for k in chaves:
        a = tempos_b.get(k)
        b = tempos_o.get(k)
        comparativo.append(
            {
                "etapa": k,
                "baseline_s": a,
                "optimized_s": b,
                "ganho_pct": _pct_antes_depois(a, b) if a and b else None,
            }
        )
    report["comparativo_etapas"] = comparativo
    report["tempo_total"] = {
        "baseline_s": tt_b,
        "optimized_s": tt_o,
        "ganho_pct": _pct_antes_depois(tt_b, tt_o),
    }

    # Auditoria carteiras + estrutura
    import importlib.util

    aud_path = REPO / "scripts" / "auditar_carteiras_geral_final.py"
    spec = importlib.util.spec_from_file_location("auditar_carteiras_geral_final", aud_path)
    if spec is None or spec.loader is None:
        raise RuntimeError("Não foi possível carregar auditar_carteiras_geral_final.py")
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    audit_carteiras = mod.audit

    aud_b = audit_carteiras(path_b)
    aud_o = audit_carteiras(path_o)
    report["auditoria_baseline"] = {
        "ok": aud_b.get("ok"),
        "falhas": aud_b.get("falhas", []),
        "avisos": aud_b.get("avisos", []),
        "somas_abas": aud_b.get("somas_abas_empreendimento"),
        "somas_resumo": aud_b.get("somas_resumo_geral"),
        "delta_resumo_vs_abas": aud_b.get("delta_resumo_vs_soma_abas"),
    }
    report["auditoria_optimized"] = {
        "ok": aud_o.get("ok"),
        "falhas": aud_o.get("falhas", []),
        "avisos": aud_o.get("avisos", []),
        "somas_abas": aud_o.get("somas_abas_empreendimento"),
        "somas_resumo": aud_o.get("somas_resumo_geral"),
        "delta_resumo_vs_abas": aud_o.get("delta_resumo_vs_soma_abas"),
    }

    metricas = [
        "Vl.Carteira",
        "Vl.Pago",
        "Vl.Vencer",
        "Vl.Principal (Encargos)",
        "Qtd.Parc.Atrasada",
    ]
    delta_fin = {}
    for m in metricas:
        sb = (aud_b.get("somas_abas_empreendimento") or {}).get(m)
        so = (aud_o.get("somas_abas_empreendimento") or {}).get(m)
        if sb is not None and so is not None:
            delta_fin[m] = round(float(so) - float(sb), 2)
        else:
            delta_fin[m] = None
    report["delta_financeiro_abas_baseline_vs_optimized"] = delta_fin

    report["estrutura_baseline"] = _checar_estrutura_openpyxl(path_b)
    report["estrutura_optimized"] = _checar_estrutura_openpyxl(path_o)

    aceite_pos = report["comparativo_etapas"]
    pos_gain = next(
        (x for x in aceite_pos if x["etapa"] == "excel_pos_formatacao_openpyxl"),
        None,
    )
    report["aceite"] = {
        "ganho_pos_formatacao_ge_40pct": bool(
            pos_gain and pos_gain.get("ganho_pct") is not None and pos_gain["ganho_pct"] >= 40.0
        ),
        "auditoria_sem_falha_critica": bool(aud_o.get("ok") and aud_b.get("ok")),
        "deltas_financeiros_zero": all(
            v is not None and abs(v) < 0.03 for v in delta_fin.values()
        ),
    }

    out_json = OUT_ROOT / "relatorio_audit_perf_scpgo.json"
    out_json.write_text(
        json.dumps(report, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )
    print(json.dumps(report, ensure_ascii=False, indent=2, default=str))
    print(f"\nRelatório salvo em: {out_json}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
