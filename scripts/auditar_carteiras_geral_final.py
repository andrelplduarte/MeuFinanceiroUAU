# -*- coding: utf-8 -*-
"""Auditoria objetiva do CARTEIRAS GERAL.xlsx."""
from __future__ import annotations

import argparse
import json
import re
import unicodedata
from decimal import Decimal
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

AZUL = "10243F"
VERDE = "92D050"
VERMELHO_ESPERADO = {"FF5E5E", "F8696B", "FF0000", "C00000", "E74C3C"}
AZUL_CLARO = "00B0F0"
AMARELO = "FFFF00"
BEGE = "FFF2CC"
BRANCO = "FFFFFF"
PRETO = "000000"

NOME_RESUMO = "RESUMO GERAL"

SIGLAS_MAPA = frozenset(
    "NVLOT LTMAG SCPTO SCPTI CIDAN VROLT ALVLT LTMON RVERD LTVIL LTMIN SCPGO ARAHF BVGWH MANHA MONTB LIFE".split()
)


def _fold(s: str) -> str:
    s = unicodedata.normalize("NFD", str(s or ""))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s.strip().upper())


def _norm_cols_df(df):
    return {_fold(c): c for c in df.columns}


def _norm_rgb(v) -> str:
    if v is None:
        return ""
    s = str(v).upper().replace("#", "")
    if len(s) == 8 and s.startswith(("FF", "00")):
        s = s[2:]
    return s


def _cell_fill_rgb(cell) -> str:
    try:
        fg = cell.fill.fgColor
        if fg and fg.type == "rgb" and fg.rgb:
            return _norm_rgb(fg.rgb)
    except Exception:
        pass
    return ""


def _merged_ranges(ws):
    return list(ws.merged_cells.ranges)


def _range_text(mr) -> str:
    return f"{get_column_letter(mr.min_col)}{mr.min_row}:{get_column_letter(mr.max_col)}{mr.max_row}"


def _sum_numeric_column_ws(ws, col_letter: str, start_row: int = 9) -> float:
    col_idx = column_index_from_string(col_letter)
    total = Decimal("0")
    for row in range(start_row, (ws.max_row or 0) + 1):
        val = ws.cell(row=row, column=col_idx).value
        if val in (None, ""):
            continue
        try:
            total += Decimal(str(val))
        except Exception:
            try:
                total += Decimal(str(float(val)))
            except Exception:
                continue
    return float(total)


def audit(path: str) -> dict:
    out: dict = {"path": path, "ok": True, "falhas": [], "avisos": []}

    wb = load_workbook(path, data_only=False)
    names = wb.sheetnames
    if NOME_RESUMO not in names:
        out["ok"] = False
        out["falhas"].append("Aba RESUMO GERAL ausente.")
        return out

    ws_r = wb[NOME_RESUMO]
    merges_r7 = [mr for mr in _merged_ranges(ws_r) if mr.min_row == mr.max_row == 7]
    out["resumo_merges_linha7"] = sorted(_range_text(m) for m in merges_r7)
    esperado_r7 = {"A7:C7", "D7:E7", "F7:G7", "H7:I7", "J7:M7"}
    if set(out["resumo_merges_linha7"]) != esperado_r7:
        out["ok"] = False
        out["falhas"].append(
            f"RESUMO linha7 merges esperado {sorted(esperado_r7)} obteve {out['resumo_merges_linha7']}"
        )

    for ref, fg_e in [
        ("A7", AZUL),
        ("D7", VERDE),
        ("F7", None),
        ("H7", AZUL_CLARO),
        ("J7", AMARELO),
        ("N7", BEGE),
    ]:
        rgb = _cell_fill_rgb(ws_r[ref])
        if fg_e is None:
            if rgb not in VERMELHO_ESPERADO and rgb != "":
                out["avisos"].append(f"RESUMO {ref} fill={rgb} fora da faixa vermelha esperada")
        elif rgb != fg_e:
            out["ok"] = False
            out["falhas"].append(f"RESUMO {ref} fill esperado {fg_e} obteve {rgb or '(vazio/theme)'}")

    emp_sheets = [n for n in names if n != NOME_RESUMO]
    out["abas_empreendimento"] = emp_sheets
    out["emp_audit"] = {}
    esp7 = {"A7:F7", "G7:H7", "I7:J7", "K7:Q7", "R7:S7", "T7:W7", "X7:AA7"}

    for sn in emp_sheets:
        ws = wb[sn]
        ad = {"nome": sn, "max_column": ws.max_column}
        if ws.max_column < column_index_from_string("AA"):
            out["ok"] = False
            out["falhas"].append(f"[{sn}] max_column {ws.max_column} < 27 (AA)")

        merges_7 = [mr for mr in _merged_ranges(ws) if mr.min_row == mr.max_row == 7]
        ad["merges_linha7"] = sorted(_range_text(m) for m in merges_7)
        if set(ad["merges_linha7"]) != esp7:
            out["ok"] = False
            out["falhas"].append(f"[{sn}] linha7 merges esperado {sorted(esp7)} obteve {ad['merges_linha7']}")

        af = ws.auto_filter
        ad["auto_filter_ref"] = str(af.ref) if af and af.ref else ""
        if not ad["auto_filter_ref"]:
            out["ok"] = False
            out["falhas"].append(f"[{sn}] autofiltro ausente")
        elif not ad["auto_filter_ref"].upper().startswith("A8"):
            out["avisos"].append(f"[{sn}] autofiltro ref={ad['auto_filter_ref']} (esperado iniciar em A8)")

        for ref, fg_e in [
            ("A7", AZUL),
            ("G7", "C5D9F1"),
            ("I7", VERDE),
            ("K7", None),
            ("R7", AZUL_CLARO),
            ("T7", AMARELO),
            ("X7", BEGE),
        ]:
            rgb = _cell_fill_rgb(ws[ref])
            if fg_e is None:
                if rgb not in VERMELHO_ESPERADO and rgb != "":
                    out["avisos"].append(f"[{sn}] {ref} fill={rgb} fora da faixa vermelha esperada")
            elif rgb != fg_e:
                out["ok"] = False
                out["falhas"].append(f"[{sn}] {ref} fill esperado {fg_e} obteve {rgb or '(vazio)'}")
        out["emp_audit"][sn] = ad

    wb.close()

    col_det = {
        "Vl.Carteira": "T",
        "Vl.Pago": "J",
        "Vl.Vencer": "S",
        "Vl.Principal (Encargos)": "Q",
        "Qtd.Parc.Atrasada": "K",
    }
    col_res = {
        "Vl.Pago": "E",
        "Vl.Vencer": "I",
        "Vl.Principal (Encargos)": "G",
        "Vl.Carteira": "J",
        "Qtd.Parc.Atrasada": "F",
    }
    metrics = list(col_det.keys())

    wb_vals = load_workbook(path, data_only=True, read_only=True)
    out["somas_abas_empreendimento"] = {
        metric: round(sum(_sum_numeric_column_ws(wb_vals[sn], col_det[metric]) for sn in emp_sheets), 2)
        for metric in metrics
    }
    out["somas_resumo_geral"] = {
        metric: round(_sum_numeric_column_ws(wb_vals[NOME_RESUMO], col_res[metric]), 2)
        for metric in metrics
    }
    out["delta_resumo_vs_soma_abas"] = {}
    tol = 0.02
    for metric in metrics:
        a = out["somas_abas_empreendimento"][metric]
        b = out["somas_resumo_geral"][metric]
        out["delta_resumo_vs_soma_abas"][metric] = round(a - b, 2)
        if abs(a - b) > tol:
            out["ok"] = False
            out["falhas"].append(f"Soma [{metric}] detalhe={a:.2f} vs resumo={b:.2f} diff={a-b:.2f}")
    out["somas_abas_empreendimento_rep"] = {
        metric: round(sum(_sum_numeric_column_ws(wb_vals[sn], col_det[metric]) for sn in emp_sheets), 2)
        for metric in metrics
    }
    out["delta_leitura_dupla"] = {
        metric: round(out["somas_abas_empreendimento"][metric] - out["somas_abas_empreendimento_rep"][metric], 6)
        for metric in metrics
    }
    wb_vals.close()

    corrupt_re = re.compile(r"BVGWH{2,}", re.I)
    out["emp_obra_corrompidos"] = []
    out["nao_informado_com_sigla"] = []
    for sn in emp_sheets:
        df = pd.read_excel(path, sheet_name=sn, header=7, engine="openpyxl")
        mc = _norm_cols_df(df)
        c_eo = mc.get(_fold("Emp/Obra")) or mc.get(_fold("EMP/OBRA"))
        c_emp = mc.get(_fold("Empreendimento")) or mc.get(_fold("EMPREENDIMENTO"))
        if not c_eo or not c_emp:
            continue
        for idx, val in df[c_eo].items():
            s = str(val).strip().upper()
            if corrupt_re.search(s) or "BVGWHHHH" in s or "BVGWHHH" in s:
                out["emp_obra_corrompidos"].append({"aba": sn, "linha": idx + 9, "valor": val})
        emp_col = df[c_emp].fillna("").astype(str).str.strip()
        eo_col = df[c_eo].fillna("").astype(str).str.strip()
        for idx, (eo, emp) in enumerate(zip(eo_col, emp_col)):
            if "NAO INFORMADO" not in _fold(emp):
                continue
            parts = eo.upper().split("/")
            sig = parts[-1].strip() if len(parts) >= 2 else ""
            sig = re.sub(r"[^A-Z0-9]", "", sig)
            if sig in SIGLAS_MAPA:
                out["nao_informado_com_sigla"].append({"aba": sn, "linha": idx + 9, "Emp/Obra": eo})

    if out["emp_obra_corrompidos"]:
        out["ok"] = False
        out["falhas"].append(f"EMP/OBRA corrompidos: {len(out['emp_obra_corrompidos'])} ocorrencias")
    if out["nao_informado_com_sigla"]:
        out["ok"] = False
        out["falhas"].append(
            f"NAO INFORMADO com sigla mapeavel: {len(out['nao_informado_com_sigla'])} linhas"
        )

    wb2 = load_workbook(path, data_only=True)
    hashes = []
    money_cols_letters = {"G", "J", "L", "M", "N", "O", "P", "Q", "S", "T"}
    for sn in emp_sheets:
        ws = wb2[sn]
        max_r = min(ws.max_row or 0, 5000)
        for row in range(9, max_r + 1):
            for col in range(1, min(ws.max_column or 0, 27) + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is None:
                    continue
                if isinstance(cell.value, str) and cell.value.strip() == "#######":
                    hashes.append({"aba": sn, "cell": cell.coordinate})
                    continue
                letter = get_column_letter(col)
                if letter in money_cols_letters and isinstance(cell.value, (int, float)) and cell.value != 0:
                    width = ws.column_dimensions[letter].width or 8.43
                    if abs(float(cell.value)) >= 1_000_000 and width < 10:
                        hashes.append({"aba": sn, "cell": cell.coordinate, "hint": "valor grande col estreita"})
    wb2.close()
    out["celulas_hash"] = hashes[:50]
    out["celulas_hash_total"] = len(hashes)
    if hashes:
        out["avisos"].append(f"Possivel exibicao truncada (#### ou col estreita): {len(hashes)} casos")

    wb3 = load_workbook(path, data_only=False)
    fmt_samples = {}
    for sn in emp_sheets[:1]:
        ws = wb3[sn]
        for letter in ("G", "J", "U", "C"):
            fmt_samples[f"{sn}!{letter}10"] = ws[f"{letter}10"].number_format or ""
    wb3.close()
    out["formato_amostra"] = fmt_samples

    return out


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx", nargs="?", default=str(Path("outputs") / "CARTEIRAS GERAL.xlsx"))
    args = parser.parse_args()
    print(json.dumps(audit(args.xlsx), ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
