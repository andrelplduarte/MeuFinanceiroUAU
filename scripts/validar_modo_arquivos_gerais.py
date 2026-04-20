# -*- coding: utf-8 -*-
"""Valida equivalencia entre modo por empreendimento e modo arquivos gerais."""
from __future__ import annotations

import json
import shutil
import sys
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

BASE = Path(__file__).resolve().parents[1]
if str(BASE) not in sys.path:
    sys.path.insert(0, str(BASE))

from services.orquestrador_lote_uau import processar_entrada_simples_ou_lote  # noqa: E402


def _listar_por_prefixo(pasta: Path, prefixo: str) -> list[str]:
    return sorted(
        str(p)
        for p in pasta.glob(f"{prefixo}_*.txt")
        if p.is_file() and "CONTAS_A_" not in p.name and "CONTAS_RECEBIDAS_" not in p.name
    )


def _sum_numeric_column_ws(ws, col_letter: str, start_row: int = 9) -> float:
    col_idx = column_index_from_string(col_letter)
    total = Decimal("0")
    for row in range(start_row, (ws.max_row or 0) + 1):
        val = ws.cell(row=row, column=col_idx).value
        if val in (None, ""):
            continue
        try:
            total += Decimal(str(val))
        except Exception:
            try:
                total += Decimal(str(float(val)))
            except Exception:
                continue
    return float(total)


def _somas_consolidado(path_xlsx: Path) -> dict:
    if not path_xlsx.is_file():
        return {}
    colunas = {
        "Vl.Carteira": "T",
        "Vl.Pago": "J",
        "Vl.Vencer": "S",
        "Vl.Principal (Encargos)": "Q",
        "Qtd.Parc.Atrasada": "K",
        "Qtd.Parc.A Vencer": "R",
    }
    wb = load_workbook(path_xlsx, data_only=True, read_only=True)
    abas = [x for x in wb.sheetnames if x != "RESUMO GERAL"]
    tot = {k: 0.0 for k in colunas}
    for aba in abas:
        ws = wb[aba]
        for nome, col_letter in colunas.items():
            tot[nome] += _sum_numeric_column_ws(ws, col_letter)
    wb.close()
    return {k: round(v, 2) for k, v in tot.items()}


def _materializar_saida(saida, destino_copia: Path) -> Path:
    caminho = Path(saida[0] if isinstance(saida, tuple) else saida)
    destino_copia.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(caminho, destino_copia)
    return destino_copia


def main() -> int:
    uploads = BASE / "uploads"
    rec = _listar_por_prefixo(uploads, "rec")
    reb = _listar_por_prefixo(uploads, "reb")
    if not rec or not reb:
        print(
            json.dumps(
                {
                    "ok": False,
                    "motivo": "Arquivos rec_*.txt / reb_*.txt nao encontrados em uploads/ para validacao comparativa."
                },
                ensure_ascii=False,
            )
        )
        return 2

    tmp_path = BASE / "outputs" / "_validar_modo_arquivos_gerais"
    if tmp_path.exists():
        shutil.rmtree(tmp_path)
    tmp_path.mkdir(parents=True, exist_ok=True)

    geral_rec = tmp_path / "GERAL_RECEBER.txt"
    geral_reb = tmp_path / "GERAL_RECEBIDOS.txt"
    geral_rec.write_text("\n".join(Path(p).read_text(encoding="utf-8", errors="ignore") for p in rec), encoding="utf-8")
    geral_reb.write_text("\n".join(Path(p).read_text(encoding="utf-8", errors="ignore") for p in reb), encoding="utf-8")

    out_sep = tmp_path / "out_sep"
    out_ger = tmp_path / "out_ger"
    out_sep.mkdir(parents=True, exist_ok=True)
    out_ger.mkdir(parents=True, exist_ok=True)

    saida_sep, tempo_sep = processar_entrada_simples_ou_lote(
        rec, reb, str(out_sep), "POR_EMPREENDIMENTO", caminhos_estoque=None
    )
    x_sep = _materializar_saida(saida_sep, tmp_path / "CARTEIRAS GERAL - SEP.xlsx")

    saida_ger, tempo_ger = processar_entrada_simples_ou_lote(
        [str(geral_rec)], [str(geral_reb)], str(out_ger), "ARQUIVOS_GERAIS", caminhos_estoque=None
    )
    x_ger = _materializar_saida(saida_ger, tmp_path / "CARTEIRAS GERAL - GERAIS.xlsx")

    s_sep = _somas_consolidado(x_sep)
    s_ger = _somas_consolidado(x_ger)
    diff = {k: round(float(s_ger.get(k, 0)) - float(s_sep.get(k, 0)), 2) for k in sorted(set(s_sep) | set(s_ger))}

    print(
        json.dumps(
            {
                "ok": True,
                "tempo_sep_s": round(float(tempo_sep), 2),
                "tempo_gerais_s": round(float(tempo_ger), 2),
                "somas_por_empreendimento": s_sep,
                "somas_arquivos_gerais": s_ger,
                "diff_gerais_menos_sep": diff,
                "arquivo_sep": str(x_sep),
                "arquivo_gerais": str(x_ger),
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
