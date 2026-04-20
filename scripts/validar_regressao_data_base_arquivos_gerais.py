#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Blindagem de regressão para DATA-BASE no modo ARQUIVOS_GERAIS.

Etapa A (split):
- Garante que os blocos temporários por empreendimento preservam a DATA-BASE da origem.
- Garante que o cabeçalho NÃO cai no dt_max das linhas quando dt_max difere da origem.

Etapa B (resultado final):
- Gera workbook final em modo ARQUIVOS_GERAIS com subset reproduzível.
- Confirma DATA-BASE nas abas-alvo.
- Confirma coerência RESUMO GERAL vs soma das abas para A VENCER.
"""
from __future__ import annotations

import json
import os
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parents[1]
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

from services.processador_uau import extrair_data_base, ler_texto_robusto  # noqa: E402
from services.orquestrador_lote_uau import (  # noqa: E402
    _escrever_bloco_temp_txt,
    _split_arquivo_geral_receber_por_emp,
    _split_arquivo_geral_recebidos_por_emp,
    processar_entrada_simples_ou_lote,
)


ALVOS = ("LTMON", "BVGWH", "LTVIL", "VROLT", "SCPGO")


def _max_data_linhas(linhas) -> datetime | None:
    dt_max = None
    for ln in linhas:
        for m in re.finditer(r"\b(\d{2})/(\d{2})/(\d{4})\b", str(ln or "")):
            try:
                d = datetime(int(m.group(3)), int(m.group(2)), int(m.group(1)))
            except Exception:
                continue
            if dt_max is None or d > dt_max:
                dt_max = d
    return dt_max


def _escrever_geral_sintetico(path: Path, data_base: datetime, linhas_agregadas, tipo: str) -> None:
    meses_pt = {
        1: "janeiro",
        2: "fevereiro",
        3: "março",
        4: "abril",
        5: "maio",
        6: "junho",
        7: "julho",
        8: "agosto",
        9: "setembro",
        10: "outubro",
        11: "novembro",
        12: "dezembro",
    }
    data_ext = f"{data_base.day} de {meses_pt[data_base.month]} de {data_base.year}"
    cab_tipo = "Contas a Receber" if tipo == "RECEBER" else "Contas Recebidas"
    cab_periodo = "Período por Vencimento" if tipo == "RECEBER" else "Período por Recebimento"
    path.write_text(
        data_ext + "\n" + cab_tipo + "\n" + cab_periodo + "\n" + "\n".join(linhas_agregadas) + "\n",
        encoding="utf-8",
    )


def _map_headers(ws):
    out = {}
    for c in ws[8]:
        if c.value is None:
            continue
        out[str(c.value).strip().upper()] = c.column
    return out


def _headers_alias(ws, *aliases):
    headers = _map_headers(ws)
    for alias in aliases:
        col = headers.get(str(alias).strip().upper())
        if col:
            return col
    return None


def _sum_col(ws, col_idx):
    if not col_idx:
        return 0.0
    s = 0.0
    for r in range(9, ws.max_row + 1):
        v = ws.cell(r, col_idx).value
        if isinstance(v, (int, float)):
            s += float(v)
    return round(s, 2)


def main() -> int:
    uploads = BASE_DIR / "uploads"
    rec_geral = uploads / "rec_00_CONTAS_A_RECEBER_14.04.2026.txt"
    reb_geral = uploads / "reb_00_CONTAS_RECEBIDAS_14.04.2026.txt"
    if not rec_geral.is_file() or not reb_geral.is_file():
        print(
            json.dumps(
                {
                    "ok": False,
                    "erro": "Arquivos gerais de referência não encontrados em uploads/",
                    "esperados": [str(rec_geral), str(reb_geral)],
                },
                ensure_ascii=False,
                indent=2,
            )
        )
        return 2

    db_origem = extrair_data_base(str(rec_geral)) or extrair_data_base(str(reb_geral))
    if db_origem is None:
        print(json.dumps({"ok": False, "erro": "Não foi possível extrair DATA-BASE da origem."}, ensure_ascii=False, indent=2))
        return 2

    grupos_r = _split_arquivo_geral_receber_por_emp(str(rec_geral))
    grupos_p = _split_arquivo_geral_recebidos_por_emp(str(reb_geral))
    inter = sorted(set(grupos_r) & set(grupos_p))
    alvos_presentes = [k for k in inter if any(sig in k.upper() for sig in ALVOS)]
    if len(alvos_presentes) < 3:
        print(
            json.dumps(
                {
                    "ok": False,
                    "erro": "Poucos alvos encontrados no split para validação confiável.",
                    "alvos_presentes": alvos_presentes,
                },
                ensure_ascii=False,
                indent=2,
            )
        )
        return 2

    report = {
        "ok": True,
        "origem": {
            "rec_geral": str(rec_geral),
            "reb_geral": str(reb_geral),
            "data_base": db_origem.strftime("%d/%m/%Y"),
        },
        "etapa_a_split": [],
        "etapa_b_resultado_final": {},
        "falhas": [],
    }

    # Etapa A: valida preservação da data-base nos temporários por empreendimento.
    td_path = BASE_DIR / "outputs" / "_tests_data_base_gerais" / "work"
    if td_path.exists():
        shutil.rmtree(td_path)
    td_path.mkdir(parents=True, exist_ok=True)
    for k in alvos_presentes:
        l_r = grupos_r[k]
        l_p = grupos_p[k]
        tmp_r = _escrever_bloco_temp_txt(
            prefixo=f"teste_rec_{k}",
            linhas=l_r,
            tipo="RECEBER",
            data_base_cabecalho=db_origem,
        )
        tmp_p = _escrever_bloco_temp_txt(
            prefixo=f"teste_reb_{k}",
            linhas=l_p,
            tipo="RECEBIDOS",
            data_base_cabecalho=db_origem,
        )
        try:
            db_tmp_r = extrair_data_base(tmp_r, texto_pre_lido=ler_texto_robusto(tmp_r))
            db_tmp_p = extrair_data_base(tmp_p, texto_pre_lido=ler_texto_robusto(tmp_p))
            dt_max_r = _max_data_linhas(l_r)
            dt_max_p = _max_data_linhas(l_p)
            item = {
                "chave": k,
                "data_base_tmp_receber": db_tmp_r.strftime("%d/%m/%Y") if db_tmp_r else None,
                "data_base_tmp_recebidos": db_tmp_p.strftime("%d/%m/%Y") if db_tmp_p else None,
                "dt_max_linhas_receber": dt_max_r.strftime("%d/%m/%Y") if dt_max_r else None,
                "dt_max_linhas_recebidos": dt_max_p.strftime("%d/%m/%Y") if dt_max_p else None,
            }
            report["etapa_a_split"].append(item)
            if db_tmp_r != db_origem or db_tmp_p != db_origem:
                report["falhas"].append(f"DATA-BASE temporária divergente da origem na chave {k}.")
            if dt_max_r and dt_max_r != db_origem and db_tmp_r == dt_max_r:
                report["falhas"].append(f"Fallback dt_max detectado no receber para {k}.")
            if dt_max_p and dt_max_p != db_origem and db_tmp_p == dt_max_p:
                report["falhas"].append(f"Fallback dt_max detectado no recebidos para {k}.")
        finally:
            try:
                os.unlink(tmp_r)
            except OSError:
                pass
            try:
                os.unlink(tmp_p)
            except OSError:
                pass

        # Etapa B: cenário reduzido reproduzível com os alvos.
        linhas_rec = []
        linhas_reb = []
        for k in alvos_presentes:
            linhas_rec.extend(grupos_r[k])
            linhas_reb.extend(grupos_p[k])
        rec_subset = td_path / "REC_GERAL_SUBSET.txt"
        reb_subset = td_path / "REB_GERAL_SUBSET.txt"
        _escrever_geral_sintetico(rec_subset, db_origem, linhas_rec, "RECEBER")
        _escrever_geral_sintetico(reb_subset, db_origem, linhas_reb, "RECEBIDOS")

        out_dir = BASE_DIR / "outputs" / "_tests_data_base_gerais" / "resultado"
        out_dir.mkdir(parents=True, exist_ok=True)
        saida, tempo = processar_entrada_simples_ou_lote(
            [str(rec_subset)],
            [str(reb_subset)],
            str((out_dir / "base_placeholder.xlsx").resolve()),
            "ARQUIVOS_GERAIS",
            caminhos_estoque=None,
        )
        xlsx = Path(saida[0] if isinstance(saida, tuple) else saida)
        wb = load_workbook(xlsx, data_only=True)
        try:
            dados_alvos = {}
            for sigla in ALVOS:
                ws = None
                for n in wb.sheetnames:
                    up = n.upper()
                    if sigla in up:
                        ws = wb[n]
                        break
                if ws is None:
                    continue
                headers = _map_headers(ws)
                dados_alvos[sigla] = {
                    "aba": ws.title,
                    "data_base": ws["B2"].value,
                    "qtd_parc_a_vencer": _sum_col(ws, _headers_alias(ws, "QTD.PARC.A VENCER", "QTD PARC. A VENCER")),
                    "vl_a_vencer": _sum_col(ws, _headers_alias(ws, "VL.A VENCER", "VL.VENCER", "VALOR A VENCER")),
                }
                if str(ws["B2"].value or "").strip() != db_origem.strftime("%d/%m/%Y"):
                    report["falhas"].append(f"DATA-BASE final divergente na aba de {sigla}.")

            if "RESUMO GERAL" in wb.sheetnames:
                ws_res = wb["RESUMO GERAL"]
                resumo_qtd = _sum_col(ws_res, _headers_alias(ws_res, "QTD PARC. A VENCER", "QTD.PARC.A VENCER"))
                resumo_vl = _sum_col(ws_res, _headers_alias(ws_res, "VL.A VENCER", "VALOR A VENCER", "VL.VENCER"))
            else:
                resumo_qtd = 0.0
                resumo_vl = 0.0
                report["falhas"].append("Aba RESUMO GERAL ausente no resultado final.")

            soma_abas_qtd = 0.0
            soma_abas_vl = 0.0
            for n in wb.sheetnames:
                if n == "RESUMO GERAL":
                    continue
                ws = wb[n]
                soma_abas_qtd += _sum_col(ws, _headers_alias(ws, "QTD.PARC.A VENCER", "QTD PARC. A VENCER"))
                soma_abas_vl += _sum_col(ws, _headers_alias(ws, "VL.A VENCER", "VL.VENCER", "VALOR A VENCER"))
            soma_abas_qtd = round(soma_abas_qtd, 2)
            soma_abas_vl = round(soma_abas_vl, 2)

            report["etapa_b_resultado_final"] = {
                "arquivo": str(xlsx),
                "tempo_s": round(float(tempo), 2),
                "alvos": dados_alvos,
                "resumo_vs_abas": {
                    "resumo_qtd_parc_a_vencer": resumo_qtd,
                    "abas_qtd_parc_a_vencer": soma_abas_qtd,
                    "delta_qtd": round(resumo_qtd - soma_abas_qtd, 2),
                    "resumo_vl_a_vencer": resumo_vl,
                    "abas_vl_a_vencer": soma_abas_vl,
                    "delta_vl": round(resumo_vl - soma_abas_vl, 2),
                },
            }
            if round(resumo_qtd - soma_abas_qtd, 2) != 0.0:
                report["falhas"].append("Divergência RESUMO GERAL x abas em QTD PARC. A VENCER.")
            if round(resumo_vl - soma_abas_vl, 2) != 0.0:
                report["falhas"].append("Divergência RESUMO GERAL x abas em VL.A VENCER.")
        finally:
            wb.close()

    report["ok"] = len(report["falhas"]) == 0
    out_json = BASE_DIR / "outputs" / "_tests_data_base_gerais" / "relatorio_validacao_data_base_gerais.json"
    out_json.parent.mkdir(parents=True, exist_ok=True)
    out_json.write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2, default=str))
    print(f"\nRelatório salvo em: {out_json}")
    return 0 if report["ok"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
