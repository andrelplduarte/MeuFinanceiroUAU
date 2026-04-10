import json
import re

from openpyxl import load_workbook


def main():
    p = r"c:\Users\USER\Desktop\MeuFinanceiroUAU\outputs\CARTEIRAS GERAL.xlsx"
    wb = load_workbook(p, data_only=False)
    res = wb["RESUMO GERAL"]

    res_rows = []
    for r in range(9, res.max_row + 1):
        eo = res.cell(r, 1).value
        emp = res.cell(r, 2).value
        se = str(eo or "").strip()
        sm = str(emp or "").strip()
        if not se and not sm:
            continue
        res_rows.append({"linha": r, "emp_obra": se, "empreendimento": sm})

    bad = []
    pat = re.compile(r"BVGWH{2,}|/BVGW($|[^H])", re.I)
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=9, max_row=min(ws.max_row, 6000), min_col=1, max_col=min(ws.max_column, 27)):
            for c in row:
                v = c.value
                if isinstance(v, str) and "BVGW" in v.upper():
                    if pat.search(v.upper()) or "BVGWHHH" in v.upper():
                        bad.append({"aba": ws.title, "celula": c.coordinate, "valor": v})
                        if len(bad) >= 80:
                            break
            if len(bad) >= 80:
                break
        if len(bad) >= 80:
            break

    emp_sheets = [n for n in wb.sheetnames if n != "RESUMO GERAL"]
    widths = {}
    for n in emp_sheets:
        ws = wb[n]
        widths[n] = {
            "D_CLIENTE": ws.column_dimensions["D"].width,
            "F_IDENTIFICADOR": ws.column_dimensions["F"].width,
            "A": ws.column_dimensions["A"].width,
            "AA": ws.column_dimensions["AA"].width,
        }

    title_resumo = {
        "merge_C1_L6": any(str(m) == "C1:L6" for m in res.merged_cells.ranges),
        "A1": res["A1"].value,
        "B1": res["B1"].value,
        "C1": res["C1"].value,
        "autofilter": res.auto_filter.ref,
    }

    mapa = {
        "NVLOT": "RES.NILSON VELOSO",
        "LTMAG": "RES.MAGALHAES",
        "SCPTO": "LOT.TOCANTINS",
        "SCPTI": "LOT.TIRADENTES",
        "CIDAN": "RES.CIDADE NOVA",
        "VROLT": "LOT.VALE DAS ROSAS",
        "ALVLT": "RES.ALVORADA",
        "LTMON": "LOT.MONTE NEGRO",
        "RVERD": "RIO VERDE",
        "LTVIL": "LOT.VILA NOVA",
        "SCPGO": "RES.GOIANIA",
        "ARAHF": "RES.ARARAS",
        "BVGWH": "COND.BELLA WHITE",
        "MANHA": "MANHATAN",
        "MONTB": "MONTBLANC",
        "LIFE": "LIFE",
    }
    titulos_emp = []
    for n in emp_sheets:
        ws = wb[n]
        eo = str(ws["A9"].value or "").strip().upper()
        sig = eo.split("/")[-1] if "/" in eo else ""
        titulos_emp.append(
            {
                "aba": n,
                "A9": eo,
                "B1": ws["B1"].value,
                "C1": ws["C1"].value,
                "merge_C1_U6": any(str(m) == "C1:U6" for m in ws.merged_cells.ranges),
                "esperado_oficial": mapa.get(sig, ""),
                "autofilter": ws.auto_filter.ref,
            }
        )

    ws0 = wb[emp_sheets[0]]
    estrutura = {
        "merges_linha7_primeira_aba": [str(m) for m in ws0.merged_cells.ranges if m.min_row == 7 and m.max_row == 7],
        "primeira_aba": emp_sheets[0],
    }

    wb.close()

    out = {
        "arquivo": p,
        "qt_abas": len(wb.sheetnames),
        "abas": wb.sheetnames,
        "resumo_emp_obra_empreendimento": res_rows,
        "variantes_bvgw_encontradas": bad,
        "titulos_resumo": title_resumo,
        "titulos_abas_empreendimento": titulos_emp,
        "larguras_fixas_cliente_identificador": widths,
        "estrutura": estrutura,
    }
    op = r"c:\Users\USER\Desktop\MeuFinanceiroUAU\outputs\_evidencia_final_layout.json"
    with open(op, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    print(op)


if __name__ == "__main__":
    main()

