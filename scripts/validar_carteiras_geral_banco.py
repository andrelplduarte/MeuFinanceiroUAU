#!/usr/bin/env python
# -*- coding: utf-8 -*-
from __future__ import annotations

import hashlib
import os
import sys
from typing import List, Tuple

import pandas as pd
from openpyxl import load_workbook


BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, BASE_DIR)
from services.orquestrador_lote_uau import aba_e_consolidado_carteiras_geral  # noqa: E402

OUT_DIR = os.path.join(BASE_DIR, "outputs")
PATH_GERAL = os.path.join(OUT_DIR, "CARTEIRAS GERAL.xlsx")
PATH_BASE = os.path.join(OUT_DIR, "CARTEIRAS BANCO DE DADOS.xlsx")
PATH_PROCESSADOR = os.path.join(BASE_DIR, "services", "processador_uau.py")

# Base opcional (SQL-like): exatamente estas duas abas, nesta ordem, cabeçalho na linha 1.
ABAS_BASE_ESPERADAS = ["DADOS_RECEBER", "DADOS_RECEBIDOS"]


def _sha256_arquivo(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _sigla_da_aba(nome_aba: str) -> str:
    nome = str(nome_aba or "").strip()
    for sep in (" – ", " - "):
        if sep in nome:
            return nome.split(sep, 1)[0].strip()
    return nome


def _eh_aba_consolidado(nome_aba: str) -> bool:
    return aba_e_consolidado_carteiras_geral(nome_aba)


def _ler_df_com_header(path: str, aba: str, header_0based: int) -> pd.DataFrame:
    try:
        return pd.read_excel(path, sheet_name=aba, header=header_0based)
    except Exception:
        return pd.DataFrame()


def _coluna_inadimplencia(df: pd.DataFrame) -> str | None:
    if df is None or df.empty:
        return None
    for c in df.columns:
        cu = str(c).upper()
        if "VL.PRINCIPAL" in cu and "ENCARG" in cu:
            return c
    for c in df.columns:
        cu = str(c).upper()
        if "PRINCIPAL" in cu and "ENCARG" in cu:
            return c
    return None


def validar_guard_rail() -> List[str]:
    erros: List[str] = []
    if not os.path.isfile(PATH_PROCESSADOR):
        erros.append(f"Arquivo do motor não encontrado: {PATH_PROCESSADOR}")
        return erros
    sha_atual = _sha256_arquivo(PATH_PROCESSADOR)
    sha_base = os.environ.get("PROCESSADOR_SHA256_BASE", "").strip().lower()
    print("\n[GUARD RAIL]")
    print(f"SHA256 atual processador_uau.py: {sha_atual}")
    if sha_base:
        print(f"SHA256 base informado:          {sha_base}")
        if sha_atual != sha_base:
            erros.append("ALERTA: processador_uau.py alterado fora de escopo.")
    else:
        print("PROCESSADOR_SHA256_BASE não informado; hash atual exibido como referência.")
    return erros


def validar_carteiras_geral() -> Tuple[List[str], List[Tuple[str, float]]]:
    erros: List[str] = []
    inad_por_aba: List[Tuple[str, float]] = []

    print("\n[VALIDAÇÃO] CARTEIRAS GERAL.xlsx")
    if not os.path.isfile(PATH_GERAL):
        return [f"Arquivo não encontrado: {PATH_GERAL}"], inad_por_aba

    wb = load_workbook(PATH_GERAL, read_only=True, data_only=True)
    abas = list(wb.sheetnames)
    wb.close()
    print(f"Abas: {abas}")

    if not abas:
        erros.append("CARTEIRAS GERAL.xlsx sem abas.")
        return erros, inad_por_aba

    if abas[0] != "RESUMO GERAL":
        erros.append("Primeira aba de CARTEIRAS GERAL.xlsx deve ser RESUMO GERAL.")

    abas_consolidadas = abas[1:]
    for aba in abas_consolidadas:
        if not _eh_aba_consolidado(aba):
            erros.append(f"Aba inválida no arquivo geral (não consolidado): {aba}")

    for aba in abas_consolidadas:
        df = _ler_df_com_header(PATH_GERAL, aba, 7)
        col_inad = _coluna_inadimplencia(df)
        if not col_inad:
            erros.append(
                f"Coluna de inadimplência não encontrada na aba {aba}."
            )
            inad = 0.0
        else:
            inad = float(
                pd.to_numeric(df[col_inad], errors="coerce")
                .fillna(0.0)
                .sum()
            )
        inad_por_aba.append((aba, inad))

    esperado = sorted(
        inad_por_aba,
        key=lambda x: (-x[1], _sigla_da_aba(x[0]).upper()),
    )
    if [a for a, _ in esperado] != [a for a, _ in inad_por_aba]:
        erros.append("Ordem das abas consolidadas não está em inadimplência desc com desempate por sigla.")

    print("\nTabela-resumo (ABA | INAD_TOTAL):")
    for aba, inad in inad_por_aba:
        print(f"- {aba} | {inad:.2f}")

    return erros, inad_por_aba


def validar_carteiras_banco() -> Tuple[List[str], List[Tuple[str, int]]]:
    erros: List[str] = []
    contagens: List[Tuple[str, int]] = []

    print("\n[VALIDAÇÃO] CARTEIRAS BANCO DE DADOS.xlsx")
    if not os.path.isfile(PATH_BASE):
        return [f"Arquivo não encontrado: {PATH_BASE}"], contagens

    wb = load_workbook(PATH_BASE, read_only=True, data_only=True)
    abas = list(wb.sheetnames)
    wb.close()
    print(f"Abas: {abas}")

    if abas != ABAS_BASE_ESPERADAS:
        erros.append(
            "CARTEIRAS BANCO DE DADOS.xlsx deve conter exatamente duas abas, nesta ordem: "
            "DADOS_RECEBER, DADOS_RECEBIDOS. "
            f"Encontrado: {abas}."
        )

    for nome_aba in ABAS_BASE_ESPERADAS:
        if nome_aba not in abas:
            contagens.append((nome_aba, -1))
            continue
        df = _ler_df_com_header(PATH_BASE, nome_aba, header_0based=0)
        contagens.append((nome_aba, len(df)))

    print("\nLinhas por aba (base, dados após cabeçalho):")
    for aba, n in contagens:
        print(f"- {aba}: {n}")

    return erros, contagens


def main() -> None:
    erros: List[str] = []

    erros.extend(validar_guard_rail())
    erros_geral, _ = validar_carteiras_geral()
    erros.extend(erros_geral)
    erros_base, _ = validar_carteiras_banco()
    erros.extend(erros_base)

    print("\n[RESULTADO FINAL]")
    if erros:
        for err in erros:
            print(f"- FALHA: {err}")
        sys.exit(1)
    print("Tudo OK.")
    sys.exit(0)


if __name__ == "__main__":
    main()
