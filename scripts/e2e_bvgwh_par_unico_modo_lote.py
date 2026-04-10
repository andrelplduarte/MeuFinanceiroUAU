# -*- coding: utf-8 -*-
"""E2E: um par BVGWH com modo POR_EMPREENDIMENTO (caminho de lote + estilo), rápido."""
from __future__ import annotations

import json
import re
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

import pandas as pd

RAIZ = Path(__file__).resolve().parents[1]
if str(RAIZ) not in sys.path:
    sys.path.insert(0, str(RAIZ))

from services.orquestrador_lote_uau import processar_entrada_simples_ou_lote  # noqa: E402


def _fold(s: str) -> str:
    s = unicodedata.normalize("NFD", str(s or ""))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s.strip().upper())


def _norm_cols(df):
    return {_fold(c): c for c in df.columns}


def main() -> int:
    uploads = RAIZ / "uploads"
    rec = str(uploads / "rec_08_BVGWH_-_EMP.BELLA_WHITE_-_RECEBER.txt")
    reb = str(uploads / "reb_11_BVGWH_-_EMP.BELLA_WHITE_-_RECEBIDOS.txt")
    if not Path(rec).is_file() or not Path(reb).is_file():
        print(json.dumps({"erro": "TXT BVGWH ausente"}, ensure_ascii=False))
        return 2

    saida = RAIZ / "outputs" / f"_e2e_bvgwh_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    saida.mkdir(parents=True, exist_ok=True)
    # dirname(caminho) deve ser a pasta de saída (mesmo padrão do app: *.xlsx dentro de outputs/).
    placeholder = saida / "consolidacao_uau.xlsx"

    out, tempo = processar_entrada_simples_ou_lote(
        [rec], [reb], str(placeholder), "POR_EMPREENDIMENTO", caminhos_estoque=None
    )
    path_novo = Path(out[0] if isinstance(out, tuple) else out)

    from openpyxl import load_workbook

    wb = load_workbook(path_novo, data_only=True)
    abas = wb.sheetnames
    ws_r = wb["RESUMO GERAL"]
    amostra_resumo = []
    for row in range(9, 12):
        amostra_resumo.append(
            {
                "linha": row,
                "A": ws_r.cell(row=row, column=1).value,
                "B": ws_r.cell(row=row, column=2).value,
                "C": ws_r.cell(row=row, column=3).value,
            }
        )
    wb.close()

    alvo = next((n for n in abas if "BVGWH" in n.upper()), None)
    bvg = {"aba": alvo, "linhas": [], "bvgwhhhhh_count": 0}
    if alvo:
        df = pd.read_excel(path_novo, sheet_name=alvo, header=7, engine="openpyxl")
        mc = _norm_cols(df)
        c_eo = mc.get(_fold("EMP/OBRA"))
        c_emp = mc.get(_fold("EMPREENDIMENTO"))
        c_vl = mc.get(_fold("VL.CARTEIRA"))
        c_pago = mc.get(_fold("VL.PAGO"))
        if c_eo:
            bvg["bvgwhhhhh_count"] = int(
                df[c_eo].fillna("").astype(str).str.upper().str.contains("BVGWHHHH", regex=False).sum()
            )
            bvg["emp_obra_unicos"] = (
                df[c_eo].fillna("").astype(str).str.strip().str.upper().unique().tolist()[:15]
            )
        if c_eo and c_emp:
            for i in range(min(8, len(df))):
                r = df.iloc[i]
                bvg["linhas"].append(
                    {
                        "EMP/OBRA": str(r[c_eo]),
                        "EMPREENDIMENTO": str(r[c_emp]),
                        "VL.CARTEIRA": float(r[c_vl]) if c_vl and pd.notna(r.get(c_vl)) else None,
                        "VL.PAGO": float(r[c_pago]) if c_pago and pd.notna(r.get(c_pago)) else None,
                    }
                )
            bvg["total_linhas"] = len(df)

    wb2 = load_workbook(path_novo, data_only=True)
    hashes = []
    for sn in abas:
        if sn == "RESUMO GERAL":
            continue
        ws = wb2[sn]
        for row in range(9, min(ws.max_row or 0, 8000) + 1):
            for col in range(1, min(ws.max_column or 0, 30) + 1):
                v = ws.cell(row=row, column=col).value
                if isinstance(v, str) and "####" in v:
                    hashes.append(f"{sn}!{ws.cell(row=row, column=col).coordinate}")
    wb2.close()

    # Somatórios neste arquivo (1 aba empreendimento + resumo)
    depois = {}
    xl = pd.ExcelFile(path_novo, engine="openpyxl")
    emp_names = [x for x in xl.sheet_names if x != "RESUMO GERAL"]
    for label, colname in [
        ("Vl.Carteira", "VL.CARTEIRA"),
        ("Vl.Pago", "VL.PAGO"),
        ("Vl.Vencer", "VL.VENCER"),
        ("Vl.Principal (Encargos)", "VL.PRINCIPAL (ENCARGOS)"),
        ("Qtd.Parc.Atrasada", "QTD.PARC.ATRASADA"),
    ]:
        cf = _fold(colname)
        t = 0.0
        for sn in emp_names:
            df = pd.read_excel(path_novo, sheet_name=sn, header=7, engine="openpyxl")
            m = _norm_cols(df)
            if cf in m:
                t += float(pd.to_numeric(df[m[cf]], errors="coerce").fillna(0).sum())
        depois[label] = t

    df_res = pd.read_excel(path_novo, sheet_name="RESUMO GERAL", header=7, engine="openpyxl")
    mr = _norm_cols(df_res)
    resumo_cols = {
        "Vl.Carteira": _fold("VL.CARTEIRA"),
        "Vl.Pago": _fold("VALOR PAGO"),
        "Vl.Vencer": _fold("VALOR A VENCER"),
        "Vl.Principal (Encargos)": _fold("VALOR INADIMPLÊNCIA"),
        "Qtd.Parc.Atrasada": _fold("QTD PARC. INADIMPLÊNCIA"),
    }
    delta = {}
    for k, rc in resumo_cols.items():
        if rc in mr:
            sr = float(pd.to_numeric(df_res[mr[rc]], errors="coerce").fillna(0).sum())
            delta[k] = round(depois.get(k, 0) - sr, 4)

    rel = {
        "nota": "Par único BVGWH com modo POR_EMPREENDIMENTO (mesmo pipeline de estilo/anexo do lote).",
        "path_gerado": str(path_novo),
        "tempo_s": round(tempo, 2),
        "abas": abas,
        "amostra_resumo_geral_linhas_9_11": amostra_resumo,
        "aba_bvgwh": bvg,
        "celulas_com_cerquilha": hashes,
        "somas_aba_empreendimento": depois,
        "delta_detalhe_vs_resumo": delta,
    }
    out_j = RAIZ / "outputs" / "_e2e_evidencia_bvgwh.json"
    out_j.write_text(json.dumps(rel, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print(json.dumps(rel, ensure_ascii=False, indent=2, default=str))
    return 0


if __name__ == "__main__":
    sys.exit(main())
