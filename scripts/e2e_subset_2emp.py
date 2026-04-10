# -*- coding: utf-8 -*-
"""E2E rápido: 2 emparelhamentos (BVGWH + ALVLT) para evidência objetiva sem lote completo."""
from __future__ import annotations

import json
import re
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

import pandas as pd

RAIZ = Path(__file__).resolve().parents[1]
if str(RAIZ) not in sys.path:
    sys.path.insert(0, str(RAIZ))

from services.orquestrador_lote_uau import processar_entrada_simples_ou_lote  # noqa: E402


def _fold(s: str) -> str:
    s = unicodedata.normalize("NFD", str(s or ""))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s.strip().upper())


def _norm_cols(df):
    return {_fold(c): c for c in df.columns}


def somas(path: str) -> dict:
    p = Path(path)
    xl = pd.ExcelFile(p, engine="openpyxl")
    names = [x for x in xl.sheet_names if x != "RESUMO GERAL"]
    keys = {
        "Vl.Carteira": _fold("VL.CARTEIRA"),
        "Vl.Pago": _fold("VL.PAGO"),
        "Vl.Vencer": _fold("VL.VENCER"),
        "Vl.Principal (Encargos)": _fold("VL.PRINCIPAL (ENCARGOS)"),
        "Qtd.Parc.Atrasada": _fold("QTD.PARC.ATRASADA"),
    }
    out = {}
    for label, col_f in keys.items():
        t = 0.0
        for sn in names:
            df = pd.read_excel(p, sheet_name=sn, header=7, engine="openpyxl")
            m = _norm_cols(df)
            if col_f not in m:
                continue
            t += float(pd.to_numeric(df[m[col_f]], errors="coerce").fillna(0).sum())
        out[label] = t
    return out


def main() -> int:
    uploads = RAIZ / "uploads"
    rec = sorted(
        [
            str(uploads / "rec_08_BVGWH_-_EMP.BELLA_WHITE_-_RECEBER.txt"),
            str(uploads / "rec_02_ALVLT_-_LOT.SPE_RESIDENCIAL_OURILANDIA_-_RECEBER.txt"),
        ]
    )
    reb = sorted(
        [
            str(uploads / "reb_11_BVGWH_-_EMP.BELLA_WHITE_-_RECEBIDOS.txt"),
            str(uploads / "reb_05_ALVLT_-_LOT.SPE_RESIDENCIAL_OURILANDIA_-_RECEBIDOS.txt"),
        ]
    )
    for p in rec + reb:
        if not Path(p).is_file():
            print(json.dumps({"erro": f"arquivo ausente: {p}"}, ensure_ascii=False))
            return 2

    saida = RAIZ / "outputs" / f"_e2e_subset_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    saida.mkdir(parents=True, exist_ok=True)
    path_antigo = RAIZ / "outputs" / "CARTEIRAS GERAL.xlsx"
    antes = somas(str(path_antigo)) if path_antigo.is_file() else {}

    out, tempo = processar_entrada_simples_ou_lote(
        rec, reb, str(saida), "POR_EMPREENDIMENTO", caminhos_estoque=None
    )
    path_novo = Path(out[0] if isinstance(out, tuple) else out)
    depois = somas(str(path_novo))

    from openpyxl import load_workbook

    wb = load_workbook(path_novo, data_only=True)
    abas = wb.sheetnames
    ws_r = wb["RESUMO GERAL"]
    amostra_resumo = []
    for row in range(9, 12):
        amostra_resumo.append(
            {
                "linha": row,
                "A": ws_r.cell(row=row, column=1).value,
                "B": ws_r.cell(row=row, column=2).value,
                "C": ws_r.cell(row=row, column=3).value,
            }
        )
    wb.close()

    alvo = next((n for n in abas if "BVGWH" in n.upper()), None)
    bvg = {"aba": alvo, "linhas": [], "bvgwhhhhh_count": 0}
    if alvo:
        df = pd.read_excel(path_novo, sheet_name=alvo, header=7, engine="openpyxl")
        mc = _norm_cols(df)
        c_eo = mc.get(_fold("EMP/OBRA"))
        c_emp = mc.get(_fold("EMPREENDIMENTO"))
        c_vl = mc.get(_fold("VL.CARTEIRA"))
        c_pago = mc.get(_fold("VL.PAGO"))
        if c_eo:
            bvg["bvgwhhhhh_count"] = int(
                df[c_eo].fillna("").astype(str).str.upper().str.contains("BVGWHHHH", regex=False).sum()
            )
        if c_eo and c_emp:
            for i in range(min(8, len(df))):
                r = df.iloc[i]
                bvg["linhas"].append(
                    {
                        "EMP/OBRA": str(r[c_eo]),
                        "EMPREENDIMENTO": str(r[c_emp]),
                        "VL.CARTEIRA": r.get(c_vl) if c_vl else None,
                        "VL.PAGO": r.get(c_pago) if c_pago else None,
                    }
                )
            bvg["total_linhas"] = len(df)

    wb2 = load_workbook(path_novo, data_only=True)
    hashes = []
    for sn in abas:
        if sn == "RESUMO GERAL":
            continue
        ws = wb2[sn]
        for row in range(9, min(ws.max_row or 0, 5000) + 1):
            for col in range(1, min(ws.max_column or 0, 30) + 1):
                v = ws.cell(row=row, column=col).value
                if isinstance(v, str) and "####" in v:
                    hashes.append(f"{sn}!{ws.cell(row=row, column=col).coordinate}")
                    if len(hashes) > 20:
                        break
            if len(hashes) > 20:
                break
        if len(hashes) > 20:
            break
    wb2.close()

    metricas = list(depois.keys())
    tabela = []
    for m in metricas:
        va = antes.get(m)
        vd = depois.get(m)
        diff = None if va is None or vd is None else round(float(vd) - float(va), 2)
        tabela.append({"metrica": m, "antes": va, "depois": vd, "diff": diff})

    rel = {
        "nota": "Subset 2 empreendimentos (BVGWH+ALVLT); totais != carteira geral completa.",
        "path_gerado": str(path_novo),
        "tempo_s": round(tempo, 2),
        "abas": abas,
        "amostra_resumo_geral_linhas_9_11": amostra_resumo,
        "aba_bvgwh": bvg,
        "celulas_com_cerquilha": hashes,
        "tabela_antes_depois": tabela,
    }
    out_j = RAIZ / "outputs" / "_e2e_evidencia_subset.json"
    out_j.write_text(json.dumps(rel, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print(json.dumps(rel, ensure_ascii=False, indent=2, default=str))
    return 0


if __name__ == "__main__":
    sys.exit(main())
