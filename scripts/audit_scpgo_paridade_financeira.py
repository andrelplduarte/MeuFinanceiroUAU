# -*- coding: utf-8 -*-
"""Paridade baseline vs otimizado: somas na aba consolidado SCPGO + freeze/autofilter/merges linha 7."""
from __future__ import annotations

import json
import re
import sys
import unicodedata

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def _fold(s: str) -> str:
    s = unicodedata.normalize("NFD", str(s or ""))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s.strip().upper())


def _aba_consolidado(names: list[str]) -> str:
    for n in names:
        if "CONSOLIDADO" in _fold(n) and "SCPGO" in _fold(n):
            return n
    for n in names:
        if "CONSOLIDADO" in _fold(n):
            return n
    raise ValueError(f"Nenhuma aba consolidado entre: {names}")


def _somas_consolidado(path: str, sheet: str) -> dict[str, float]:
    df = pd.read_excel(path, sheet_name=sheet, header=7, engine="openpyxl")
    m = {_fold(c): c for c in df.columns}
    want = {
        "Vl.Carteira": "VL.CARTEIRA",
        "Vl.Pago": "VL.PAGO",
        "Vl.Vencer": "VL.VENCER",
        "Vl.Principal (Encargos)": "VL.PRINCIPAL (ENCARGOS)",
        "Qtd.Parc.Atrasada": "QTD.PARC.ATRASADA",
    }
    out: dict[str, float] = {}
    for label, key in want.items():
        col = m.get(_fold(key))
        if not col:
            out[label] = float("nan")
            continue
        s = pd.to_numeric(df[col], errors="coerce").fillna(0)
        out[label] = float(s.sum())
    return out


def _estrutura_consolidado(path: str, sheet: str) -> dict:
    wb = load_workbook(path, data_only=False)
    try:
        ws = wb[sheet]
        esp7 = {"A7:G7", "H7:K7", "L7:R7", "S7:T7", "U7:X7", "Y7:AA7"}
        merges7 = sorted(
            f"{get_column_letter(m.min_col)}{m.min_row}:{get_column_letter(m.max_col)}{m.max_row}"
            for m in ws.merged_cells.ranges
            if m.min_row == m.max_row == 7
        )
        af = ws.auto_filter.ref if ws.auto_filter else None
        fp = ws.freeze_panes
        ok_merges = set(merges7) == esp7
        ok_af = bool(af) and str(af).upper().startswith("A8")
        ok_fp = str(fp or "").upper().replace("$", "") == "A9"
        return {
            "merges_linha7": merges7,
            "merges_ok": ok_merges,
            "auto_filter_ref": str(af or ""),
            "auto_filter_ok": ok_af,
            "freeze_panes": str(fp or ""),
            "freeze_ok": ok_fp,
            "max_column": ws.max_column,
        }
    finally:
        wb.close()


def main() -> int:
    if len(sys.argv) != 3:
        print("Uso: audit_scpgo_paridade_financeira.py <baseline.xlsx> <optimized.xlsx>", file=sys.stderr)
        return 2
    p_b, p_o = sys.argv[1], sys.argv[2]
    wb_b = load_workbook(p_b, read_only=True)
    names_b = list(wb_b.sheetnames)
    wb_b.close()
    wb_o = load_workbook(p_o, read_only=True)
    names_o = list(wb_o.sheetnames)
    wb_o.close()

    sb = _aba_consolidado(names_b)
    so = _aba_consolidado(names_o)

    somas_b = _somas_consolidado(p_b, sb)
    somas_o = _somas_consolidado(p_o, so)
    est_b = _estrutura_consolidado(p_b, sb)
    est_o = _estrutura_consolidado(p_o, so)

    tol = 0.02
    deltas = {k: round(somas_o[k] - somas_b[k], 2) for k in somas_b}
    finance_ok = all(abs(deltas[k]) <= tol for k in deltas)

    struct_ok = (
        est_b["merges_ok"]
        and est_o["merges_ok"]
        and est_b["auto_filter_ok"]
        and est_o["auto_filter_ok"]
        and est_b["freeze_ok"]
        and est_o["freeze_ok"]
    )

    report = {
        "aba_consolidado_baseline": sb,
        "aba_consolidado_optimized": so,
        "somas_baseline": somas_b,
        "somas_optimized": somas_o,
        "delta_optimized_minus_baseline": deltas,
        "estrutura_baseline": est_b,
        "estrutura_optimized": est_o,
        "pass_financeiro": finance_ok,
        "pass_estrutura_consolidado": struct_ok,
    }
    print(json.dumps(report, ensure_ascii=False, indent=2, default=str))
    return 0 if finance_ok and struct_ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
