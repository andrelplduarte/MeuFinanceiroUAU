# -*- coding: utf-8 -*-
"""Gera CARTEIRAS GERAL via lote real e extrai evidências objetivas."""
from __future__ import annotations

import json
import re
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

import pandas as pd

RAIZ = Path(__file__).resolve().parents[1]
if str(RAIZ) not in sys.path:
    sys.path.insert(0, str(RAIZ))

from services.orquestrador_lote_uau import processar_entrada_simples_ou_lote  # noqa: E402


def _fold(s: str) -> str:
    s = unicodedata.normalize("NFD", str(s or ""))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s.strip().upper())


def _norm_cols(df):
    return {_fold(c): c for c in df.columns}


def somas_carteiras_geral(path: str) -> dict:
    p = Path(path)
    if not p.is_file():
        return {}
    xl = pd.ExcelFile(p, engine="openpyxl")
    names = [x for x in xl.sheet_names if x != "RESUMO GERAL"]
    COL_DET = {
        "Vl.Carteira": _fold("VL.CARTEIRA"),
        "Vl.Pago": _fold("VL.PAGO"),
        "Vl.Vencer": _fold("VL.VENCER"),
        "Vl.Principal (Encargos)": _fold("VL.PRINCIPAL (ENCARGOS)"),
        "Qtd.Parc.Atrasada": _fold("QTD.PARC.ATRASADA"),
    }
    COL_RES = {
        "Vl.Pago": _fold("VALOR PAGO"),
        "Vl.Vencer": _fold("VALOR A VENCER"),
        "Vl.Principal (Encargos)": _fold("VALOR INADIMPLÊNCIA"),
        "Vl.Carteira": _fold("VL.CARTEIRA"),
        "Qtd.Parc.Atrasada": _fold("QTD PARC. INADIMPLÊNCIA"),
    }
    out = {}
    for label, col_f in COL_DET.items():
        total = 0.0
        for sn in names:
            df = pd.read_excel(p, sheet_name=sn, header=7, engine="openpyxl")
            m = _norm_cols(df)
            if col_f not in m:
                continue
            total += float(pd.to_numeric(df[m[col_f]], errors="coerce").fillna(0).sum())
        out[label] = total
    df_res = pd.read_excel(p, sheet_name="RESUMO GERAL", header=7, engine="openpyxl")
    mr = _norm_cols(df_res)
    out["_resumo"] = {}
    for label, rc in COL_RES.items():
        if rc in mr:
            out["_resumo"][label] = float(
                pd.to_numeric(df_res[mr[rc]], errors="coerce").fillna(0).sum()
            )
    return out


def main() -> int:
    uploads = RAIZ / "uploads"
    rec = sorted(str(p) for p in uploads.glob("rec_*.txt"))
    reb = sorted(str(p) for p in uploads.glob("reb_*.txt"))
    if len(rec) < 1 or len(reb) < 1:
        print(json.dumps({"erro": "Faltam rec_*.txt / reb_*.txt em uploads/"}, ensure_ascii=False))
        return 2

    saida = RAIZ / "outputs" / f"_e2e_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    saida.mkdir(parents=True, exist_ok=True)
    path_antigo = RAIZ / "outputs" / "CARTEIRAS GERAL.xlsx"
    antes = somas_carteiras_geral(str(path_antigo)) if path_antigo.is_file() else {}

    out, tempo = processar_entrada_simples_ou_lote(
        rec,
        reb,
        str(saida),
        "POR_EMPREENDIMENTO",
        caminhos_estoque=None,
    )
    path_novo = Path(out[0] if isinstance(out, tuple) else out)
    depois = somas_carteiras_geral(str(path_novo))

    # Evidências XLSX
    from openpyxl import load_workbook

    wb = load_workbook(path_novo, data_only=True)
    abas = wb.sheetnames
    ws_r = wb["RESUMO GERAL"]
    amostra_resumo = []
    for row in range(9, 12):
        a = ws_r.cell(row=row, column=1).value
        b = ws_r.cell(row=row, column=2).value
        c = ws_r.cell(row=row, column=3).value
        amostra_resumo.append({"linha": row, "A": a, "B": b, "C": c})
    wb.close()

    # Aba BVGWH
    alvo = next((n for n in abas if "BVGWH" in n.upper()), None)
    bvg = {"aba": alvo, "linhas": [], "bvgwhhhhh_count": 0}
    if alvo:
        df = pd.read_excel(path_novo, sheet_name=alvo, header=7, engine="openpyxl")
        mc = _norm_cols(df)
        c_eo = mc.get(_fold("EMP/OBRA"))
        c_emp = mc.get(_fold("EMPREENDIMENTO"))
        c_vl = mc.get(_fold("VL.CARTEIRA"))
        c_pago = mc.get(_fold("VL.PAGO"))
        if c_eo and c_emp:
            for i in range(min(8, len(df))):
                r = df.iloc[i]
                eo = str(r[c_eo] if c_eo in r.index else "")
                emp = str(r[c_emp] if c_emp in r.index else "")
                vc = r.get(c_vl) if c_vl else None
                vp = r.get(c_pago) if c_pago else None
                bvg["linhas"].append(
                    {"EMP/OBRA": eo, "EMPREENDIMENTO": emp, "VL.CARTEIRA": vc, "VL.PAGO": vp}
                )
            bvg["total_linhas"] = len(df)
        if c_eo:
            s = df[c_eo].fillna("").astype(str)
            bvg["bvgwhhhhh_count"] = int(s.str.upper().str.contains("BVGWHHHH", regex=False).sum())
            bvg["bvgwh_extra_h_count"] = int(s.str.upper().str.contains(r"BVGWH{2,}", regex=True).sum())

    # #######
    wb2 = load_workbook(path_novo, data_only=True)
    hashes = []
    for sn in abas:
        if sn == "RESUMO GERAL":
            continue
        ws = wb2[sn]
        for row in range(9, min(ws.max_row or 0, 8000) + 1):
            for col in range(1, min(ws.max_column or 0, 30) + 1):
                v = ws.cell(row=row, column=col).value
                if isinstance(v, str) and "####" in v:
                    hashes.append(f"{sn}!{ws.cell(row=row, column=col).coordinate}")
                    if len(hashes) > 30:
                        break
            if len(hashes) > 30:
                break
        if len(hashes) > 30:
            break
    wb2.close()

    metricas = [
        "Vl.Carteira",
        "Vl.Pago",
        "Vl.Vencer",
        "Vl.Principal (Encargos)",
        "Qtd.Parc.Atrasada",
    ]
    tabela = []
    for m in metricas:
        va = antes.get(m)
        vd = depois.get(m)
        diff = None if va is None or vd is None else round(float(vd) - float(va), 2)
        tabela.append({"metrica": m, "antes": va, "depois": vd, "diff": diff})

    rel = {
        "path_gerado": str(path_novo),
        "tempo_s": round(tempo, 2),
        "abas": abas,
        "amostra_resumo_geral_linhas_9_11": amostra_resumo,
        "aba_bvgwh": bvg,
        "celulas_com_cerquilha": hashes,
        "total_cerquilha_amostra": len(hashes),
        "tabela_antes_depois": tabela,
        "delta_resumo_interno": {
            m: (
                round(depois.get(m, 0) - depois.get("_resumo", {}).get(m, 0), 4)
                if depois.get("_resumo")
                else None
            )
            for m in metricas
        },
    }
    out_json = RAIZ / "outputs" / "_e2e_evidencia_geral.json"
    out_json.parent.mkdir(parents=True, exist_ok=True)
    out_json.write_text(json.dumps(rel, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print(json.dumps(rel, ensure_ascii=False, indent=2, default=str))
    return 0


if __name__ == "__main__":
    sys.exit(main())
