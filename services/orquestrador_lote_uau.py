# -*- coding: utf-8 -*-
"""
Orquestração em lote acima do motor processar_e_gerar_excel.
Não altera regras financeiras — apenas funde TXT, agrupa por empreendimento e compõe workbooks.
"""
from __future__ import annotations

import os
import re
import shutil
import tempfile
import time
import uuid
import unicodedata
from collections import defaultdict
from copy import copy
from datetime import datetime
from typing import Dict, List, Sequence, Tuple

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from services.processador_uau import (
    NOME_ABA_RESUMO_GERAL,
    ProcessamentoUAUErro,
    aplicar_estilo_excel,
    aplicar_estilo_arquivo_so_aba_consolidado_estoque,
    aplicar_estilo_arquivo_so_aba_resumo_geral,
    extrair_nome_empreendimento_nome_arquivo,
    extrair_nome_empreendimento_txt,
    identificar_tipo_relatorio_uau_por_texto,
    ler_dataframe_consolidado_de_xlsx_motor,
    ler_texto_robusto,
    limpar_nome_empreendimento,
    montar_dataframe_resumo_geral,
    processar_e_gerar_excel,
    sanitizar_nome_arquivo,
    _estrutura_minima_uau_ok,
    _ler_texto_validacao_entrada,
)

MODO_POR_EMPREENDIMENTO = "POR_EMPREENDIMENTO"

# Nome de exibição da aba consolidada em CARTEIRAS GERAL.xlsx (lote por empreendimento).
# Chave = sigla normalizada (maiúsculas). Ausente na tabela → fallback "<SIGLA> – CONSOLIDADO".
MAPA_TITULO_ABA_CONSOLIDADO_LOTE: Dict[str, str] = {
    "NVLOT": "NVLOT.NIL.VELOSO.RVD-GO",
    "LTMAG": "LTMAG.MAGALHAES.MAB-PA",
    "SCPTO": "SCPTO.TOCANTINS.MAB-PA",
    "SCPTI": "SCPTI.TIRADENTES.MAB-PA",
    "CIDAN": "CIDAN.CID.NOVA.ITP-PA",
    "VROLT": "VROLT.VLE.D.ROSAS.TCM-PA",
    "ALVLT": "ALVLT.ALVORADA.OUR-PA",
    "LTMON": "LTMON.MON.NEGRO.SFX-PA",
    "RVERD": "RVERD.R.VERDE.PPB-PA",
    "LTVIL": "LTVIL.VIL.NOVA.PPB-PA",
    "LTMIN": "LTMIN.MINERIOS.PB-PA",
    "SCPGO": "SCPGO.GOIANIA.CTP-PA",
    "ARAHF": "ARAHF.ARARAS.MIN-GO",
    "BVGWH": "BVGWH.B.WHITE.BVG-GO",
    "MANHA": "MANHA.MANHATAN.RVD-GO",
    "MONTB": "MONTB.MONTBLANC.RVD-GO",
    "LIFE": "LIFE.LIFE.RVD-GO",
}


def _titulo_aba_consolidado_carteiras_geral(sigla: str) -> str:
    k = str(sigla or "").strip().upper()
    if k in MAPA_TITULO_ABA_CONSOLIDADO_LOTE:
        return MAPA_TITULO_ABA_CONSOLIDADO_LOTE[k]
    return f"{sigla} – CONSOLIDADO"


def aba_e_consolidado_carteiras_geral(nome_aba: str) -> bool:
    """True se o título é aba consolidada por empreendimento (nome oficial ou fallback)."""
    s = str(nome_aba or "").strip()
    if not s:
        return False
    up = s.upper()
    # Aceita hífen ASCII, travessão (U+2013) e sufixo legado com caracteres corrompidos em versões antigas.
    if (
        up.endswith("- CONSOLIDADO")
        or up.endswith("– CONSOLIDADO")
        or up.endswith("â€“ CONSOLIDADO")
    ):
        return True
    return s in MAPA_TITULO_ABA_CONSOLIDADO_LOTE.values()


# Prefixos genéricos no nome do arquivo que não são sigla de obra (ex.: REC_*, RECEBER_*).
# Se o primeiro segmento for só isso, o regex antigo colapsava todo o lote numa única chave "REC".
# REC/REB/EST/UPL: prefixos técnicos do app ao salvar uploads (rec_/reb_/est_/upl_ + índice).
_PREFIXOS_NAO_SIGLA = frozenset(
    {
        "REC",
        "REB",
        "EST",
        "UPL",
        "RECEBER",
        "RECEBIDOS",
        "RECEB",
        "PAG",
        "PAGOS",
        "CONTAS",
        "LOT",
        "LOTE",
        "DADOS",
        "UAC",
        "UAU",
        "EMP",
        "FILE",
        "TEMP",
        "TXT",
        "REL",
        "RELATORIO",
    }
)

# Padrão app.py: <prefixo>_<índice>_<nome original...>
_RE_PREFIXO_UPLOAD_APP = re.compile(r"^(REC|REB|EST|UPL)_\d+_", re.IGNORECASE)


def _sigla_curta_do_caminho(caminho: str) -> str:
    """Prefixo tipo SCPGO a partir do nome do arquivo (ex.: ALVLT, SCPGO)."""
    try:
        b = os.path.basename(str(caminho or "")).upper()
    except Exception:
        b = ""
    # Uploads Flask: 00_SCPGO_-LOT... → ignorar prefixo numérico
    b = re.sub(r"^\d+_", "", b)
    # reb_00_ALVLT_... / rec_03_CIDAN_... (repetir: encadeamentos raros rec_/reb_ no nome)
    for _ in range(4):
        nb = _RE_PREFIXO_UPLOAD_APP.sub("", b)
        if nb == b:
            break
        b = nb
    base_sem_ext = os.path.splitext(b)[0]
    # Preferir o primeiro token alfanumérico que não seja prefixo de ruído (evita REC, RECEBER, LOT...).
    partes = [p for p in re.split(r"[_\-\s]+", base_sem_ext) if p]
    for raw in partes:
        token = sanitizar_nome_arquivo(raw.split(".")[0])
        if not token or token.isdigit():
            continue
        if token in _PREFIXOS_NAO_SIGLA:
            continue
        if len(token) < 2:
            continue
        if not re.match(r"^[A-Z0-9]+$", token):
            continue
        return token[:20]
    m = re.match(r"^([A-Z0-9]+)[_-]", b)
    if m:
        return sanitizar_nome_arquivo(m.group(1))[:20]
    base = os.path.splitext(b)[0][:12]
    return sanitizar_nome_arquivo(base) if base else "EMP"


def _chave_pareamento_por_prefixo_arquivo(caminho: str) -> str:
    """
    Pareamento estável no modo por empreendimento: mesma sigla no nome do arquivo
    (Receber e Recebidos do mesmo código, ex. LTMIN / LTMIN).
    Se não houver sigla reconhecível, cai no agrupamento por nome canônico do TXT.
    """
    s = _sigla_curta_do_caminho(caminho)
    if s and s != "EMP":
        return s
    return _chave_grupo_empreendimento(caminho)


def _diagnostico_pareamento_basename_chave(caminhos: Sequence[str], limite: int = 5) -> str:
    """Uma linha por arquivo: basename->chave (para mensagens de pareamento)."""
    linhas: List[str] = []
    n = len(caminhos)
    for p in caminhos[:limite]:
        bn = os.path.basename(str(p or ""))
        ch = _chave_pareamento_por_prefixo_arquivo(str(p or ""))
        linhas.append(f"  {bn}->{ch}")
    if n > limite:
        linhas.append(f"  … (+{n - limite} arquivo(s) não listados)")
    return "\n".join(linhas) if linhas else "  (nenhum arquivo)"


def _chave_grupo_empreendimento(caminho: str) -> str:
    k = extrair_nome_empreendimento_txt(caminho) or extrair_nome_empreendimento_nome_arquivo(caminho)
    k = limpar_nome_empreendimento(k)
    if not k:
        k = _sigla_curta_do_caminho(caminho)
    return sanitizar_nome_arquivo(k)


def _validar_tipo_arquivo(caminho: str, esperado: str, campo: str) -> None:
    texto = _ler_texto_validacao_entrada(caminho)
    tipo = identificar_tipo_relatorio_uau_por_texto(texto)
    if esperado == "RECEBER" and tipo != "RECEBER":
        raise ProcessamentoUAUErro(
            etapa="validação de entrada",
            funcao="orquestrador_lote_uau",
            validacao="tipo de relatório",
            mensagem=f"Arquivo em {campo} não foi reconhecido como Contas a Receber: {os.path.basename(caminho)}",
            campo_ou_aba=campo,
        )
    if esperado == "RECEBIDOS" and tipo != "RECEBIDOS":
        raise ProcessamentoUAUErro(
            etapa="validação de entrada",
            funcao="orquestrador_lote_uau",
            validacao="tipo de relatório",
            mensagem=f"Arquivo em {campo} não foi reconhecido como Contas Recebidas: {os.path.basename(caminho)}",
            campo_ou_aba=campo,
        )
    if not _estrutura_minima_uau_ok(texto, esperado):
        msg = (
            "Estrutura mínima incompatível (Contas a Receber)."
            if esperado == "RECEBER"
            else "Estrutura mínima incompatível (Contas Recebidas)."
        )
        raise ProcessamentoUAUErro(
            etapa="validação de entrada",
            funcao="orquestrador_lote_uau",
            validacao="estrutura mínima",
            mensagem=f"{msg} Arquivo: {os.path.basename(caminho)}",
            campo_ou_aba=campo,
        )


def _fundir_textos_em_temp(caminhos: Sequence[str], prefixo: str) -> str:
    partes = []
    for p in caminhos:
        partes.append(ler_texto_robusto(p))
    corpo = "\n".join(partes)
    fd, path = tempfile.mkstemp(prefix=prefixo + "_", suffix=".txt", text=True)
    try:
        with os.fdopen(fd, "w", encoding="utf-8", newline="\n") as f:
            f.write(corpo)
    except Exception:
        try:
            os.unlink(path)
        except OSError:
            pass
        raise
    return path


def _remover_seguro(path: str) -> None:
    try:
        if path and os.path.isfile(path):
            os.unlink(path)
    except OSError:
        pass


def _resolver_estoque_unificado_lote(caminhos_est: Sequence[str], temporarios: List[str]) -> str | None:
    """Funde vários TXT de estoque num único ficheiro temporário (modo unificado / par único com N estoques)."""
    paths = [os.path.abspath(os.path.normpath(p)) for p in (caminhos_est or []) if p]
    if not paths:
        return None
    if len(paths) == 1:
        return paths[0]
    m = _fundir_textos_em_temp(paths, "uau_est_u_")
    temporarios.append(m)
    return m


def _resolver_estoque_por_chave_lote(
    chave: str, caminhos_est: Sequence[str], temporarios: List[str]
) -> str | None:
    """
    Um ficheiro de estoque → aplicado a todos os grupos.
    Vários ficheiros → agrupa pela mesma chave de prefixo do nome (paralelo ao Receber/Recebidos).
    """
    paths = [os.path.abspath(os.path.normpath(p)) for p in (caminhos_est or []) if p]
    if not paths:
        return None
    if len(paths) == 1:
        return paths[0]
    grupos: Dict[str, List[str]] = defaultdict(list)
    for p in paths:
        grupos[_chave_pareamento_por_prefixo_arquivo(p)].append(p)
    lst = sorted(grupos.get(chave, []))
    if not lst:
        return None
    if len(lst) == 1:
        return lst[0]
    m = _fundir_textos_em_temp(lst, f"est_{chave}_")
    temporarios.append(m)
    return m


def _data_base_de_primeiro_xlsx_motor(caminho: str):
    """Lê B2 do painel do Consolidado no workbook gerado pelo motor."""
    if not caminho or not os.path.isfile(caminho):
        return None
    wb = None
    try:
        wb = load_workbook(caminho, data_only=True)
        for nm in wb.sheetnames:
            su = str(nm).upper().replace("Í", "I")
            if "CONSOLIDADO" in su and "CRIT" not in su:
                v = wb[nm]["B2"].value
                if isinstance(v, datetime):
                    return v
                if isinstance(v, str) and str(v).strip():
                    try:
                        return datetime.strptime(str(v).strip(), "%d/%m/%Y")
                    except ValueError:
                        return None
                return None
    except Exception:
        return None
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
    return None


def _mapear_titulo_aba_por_empreendimento(nome_original: str, sigla: str) -> str:
    n = str(nome_original or "").strip()
    nu = n.upper()
    if nu == "CRITÉRIOS" or nu == "CRITERIOS" or nu == "CRITERIOS ANALISES":
        return "CRITERIOS ANALISES"
    if "CONSOLIDADO" in nu and "CRIT" not in nu:
        return _titulo_aba_consolidado_carteiras_geral(sigla)
    if n == "Dados Receber":
        return f"DADOS RECEBER - {sigla}"
    if n == "Dados Recebidos":
        return f"DADOS RECEBIDOS - {sigla}"
    if n == "Pendencias_Parcelas" or n == "PENDENCIAS_PARCELAS" or n == "PEND.PARCELAS":
        return f"PEND.PARCELAS - {sigla}"
    if n == "RELATORIO ANALITICO" or "RELATORIO" in nu:
        return f"RELATORIO ANALITICO - {sigla}"
    return f"{sigla} – {n}"


def _titulo_aba_unico(_wb: Workbook, titulo: str, usados: set) -> str:
    base = (titulo or "ABA")[:31]
    nome = base
    i = 2
    while nome in usados:
        suf = f"_{i}"
        nome = (base[: 31 - len(suf)] + suf)[:31]
        i += 1
    usados.add(nome)
    return nome


def _copiar_planilha_estilizada(ws_src, wb_dst, titulo: str):
    ws = wb_dst.create_sheet(title=titulo)
    max_r = ws_src.max_row or 1
    max_c = ws_src.max_column or 1
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            src = ws_src.cell(row=r, column=c)
            dst = ws.cell(row=r, column=c)
            dst.value = src.value
            if src.has_style:
                try:
                    dst.font = copy(src.font)
                    dst.fill = copy(src.fill)
                    dst.border = copy(src.border)
                    dst.alignment = copy(src.alignment)
                    dst.number_format = src.number_format
                except Exception:
                    pass
    for col_letter, dim in ws_src.column_dimensions.items():
        if dim.width is not None:
            ws.column_dimensions[col_letter].width = dim.width
        if dim.hidden:
            ws.column_dimensions[col_letter].hidden = True
    for row_idx, rd in ws_src.row_dimensions.items():
        if rd.height is not None:
            ws.row_dimensions[row_idx].height = rd.height
    try:
        merges = sorted(
            ws_src.merged_cells.ranges,
            key=lambda x: (x.min_row, x.min_col, x.max_row, x.max_col),
        )
        for mc in merges:
            ws.merge_cells(str(mc))
    except Exception:
        pass
    if ws_src.freeze_panes:
        ws.freeze_panes = ws_src.freeze_panes
    if ws_src.auto_filter and ws_src.auto_filter.ref:
        ws.auto_filter.ref = ws_src.auto_filter.ref


def _nome_aba_consolidado_motor(wb) -> str | None:
    """Aba gerencial por obra (exclui estoque e critérios)."""
    for s in wb.sheetnames:
        su = str(s).upper().replace("Í", "I")
        if "CONSOLIDADO" in su and "CRIT" not in su and "ESTOQUE" not in su:
            return s
    return None


def _anexar_somente_consolidado_por_sigla(
    wb_dest: Workbook, caminho_xlsx: str, sigla: str, titulos_usados: set
) -> None:
    """No lote por empreendimento: só a aba consolidada da obra entra separada no workbook final."""
    wb = load_workbook(caminho_xlsx, data_only=False)
    try:
        nome = _nome_aba_consolidado_motor(wb)
        if not nome:
            return
        titulo = _mapear_titulo_aba_por_empreendimento("Consolidado Venda", sigla)
        titulo_final = _titulo_aba_unico(wb_dest, titulo[:31], titulos_usados)
        _copiar_planilha_estilizada(wb[nome], wb_dest, titulo_final)
    finally:
        wb.close()


def _ler_df_aba_xlsx_motor(caminho: str, nome_aba: str, header_row_0based: int) -> pd.DataFrame:
    """Lê aba já formatada pelo motor (cabeçalho tabular na linha Excel = header_row_0based + 1)."""
    if not caminho or not os.path.isfile(caminho):
        return pd.DataFrame()
    sinais = {"EMP_OBRA", "VENDA", "CLIENTE", "PARCELA", "PARC_NUM", "PARC_TOTAL", "DATA_REC", "VENCIMENTO"}

    def _score(df: pd.DataFrame) -> int:
        if df is None or df.empty:
            return -1
        cols = [_normalizar_nome_coluna_sql(c) for c in df.columns.tolist()]
        return sum(1 for c in cols if c in sinais)

    def _headers_candidatos() -> List[int]:
        base = max(0, int(header_row_0based))
        return list(dict.fromkeys([base, base + 1, base + 2, base + 3, max(0, base - 1)]))

    def _tentar_ler(sheet_name: str) -> List[pd.DataFrame]:
        out = []
        for hdr in _headers_candidatos():
            try:
                out.append(pd.read_excel(caminho, sheet_name=sheet_name, header=hdr))
            except Exception:
                pass
        return out

    candidatos: List[pd.DataFrame] = []
    try:
        candidatos.extend(_tentar_ler(nome_aba))
    except Exception:
        pass
    try:
        xl = pd.ExcelFile(caminho)
        alvo = str(nome_aba or "").strip().upper()
        aliases = {alvo}
        if alvo == "DADOS RECEBER":
            aliases.add("DADOS RECEBER")
            aliases.add("DADOS RECEBER".title())
        elif alvo == "DADOS RECEBIDOS":
            aliases.add("DADOS RECEBIDOS")
            aliases.add("DADOS RECEBIDOS".title())
        elif alvo == "DADOS GERAL":
            aliases.update({"RELATORIO ANALITICO", "DADOS GERAL"})
        elif alvo == "PENDENCIAS_PARCELAS" or alvo == "PEND.PARCELAS":
            aliases.update({"PEND.PARCELAS", "PENDENCIAS_PARCELAS", "PENDENCIAS_PARCELAS".title(), "Pendencias_Parcelas"})
        elif alvo == "CRITERIOS" or alvo == "CRITERIOS ANALISES":
            aliases.update({"CRITERIOS ANALISES", "CRITERIOS", "CRITÉRIOS", "CRITERIOS".title(), "Critérios"})
        for sn in xl.sheet_names:
            su = str(sn or "").strip().upper()
            if su in aliases:
                candidatos.extend(_tentar_ler(sn))
    except Exception:
        pass
    if not candidatos:
        return pd.DataFrame()
    melhor = max(candidatos, key=_score)
    return melhor if _score(melhor) >= 2 else candidatos[0]


def _concat_dfs_vertical(partes: List[pd.DataFrame]) -> pd.DataFrame:
    ok = [x for x in partes if x is not None and not x.empty]
    if not ok:
        return pd.DataFrame()
    return pd.concat(ok, ignore_index=True, sort=False)


def _normalizar_nome_coluna_sql(nome: str) -> str:
    s = str(nome or "").strip()
    if not s:
        return "COL"
    s = (
        unicodedata.normalize("NFKD", s)
        .encode("ascii", "ignore")
        .decode("ascii")
    )
    s = s.upper()
    s = re.sub(r"[^A-Z0-9]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s or "COL"


def _df_sql_like(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    out = df.copy()
    usados = set()
    novas = []
    for c in out.columns:
        base = _normalizar_nome_coluna_sql(c)
        nome = base
        i = 2
        while nome in usados:
            nome = f"{base}_{i}"
            i += 1
        usados.add(nome)
        novas.append(nome)
    out.columns = novas
    return out


def _ordenar_df_base_sql(df: pd.DataFrame, candidatos: Sequence[str]) -> pd.DataFrame:
    """
    Ordena dados da base opcional com chave estável e consistente.
    Prioriza a primeira coluna de negócio disponível.
    """
    if df is None or df.empty:
        return pd.DataFrame()
    out = df.copy()
    colunas_existentes = [c for c in candidatos if c in out.columns]
    if not colunas_existentes:
        return out
    chave_principal = colunas_existentes[0]
    extras = [c for c in colunas_existentes[1:] if c != chave_principal]
    chaves = [chave_principal] + extras
    return out.sort_values(by=chaves, ascending=True, kind="mergesort").reset_index(drop=True)


def _padronizar_colunas_base_para_negocio(df: pd.DataFrame, origem: str) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    out = df.copy()
    col_norm = {c: _normalizar_nome_coluna_sql(c) for c in out.columns}

    def _find_col(*aliases: str) -> str | None:
        als = [a for a in aliases if a]
        for alias in als:
            a_norm = _normalizar_nome_coluna_sql(alias)
            for col, c_norm in col_norm.items():
                if c_norm == a_norm or c_norm.startswith(a_norm + "_"):
                    return col
        return None

    if origem == "RECEBER":
        schema = [
            ("EMP/OBRA", ("EMP_OBRA",)),
            ("VENDA", ("VENDA",)),
            ("CLIENTE", ("CLIENTE",)),
            ("CLI.BASE", ("CLI_BASE", "CLIENTE_BASE")),
            ("IDENTIFICADOR", ("IDENTIFICADOR", "IDENTIFICADOR_PRODUTO")),
            ("PARC.(GERAL)", ("PARC_GERAL", "PARCELA")),
            ("PARC.NUM", ("PARC_NUM",)),
            ("PARC.TOTAL", ("PARC_TOTAL",)),
            ("VENC.DATA", ("VENC_DATA", "VENCIMENTO")),
            ("STATUS", ("STATUS", "STATUS_VENCIMENTO")),
            ("DIA.VENC.", ("DIA_VENC", "DIA_VENCIMENTO_BOLETO")),
            ("MES.VENC.", ("MES_VENC", "MES_VENCIMENTO")),
            ("ANO.VENC.", ("ANO_VENC", "ANO_VENCIMENTO")),
            ("CLASSIFICAÇÃO", ("CLASSIFICACAO", "CLASSIFICACAO_ADIMPLENCIA")),
            ("PRINCIPAL", ("PRINCIPAL",)),
            ("CORREÇÃO", ("CORRECAO",)),
            ("JUROS ATRASO", ("JUROS_ATRASO",)),
            ("MULTA ATRASO", ("MULTA_ATRASO",)),
            ("CORREÇÃO ATRASO", ("CORRECAO_ATRASO",)),
            ("VL.PARCELA", ("VL_PARCELA", "VLR_PARCELA")),
        ]
    else:
        schema = [
            ("EMP/OBRA", ("EMP_OBRA",)),
            ("VENDA", ("VENDA",)),
            ("CLIENTE", ("CLIENTE",)),
            ("CLI.BASE", ("CLI_BASE", "CLIENTE_BASE")),
            ("IDENTIFICADOR", ("IDENTIFICADOR", "IDENTIFICADOR_PRODUTO")),
            ("PARC.(GERAL)", ("PARC_GERAL", "PARCELA")),
            ("PARC.NUM", ("PARC_NUM",)),
            ("PARC.TOTAL", ("PARC_TOTAL",)),
            ("DATA.REC.", ("DATA_REC",)),
            ("VL.PARCELA", ("VL_PARCELA", "VLR_PARCELA", "TOTAL_DEP")),
        ]

    out_final = pd.DataFrame(index=out.index)
    for destino, aliases in schema:
        src = _find_col(*aliases)
        if src is None:
            out_final[destino] = ""
        else:
            out_final[destino] = out[src]
    return out_final


def _promover_primeira_linha_cabecalho_se_necessario(df: pd.DataFrame) -> pd.DataFrame:
    """
    Alguns lotes trazem uma linha de cabeçalho repetida como primeiro registro.
    Quando detectado, promove essa linha para cabeçalho real na base SQL-like.
    """
    if df is None or df.empty:
        return pd.DataFrame()
    out = df.copy()
    if len(out) < 1:
        return out

    linhas_candidatas = [(-1, [str(c or "").strip() for c in out.columns.tolist()])]
    limite = min(30, len(out))
    for i in range(limite):
        linhas_candidatas.append((i, [str(v or "").strip() for v in out.iloc[i].tolist()]))
    sinais = {"EMP_OBRA", "VENDA", "CLIENTE", "PARCELA", "VENCIMENTO", "DATA_REC", "STATUS_VENCIMENTO"}
    melhor_idx = -1
    melhor_score = -1
    melhor_norm = []
    for idx, vals in linhas_candidatas:
        vals_norm = [_normalizar_nome_coluna_sql(v) for v in vals]
        score = sum(1 for x in vals_norm if x in sinais)
        if score > melhor_score:
            melhor_score = score
            melhor_idx = idx
            melhor_norm = vals_norm
    if melhor_score < 4:
        return out

    usados = set()
    novas = []
    for c in melhor_norm:
        base = c or "COL"
        nome = base
        i = 2
        while nome in usados:
            nome = f"{base}_{i}"
            i += 1
        usados.add(nome)
        novas.append(nome)
    out.columns = novas
    if melhor_idx >= 0:
        out = out.iloc[melhor_idx + 1 :].reset_index(drop=True)
    return out


def _normalizar_schema_final_base(df: pd.DataFrame, origem: str) -> pd.DataFrame:
    if df is None:
        df = pd.DataFrame()
    base = df.copy()
    norm_cols = {c: _normalizar_nome_coluna_sql(c) for c in base.columns}

    def pick(*aliases: str):
        als = [_normalizar_nome_coluna_sql(a) for a in aliases if a]
        for a in als:
            for c, nc in norm_cols.items():
                if nc == a or nc.startswith(a + "_"):
                    return base[c]
        return pd.Series([""] * len(base), index=base.index)

    if origem == "RECEBER":
        out = pd.DataFrame(
            {
                "EMP/OBRA": pick("EMP_OBRA"),
                "VENDA": pick("VENDA"),
                "CLIENTE": pick("CLIENTE"),
                "CLI.BASE": pick("CLI_BASE", "CLIENTE_BASE"),
                "IDENTIFICADOR": pick("IDENTIFICADOR", "IDENTIFICADOR_PRODUTO"),
                "PARC.(GERAL)": pick("PARC_GERAL", "PARCELA"),
                "PARC.NUM": pick("PARC_NUM"),
                "PARC.TOTAL": pick("PARC_TOTAL"),
                "VENC.DATA": pick("VENC_DATA", "VENCIMENTO"),
                "STATUS": pick("STATUS", "STATUS_VENCIMENTO"),
                "DIA.VENC.": pick("DIA_VENC", "DIA_VENCIMENTO_BOLETO"),
                "MES.VENC.": pick("MES_VENC", "MES_VENCIMENTO"),
                "ANO.VENC.": pick("ANO_VENC", "ANO_VENCIMENTO"),
                "CLASSIFICAÇÃO": pick("CLASSIFICACAO", "CLASSIFICACAO_ADIMPLENCIA"),
                "PRINCIPAL": pick("PRINCIPAL"),
                "CORREÇÃO": pick("CORRECAO"),
                "JUROS ATRASO": pick("JUROS_ATRASO"),
                "MULTA ATRASO": pick("MULTA_ATRASO"),
                "CORREÇÃO ATRASO": pick("CORRECAO_ATRASO"),
                "VL.PARCELA": pick("VL_PARCELA", "VLR_PARCELA"),
            }
        )
        return out

    out = pd.DataFrame(
        {
            "EMP/OBRA": pick("EMP_OBRA"),
            "VENDA": pick("VENDA"),
            "CLIENTE": pick("CLIENTE"),
            "CLI.BASE": pick("CLI_BASE", "CLIENTE_BASE"),
            "IDENTIFICADOR": pick("IDENTIFICADOR", "IDENTIFICADOR_PRODUTO"),
            "PARC.(GERAL)": pick("PARC_GERAL", "PARCELA"),
            "PARC.NUM": pick("PARC_NUM"),
            "PARC.TOTAL": pick("PARC_TOTAL"),
            "DATA.REC.": pick("DATA_REC"),
            "VL.PARCELA": pick("VL_PARCELA", "VLR_PARCELA", "TOTAL_DEP"),
        }
    )
    return out


def _aplicar_schema_e_formato_base_final(
    destino_base: str,
    dr_in: pd.DataFrame | None = None,
    dp_in: pd.DataFrame | None = None,
) -> None:
    """
    Publica a base final com schema padronizado e estilo mínimo.
    Mantém os mesmos dados, reduzindo formatação massiva célula-a-célula
    para melhorar fluidez de abertura no Excel.
    """
    dr = dr_in.copy() if isinstance(dr_in, pd.DataFrame) else pd.DataFrame()
    dp = dp_in.copy() if isinstance(dp_in, pd.DataFrame) else pd.DataFrame()
    if dr.empty and dp.empty:
        try:
            dr = pd.read_excel(destino_base, sheet_name="DADOS_RECEBER")
        except Exception:
            dr = pd.DataFrame()
        try:
            dp = pd.read_excel(destino_base, sheet_name="DADOS_RECEBIDOS")
        except Exception:
            dp = pd.DataFrame()

    dr_f = _normalizar_schema_final_base(dr, "RECEBER")
    dp_f = _normalizar_schema_final_base(dp, "RECEBIDOS")
    with pd.ExcelWriter(destino_base, engine="openpyxl") as wr:
        dr_f.to_excel(wr, sheet_name="DADOS_RECEBER", index=False)
        dp_f.to_excel(wr, sheet_name="DADOS_RECEBIDOS", index=False)

    # Estilo enxuto: apenas autofiltro na linha de cabeçalho.
    wb = load_workbook(destino_base)
    try:
        for aba in ("DADOS_RECEBER", "DADOS_RECEBIDOS"):
            if aba not in wb.sheetnames:
                continue
            ws = wb[aba]
            if ws.max_column > 0:
                ult_col = get_column_letter(ws.max_column)
                ws.auto_filter.ref = f"A1:{ult_col}1"
    finally:
        wb.save(destino_base)
        wb.close()


def _copiar_abas_ordenadas_para_destino(
    wb_fonte: Workbook, wb_dest: Workbook, nomes_origem: List[str], titulos_destino: List[str], titulos_usados: set
) -> None:
    for orig, tit in zip(nomes_origem, titulos_destino):
        if orig not in wb_fonte.sheetnames:
            continue
        titulo_final = _titulo_aba_unico(wb_dest, tit[:31], titulos_usados)
        _copiar_planilha_estilizada(wb_fonte[orig], wb_dest, titulo_final)


def processar_lote_uau(
    caminhos_receber: Sequence[str],
    caminhos_recebidos: Sequence[str],
    caminho_saida_base: str,
    modo_geracao: str,
    caminhos_estoque: Sequence[str] | None = None,
    progresso_cb=None,
) -> Tuple[Tuple[str, str], float]:
    """
    Entrada: listas de caminhos absolutos já salvos em disco.
    Saída: ((caminho_xlsx_principal, caminho_xlsx_base_opcional), tempo_total_segundos).
    """
    def _emitir_progresso(**payload):
        if callable(progresso_cb):
            try:
                progresso_cb(payload)
            except Exception:
                pass

    def _cor_tempo(segundos: float, media_ref: float) -> str:
        base = max(1.0, float(media_ref or 0.0))
        s = max(0.0, float(segundos or 0.0))
        if s <= base * 0.85:
            return "verde"
        if s <= base * 1.05:
            return "azul"
        if s <= base * 1.30:
            return "amarelo"
        return "vermelho"

    cr = [os.path.abspath(os.path.normpath(p)) for p in caminhos_receber if p]
    cp = [os.path.abspath(os.path.normpath(p)) for p in caminhos_recebidos if p]
    if len(cr) < 1 or len(cp) < 1:
        raise ProcessamentoUAUErro(
            etapa="validação",
            funcao="processar_lote_uau",
            validacao="arquivos insuficientes",
            mensagem="Envie ao menos um TXT em Contas a Receber e outro em Contas Recebidas.",
            campo_ou_aba="upload",
        )

    modo = str(modo_geracao or "").strip().upper()
    if modo in ("POR_EMPREENDIMENTO", "POR EMPREENDIMENTO", "EMP", "2"):
        modo = MODO_POR_EMPREENDIMENTO
    else:
        raise ProcessamentoUAUErro(
            etapa="validação",
            funcao="processar_lote_uau",
            validacao="modo de geração",
            mensagem="Modo de geração inválido. Use POR_EMPREENDIMENTO.",
            campo_ou_aba="modo_geracao",
        )

    todos = set(cr) | set(cp)
    if len(todos) < len(cr) + len(cp):
        raise ProcessamentoUAUErro(
            etapa="validação",
            funcao="processar_lote_uau",
            validacao="arquivos duplicados",
            mensagem="Há caminhos duplicados entre os anexos. Remova duplicatas.",
            campo_ou_aba="upload",
        )

    t0 = time.perf_counter()
    pasta_saida = os.path.dirname(os.path.abspath(caminho_saida_base)) or "."
    os.makedirs(pasta_saida, exist_ok=True)
    pasta_temp_local = os.path.join(pasta_saida, "_tmp_lote_uau")
    os.makedirs(pasta_temp_local, exist_ok=True)

    for p in cr:
        _validar_tipo_arquivo(p, "RECEBER", "Contas a Receber")
    for p in cp:
        _validar_tipo_arquivo(p, "RECEBIDOS", "Contas Recebidas")

    grupos_r: Dict[str, List[str]] = {}
    for p in cr:
        k = _chave_pareamento_por_prefixo_arquivo(p)
        grupos_r.setdefault(k, []).append(p)
    grupos_p: Dict[str, List[str]] = {}
    for p in cp:
        k = _chave_pareamento_por_prefixo_arquivo(p)
        grupos_p.setdefault(k, []).append(p)

    chaves_ok = sorted(set(grupos_r) & set(grupos_p))
    if not chaves_ok:
        amostra_r = ", ".join(sorted(list(grupos_r.keys()))[:8]) if grupos_r else "nenhuma"
        amostra_p = ", ".join(sorted(list(grupos_p.keys()))[:8]) if grupos_p else "nenhuma"
        aux_pareamento = (
            "Amostra basename->chave extraída (máx. 5 por lado):\n"
            "Receber:\n"
            f"{_diagnostico_pareamento_basename_chave(cr)}\n"
            "Recebidos:\n"
            f"{_diagnostico_pareamento_basename_chave(cp)}\n"
            f"Chaves agregadas — Receber: {sorted(grupos_r.keys())} | Recebidos: {sorted(grupos_p.keys())}"
        )
        raise ProcessamentoUAUErro(
            etapa="validação",
            funcao="processar_lote_uau",
            validacao="pareamento por empreendimento",
            mensagem=(
                "Não foi possível parear Contas a Receber e Contas Recebidas pelo mesmo empreendimento. "
                "Verifique se cada empreendimento tem um par de arquivos (nomes/cabeçalhos coerentes). "
                f"Chaves detectadas em Receber: [{amostra_r}] | "
                f"Chaves detectadas em Recebidos: [{amostra_p}]"
            ),
            campo_ou_aba="lote",
            contexto={"Dados_Auxiliares": aux_pareamento},
        )

    apenas_r = set(grupos_r) - set(grupos_p)
    apenas_p = set(grupos_p) - set(grupos_r)
    if apenas_r or apenas_p:
        msg_extra = []
        if apenas_r:
            msg_extra.append(f"Somente Receber: {', '.join(list(apenas_r)[:5])}")
        if apenas_p:
            msg_extra.append(f"Somente Recebidos: {', '.join(list(apenas_p)[:5])}")
        aux_pareamento = (
            "Amostra basename->chave extraída (máx. 5 por lado):\n"
            "Receber:\n"
            f"{_diagnostico_pareamento_basename_chave(cr)}\n"
            "Recebidos:\n"
            f"{_diagnostico_pareamento_basename_chave(cp)}\n"
            f"Chaves agregadas — Receber: {sorted(grupos_r.keys())} | Recebidos: {sorted(grupos_p.keys())}"
        )
        raise ProcessamentoUAUErro(
            etapa="validação",
            funcao="processar_lote_uau",
            validacao="pareamento por empreendimento",
            mensagem="Empreendimentos sem par completo. " + " | ".join(msg_extra),
            campo_ou_aba="lote",
            contexto={"Dados_Auxiliares": aux_pareamento},
        )

    _emitir_progresso(
        status="iniciado",
        total_empreendimentos=len(chaves_ok),
        concluidos=0,
        empreendimento_atual=None,
    )

    wb_exec = Workbook()
    wb_exec.remove(wb_exec.active)
    titulos_exec: set = set()
    temporarios: List[str] = []
    pastas_temp_workbook: List[str] = []
    caminhos_motor_para_resumo: List[str] = []
    consolidado_por_sigla: List[Tuple[str, str, float]] = []
    consolidado_cache_por_path: Dict[str, pd.DataFrame] = {}
    pares_motor_est: List[Tuple[str, str | None]] = []
    partes_dr: List[pd.DataFrame] = []
    partes_dp: List[pd.DataFrame] = []
    df_criterios_ref: pd.DataFrame | None = None
    tempos_por_item: List[Dict[str, object]] = []
    inicio_item_ts: float | None = None

    try:
        for idx, chave in enumerate(chaves_ok, start=1):
            inicio_item_ts = time.perf_counter()
            tempo_total_decorrido = max(0.0, time.perf_counter() - t0)
            lista_r = sorted(grupos_r[chave])
            lista_p = sorted(grupos_p[chave])
            sigla = _sigla_curta_do_caminho(lista_r[0])
            nome_aba_consolidado = _titulo_aba_consolidado_carteiras_geral(sigla)
            _emitir_progresso(
                status="processando",
                total_empreendimentos=len(chaves_ok),
                concluidos=idx - 1,
                empreendimento_atual=str(chave),
                item_atual_abas=nome_aba_consolidado,
                abas_item=[nome_aba_consolidado],
                itens_tempo=tempos_por_item,
                tempo_decorrido_segundos=tempo_total_decorrido,
            )
            tmp_r = _fundir_textos_em_temp(lista_r, "rec_")
            tmp_p = _fundir_textos_em_temp(lista_p, "pag_")
            temporarios.extend([tmp_r, tmp_p])
            # Pasta exclusiva: processar_e_gerar_excel limpa *todos* os .xlsx em dirname(caminho_saida).
            # Se o placeholder estiver em %TEMP%, o mkstemp .xlsx era apagado antes do anexo.
            wdir = os.path.join(pasta_temp_local, f"uau_lote_wk_{uuid.uuid4().hex[:10]}")
            os.makedirs(wdir, exist_ok=True)
            pastas_temp_workbook.append(wdir)
            placeholder = os.path.join(wdir, "base.xlsx")
            ce_chave = _resolver_estoque_por_chave_lote(chave, caminhos_estoque or [], temporarios)
            caminho_xlsx_motor, _ = processar_e_gerar_excel(
                tmp_r,
                tmp_p,
                placeholder,
                gerar_aba_resumo_geral=False,
                gerar_aba_consolidado_estoque=False,
                caminho_estoque=ce_chave,
                progresso_cb=lambda payload, _ch=chave, _i=idx: _emitir_progresso(
                    status=str((payload or {}).get("status") or "processando"),
                    total_empreendimentos=len(chaves_ok),
                    concluidos=_i - 1,
                    empreendimento_atual=str(_ch),
                    item_atual_abas=str((payload or {}).get("item_atual_abas") or nome_aba_consolidado),
                    abas_item=(payload or {}).get("abas_item") or [nome_aba_consolidado],
                    itens_tempo=tempos_por_item,
                    tempo_decorrido_segundos=(payload or {}).get("tempo_decorrido_segundos") or max(0.0, time.perf_counter() - t0),
                    mensagem=(payload or {}).get("mensagem"),
                ),
            )
            _emitir_progresso(
                status="processando",
                total_empreendimentos=len(chaves_ok),
                concluidos=idx - 1,
                empreendimento_atual=str(chave),
                item_atual_abas="DADOS_RECEBER",
                abas_item=["DADOS_RECEBER"],
                itens_tempo=tempos_por_item,
                tempo_decorrido_segundos=max(0.0, time.perf_counter() - t0),
            )
            caminhos_motor_para_resumo.append(caminho_xlsx_motor)
            pares_motor_est.append((caminho_xlsx_motor, ce_chave))
            dfc_sigla = ler_dataframe_consolidado_de_xlsx_motor(caminho_xlsx_motor)
            consolidado_cache_por_path[caminho_xlsx_motor] = (
                dfc_sigla if dfc_sigla is not None else pd.DataFrame()
            )
            soma_inad = 0.0
            if dfc_sigla is not None and not dfc_sigla.empty and "Vl.Principal (Encargos)" in dfc_sigla.columns:
                try:
                    serie_inad = pd.to_numeric(
                        dfc_sigla["Vl.Principal (Encargos)"],
                        errors="coerce",
                    ).fillna(0.0)
                except Exception:
                    serie_inad = pd.to_numeric(
                        dfc_sigla["Vl.Principal (Encargos)"]
                        .astype(str)
                        .str.replace(".", "", regex=False)
                        .str.replace(",", ".", regex=False),
                        errors="coerce",
                    ).fillna(0.0)
                soma_inad = float(serie_inad.sum())
            consolidado_por_sigla.append((sigla, caminho_xlsx_motor, soma_inad))

            partes_dr.append(_ler_df_aba_xlsx_motor(caminho_xlsx_motor, "DADOS RECEBER", 7))
            _emitir_progresso(
                status="processando",
                total_empreendimentos=len(chaves_ok),
                concluidos=idx - 1,
                empreendimento_atual=str(chave),
                item_atual_abas="DADOS_RECEBIDOS",
                abas_item=["DADOS_RECEBIDOS"],
                itens_tempo=tempos_por_item,
                tempo_decorrido_segundos=max(0.0, time.perf_counter() - t0),
            )
            partes_dp.append(_ler_df_aba_xlsx_motor(caminho_xlsx_motor, "DADOS RECEBIDOS", 7))
            if df_criterios_ref is None or (getattr(df_criterios_ref, "empty", True)):
                dc = _ler_df_aba_xlsx_motor(caminho_xlsx_motor, "CRITERIOS ANALISES", 0)
                if dc is not None and not dc.empty:
                    df_criterios_ref = dc

            dur_item = max(0.0, time.perf_counter() - (inicio_item_ts or time.perf_counter()))
            tempos_base = [float(x.get("segundos", 0.0) or 0.0) for x in tempos_por_item if float(x.get("segundos", 0.0) or 0.0) > 0]
            media_ref = (sum(tempos_base) / len(tempos_base)) if tempos_base else dur_item
            registro_item = {
                "empreendimento": str(chave),
                "segundos": round(dur_item, 2),
                "cor": _cor_tempo(dur_item, media_ref),
                "abas": ["CONSOLIDADO", "DADOS RECEBER", "DADOS RECEBIDOS"],
            }
            tempos_por_item.append(registro_item)

            concluidos_report = idx if idx < len(chaves_ok) else (idx - 1)
            soma_real = sum(float(x.get("segundos", 0.0) or 0.0) for x in tempos_por_item)
            media_exec = soma_real / max(1, len(tempos_por_item))
            estim_total = max(soma_real + (len(chaves_ok) - len(tempos_por_item)) * media_exec, soma_real + 1.0)
            percentual_tempo = int(min(99, max(1, (soma_real / estim_total) * 100.0)))
            _emitir_progresso(
                status="processando",
                total_empreendimentos=len(chaves_ok),
                concluidos=concluidos_report,
                empreendimento_atual=str(chave),
                item_atual_abas="DADOS_RECEBIDOS",
                abas_item=["DADOS_RECEBIDOS"],
                itens_tempo=tempos_por_item,
                estimativa_total_segundos=round(float(estim_total), 2),
                tempo_decorrido_segundos=round(max(0.0, time.perf_counter() - t0), 2),
                percentual_tempo=percentual_tempo,
            )

        df_dr_u = _concat_dfs_vertical(partes_dr)
        df_dp_u = _concat_dfs_vertical(partes_dp)
        if df_criterios_ref is None:
            df_criterios_ref = pd.DataFrame()

        df_dr_sql = _ordenar_df_base_sql(
            _promover_primeira_linha_cabecalho_se_necessario(
                _df_sql_like(df_dr_u if not df_dr_u.empty else pd.DataFrame())
            ),
            ("EMP_OBRA", "EMPREENDIMENTO", "VENDA", "CLIENTE", "IDENTIFICADOR_PRODUTO"),
        )
        df_dp_sql = _ordenar_df_base_sql(
            _promover_primeira_linha_cabecalho_se_necessario(
                _df_sql_like(df_dp_u if not df_dp_u.empty else pd.DataFrame())
            ),
            ("EMP_OBRA", "EMPREENDIMENTO", "VENDA", "CLIENTE", "IDENTIFICADOR_PRODUTO"),
        )
        df_dr_sql = _padronizar_colunas_base_para_negocio(df_dr_sql, "RECEBER")
        df_dp_sql = _padronizar_colunas_base_para_negocio(df_dp_sql, "RECEBIDOS")

        partes_c = []
        for pth in caminhos_motor_para_resumo:
            dfc = consolidado_cache_por_path.get(pth)
            if dfc is not None and not dfc.empty:
                partes_c.append(dfc)
        df_resumo_lote = pd.DataFrame()
        titulo_rg: str | None = None
        if partes_c:
            df_resumo_lote = montar_dataframe_resumo_geral(pd.concat(partes_c, ignore_index=True))
        if not df_resumo_lote.empty:
            tmp_resumo = os.path.join(pasta_temp_local, f"uau_resumo_lote_{uuid.uuid4().hex}.xlsx")
            try:
                with pd.ExcelWriter(tmp_resumo, engine="openpyxl") as wr:
                    df_resumo_lote.to_excel(
                        wr, sheet_name=NOME_ABA_RESUMO_GERAL, index=False, startrow=7
                    )
                db0 = _data_base_de_primeiro_xlsx_motor(caminhos_motor_para_resumo[0])
                aplicar_estilo_arquivo_so_aba_resumo_geral(
                    tmp_resumo,
                    db0,
                    "LOTE — TODOS OS EMPREENDIMENTOS",
                )
                wb_r = load_workbook(tmp_resumo)
                try:
                    titulo_rg = _titulo_aba_unico(
                        wb_exec, (NOME_ABA_RESUMO_GERAL or "RESUMO")[:31], titulos_exec
                    )
                    _copiar_planilha_estilizada(wb_r[NOME_ABA_RESUMO_GERAL], wb_exec, titulo_rg)
                finally:
                    wb_r.close()
            finally:
                _remover_seguro(tmp_resumo)

        ordem_consolidados = sorted(
            consolidado_por_sigla,
            key=lambda x: (-float(x[2]), str(x[0]).upper()),
        )
        for sigla_ord, caminho_xlsx_ord, _ in ordem_consolidados:
            _anexar_somente_consolidado_por_sigla(
                wb_exec,
                caminho_xlsx_ord,
                sigla_ord,
                titulos_exec,
            )

        if (
            NOME_ABA_RESUMO_GERAL in wb_exec.sheetnames
            and wb_exec.sheetnames[0] != NOME_ABA_RESUMO_GERAL
        ):
            try:
                idx_rg = wb_exec.sheetnames.index(NOME_ABA_RESUMO_GERAL)
                wb_exec.move_sheet(wb_exec[NOME_ABA_RESUMO_GERAL], offset=-idx_rg)
            except Exception:
                pass
        destino_exec = os.path.join(pasta_saida, "CARTEIRAS GERAL.xlsx")
        destino_base = os.path.join(pasta_saida, "CARTEIRAS BANCO DE DADOS.xlsx")
        wb_exec.save(destino_exec)

        # Lote: gera CONSOLIDADO ESTOQUE somente no arquivo final
        # (sem criar aba de estoque nos workbooks intermediários por empreendimento).
        if caminhos_estoque:
            from services.estoque_uau import (
                COLUNAS_SAIDA_CONSOLIDADO_ESTOQUE,
                CONSOLIDADO_ESTOQUE_PANDAS_STARTROW,
                NOME_ABA_CONSOLIDADO_ESTOQUE,
                calcular_indicadores_painel_consolidado_estoque,
                carregar_estoque_bruto,
                montar_dataframe_consolidado_estoque,
            )

            partes_consolidado_lote: List[pd.DataFrame] = []
            for pth in caminhos_motor_para_resumo:
                dfc = consolidado_cache_por_path.get(pth)
                if isinstance(dfc, pd.DataFrame) and not dfc.empty:
                    partes_consolidado_lote.append(dfc)
            df_consolidado_lote = (
                pd.concat(partes_consolidado_lote, ignore_index=True, sort=False)
                if partes_consolidado_lote
                else pd.DataFrame()
            )
            caminho_estoque_lote = _resolver_estoque_unificado_lote(
                caminhos_estoque or [], temporarios
            )
            df_estoque_in = carregar_estoque_bruto(caminho_estoque_lote or "")
            df_consolidado_estoque = montar_dataframe_consolidado_estoque(
                df_consolidado_lote, df_estoque_in
            )
            if df_consolidado_estoque.empty:
                df_consolidado_estoque = pd.DataFrame(
                    columns=COLUNAS_SAIDA_CONSOLIDADO_ESTOQUE
                )

            with pd.ExcelWriter(
                destino_exec,
                engine="openpyxl",
                mode="a",
                if_sheet_exists="replace",
            ) as wr_exec:
                df_consolidado_estoque.to_excel(
                    wr_exec,
                    sheet_name=NOME_ABA_CONSOLIDADO_ESTOQUE,
                    index=False,
                    startrow=CONSOLIDADO_ESTOQUE_PANDAS_STARTROW,
                )
            data_base_lote = (
                _data_base_de_primeiro_xlsx_motor(caminhos_motor_para_resumo[0])
                if caminhos_motor_para_resumo
                else None
            )
            indicadores_estoque = calcular_indicadores_painel_consolidado_estoque(
                df_consolidado_estoque
            )
            aplicar_estilo_arquivo_so_aba_consolidado_estoque(
                destino_exec,
                data_base_lote,
                "LOTE - TODOS OS EMPREENDIMENTOS",
                indicadores_painel=indicadores_estoque,
            )

        # Windows-safe: escreve primeiro em arquivo temporário e só então publica no destino final.
        tmp_base = os.path.join(pasta_temp_local, f"uau_base_sql_{uuid.uuid4().hex}.xlsx")
        with pd.ExcelWriter(tmp_base, engine="openpyxl") as wr_base:
            (df_dr_sql if not df_dr_sql.empty else pd.DataFrame()).to_excel(
                wr_base, sheet_name="DADOS_RECEBER", index=False
            )
            (df_dp_sql if not df_dp_sql.empty else pd.DataFrame()).to_excel(
                wr_base, sheet_name="DADOS_RECEBIDOS", index=False
            )
        _aplicar_schema_e_formato_base_final(tmp_base, df_dr_sql, df_dp_sql)
        try:
            shutil.copy2(tmp_base, destino_base)
        except OSError:
            # Fallback de publicação quando o caminho padrão está inválido/bloqueado no momento.
            destino_base = os.path.join(
                pasta_saida,
                f"CARTEIRAS BANCO DE DADOS - {datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            )
            shutil.copy2(tmp_base, destino_base)
        _emitir_progresso(
            status="concluido",
            total_empreendimentos=len(chaves_ok),
            concluidos=len(chaves_ok),
            empreendimento_atual=None,
            item_atual_abas="DOWNLOAD DISPONÍVEL",
            abas_item=["CARTEIRAS GERAL", "CARTEIRAS BANCO DE DADOS"],
            itens_tempo=tempos_por_item,
            tempo_decorrido_segundos=round(max(0.0, time.perf_counter() - t0), 2),
            estimativa_total_segundos=round(max(0.0, time.perf_counter() - t0), 2),
            percentual_tempo=100,
        )
        return (destino_exec, destino_base), time.perf_counter() - t0
    finally:
        for p in temporarios:
            _remover_seguro(p)
        for d in pastas_temp_workbook:
            try:
                shutil.rmtree(d, ignore_errors=True)
            except OSError:
                pass


def processar_entrada_simples_ou_lote(
    caminhos_receber: Sequence[str],
    caminhos_recebidos: Sequence[str],
    caminho_saida_base: str,
    modo_geracao: str | None,
    caminhos_estoque: Sequence[str] | None = None,
    progresso_cb=None,
) -> Tuple[str | Tuple[str, str], float]:
    """
    Compatível com fluxo antigo: 1+1 arquivos.
    Se um arquivo em cada lista e modo None ou vazio → delega ao processar_e_gerar_excel original.
    """
    cr = [p for p in caminhos_receber if p]
    cp = [p for p in caminhos_recebidos if p]
    modo = (modo_geracao or "").strip()
    if len(cr) == 1 and len(cp) == 1 and not modo:
        tmp_est_par: List[str] = []
        try:
            ce_par = _resolver_estoque_unificado_lote(caminhos_estoque or [], tmp_est_par)
            if callable(progresso_cb):
                try:
                    progresso_cb(
                        {
                            "status": "iniciado",
                            "total_empreendimentos": 1,
                            "concluidos": 0,
                            "empreendimento_atual": "PAR_UNICO",
                        }
                    )
                except Exception:
                    pass
            return processar_e_gerar_excel(
                cr[0], cp[0], caminho_saida_base, caminho_estoque=ce_par, progresso_cb=progresso_cb
            )
        finally:
            for p in tmp_est_par:
                _remover_seguro(p)
    if len(cr) == 1 and len(cp) == 1 and modo:
        # Um par explícito com modo: usa o fluxo de lote POR_EMPREENDIMENTO.
        return processar_lote_uau(
            cr,
            cp,
            caminho_saida_base,
            modo,
            caminhos_estoque=caminhos_estoque,
            progresso_cb=progresso_cb,
        )
    if len(cr) >= 1 and len(cp) >= 1:
        if not modo:
            raise ProcessamentoUAUErro(
                etapa="validação",
                funcao="processar_entrada_simples_ou_lote",
                validacao="modo de geração",
                mensagem="Com vários arquivos, selecione o modo: Por empreendimento.",
                campo_ou_aba="modo_geracao",
            )
        return processar_lote_uau(
            cr,
            cp,
            caminho_saida_base,
            modo,
            caminhos_estoque=caminhos_estoque,
            progresso_cb=progresso_cb,
        )
    raise ProcessamentoUAUErro(
        etapa="validação",
        funcao="processar_entrada_simples_ou_lote",
        validacao="arquivos",
        mensagem="Envie os arquivos TXT necessários.",
        campo_ou_aba="upload",
    )
