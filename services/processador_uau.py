import os
import re
import time
import unicodedata
import filecmp
from datetime import datetime
from collections import Counter, defaultdict

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import quote_sheetname

from services.etl_preprocessamento_uau import (
    obter_etl_stats_acumulado,
    preprocessar_texto_uau_bruto,
    reset_etl_stats_acumulado,
)
from services.auditoria_confianca_uau import (
    auditoria_integridade_financeira_obrigatoria,
    calcular_confianca_final_por_venda,
    classificar_alertas_confiabilidade,
    coletar_alertas_cliente_base,
    coletar_alertas_conflito_duplicidade_flag,
    coletar_alertas_grupos_deduplicacao,
    montar_alertas_etl_de_metricas,
    resumo_confianca_executivo,
)
from services.auditoria_parcelas_uau import (
    auditoria_alertas_qtd_parcelas_consolidado,
    auditoria_sequencia_parcelas_receber,
    ajustar_scores_com_alertas_tot,
    calcular_score_qualidade_parcelas_por_venda,
)
from services.auditoria_executiva_uau import (
    formatar_resumo_auditoria_para_log,
    gerar_resumo_auditoria_consolidado,
)

# Auditoria detalhada da validação pré-exportação (parcelas, subgrupos, divergências).
# False reduz drasticamente I/O no terminal e acelera a execução; regras e resultados não mudam.
DEBUG_VALIDACAO = False

# Aba Pendencias_Parcelas (referência Excel 1-based, após insert_rows em aplicar_estilo_excel):
# linhas 1–4 = resumo visual; linha 5 = cabeçalho tabular; linha 6+ = dados (pandas: header=4).
PENDENCIAS_PARCELAS_HEADER_ROW = 5
PENDENCIAS_PARCELAS_DATA_START_ROW = 6

# Dump de amostras/to_dict/listas grandes no terminal (principalmente montar_consolidado).
# False evita serialização de DataFrames e deixa o fluxo muito mais rápido.
DEBUG_DADOS = False

# Contrato funcional oficial de Vl.Carteira:
# - "SALDO_ABERTO": Encargos + Vl.Vencer
# - "POSICAO_TOTAL": Vl.Pago + Encargos + Vl.Vencer
# Trocar apenas aqui para evitar ambiguidade de semântica no motor e no BI.
CARTEIRA_MODO_OFICIAL = "POSICAO_TOTAL"

# Aba executiva agregada a partir do Consolidado (sem nova regra financeira por linha).
NOME_ABA_RESUMO_GERAL = "RESUMO GERAL"

# Mapeamento opcional de links da pasta/documentação por empreendimento.
# Chave: SIGLA (ex.: ALVLT) OU EMP/OBRA normalizado OU nome do empreendimento normalizado.
# Valor: URL completa (https://...).
MAPA_LINKS_DRIVE_EMPREENDIMENTO = {
    "NVLOT": "",
    "LTMAG": "https://drive.google.com/drive/folders/1Jv7nUHigYLQyvDx_RsoBt1zX4R_adlrB",
    "SCPTO": "https://drive.google.com/drive/folders/1nOMBP8udfjtEokkHWub9G_5nrjwIoizs",
    "SCPTI": "https://drive.google.com/drive/folders/1iDISqG4mQLIfcUrf3ZBRumJLgynI4T3a",
    "CIDAN": "https://drive.google.com/drive/folders/1U_R0vZCtBUnzRrZmThq1sBA5KGB_usW3",
    "VROLT": "https://drive.google.com/drive/folders/1KphscK3Kv3tp_HC-FzVvjLgLDaL_Z4kB",
    "ALVLT": "https://drive.google.com/drive/folders/1cPDgewWBUiWzRl_l5g0zHMKnIIXuoDYs",
    "LTMON": "https://drive.google.com/drive/folders/1s8xmoz8G9ByTOqCQzcOEhN-LVfst0APH",
    "RVERD": "",
    "LTVIL": "https://drive.google.com/drive/folders/1csTlqosfRlf6W1-jZRfgxRq1sSt_f5Ik",
    "LTMIN": "https://drive.google.com/drive/folders/1ZC2UAr5Cj8VIwB9MQ1le4OfObdEDUPhe",
    "SCPGO": "",
    "ARAHF": "https://drive.google.com/drive/folders/1AcP51anpZX8l3WwnTkKYZgOU7JwGz-Es",
    "BVGWH": "https://drive.google.com/drive/folders/1M8FaqJM5ps405xB9pKrOskzsWiahXLXb",
    "MANHA": "",
    "MONTB": "",
    "LIFE": "",
}

# De-para oficial de empreendimento (fonte única para exibição no topo e colunas).
# Chave: sigla normalizada (ex.: ALVLT). Valor: nome oficial de exibição.
MAPA_EMPREENDIMENTO_OFICIAL_POR_SIGLA = {
    "NVLOT": "LOT.RES.NILSON VELOSO",
    "LTMAG": "LOT.RES.MAGALHAES",
    "SCPTO": "LOT.RES.VALE DO TOCANTINS",
    "SCPTI": "LOT.RES.TIRADENTES",
    "CIDAN": "LOT.RES.CIDADE NOVA",
    "VROLT": "LOT.RES.VALE DAS ROSAS",
    "ALVLT": "LOT.RES.ALVORADA",
    "LTMON": "LOT.RES.MONTE NEGRO",
    "RVERD": "LOT.RES.IO VERDE",
    "LTVIL": "LOT.RES.VILA NOVA",
    "LTMIN": "LOT.RES.BAIRRO DOS MINERIOS",
    "SCPGO": "LOT.RES.GOIANIA",
    "ARAHF": "INC.RES.ARARAS",
    "BVGWH": "INC.COND.BELLA WHITE",
    "MANHA": "INC.MANHATAN",
    "MONTB": "INC.MONTBLANC",
    "LIFE": "INC.LIFE",
}


class ProcessamentoUAUErro(Exception):
    """
    Exceção de processamento com contexto auditável (validação, exportação, leitura).
    `contexto`: dict opcional com Venda, Cliente, valores esperado/encontrado, regra, causa raiz, etc.
    """

    def __init__(
        self,
        etapa,
        funcao,
        validacao,
        mensagem,
        campo_ou_aba="",
        erro_tecnico=None,
        contexto=None,
    ):
        self.etapa = str(etapa or "").strip()
        self.funcao = str(funcao or "").strip()
        self.validacao = str(validacao or "").strip()
        self.campo_ou_aba = str(campo_ou_aba or "").strip()
        self.mensagem = str(mensagem or "").strip()
        self.erro_tecnico = erro_tecnico
        self.contexto = dict(contexto) if contexto else {}

        partes = [
            f"Erro na etapa: {self.etapa}",
            f"Função: {self.funcao}",
            f"Validação: {self.validacao}",
        ]
        if self.campo_ou_aba:
            partes.append(f"Campo/Aba: {self.campo_ou_aba}")
        partes.append(f"Mensagem: {self.mensagem}")
        if erro_tecnico is not None:
            partes.append(f"Erro técnico: {type(erro_tecnico).__name__}: {erro_tecnico}")
        super().__init__(" | ".join(partes))

    def formatar_relatorio_completo(self) -> str:
        """Formato padrão para log e auditoria (multi-linha)."""
        linhas = [
            "========== ERRO PROCESSAMENTO UAU ==========",
            f"Etapa: {self.etapa}",
            f"Função: {self.funcao}",
            f"Validação: {self.validacao}",
            f"Campo/Aba: {self.campo_ou_aba or '(não especificado)'}",
            "",
            f"Venda: {self.contexto.get('Venda', '')}",
            f"Cliente: {self.contexto.get('Cliente', '')}",
            f"Identificador: {self.contexto.get('Identificador', '')}",
            "",
            f"Valor encontrado: {self.contexto.get('Valor_Encontrado', '')}",
            f"Valor esperado: {self.contexto.get('Valor_Esperado', '')}",
            f"Diferença: {self.contexto.get('Diferenca', '')}",
            "",
            f"Regra violada: {self.contexto.get('Regra_Violada', '')}",
            f"Possível causa raiz: {self.contexto.get('Causa_Raiz', '')}",
            "",
            "Detalhe técnico:",
        ]
        if self.erro_tecnico is not None:
            linhas.append(f"  {type(self.erro_tecnico).__name__}: {self.erro_tecnico}")
        else:
            linhas.append("  sem exceção Python associada")
        aux = self.contexto.get("Dados_Auxiliares", "")
        if aux:
            linhas.extend(["", "Dados auxiliares (resumo):", str(aux)])
        linhas.extend(["", "Mensagem:", self.mensagem, "============================================"])
        return "\n".join(linhas)


MESES_PT = {
    "janeiro": 1,
    "fevereiro": 2,
    "março": 3,
    "marco": 3,
    "abril": 4,
    "maio": 5,
    "junho": 6,
    "julho": 7,
    "agosto": 8,
    "setembro": 9,
    "outubro": 10,
    "novembro": 11,
    "dezembro": 12,
}

CLIENTES_NAO_APORTE = [
    "ARCIDIO",
]

DATE_RE = re.compile(r"^\d{2}/\d{2}/\d{4}$")
EMP_RE = re.compile(r"^\d+/\S+")

STOPWORDS_NOME = {
    "DA", "DE", "DI", "DO", "DU", "DAS", "DOS", "E"
}

TERMOS_CONTAMINACAO = [
    "SPE RESIDENCIAL",
    "SEGUNDA-FEIRA",
    "TERCA-FEIRA",
    "QUARTA-FEIRA",
    "QUINTA-FEIRA",
    "SEXTA-FEIRA",
    "SABADO",
    "DOMINGO",
    "PERIODO POR RECEBIMENTO",
    "PERIODO POR VENCIMENTO",
    "UAU",
    "PAGINA",
    "RELATORIO",
    "TOTAL CLIENTE",
    "TOTAL POR CLIENTE",
]

NOMES_APORTE_EXATOS = {
    "SICOOB",
    "CAIXA ECONOMICA FEDERAL",
    "BANCO DO BRASIL",
    "ITAU",
    "BRADESCO",
    "SANTANDER",
    "CEF",
}

TERMOS_APORTE = [
    "CAIXA",
    "ECONOMICA",
    "CEF",
    "BANCO",
    "ITAU",
    "BRADESCO",
    "SANTANDER",
    "SICOOB",
    "APORTE",
    "REPASSE",
    "FINANCIAMENTO",
]


# =========================
# FUNÇÕES BÁSICAS
# =========================
def extrair_data_base(caminho_arquivo, texto_pre_lido=None):
    texto = texto_pre_lido if texto_pre_lido is not None else ler_texto_robusto(caminho_arquivo)

    padrao = r"(\d{1,2}) de ([A-Za-zçÇãÃáÁéÉêÊíÍóÓôÔõÕúÚ]+) de (\d{4})"
    achado = re.search(padrao, texto, flags=re.IGNORECASE)

    if not achado:
        return None

    dia = int(achado.group(1))
    mes_nome = achado.group(2).strip().lower()
    ano = int(achado.group(3))
    mes = MESES_PT.get(mes_nome)

    if not mes:
        return None

    return datetime(ano, mes, dia)


def extrair_nome_empreendimento_txt(caminho_arquivo, texto_pre_lido=None):
    texto = texto_pre_lido if texto_pre_lido is not None else ler_texto_robusto(caminho_arquivo)
    linhas = [str(x or "").strip() for x in texto.splitlines()]

    termos_invalidos = [
        "UAU! SOFTWARE",
        "UAU! SOFTWARE LTDA",
        "CONTAS A RECEBER",
        "CONTAS RECEBIDAS",
        "PERÍODO POR VENCIMENTO",
        "PERIODO POR VENCIMENTO",
        "PERÍODO POR RECEBIMENTO",
        "PERIODO POR RECEBIMENTO",
        "PÁGINA",
        "PAGINA",
        "TOTAL CLIENTE",
        "TOTAL POR CLIENTE",
        "EMP/OBRA",
        "ANDREDUA",
        "IDENTIFICADOR QTDE",
        "NR PERSON. IDENTIFICADOR",
        "DESCRICAO NR PERSON",
        "DESCRIÇÃO NR PERSON",
    ]

    candidatos = []
    for linha in linhas:
        if not linha:
            continue
        if is_main_receber_line(linha) or is_main_recebidos_line(linha):
            continue

        parts = split_linha_tabular(linha)
        cod_prod = (parts[0] or "").strip() if parts else ""
        if len(parts) >= 2 and re.fullmatch(r"\d{2,4}", cod_prod):
            desc_tab = (parts[1] or "").strip()
            if desc_tab:
                linha_prod = limpar_texto_nome(desc_tab)
                if not linha_prod.startswith("PROD."):
                    if not any(t in linha_prod for t in termos_invalidos):
                        if "SPE" in linha_prod or "RESIDENCIAL" in linha_prod:
                            nome_limpo = limpar_nome_empreendimento(linha_prod)
                            if nome_limpo:
                                candidatos.append(nome_limpo)
            continue

        linha_up = limpar_texto_nome(linha)
        if linha_up.startswith("PROD."):
            continue
        if any(t in linha_up for t in termos_invalidos):
            continue
        if re.search(r"^\d{2}/\d{2}/\d{4}$", linha_up):
            continue
        if re.search(r"^\d+/\S+$", linha_up):
            continue
        if re.search(r"^\d+(\t|\s{2,})", linha):
            continue
        if "SPE" in linha_up or "RESIDENCIAL" in linha_up:
            nome_limpo = limpar_nome_empreendimento(linha_up)
            if nome_limpo:
                candidatos.append(nome_limpo)

    if candidatos:
        return escolher_moda_texto(candidatos)

    return ""


_RE_DIA_SEMANA_CABECALHO_UAU = re.compile(
    r"(segunda|ter[çc]a|quarta|quinta|sexta|s[áa]bado|domingo)\s*-\s*feira",
    flags=re.IGNORECASE,
)


def mapa_emp_obra_nome_legal_de_texto_receber_multibloco(texto_receber: str) -> dict:
    """
    TXT de Contas a Receber com vários relatórios concatenados (ex.: lote unificado):
    percorre em ordem e associa cada Emp/Obra de linha tabular ao último cabeçalho de
    empresa (linha típica: RAZÃO SOCIAL <tab> Segunda-feira, ...).
    """
    if texto_receber is None:
        return {}
    texto = str(texto_receber)
    if not texto.strip():
        return {}
    out: dict[str, str] = {}
    ultimo_legal = ""
    for linha in texto.splitlines():
        s = str(linha or "").rstrip("\r\n")
        st = s.strip()
        if not st:
            continue
        if "\t" in s and _RE_DIA_SEMANA_CABECALHO_UAU.search(s):
            left = s.split("\t", 1)[0].strip()
            cand = limpar_texto_nome(left)
            if not cand or len(cand) < 4:
                continue
            up = cand.upper()
            if "UAU" in up and "SOFTWARE" in up:
                continue
            nome_l = limpar_nome_empreendimento(cand)
            if nome_l:
                ultimo_legal = nome_l
        if is_main_receber_line(st):
            parts = split_linha_tabular(st)
            if len(parts) < 2:
                continue
            eo = normalizar_emp_obra(get_coluna(parts, 0))
            if eo and ultimo_legal:
                out[eo] = ultimo_legal
    return out


def extrair_nome_empreendimento_nome_arquivo(caminho_arquivo):
    """
    Fallback pelo nome do arquivo de upload UAU.
    Ex.: SCPGO_-LOT.SCP_RESIDENCIAL_GOIANIA_-_RECEBER.txt -> SCP RESIDENCIAL GOIANIA
    """
    try:
        nome = os.path.basename(str(caminho_arquivo or "")).strip()
    except Exception:
        nome = ""
    if not nome:
        return ""
    nome_up = nome.upper()
    m = re.search(r"-_?LOT\.([A-Z0-9_\.]+?)_-_", nome_up)
    if not m:
        return ""
    bruto = m.group(1).replace(".", " ").replace("_", " ").strip()
    return limpar_nome_empreendimento(bruto)


def limpar_nome_empreendimento(txt):
    txt = limpar_texto_nome(txt)
    txt = re.sub(
        r"\b(SEGUNDA-FEIRA|TERCA-FEIRA|TERÇA-FEIRA|QUARTA-FEIRA|QUINTA-FEIRA|SEXTA-FEIRA|SABADO|SÁBADO|DOMINGO)\b[, ]*",
        " ",
        txt,
        flags=re.IGNORECASE,
    )
    txt = re.sub(
        r"\b\d{1,2}\s+DE\s+[A-ZÇÃÁÉÊÍÓÔÕÚ]+\s+DE\s+\d{4}\b",
        " ",
        txt,
        flags=re.IGNORECASE,
    )
    # Remove sufixos comuns de cabecalho apos o nome da SPE.
    txt = re.sub(
        r"\b(PERIODO|PERÍODO|CONTAS A RECEBER|CONTAS RECEBIDAS|PAGINA|PÁGINA|UAU! SOFTWARE)\b.*$",
        "",
        txt,
        flags=re.IGNORECASE,
    )
    # Mantem apenas o bloco principal ate o sufixo juridico, quando existir.
    txt = re.sub(r"^(.*?\b(LTDA|S/A|EIRELI)\b).*$", r"\1", txt, flags=re.IGNORECASE)
    txt = re.sub(r"\s+", " ", txt).strip(" ,;-")
    return txt


def ler_texto_robusto(caminho_arquivo):
    # Tenta encodings comuns de exportacao TXT do UAU e afins.
    encodings = ["utf-8-sig", "utf-8", "cp1252", "latin-1"]

    for enc in encodings:
        try:
            with open(caminho_arquivo, "r", encoding=enc, errors="strict") as f:
                return preprocessar_texto_uau_bruto(f.read())
        except UnicodeDecodeError:
            continue
        except Exception:
            break

    # Fallback final para evitar quebra por caracteres isolados corrompidos.
    with open(caminho_arquivo, "r", encoding="latin-1", errors="ignore") as f:
        texto = f.read()
    return preprocessar_texto_uau_bruto(texto)


def split_linha_tabular(line):
    raw = str(line or "").replace("\ufeff", "").rstrip("\r\n")
    if not raw:
        return []

    if "\t" in raw:
        return [p.strip() for p in raw.split("\t")]

    # Fallback para exportacoes com colunas desalinhadas por espacos.
    return [p.strip() for p in re.split(r"\s{2,}", raw)]


def get_coluna(partes, idx, padrao=""):
    if idx < 0 or idx >= len(partes):
        return padrao
    return str(partes[idx]).strip()


def converter_valor(valor):
    if pd.isna(valor):
        return 0.0

    valor = str(valor).strip()
    if not valor:
        return 0.0

    negativo = valor.startswith("(") and valor.endswith(")")
    valor = valor.replace("(", "").replace(")", "")
    valor = valor.replace("R$", "").replace(" ", "")
    valor = valor.replace(".", "").replace(",", ".")

    try:
        numero = float(valor)
        return -numero if negativo else numero
    except Exception:
        return 0.0


def normalizar_parcela(parcela):
    """
    Canoniza rótulo de parcela (n/total) para deduplicação, contagens e universo distinto.
    Remove ruídos (PARC, PARCELA, etc.), unifica separadores e elimina zeros à esquerda no numerador/denominador.
    """
    s = str(parcela or "").strip().upper()
    if not s:
        return ""

    s = s.replace("\\", "/")
    s = s.replace("-", "/")
    s = re.sub(r"\b(PARCELA|PARCELAS|PARC|PCL|P)\b", "", s)
    s = re.sub(r"\s+", "", s)

    m = re.search(r"(\d{1,4})/(\d{1,4})", s)
    if not m:
        return ""

    atual = int(m.group(1))
    total = int(m.group(2))

    if atual <= 0 or total <= 0:
        return ""

    return f"{atual}/{total}"


def extrair_numero_parcela(parcela):
    s = normalizar_parcela(parcela)
    if not s:
        return 0
    try:
        return int(s.split("/")[0])
    except Exception:
        return 0


def extrair_total_parcela(parcela):
    s = normalizar_parcela(parcela)
    if not s:
        return 0
    try:
        return int(s.split("/")[1])
    except Exception:
        return 0


def calcular_qtd_parc_total(df_venda, col_parcela_total="Parc_Total"):
    """
    Denominador contratual dominante na venda: frequência de Parc_Total nas linhas;
    empate → maior denominador. Coluna padrão Parc_Total (equivalente a parcela_total).
    """
    if df_venda is None or len(df_venda) == 0:
        return 0
    if col_parcela_total not in df_venda.columns:
        return 0
    denominadores = pd.to_numeric(df_venda[col_parcela_total], errors="coerce").dropna()
    if denominadores.empty:
        return 0
    denominadores = denominadores.astype(int)
    denominadores = denominadores[denominadores > 0]
    if denominadores.empty:
        return 0
    # Ruído de entrada: se existir contrato “grande” (>=10), ignora denominadores <10 na moda.
    if denominadores.ge(10).any():
        denominadores = denominadores[denominadores >= 10]
        if denominadores.empty:
            return 0
    freq = denominadores.value_counts()
    mx = int(freq.max())
    maiores = [int(x) for x in freq[freq == mx].index.tolist()]
    return int(max(maiores))


def _mapa_qtd_parc_total_por_venda(df):
    """Mapa Venda -> denominador dominante (mesma regra de calcular_qtd_parc_total por grupo)."""
    out = {}
    if df is None or df.empty or "Venda" not in df.columns or "Parc_Total" not in df.columns:
        return out
    vk = df["Venda"].fillna("").astype(str).str.strip()
    for v_key, g in df.groupby(vk, sort=False):
        vs = str(v_key).strip()
        if vs:
            out[vs] = calcular_qtd_parc_total(g)
    return out


def ajustar_total_confiavel_global(total_base, pago_base, atr_base, av_base, tem_sobreposicao):
    """
    Faixa de coerência para Qtd.Parc.Total no escopo GLOBAL (universo homogêneo), alinhada ao
    consolidado (denominador dominante por Parc_Total) sem igualdade rígida a mapa_total_confiavel.

    Retorna (ref_min, ref_max):
    - ref_min: piso operacional (componentes; se universos disjuntos, também soma pago+atr+av).
    - ref_max: teto estrutural max(mapa_total_confiavel, ref_min), evita total inflado acima do mapa.
    """
    pb = int(pago_base or 0)
    ab = int(atr_base or 0)
    vb = int(av_base or 0)
    ref_min = max(pb, ab, vb)
    if not tem_sobreposicao:
        ref_min = max(ref_min, pb + ab + vb)
    tb = int(total_base or 0)
    ref_max = max(tb, ref_min)
    return ref_min, ref_max


def _mapa_universo_parcelas_distintas_por_venda(df):
    """Venda (str) -> contagem de parcelas canônicas distintas."""
    out = {}
    if df is None or df.empty or "Venda" not in df.columns or "Parcela" not in df.columns:
        return out
    vk = df["Venda"].fillna("").astype(str).str.strip()
    for v_key, gv in df.groupby(vk, sort=False):
        vs = str(v_key).strip()
        if not vs:
            continue
        out[vs] = int(contar_parcelas_distintas_padrao(gv["Parcela"]))
    return out


def _receber_confiavel_qtd_parc_total(qr, qp, ur, up, piso_M):
    """
    Critério explícito de confiança do Receber para o denominador Qtd.Parc.Total.
    qr, qp: modas QtdTotal (Parc_Total) por venda; ur, up: universos distintos de parcelas;
    piso_M: max(Paga, Atrasada, A Vencer) na venda.
    Retorna (confiável, lista de motivos técnicos quando não confiável).
    """
    motivos = []
    qr = int(qr or 0)
    qp = int(qp or 0)
    ur = int(ur or 0)
    up = int(up or 0)
    piso_M = int(piso_M or 0)

    if qr <= 0 and ur <= 0:
        return True, []

    if qr > 0 and piso_M > 0 and qr < piso_M:
        motivos.append("QtdTotal_Receber<piso_operacional")

    limite_gap = max(8, up // 4) if up else 0
    if up >= 8 and ur > 0 and ur < up and (up - ur) >= limite_gap:
        motivos.append("uni_Receber inferior a uni_Recebidos")

    if qr > 0 and qp > 0 and (qp - qr) >= 15 and piso_M >= max(20, int(qr * 1.5)):
        motivos.append("moda_Recebidos>>moda_Receber com trilha operacional")

    return (len(motivos) == 0), motivos


def normalizar_emp_obra(txt):
    txt = str(txt or "").strip().upper()
    txt = re.sub(r"\s+", "", txt)
    txt = txt.replace("SCPG O", "SCPGO")
    txt = txt.replace("SCPG0", "SCPGO")
    if txt == "69/LTMO":
        txt = "69/LTMON"
    txt = txt.replace("51/BVGW", "51/BVGWH")
    txt = txt.replace("/BVGW", "/BVGWH")
    # Bella White: colapsa 51/BVGWH, 51/BVGWHHH, 51/BVGWHHHH etc. → 51/BVGWH (exibição/chave única).
    m_bb = re.match(r"^(\d+/)BVGWH+$", txt)
    if m_bb:
        txt = f"{m_bb.group(1)}BVGWH"
    # Segurança extra: qualquer variação com sufixo BVGW(H...) vira canônico 51/BVGWH.
    if re.search(r"(^|/)BVGW(H+)?$", txt):
        txt = "51/BVGWH"
    if txt == "27/SCPG":
        txt = "27/SCPGO"
    return txt


def empreendimento_oficial_para_emp_obra(emp_obra: str) -> str:
    """Nome de exibição oficial a partir de Emp/Obra normalizado (mapa homologado por sigla)."""
    eo = normalizar_emp_obra(emp_obra)
    if not eo or "/" not in eo:
        return ""
    sigla = eo.split("/")[-1].strip()
    return str(MAPA_EMPREENDIMENTO_OFICIAL_POR_SIGLA.get(sigla, "") or "").strip()


def _aplicar_nome_oficial_em_series(emp_obra_series: pd.Series, empreendimento_series: pd.Series) -> pd.Series:
    """Padroniza nome de empreendimento pela sigla oficial quando conhecida."""
    eo_norm = emp_obra_series.fillna("").astype(str).apply(normalizar_emp_obra)
    emp_cur = empreendimento_series.fillna("").astype(str).apply(limpar_nome_empreendimento).astype(str).str.strip()
    emp_of = eo_norm.map(empreendimento_oficial_para_emp_obra).fillna("").astype(str).str.strip()
    m_of = emp_of.ne("")
    emp_cur.loc[m_of] = emp_of.loc[m_of]
    return emp_cur


def normalizar_identificador(texto):
    texto = str(texto or "").strip().upper()
    texto = texto.replace("\xa0", " ")
    # Ruído OCR/TXT: recompõe palavras-chave de endereço fragmentadas (ex.: "L OTE" -> "LOTE").
    for token in ("QUADRA", "LOTE", "UNIDADE", "APTO", "APT", "BLOCO", "TORRE", "CASA"):
        pad = r"".join([re.escape(ch) + r"\s*" for ch in token])
        texto = re.sub(rf"\b{pad}\b", token, texto)
    texto = re.sub(r"\s+", " ", texto)
    texto = texto.replace("| ", "|").replace(" |", "|")
    texto = texto.replace("QUADR A", "QUADRA")
    texto = texto.replace("QDRA", "QUADRA")
    texto = texto.replace("QD ", "QUADRA ")
    texto = texto.replace("LOT E", "LOTE")
    texto = re.sub(r"\bL\s+OTE\b", "LOTE", texto)
    texto = re.sub(r"\bQ\s+UADRA\b", "QUADRA", texto)
    texto = re.sub(r"\b(QUADRA|LOTE|UNIDADE|APTO|APT|BLOCO|TORRE|CASA)(\d+[A-Z]?)\b", r"\1 \2", texto)
    texto = texto.replace("/ LOTE", "/LOTE")
    texto = texto.replace(" /LOTE", "/LOTE")
    texto = texto.replace("/ LOTE ", "/LOTE ")
    texto = texto.replace(" / ", "/")
    texto = texto.replace(" /", "/")
    texto = texto.replace("/ ", "/")
    texto = re.sub(r"\s+", " ", texto)
    texto = _limpar_sufixos_operacionais_identificador(texto)
    return texto.strip(" |")


def _limpar_sufixos_operacionais_identificador(texto):
    texto = str(texto or "").strip()
    if not texto:
        return ""

    def _limpar_parte(parte):
        parte = str(parte or "").strip()
        parte = re.sub(
            r"\s+(?:"
            r"CUSTAS?\b|"
            r"FINANC(?:IAMENTO)?\s+BANCARIO\b|"
            r"TX\s+EVOLU.*|"
            r"TAXA\s+EVOLU.*|"
            r"PARCELA\b.*|"
            r"ENTRADA\b.*|"
            r"SINAL\b.*|"
            r"INTERMED(?:IARIA)?\b.*|"
            r"BAL[AÃ]O\b.*"
            r").*$",
            "",
            parte,
            flags=re.IGNORECASE,
        )
        if re.fullmatch(r"QUADR|QUADRA|LOTE|QUADRA/LOTE", parte, flags=re.IGNORECASE):
            return ""
        return parte.strip()

    partes = [_limpar_parte(p) for p in texto.split("|")]
    return "|".join(p for p in partes if p).strip()


def limpar_texto_nome(txt):
    txt = str(txt or "").strip().upper()
    txt = txt.replace("\xa0", " ")
    txt = re.sub(r"\s+", " ", txt)
    return txt


def texto_contaminado(txt):
    txt = limpar_texto_nome(txt)
    if not txt:
        return False
    return any(t in txt for t in TERMOS_CONTAMINACAO)


def identificador_tem_formato_endereco(txt):
    txt = normalizar_identificador(txt)

    if not txt:
        return False

    if texto_contaminado(txt):
        return False

    padroes = [
        r"QUADRA\s*\w+.*LOTE\s*\w+",
        r"\bQD\b\s*\w+.*\bLT\b\s*\w+",
        r"\bLOTE\s*\w+",
        r"\bCASA\s*\w+",
        r"\bAPTO\s*\w+",
        r"\bAPT\s*\w+",
        r"\bBLOCO\s*\w+.*\bAPTO\s*\w+",
        r"\bUNIDADE\s*\w+",
        r"\bTORRE\s*\w+.*\bAPTO\s*\w+",
        r"\b\d+\s*-\s*TORRE\s*\w+",
        r"\b\d+\s*-\s*BLOCO\s*\w+",
    ]

    return any(re.search(p, txt, flags=re.IGNORECASE) for p in padroes)


def identificador_truncado(txt):
    txt = normalizar_identificador(txt)

    if txt in ["", "QUADRA", "LOTE", "QUADRA/LOTE", "NAN", "NONE", "-1", "1"]:
        return True

    if texto_contaminado(txt):
        return True

    if txt.endswith("|QUADR") or txt.endswith("|QUADRA"):
        return True

    if txt.endswith("/LOTE") or re.search(r"/LOTE\s*$", txt):
        return True

    if txt.startswith("QUADRA ") and "/LOTE" in txt and re.search(r"/LOTE\s*$", txt):
        return True

    if not identificador_tem_formato_endereco(txt):
        return True

    return False


def score_identificador(txt):
    txt = str(txt or "").strip().upper()
    if not txt:
        return 0

    score = 0
    if not identificador_truncado(txt):
        score += 1000

    score += len(txt)
    score += txt.count("|") * 20
    score += len(re.findall(r"LOTE\s*\d+", txt)) * 30
    score += len(re.findall(r"QUADRA\s*\d+[A-Z]?", txt)) * 20

    return score


def escolher_moda_texto(valores):
    valores_limpos = [str(v).strip() for v in valores if str(v).strip() != ""]
    if not valores_limpos:
        return ""
    return Counter(valores_limpos).most_common(1)[0][0]


def moda_valor_numerico_positivo(series):
    vals = [round(float(v), 2) for v in pd.to_numeric(series, errors="coerce").fillna(0).tolist() if float(v) > 0]
    if not vals:
        return 0.0
    freq = Counter(vals)
    return sorted(freq.keys(), key=lambda x: (freq[x], x), reverse=True)[0]


def moda_valor_parcela_por_df_ou_grupo(df_ou_grupo):
    """
    Regra madura do projeto: Valor Da Parcela = moda de Principal por venda.
    Aceita DataFrame/grupo com coluna Principal.
    """
    if df_ou_grupo is None or len(df_ou_grupo) == 0:
        return 0.0
    if "Principal" not in df_ou_grupo.columns:
        return 0.0
    return float(moda_valor_numerico_positivo(df_ou_grupo["Principal"]))


def contar_parcelas_distintas_padrao(series):
    vals = set()
    for v in series:
        n = normalizar_parcela(v)
        if n:
            vals.add(n)
    return len(vals)


def mapa_conjunto_parcelas_por_venda(df):
    """
    Por Venda, conjunto de parcelas canônicas (normalizar_parcela), para universo distinto na validação.
    """
    if df is None or df.empty or "Venda" not in df.columns or "Parcela" not in df.columns:
        return {}
    d = df.copy()
    d["Venda"] = d["Venda"].fillna("").astype(str).str.strip()
    d["Parcela_Norm"] = d["Parcela"].apply(normalizar_parcela)
    d = d.loc[d["Venda"] != ""]
    if d.empty:
        return {}
    out = {}
    for v, g in d.groupby("Venda", dropna=False):
        venda = str(v).strip()
        if not venda:
            continue
        out[venda] = {p for p in g["Parcela_Norm"].tolist() if p != ""}
    return out


def normalizar_tipo_base(tipo):
    """Normaliza rótulo de Tipo para chave de subuniverso contratual (sem alterar regras de negócio)."""
    t = str(tipo if tipo is not None else "").strip().upper()
    if not t:
        return ""
    t = " ".join(t.split())
    return re.sub(r"\s+", "_", t)


def chave_financeira_venda_cliente(row):
    """
    Chave mestre financeira: Venda + Cliente equiparado (Cliente_Base).
    Aceita Series (iterrows/apply) ou dict-like.
    """
    r = row if isinstance(row, pd.Series) else pd.Series(row)
    v = ""
    if "Venda" in r.index and pd.notna(r["Venda"]):
        v = str(r["Venda"]).strip()
    cb = ""
    if "Cliente_Base" in r.index and pd.notna(r["Cliente_Base"]):
        cb = str(r["Cliente_Base"]).strip()
    if not cb and "Cliente" in r.index and pd.notna(r["Cliente"]) and str(r["Cliente"]).strip():
        cb = gerar_cliente_base(limpar_texto_nome(str(r["Cliente"])))
    return f"{v}||{cb}"


def chave_subgrupo_contratual(row):
    """
    Subuniverso contratual: tipo_base ||| denominador_parcela ||| identificador_base.
    Separador ||| evita colisão com valores normalizados usuais.
    """
    r = row if isinstance(row, pd.Series) else pd.Series(row)
    tipo_raw = r["Tipo"] if "Tipo" in r.index else ""
    if tipo_raw is None or (isinstance(tipo_raw, float) and pd.isna(tipo_raw)):
        tipo_raw = ""
    tipo_b = normalizar_tipo_base(tipo_raw)
    parcela = r["Parcela"] if "Parcela" in r.index else ""
    if parcela is None or (isinstance(parcela, float) and pd.isna(parcela)):
        parcela = ""
    den = _denominador_parcela_audit(str(parcela))
    den_s = str(int(den)) if den > 0 else "0"
    vals_id = []
    for col in ("Unidades", "Identificador_Produto"):
        if col not in r.index:
            continue
        v = r[col]
        if v is None or (isinstance(v, float) and pd.isna(v)):
            continue
        if isinstance(v, (list, tuple)):
            for item in v:
                if item is None or str(item).strip() == "":
                    continue
                if col == "Identificador_Produto" and identificador_truncado(item):
                    continue
                vals_id.append(item)
            continue
        if str(v).strip() == "":
            continue
        if col == "Identificador_Produto" and identificador_truncado(v):
            continue
        vals_id.append(v)
    id_base = escolher_identificador_melhor(vals_id) if vals_id else ""
    if not id_base:
        id_base = "_SEM_ID"
    tb = tipo_b if tipo_b else "__SEM_TIPO__"
    return f"{tb}|||{den_s}|||{id_base}"


def metricas_universo_parcelas_operacionais(set_pag, set_venc, set_av):
    """
    União e interseções entre parcelas pagas (Recebidos), vencidas e a vencer (Receber).
    `tem_sobreposicao`: qualquer parcela contada em mais de um universo operacional.
    """
    sp = set(set_pag or set())
    sv = set(set_venc or set())
    sa = set(set_av or set())
    uni = sp | sv | sa
    ipv = sp & sv
    ipa = sp & sa
    iva = sv & sa
    return {
        "n_pag": len(sp),
        "n_venc": len(sv),
        "n_av": len(sa),
        "universo_total_distinto": len(uni),
        "intersec_pag_venc": ipv,
        "intersec_pag_av": ipa,
        "intersec_venc_av": iva,
        "tem_sobreposicao": bool(ipv or ipa or iva),
    }


def _amostra_lista_parcelas_norm(conjunto, lim=40):
    if not conjunto:
        return "[]"
    lst = sorted(conjunto)
    if len(lst) <= lim:
        return str(lst)
    return str(lst[:lim]) + f" ... (+{len(lst) - lim} parcelas)"


def _pct_texto_qtd_vs_base(encontrado, esperado):
    try:
        e = int(esperado)
        f = int(encontrado)
    except (TypeError, ValueError):
        return "N/A"
    if e == 0:
        return "N/A (base zero)"
    return f"{100.0 * abs(f - e) / float(abs(e)):.4f}%"


def _diagnostico_parcelas_venda(venda, df_receber, df_recebidos, df_consolidado):
    """
    Diagnóstico estrutural de parcelas por venda (somente leitura; mesma canonização da validação).
    Retorna conjuntos normalizados, interseções e contagens — sem alterar DataFrames.
    """
    v = str(venda or "").strip()
    vazio = {
        "venda": v,
        "parcelas_pagas_norm": set(),
        "parcelas_vencidas_norm": set(),
        "parcelas_a_vencer_norm": set(),
        "universo_distinto_union": set(),
        "intersec_pag_venc": set(),
        "intersec_pag_av": set(),
        "intersec_venc_av": set(),
        "tem_sobreposicao": False,
        "contagens": {},
        "consolidado_qtd": {},
        "inconsistencias": [],
    }
    if not v:
        return vazio

    set_pag = set()
    if df_recebidos is not None and not df_recebidos.empty:
        if "Venda" in df_recebidos.columns and "Parcela" in df_recebidos.columns:
            m_p = df_recebidos["Venda"].fillna("").astype(str).str.strip() == v
            if m_p.any():
                for x in df_recebidos.loc[m_p, "Parcela"].to_numpy():
                    n = normalizar_parcela(x)
                    if n:
                        set_pag.add(n)

    set_venc, set_av = set(), set()
    if df_receber is not None and not df_receber.empty:
        if "Venda" in df_receber.columns and "Parcela" in df_receber.columns:
            m_r = df_receber["Venda"].fillna("").astype(str).str.strip() == v
            if m_r.any() and "Status_Vencimento" in df_receber.columns:
                subp_r = df_receber.loc[m_r]
                st = subp_r["Status_Vencimento"].fillna("").astype(str).str.strip().str.upper()
                for px, stx in zip(subp_r["Parcela"].to_numpy(), st.to_numpy()):
                    n = normalizar_parcela(px)
                    if not n:
                        continue
                    if stx == "VENCIDO":
                        set_venc.add(n)
                    elif stx == "A VENCER":
                        set_av.add(n)

    muni = metricas_universo_parcelas_operacionais(set_pag, set_venc, set_av)
    uni = set_pag | set_venc | set_av

    inconsistencias = []
    if muni["intersec_venc_av"]:
        inconsistencias.append("MESMA_PARCELA_EM_VENCIDO_E_A_VENCER")
    if muni["intersec_pag_venc"]:
        inconsistencias.append("PARCELA_PAGA_E_VENCIDA_MESMA_REF")
    if muni["intersec_pag_av"]:
        inconsistencias.append("PARCELA_PAGA_E_A_VENCER_MESMA_REF")

    cons = {}
    if df_consolidado is not None and not df_consolidado.empty and "Venda" in df_consolidado.columns:
        m = df_consolidado["Venda"].fillna("").astype(str).str.strip() == v
        if m.any():
            r0 = df_consolidado.loc[m].iloc[0]
            cons = {
                "Qtd.Parc.Paga": int(float(r0.get("Qtd.Parc.Paga", 0) or 0)),
                "Qtd.Parc.Atrasada": int(float(r0.get("Qtd.Parc.Atrasada", 0) or 0)),
                "Qtd.Parc.A Vencer": int(float(r0.get("Qtd.Parc.A Vencer", 0) or 0)),
                "Qtd.Parc.Total": int(float(r0.get("Qtd.Parc.Total", 0) or 0)),
            }

    return {
        "venda": v,
        "parcelas_pagas_norm": set_pag,
        "parcelas_vencidas_norm": set_venc,
        "parcelas_a_vencer_norm": set_av,
        "universo_distinto_union": uni,
        "intersec_pag_venc": muni["intersec_pag_venc"],
        "intersec_pag_av": muni["intersec_pag_av"],
        "intersec_venc_av": muni["intersec_venc_av"],
        "tem_sobreposicao": muni["tem_sobreposicao"],
        "contagens": {
            "distintas_pagas_recebidos": len(set_pag),
            "distintas_vencidas_receber": len(set_venc),
            "distintas_a_vencer_receber": len(set_av),
            "universo_total_distinto": muni["universo_total_distinto"],
        },
        "consolidado_qtd": cons,
        "inconsistencias": inconsistencias,
    }


def _montar_erro_val_parcelas_qtd(
    titulo_curto,
    venda,
    nome_metrica,
    valor_consolidado,
    valor_base_validacao,
    origem_recalculo,
    row_consolidado,
    diag,
    regra,
    causa_raiz,
    extras_linhas=None,
):
    """Monta texto longo para lista `erros` + dict de contexto para ProcessamentoUAUErro."""
    extras_linhas = extras_linhas or []
    try:
        diff = int(valor_consolidado) - int(valor_base_validacao)
    except (TypeError, ValueError):
        diff = 0
    cliente = ""
    ident = ""
    if row_consolidado is not None:
        cliente = str(row_consolidado.get("Cliente", "") or "").strip()
        ident = str(row_consolidado.get("Identificador", "") or "").strip()
    cont = diag.get("contagens") or {}
    msg_lines = [
        f"VAL-PARCELAS [{titulo_curto}] venda={venda}",
        f"  Métrica: {nome_metrica}",
        f"  Valor consolidado (Consolidado Venda): {valor_consolidado}",
        f"  Valor recalculado na validação ({origem_recalculo}): {valor_base_validacao}",
        f"  Diferença absoluta: {diff}",
        f"  Diferença percentual (|diff|/|base|): {_pct_texto_qtd_vs_base(valor_consolidado, valor_base_validacao)}",
        f"  Parcelas canônicas em Dados Recebidos (pagas), amostra: {_amostra_lista_parcelas_norm(diag.get('parcelas_pagas_norm'))}",
        f"  Parcelas canônicas em Dados Receber VENCIDO, amostra: {_amostra_lista_parcelas_norm(diag.get('parcelas_vencidas_norm'))}",
        f"  Parcelas canônicas em Dados Receber A VENCER, amostra: {_amostra_lista_parcelas_norm(diag.get('parcelas_a_vencer_norm'))}",
        f"  Total parcelas distintas (união operacional): {cont.get('universo_total_distinto', '')}",
        f"  Sobreposição de universos: {diag.get('tem_sobreposicao')} | "
        f"n(pag∩venc)={len(diag.get('intersec_pag_venc') or set())} "
        f"n(pag∩av)={len(diag.get('intersec_pag_av') or set())} "
        f"n(venc∩av)={len(diag.get('intersec_venc_av') or set())}",
    ]
    inc = diag.get("inconsistencias") or []
    if inc:
        msg_lines.append(f"  Inconsistências estruturais detectadas: {', '.join(inc)}")
    msg_lines.extend(extras_linhas)
    msg_lines.append(f"  Regra violada: {regra}")
    msg_lines.append(f"  Possível causa raiz: {causa_raiz}")
    msg = "\n".join(msg_lines)
    aux = f"contagens={cont}; inconsistencias={inc}"
    ctx = {
        "Venda": venda,
        "Cliente": cliente,
        "Identificador": ident,
        "Valor_Encontrado": str(valor_consolidado),
        "Valor_Esperado": str(valor_base_validacao),
        "Diferenca": str(diff),
        "Regra_Violada": regra,
        "Causa_Raiz": causa_raiz,
        "Dados_Auxiliares": aux[:4000],
    }
    return msg, ctx


def moda_identificador_final_serie(series):
    vals = [str(x).strip() for x in series.tolist() if str(x).strip()]
    if not vals:
        return ""
    freq = Counter(vals)
    max_f = max(freq.values())
    cands = [k for k, v in freq.items() if v == max_f]
    return sorted(
        cands,
        key=lambda x: (-score_identificador(x), -len(str(x)), str(x)),
    )[0]


def _moda_identificador_grupo_venda_cliente(vals):
    """Moda entre identificadores já normalizados; empate → menor string canônica (ordem lexicográfica estável)."""
    vals = [str(x).strip() for x in vals if str(x).strip()]
    if not vals:
        return ""
    freq = Counter(vals)
    max_f = max(freq.values())
    cands = [k for k, v in freq.items() if v == max_f]
    return sorted(cands)[0]


def harmonizar_identificador_por_venda_cliente_dataframe(consolidado, alertas_out=None):
    """
    Por grupo (Venda, Cliente), unifica Identificador com a moda dos valores não vazios do próprio grupo.
    Empate de frequência: desempate lexicográfico estável. Não inventa valor se todos vazios.
    """
    if consolidado is None or consolidado.empty:
        return 0
    if "Identificador" not in consolidado.columns or "Cliente" not in consolidado.columns:
        return 0
    d = consolidado
    vk = d["Venda"].fillna("").astype(str).str.strip()
    ck = d["Cliente"].fillna("").astype(str).str.strip()
    work = d.assign(__vk=vk, __ck=ck)
    mud = 0
    grupos_conflito = 0
    grupos_com_vazio = 0
    for (_, _), g in work.groupby(["__vk", "__ck"], sort=False, dropna=False):
        idx = g.index
        antes = d.loc[idx, "Identificador"].fillna("").astype(str).str.strip()
        vals = [str(x).strip() for x in antes.tolist() if str(x).strip()]
        if not vals:
            continue
        w = _moda_identificador_grupo_venda_cliente(vals)
        if not w:
            continue
        if len(set(vals)) > 1:
            grupos_conflito += 1
        if (antes == "").any():
            grupos_com_vazio += 1
        alvo = d.loc[idx, "Identificador"].fillna("").astype(str).str.strip()
        mud += int((alvo != w).sum())
        d.loc[idx, "Identificador"] = w
    if alertas_out is not None and mud > 0:
        alertas_out.append({
            "Venda": "GERAL",
            "Cliente_Base": "",
            "Tipo_Alerta": "IDENTIFICADOR_HARMONIZACAO_VENDA_CLIENTE",
            "Mensagem": (
                f"Identificador alinhado por moda em grupos (Venda+Cliente): "
                f"grupos com mais de um identificador distinto={grupos_conflito}; "
                f"grupos com linhas vazias preenchidas por predominância={grupos_com_vazio}; "
                f"linhas atualizadas={mud}."
            ),
            "Divergencia": int(grupos_conflito + grupos_com_vazio),
            "Valor_Esperado": "",
            "Valor_Encontrado": str(int(mud)),
            "Regra": "Moda por Venda+Cliente; empate lexicográfico; sem invenção de texto",
            "Observacao": "Nao bloqueante",
        })
    return mud


def mapas_identificador_moda_de_bases_brutas(dfr, dfp):
    """
    Retorna (mapa_Venda+Cliente, mapa_somente_Venda) com modas de identificador canônico nas linhas brutas.
    O mapa por Venda é reserva estável quando o nome do cliente no consolidado não casa com o do TXT.
    """
    acc_vc = defaultdict(list)
    acc_v = defaultdict(list)
    for df in (dfr, dfp):
        if df is None or df.empty:
            continue
        if "Venda" not in df.columns or "Cliente" not in df.columns:
            continue
        v = df["Venda"].fillna("").astype(str).str.strip().reset_index(drop=True)
        c = df["Cliente"].fillna("").astype(str).str.strip().reset_index(drop=True)
        u = (
            df["Unidades"].fillna("").astype(str).reset_index(drop=True)
            if "Unidades" in df.columns
            else pd.Series([""] * len(df))
        )
        ip = (
            df["Identificador_Produto"].fillna("").astype(str).reset_index(drop=True)
            if "Identificador_Produto" in df.columns
            else pd.Series([""] * len(df))
        )
        for j in range(len(df)):
            vi = str(v.iloc[j]).strip()
            if not vi:
                continue
            ci = str(c.iloc[j]).strip()
            nid = ""
            for val in (u.iloc[j], ip.iloc[j]):
                if val is None or str(val).strip() == "":
                    continue
                if identificador_truncado(val):
                    continue
                n = normalizar_identificador(val)
                if n:
                    nid = n
                    break
            if nid:
                acc_v[vi].append(nid)
                if ci:
                    acc_vc[(vi, ci)].append(nid)
    d_vc = {k: _moda_identificador_grupo_venda_cliente(v) for k, v in acc_vc.items() if v}
    d_v = {k: _moda_identificador_grupo_venda_cliente(v) for k, v in acc_v.items() if v}
    return d_vc, d_v


def preencher_identificador_vazio_de_mapas_brutos(consolidado, mapa_vc, mapa_v, alertas_out=None):
    """
    Preenche Identificador vazio: 1) moda (Venda, Cliente) no bruto; 2) se vazio, moda só por Venda no bruto.
    """
    if consolidado is None or consolidado.empty:
        return 0, 0
    n_vc = 0
    n_v = 0
    for idx, row in consolidado.iterrows():
        cur = str(row.get("Identificador", "") or "").strip()
        if cur:
            continue
        v = str(row.get("Venda", "") or "").strip()
        cli = str(row.get("Cliente", "") or "").strip()
        w = ""
        if mapa_vc:
            w = mapa_vc.get((v, cli), "")
            if w:
                n_vc += 1
        if not w and mapa_v:
            w = mapa_v.get(v, "")
            if w:
                n_v += 1
        if w:
            consolidado.at[idx, "Identificador"] = w
    if alertas_out is not None:
        if n_vc > 0:
            alertas_out.append({
                "Venda": "GERAL",
                "Cliente_Base": "",
                "Tipo_Alerta": "IDENTIFICADOR_PREENCHIDO_VENDA_CLIENTE_BRUTO",
                "Mensagem": f"Identificador vazio preenchido pela moda (Venda+Cliente) nas bases brutas: {n_vc} linha(s).",
                "Divergencia": int(n_vc),
                "Valor_Esperado": "",
                "Valor_Encontrado": str(int(n_vc)),
                "Regra": "Moda nas linhas brutas por Venda+Cliente",
                "Observacao": "Nao bloqueante",
            })
        if n_v > 0:
            alertas_out.append({
                "Venda": "GERAL",
                "Cliente_Base": "",
                "Tipo_Alerta": "IDENTIFICADOR_PREENCHIDO_VENDA_BRUTO",
                "Mensagem": f"Identificador vazio preenchido pela moda somente por Venda nas bases brutas (fallback de casamento de nome): {n_v} linha(s).",
                "Divergencia": int(n_v),
                "Valor_Esperado": "",
                "Valor_Encontrado": str(int(n_v)),
                "Regra": "Moda nas linhas brutas por Venda quando Venda+Cliente não casou",
                "Observacao": "Nao bloqueante",
            })
    return n_vc, n_v


def mapa_vl_vencer_por_venda_receber_tratado(df_receber):
    """
    Soma numérica de Vlr_Parcela (coluna R) por Venda, somente Status_Vencimento = A VENCER.
    Espera a mesma base Receber já tratada pelo motor (deduplicação + classificação), para bater com o consolidado.
    """
    if df_receber is None or df_receber.empty or "Vlr_Parcela" not in df_receber.columns:
        return {}
    d = df_receber.copy()
    d["Venda"] = d["Venda"].fillna("").astype(str).str.strip()
    if "Status_Vencimento" not in d.columns:
        d["Status_Vencimento"] = ""
    st = d["Status_Vencimento"].astype(str).str.strip().str.upper()
    mask = st == "A VENCER"
    if not bool(mask.any()):
        return {}
    vp = pd.to_numeric(d["Vlr_Parcela"], errors="coerce").fillna(0)
    sub = d.loc[mask].copy()
    sub["_vp"] = vp.loc[mask]
    raw = sub.groupby("Venda")["_vp"].sum()
    return {str(k).strip(): float(v) for k, v in raw.items()}


def _denominador_parcela_audit(s):
    n = normalizar_parcela(s)
    if not n or "/" not in n:
        return 0
    try:
        den = int(n.split("/")[1])
        return den if den > 0 else 0
    except Exception:
        return 0


def escolher_identificador_melhor(valores):
    candidatos = [normalizar_identificador(v) for v in valores if str(v).strip() != ""]
    candidatos = [v for v in candidatos if v and not identificador_truncado(v)]

    if not candidatos:
        return ""

    contagem = Counter(candidatos)
    melhor = sorted(
        contagem.items(),
        key=lambda x: (x[1], score_identificador(x[0]), len(x[0])),
        reverse=True
    )[0][0]
    return melhor


def extrair_tokens_nome_relevantes(nome):
    nome = limpar_texto_nome(nome)
    if not nome:
        return []

    tokens = []
    for t in nome.split():
        t = t.strip()
        if not t:
            continue
        if t in STOPWORDS_NOME:
            continue
        tokens.append(t)

    return tokens


def gerar_cliente_base(nome):
    tokens = extrair_tokens_nome_relevantes(nome)

    if len(tokens) >= 2:
        return f"{tokens[0]} {tokens[1]}"
    if len(tokens) == 1:
        return tokens[0]
    return ""


def escolher_cliente_exibicao(series_clientes):
    valores = [limpar_texto_nome(v) for v in series_clientes if limpar_texto_nome(v) != ""]
    if not valores:
        return ""

    freq = Counter(valores)
    return sorted(freq.keys(), key=lambda x: (freq[x], len(x)), reverse=True)[0]


def adicionar_chave_cliente(df):
    if df is None:
        return pd.DataFrame(columns=["Venda", "Cliente", "Cliente_Norm", "Cliente_Base", "Chave_Cliente"])

    base = df.copy()

    if "Venda" not in base.columns:
        base["Venda"] = ""

    if "Cliente" not in base.columns:
        base["Cliente"] = ""

    base["Venda"] = base["Venda"].fillna("").astype(str).str.strip()
    base["Cliente"] = base["Cliente"].fillna("").astype(str).str.strip()

    base["Cliente_Norm"] = base["Cliente"].apply(limpar_texto_nome)
    base["Cliente_Base"] = base["Cliente_Norm"].apply(gerar_cliente_base)
    base["Chave_Cliente"] = (
        base["Venda"].astype(str).str.strip()
        + "||"
        + base["Cliente_Base"].astype(str).str.strip()
    )

    return base


def harmonizar_cliente_por_venda(df):
    if df is None or df.empty or "Venda" not in df.columns or "Cliente" not in df.columns:
        return df

    out = df.copy()
    out["Venda"] = out["Venda"].fillna("").astype(str).str.strip()
    out["Cliente"] = out["Cliente"].fillna("").astype(str).str.strip()

    mapa_cliente = {}
    for venda, grp in out.groupby("Venda", dropna=False):
        nomes = [limpar_texto_nome(x) for x in grp["Cliente"].tolist() if str(x).strip() != ""]
        if not nomes:
            continue
        freq = Counter(nomes)
        escolhido = sorted(freq.keys(), key=lambda x: (freq[x], len(x)), reverse=True)[0]
        mapa_cliente[str(venda).strip()] = escolhido

    if mapa_cliente:
        out["Cliente"] = out["Venda"].map(mapa_cliente).fillna(out["Cliente"])
    # Recalcula Cliente_Base apos unificar nome por venda (evita fragmentacao).
    return adicionar_chave_cliente(out)


def extrair_sigla_empreendimento(emp_obra):
    emp_obra = str(emp_obra or "").strip().upper()
    if "/" in emp_obra:
        return emp_obra.split("/")[-1]
    return emp_obra if emp_obra else "CONSOLIDADO"


def sanitizar_nome_arquivo(txt):
    txt = str(txt or "").strip().upper()
    txt = re.sub(r'[\\/:*?"<>|]', "", txt)
    txt = re.sub(r"\s+", " ", txt)
    return txt.strip()


def montar_nome_arquivo_empreendimento(df_consolidado, nome_empreendimento_canonico=None):
    if df_consolidado.empty:
        return "CONSOLIDADO_UAU.xlsx"

    emp_obra = ""
    empreendimento = ""

    if "Emp/Obra" in df_consolidado.columns and not df_consolidado["Emp/Obra"].dropna().empty:
        emp_obra = str(df_consolidado["Emp/Obra"].dropna().astype(str).iloc[0]).strip()

    canon = str(nome_empreendimento_canonico or "").strip()
    if canon:
        empreendimento = canon
    elif "Empreendimento" in df_consolidado.columns:
        descricoes = [
            str(v).strip()
            for v in df_consolidado["Empreendimento"].tolist()
            if str(v).strip() != ""
        ]
        empreendimento = escolher_moda_texto(descricoes) if descricoes else ""

    sigla = extrair_sigla_empreendimento(emp_obra)
    sigla = sanitizar_nome_arquivo(sigla if sigla else "CONSOLIDADO")
    empreendimento = sanitizar_nome_arquivo(empreendimento if empreendimento else "EMPREENDIMENTO")

    return f"{sigla} - {empreendimento}.xlsx"


def limpar_pasta_saida_excel_antigos(pasta_saida):
    """
    Remove arquivos Excel (.xlsx / .xlsm) existentes na pasta de destino antes de uma nova exportação.
    Mantém apenas o arquivo que será gerado em seguida (a pasta fica vazia de Excels até o write atual).
    Não remove subpastas nem outros tipos de arquivo. Ignora falhas de permissão em itens pontuais.
    """
    if not pasta_saida or not str(pasta_saida).strip():
        return
    pasta = os.path.abspath(os.path.normpath(pasta_saida))
    if not os.path.isdir(pasta):
        return
    if pasta == os.path.abspath(os.getcwd()):
        return
    try:
        for nome in os.listdir(pasta):
            caminho = os.path.join(pasta, nome)
            if not os.path.isfile(caminho):
                continue
            nl = nome.lower()
            if nl.endswith(".xlsx") or nl.endswith(".xlsm"):
                try:
                    os.remove(caminho)
                except OSError:
                    pass
    except OSError:
        pass


def normalizar_cliente_para_aporte(cliente):
    cliente = limpar_texto_nome(cliente)
    cliente = cliente.replace(".", "")
    cliente = cliente.replace(",", "")
    cliente = re.sub(r"\s+", " ", cliente)
    return cliente.strip()


def eh_aporte_financeiro(cliente, valor_total, identificador, descricao):
    cliente_up = normalizar_cliente_para_aporte(cliente)
    identificador_up = normalizar_identificador(identificador)
    descricao_up = limpar_texto_nome(descricao)

    if not cliente_up:
        return False

    if any(nome in cliente_up for nome in CLIENTES_NAO_APORTE):
        return False

    if cliente_up in NOMES_APORTE_EXATOS:
        return True

    if any(t in cliente_up for t in TERMOS_APORTE):
        return True

    if "APORTE" in descricao_up:
        return True

    if "REPASSE" in descricao_up:
        return True

    qtd_lotes = identificador_up.count("LOTE")
    if qtd_lotes >= 5:
        return True

    return False


def linha_ignorada_generica(linha):
    s = str(linha).strip()
    if not s:
        return True

    termos = [
        "UAU! Software",
        "UAU! SOFTWARE",
        "Página",
        "Contas a Receber",
        "Contas Recebidas",
        "Período por Vencimento",
        "Período por Recebimento",
        "Emp/Obra\tVenda\tCliente",
        "ANDREDUA",
        "SCP GOIÂNIA",
        "SCP GOIANIA",
        "Total cliente:",
        "Total por Cliente:",
        "Elet. dep.:",
        "Cheque dep.:",
        "Din. dep.:",
        "Bem dep.:",
        "Elet. ñ dep.:",
        "Cheque ñ dep.:",
        "Din. ñ dep.:",
        "Bem ñ dep.:",
    ]
    return any(t in s for t in termos)


def is_main_receber_line(line):
    parts = split_linha_tabular(line)
    if len(parts) < 7:
        return False
    if not EMP_RE.match(parts[0].strip()):
        return False
    for idx in range(3, min(len(parts), 9)):
        if DATE_RE.match(parts[idx].strip()):
            return True
    return False


def is_main_recebidos_line(line):
    parts = split_linha_tabular(line)
    if len(parts) < 5:
        return False
    if not EMP_RE.match(parts[0].strip()):
        return False

    for idx in range(3, min(len(parts), 8)):
        if DATE_RE.match(parts[idx].strip()):
            return True
    return False


# =========================
# VALIDAÇÃO DE ENTRADA (PRÉ-PROCESSAMENTO UAU)
# =========================
MSG_TXT_NAO_CONFIAVEL = (
    "Não foi possível interpretar o arquivo TXT de forma confiável. "
    "Verifique o encoding ou gere uma nova exportação do UAU."
)
MSG_ESTRUTURA_MINIMA_RECEBER = (
    "O arquivo de Contas a Receber não contém estrutura mínima reconhecível do UAU "
    "(cabeçalho ou linhas de dados esperadas)."
)
MSG_ESTRUTURA_MINIMA_RECEBIDOS = (
    "O arquivo de Contas Recebidas não contém estrutura mínima reconhecível do UAU "
    "(cabeçalho ou linhas de dados esperadas)."
)


def _texto_ascii_fold_upper(texto: str, max_chars: int = 400_000) -> str:
    """Remove acentos para busca robusta de marcadores no cabeçalho UAU."""
    amostra = str(texto or "")[:max_chars]
    nk = unicodedata.normalize("NFKD", amostra)
    return "".join(c for c in nk if not unicodedata.combining(c)).upper()


def _contar_linhas_principais_uau(texto: str, limite_linhas: int = 30_000):
    """Conta linhas que batem com o parser principal de cada relatório."""
    n_rec = 0
    n_receb = 0
    for i, raw in enumerate(texto.splitlines()):
        if i >= limite_linhas:
            break
        linha = str(raw).strip()
        if not linha:
            continue
        if is_main_receber_line(linha):
            n_rec += 1
        if is_main_recebidos_line(linha):
            n_receb += 1
    return n_rec, n_receb


def identificar_tipo_relatorio_uau_por_texto(texto: str) -> str:
    """
    Identifica pelo conteúdo se o TXT é Contas a Receber ou Contas Recebidas (UAU).
    Retorno: 'RECEBER' | 'RECEBIDOS' | 'INDETERMINADO' (universal, sem exceção por empreendimento).
    """
    if not texto or not str(texto).strip():
        return "INDETERMINADO"

    busca = _texto_ascii_fold_upper(texto)
    score_receber = 0
    score_recebidos = 0

    if "CONTAS A RECEBER" in busca:
        score_receber += 4
    if "PERIODO POR VENCIMENTO" in busca:
        score_receber += 2
    if "CONTAS RECEBIDAS" in busca:
        score_recebidos += 4
    if "PERIODO POR RECEBIMENTO" in busca:
        score_recebidos += 2

    n_rec, n_receb = _contar_linhas_principais_uau(texto)

    if score_receber >= 4 and score_recebidos <= 1:
        return "RECEBER"
    if score_recebidos >= 4 and score_receber <= 1:
        return "RECEBIDOS"

    if score_receber >= 6 and score_recebidos >= 6:
        if n_rec >= n_receb + 3:
            return "RECEBER"
        if n_receb >= n_rec + 3:
            return "RECEBIDOS"
        return "INDETERMINADO"

    if score_receber > score_recebidos and score_receber >= 3:
        return "RECEBER"
    if score_recebidos > score_receber and score_recebidos >= 3:
        return "RECEBIDOS"

    if n_rec >= 8 and n_rec > n_receb * 2:
        return "RECEBER"
    if n_receb >= 8 and n_receb > n_rec * 2:
        return "RECEBIDOS"

    if score_receber >= 2 and score_recebidos == 0 and n_rec >= 3:
        return "RECEBER"
    if score_recebidos >= 2 and score_receber == 0 and n_receb >= 3:
        return "RECEBIDOS"

    return "INDETERMINADO"


def identificar_tipo_relatorio_uau(caminho_arquivo: str) -> str:
    """Lê o arquivo com a mesma estratégia do motor e identifica o tipo de relatório."""
    if not caminho_arquivo or not os.path.isfile(caminho_arquivo):
        return "INDETERMINADO"
    try:
        texto = ler_texto_robusto(caminho_arquivo)
    except Exception:
        return "INDETERMINADO"
    return identificar_tipo_relatorio_uau_por_texto(texto)


def _ler_texto_validacao_entrada(caminho_arquivo: str) -> str:
    """
    Leitura alinhada a ler_texto_robusto, com checagens de confiabilidade antes do motor.
    Não exige UTF-8 puro: aceita cp1252/latin-1 como o fluxo principal.
    """
    if not caminho_arquivo or not os.path.isfile(caminho_arquivo):
        raise ProcessamentoUAUErro(
            etapa="validação de entrada",
            funcao="_ler_texto_validacao_entrada",
            validacao="arquivo inexistente",
            mensagem="Um dos arquivos TXT não foi encontrado no caminho informado.",
            campo_ou_aba="upload",
        )
    try:
        texto = ler_texto_robusto(caminho_arquivo)
    except OSError as e:
        raise ProcessamentoUAUErro(
            etapa="validação de entrada",
            funcao="_ler_texto_validacao_entrada",
            validacao="leitura do arquivo",
            mensagem=MSG_TXT_NAO_CONFIAVEL,
            campo_ou_aba="TXT",
            erro_tecnico=e,
        ) from e
    except Exception as e:
        raise ProcessamentoUAUErro(
            etapa="validação de entrada",
            funcao="_ler_texto_validacao_entrada",
            validacao="leitura do arquivo",
            mensagem=MSG_TXT_NAO_CONFIAVEL,
            campo_ou_aba="TXT",
            erro_tecnico=e,
        ) from e

    if texto is None or not str(texto).strip():
        raise ProcessamentoUAUErro(
            etapa="validação de entrada",
            funcao="_ler_texto_validacao_entrada",
            validacao="conteúdo vazio",
            mensagem="O arquivo TXT está vazio ou contém apenas espaços em branco.",
            campo_ou_aba="TXT",
        )

    if "\x00" in texto:
        raise ProcessamentoUAUErro(
            etapa="validação de entrada",
            funcao="_ler_texto_validacao_entrada",
            validacao="conteúdo binário/corrompido",
            mensagem=MSG_TXT_NAO_CONFIAVEL,
            campo_ou_aba="TXT",
        )

    amostra = texto if len(texto) <= 100_000 else texto[:100_000]
    controle = sum(1 for c in amostra if ord(c) < 32 and c not in "\n\r\t")
    if len(amostra) > 200 and controle > len(amostra) * 0.08:
        raise ProcessamentoUAUErro(
            etapa="validação de entrada",
            funcao="_ler_texto_validacao_entrada",
            validacao="caracteres de controle excessivos",
            mensagem=MSG_TXT_NAO_CONFIAVEL,
            campo_ou_aba="TXT",
        )

    return texto


def _estrutura_minima_uau_ok(texto: str, tipo: str) -> bool:
    if tipo == "RECEBER":
        return any(is_main_receber_line(str(l).strip()) for l in texto.splitlines()[:35_000])
    if tipo == "RECEBIDOS":
        return any(is_main_recebidos_line(str(l).strip()) for l in texto.splitlines()[:35_000])
    return False


def validar_arquivos_entrada_uau(caminho_receber: str, caminho_recebidos: str) -> dict:
    """
    Valida os dois TXT antes de carregar_receber_bruto / carregar_recebidos_bruto.
    Retorna dict com tipos identificados ou levanta ProcessamentoUAUErro com mensagem clara ao usuário.
    """
    ap1 = os.path.abspath(os.path.normpath(caminho_receber))
    ap2 = os.path.abspath(os.path.normpath(caminho_recebidos))
    if ap1 == ap2:
        raise ProcessamentoUAUErro(
            etapa="validação de entrada",
            funcao="validar_arquivos_entrada_uau",
            validacao="arquivos duplicados",
            mensagem="Os arquivos anexados são duplicados.",
            campo_ou_aba="Contas a Receber / Contas Recebidas",
        )

    try:
        if filecmp.cmp(caminho_receber, caminho_recebidos, shallow=False):
            raise ProcessamentoUAUErro(
                etapa="validação de entrada",
                funcao="validar_arquivos_entrada_uau",
                validacao="arquivos duplicados",
                mensagem="Os arquivos anexados são duplicados.",
                campo_ou_aba="Contas a Receber / Contas Recebidas",
            )
    except OSError:
        pass

    texto_r = _ler_texto_validacao_entrada(caminho_receber)
    texto_p = _ler_texto_validacao_entrada(caminho_recebidos)

    tipo_r = identificar_tipo_relatorio_uau_por_texto(texto_r)
    tipo_p = identificar_tipo_relatorio_uau_por_texto(texto_p)

    if tipo_r == "RECEBER" and tipo_p == "RECEBER":
        raise ProcessamentoUAUErro(
            etapa="validação de entrada",
            funcao="validar_arquivos_entrada_uau",
            validacao="tipo de relatório",
            mensagem=(
                "Os dois arquivos anexados foram identificados como Contas a Receber. "
                "Anexe um arquivo de Contas a Receber e outro de Contas Recebidas."
            ),
            campo_ou_aba="Contas a Receber / Contas Recebidas",
        )

    if tipo_r == "RECEBIDOS" and tipo_p == "RECEBIDOS":
        raise ProcessamentoUAUErro(
            etapa="validação de entrada",
            funcao="validar_arquivos_entrada_uau",
            validacao="tipo de relatório",
            mensagem=(
                "Os dois arquivos anexados foram identificados como Contas Recebidas. "
                "Anexe um arquivo de Contas a Receber e outro de Contas Recebidas."
            ),
            campo_ou_aba="Contas a Receber / Contas Recebidas",
        )

    if tipo_r == "RECEBIDOS" and tipo_p == "RECEBER":
        raise ProcessamentoUAUErro(
            etapa="validação de entrada",
            funcao="validar_arquivos_entrada_uau",
            validacao="arquivos invertidos",
            mensagem=(
                "Os arquivos parecem invertidos: o anexo do campo Contas a Receber corresponde a "
                "Contas Recebidas e o do campo Contas Recebidas corresponde a Contas a Receber. Troque os anexos."
            ),
            campo_ou_aba="Contas a Receber / Contas Recebidas",
        )

    if tipo_r != "RECEBER":
        if tipo_r == "RECEBIDOS":
            msg = (
                "O arquivo enviado no campo Contas a Receber foi identificado como Contas Recebidas. "
                "Verifique os anexos."
            )
        else:
            msg = (
                "O arquivo enviado no campo Contas a Receber não foi reconhecido como um relatório UAU "
                "de Contas a Receber (cabeçalho ou estrutura incompatível)."
            )
        raise ProcessamentoUAUErro(
            etapa="validação de entrada",
            funcao="validar_arquivos_entrada_uau",
            validacao="tipo de relatório",
            mensagem=msg,
            campo_ou_aba="Contas a Receber",
        )

    if tipo_p != "RECEBIDOS":
        if tipo_p == "RECEBER":
            msg = (
                "O arquivo enviado no campo Contas Recebidas foi identificado como Contas a Receber. "
                "Verifique os anexos."
            )
        else:
            msg = (
                "O arquivo enviado no campo Contas Recebidas não foi reconhecido como um relatório UAU "
                "de Contas Recebidas (cabeçalho ou estrutura incompatível)."
            )
        raise ProcessamentoUAUErro(
            etapa="validação de entrada",
            funcao="validar_arquivos_entrada_uau",
            validacao="tipo de relatório",
            mensagem=msg,
            campo_ou_aba="Contas Recebidas",
        )

    if not _estrutura_minima_uau_ok(texto_r, "RECEBER"):
        raise ProcessamentoUAUErro(
            etapa="validação de entrada",
            funcao="validar_arquivos_entrada_uau",
            validacao="estrutura mínima",
            mensagem=MSG_ESTRUTURA_MINIMA_RECEBER,
            campo_ou_aba="Contas a Receber",
        )

    if not _estrutura_minima_uau_ok(texto_p, "RECEBIDOS"):
        raise ProcessamentoUAUErro(
            etapa="validação de entrada",
            funcao="validar_arquivos_entrada_uau",
            validacao="estrutura mínima",
            mensagem=MSG_ESTRUTURA_MINIMA_RECEBIDOS,
            campo_ou_aba="Contas Recebidas",
        )

    return {
        "ok": True,
        "tipo_receber": tipo_r,
        "tipo_recebidos": tipo_p,
        "texto_receber": texto_r,
        "texto_recebidos": texto_p,
        "erros": [],
    }


# =========================
# LEITURA RECEBIDOS
# =========================
def _parse_recebidos_main_line(line):
    parts = split_linha_tabular(line)
    if len(parts) < 5:
        return None

    date_idx = None
    for idx in range(3, min(len(parts), 8)):
        if DATE_RE.match(parts[idx]):
            date_idx = idx
            break

    if date_idx is None:
        return None

    emp = get_coluna(parts, 0)
    venda = get_coluna(parts, 1)
    cliente = get_coluna(parts, 2)
    unidades = "\t".join(parts[3:date_idx]).strip()
    rest = parts[date_idx:]

    if len(rest) >= 7 and len(rest) < 10:
        # Layout reduzido (sem correcao/multa/juros separados).
        return {
            "Emp/Obra": emp,
            "Venda": venda,
            "Cliente": cliente,
            "Unidades": unidades,
            "Data_Rec": rest[0],
            "Tipo": rest[1],
            "Parcela": rest[2],
            "Vlr_Parcela": rest[3],
            "Principal": rest[4],
            "Correcao": "0,00",
            "Multa_Atraso": "0,00",
            "Juros_Atraso": "0,00",
            "Total_Dep": get_coluna(rest, 5),
            "Total_Nao_Dep": get_coluna(rest, 6),
        }

    if len(rest) >= 10:
        return {
            "Emp/Obra": emp,
            "Venda": venda,
            "Cliente": cliente,
            "Unidades": unidades,
            "Data_Rec": rest[0],
            "Tipo": rest[1],
            "Parcela": rest[2],
            "Vlr_Parcela": rest[3],
            "Principal": rest[4],
            "Correcao": get_coluna(rest, 5),
            "Multa_Atraso": get_coluna(rest, 6),
            "Juros_Atraso": get_coluna(rest, 7),
            "Total_Dep": get_coluna(rest, 8),
            "Total_Nao_Dep": get_coluna(rest, 9),
        }

    return None


def carregar_recebidos_bruto(caminho_arquivo):
    texto = ler_texto_robusto(caminho_arquivo)

    linhas = [x.rstrip("\r").rstrip("\n") for x in texto.splitlines()]

    registros = []
    pending_prod_lines = []
    prod_mode = False

    i = 0
    while i < len(linhas):
        raw = linhas[i]
        linha = str(raw).strip()

        if not linha:
            i += 1
            continue

        if linha.startswith("Prod.\tDescrição\tNº Person.\tIdentificador\tQtde.") or \
           linha.startswith("Prod.\tDescricao\tNº Person.\tIdentificador\tQtde.") or \
           linha.startswith("Prod.\tDescricao\tN Person.\tIdentificador\tQtde.") or \
           linha.startswith("Prod.\tDescrição\tN Person.\tIdentificador\tQtde."):
            pending_prod_lines = []
            prod_mode = True
            i += 1
            continue

        if prod_mode:
            if linha_ignorada_generica(linha):
                i += 1
                continue

            if is_main_recebidos_line(linha):
                prod_mode = False
                continue

            partes_prod = split_linha_tabular(linha)
            if len(partes_prod) >= 5:
                primeiro = get_coluna(partes_prod, 0)
                if primeiro.isdigit() or primeiro == "300":
                    pending_prod_lines.append({
                        "Produto": get_coluna(partes_prod, 0),
                        "Descricao_Produto": get_coluna(partes_prod, 1),
                        "Nr_Person": get_coluna(partes_prod, 2),
                        "Identificador_Produto": get_coluna(partes_prod, 3),
                        "Qtde_Produto": get_coluna(partes_prod, 4),
                    })
                    i += 1
                    continue

            prod_mode = False

        if linha_ignorada_generica(linha):
            i += 1
            continue

        if is_main_recebidos_line(linha):
            parsed = _parse_recebidos_main_line(linha)

            if parsed is not None:
                if pending_prod_lines:
                    for prod in pending_prod_lines:
                        reg = parsed.copy()
                        reg.update(prod)
                        registros.append(reg)
                    pending_prod_lines = []
                else:
                    reg = parsed.copy()
                    reg["Produto"] = ""
                    reg["Descricao_Produto"] = ""
                    reg["Nr_Person"] = ""
                    reg["Identificador_Produto"] = ""
                    reg["Qtde_Produto"] = ""
                    registros.append(reg)

            i += 1
            continue

        if registros and not prod_mode:
            if not linha_ignorada_generica(linha) and not texto_contaminado(linha):
                ultimo = registros[-1]
                unidades_atual = str(ultimo.get("Unidades", "")).strip()
                ultimo["Unidades"] = (unidades_atual + " " + linha).strip()

        i += 1

    colunas_base = [
        "Emp/Obra", "Venda", "Cliente", "Unidades", "Data_Rec", "Tipo", "Parcela",
        "Vlr_Parcela", "Principal", "Correcao", "Multa_Atraso", "Juros_Atraso",
        "Total_Dep", "Total_Nao_Dep", "Produto", "Descricao_Produto",
        "Nr_Person", "Identificador_Produto", "Qtde_Produto"
    ]

    if registros:
        df = pd.DataFrame(registros)
    else:
        df = pd.DataFrame(columns=colunas_base)

    for col in colunas_base:
        if col not in df.columns:
            df[col] = ""

    if df.empty:
        df["Parc_Num"] = pd.Series(dtype="int64")
        df["Parc_Total"] = pd.Series(dtype="int64")
        df["is_paid"] = pd.Series(dtype="int64")
        return df

    df.columns = [str(c).strip().replace("\ufeff", "") for c in df.columns]

    df["Emp/Obra"] = df["Emp/Obra"].fillna("").astype(str).apply(normalizar_emp_obra)
    df["Venda"] = df["Venda"].fillna("").astype(str).str.strip()
    df["Cliente"] = df["Cliente"].fillna("").astype(str).str.strip()
    df["Unidades"] = df["Unidades"].fillna("").astype(str).apply(normalizar_identificador)
    df["Identificador_Produto"] = df["Identificador_Produto"].fillna("").astype(str).apply(normalizar_identificador)
    mask_id_ruim = df["Unidades"].astype(str).str.strip().eq("") | df["Unidades"].apply(identificador_truncado)
    df.loc[mask_id_ruim, "Unidades"] = (
        df.loc[mask_id_ruim, "Identificador_Produto"].fillna(df.loc[mask_id_ruim, "Unidades"]).astype(str)
    )
    df["Descricao_Produto"] = df["Descricao_Produto"].fillna("").astype(str).str.strip()
    df["Nr_Person"] = df["Nr_Person"].fillna("").astype(str).str.strip()
    df["Qtde_Produto"] = df["Qtde_Produto"].fillna("").astype(str).str.strip()

    for col in [
        "Vlr_Parcela", "Principal", "Correcao",
        "Multa_Atraso", "Juros_Atraso", "Total_Dep", "Total_Nao_Dep"
    ]:
        df[col] = df[col].apply(converter_valor)

    df["Parc_Num"] = df["Parcela"].apply(extrair_numero_parcela)
    df["Parc_Total"] = df["Parcela"].apply(extrair_total_parcela)
    df["Data_Rec"] = pd.to_datetime(df["Data_Rec"], format="%d/%m/%Y", errors="coerce")
    df["is_paid"] = 1

    return df


# =========================
# LEITURA RECEBER
# =========================
def carregar_receber_bruto(caminho_arquivo):
    texto = ler_texto_robusto(caminho_arquivo)

    linhas = [x.rstrip("\r").rstrip("\n") for x in texto.splitlines()]

    registros = []
    ultimo_prod = None
    i = 0

    while i < len(linhas):
        linha = str(linhas[i]).strip()

        if not linha or linha_ignorada_generica(linha):
            i += 1
            continue

        if linha.startswith("Prod.\tDescrição\tNº Person.\tIdentificador\tQtde.") or \
           linha.startswith("Prod.\tDescricao\tNº Person.\tIdentificador\tQtde.") or \
           linha.startswith("Prod.\tDescricao\tN Person.\tIdentificador\tQtde.") or \
           linha.startswith("Prod.\tDescrição\tN Person.\tIdentificador\tQtde."):
            ultimo_prod = None

            if i + 1 < len(linhas):
                prox = str(linhas[i + 1]).strip()
                partes_prod = split_linha_tabular(prox)

                if len(partes_prod) >= 5:
                    ultimo_prod = {
                        "Produto": get_coluna(partes_prod, 0),
                        "Descricao_Produto": get_coluna(partes_prod, 1),
                        "Nr_Person": get_coluna(partes_prod, 2),
                        "Identificador_Produto": get_coluna(partes_prod, 3),
                        "Qtde_Produto": get_coluna(partes_prod, 4),
                    }
                    i += 2
                    continue

            i += 1
            continue

        if is_main_receber_line(linha):
            partes = split_linha_tabular(linha)

            if len(partes) < 7:
                i += 1
                continue

            date_idxs = []
            for idx in range(3, min(len(partes), 10)):
                if DATE_RE.match(get_coluna(partes, idx)):
                    date_idxs.append(idx)

            if not date_idxs:
                i += 1
                continue

            venc_idx = date_idxs[0]
            venc_pror_idx = date_idxs[1] if len(date_idxs) > 1 else venc_idx

            emp = get_coluna(partes, 0)
            venda = get_coluna(partes, 1)
            cliente = get_coluna(partes, 2)
            unidades = " ".join(partes[3:venc_idx]).strip()
            tipo = get_coluna(partes, venc_idx - 2) if venc_idx - 2 >= 3 else ""
            parcela = get_coluna(partes, venc_idx - 1) if venc_idx - 1 >= 3 else ""

            valores = partes[venc_pror_idx + 1:]
            # Alinha monetarios pela direita para tolerar colunas faltando/extras.
            valores = [""] * max(0, 5 - len(valores)) + valores
            valores = valores[-5:]

            reg = {
                "Emp/Obra": emp,
                "Venda": venda,
                "Cliente": cliente,
                "Unidades": unidades,
                "Tipo": tipo,
                "Parcela": parcela,
                "Vencimento": get_coluna(partes, venc_idx),
                "Venc. Pror.": get_coluna(partes, venc_pror_idx),
                "Principal": get_coluna(valores, 0),
                "Correcao": get_coluna(valores, 1),
                "Juros_Atraso": get_coluna(valores, 2),
                "Multa_Atraso": get_coluna(valores, 3),
                "Vlr_Parcela": get_coluna(valores, 4),
            }

            if ultimo_prod:
                reg["Produto"] = ultimo_prod.get("Produto", "")
                reg["Descricao_Produto"] = ultimo_prod.get("Descricao_Produto", "")
                reg["Nr_Person"] = ultimo_prod.get("Nr_Person", "")
                reg["Identificador_Produto"] = ultimo_prod.get("Identificador_Produto", "")
                reg["Qtde_Produto"] = ultimo_prod.get("Qtde_Produto", "")
            else:
                reg["Produto"] = ""
                reg["Descricao_Produto"] = ""
                reg["Nr_Person"] = ""
                reg["Identificador_Produto"] = ""
                reg["Qtde_Produto"] = ""

            registros.append(reg)

        i += 1

    colunas_base = [
        "Emp/Obra", "Venda", "Cliente", "Unidades", "Tipo", "Parcela",
        "Vencimento", "Venc. Pror.", "Principal", "Correcao",
        "Juros_Atraso", "Multa_Atraso", "Vlr_Parcela",
        "Produto", "Descricao_Produto", "Nr_Person",
        "Identificador_Produto", "Qtde_Produto"
    ]

    if registros:
        df = pd.DataFrame(registros)
    else:
        df = pd.DataFrame(columns=colunas_base)

    for col in colunas_base:
        if col not in df.columns:
            df[col] = ""

    if df.empty:
        df["Parc_Num"] = pd.Series(dtype="int64")
        df["Parc_Total"] = pd.Series(dtype="int64")
        df["Correcao_Atraso"] = pd.Series(dtype="float64")
        return df

    df.columns = [str(c).strip().replace("\ufeff", "") for c in df.columns]

    df["Emp/Obra"] = df["Emp/Obra"].fillna("").astype(str).apply(normalizar_emp_obra)
    df["Venda"] = df["Venda"].fillna("").astype(str).str.strip()
    df["Cliente"] = df["Cliente"].fillna("").astype(str).str.strip()
    df["Unidades"] = df["Unidades"].fillna("").astype(str).apply(normalizar_identificador)
    df["Identificador_Produto"] = df["Identificador_Produto"].fillna("").astype(str).apply(normalizar_identificador)
    mask_id_ruim = df["Unidades"].astype(str).str.strip().eq("") | df["Unidades"].apply(identificador_truncado)
    df.loc[mask_id_ruim, "Unidades"] = (
        df.loc[mask_id_ruim, "Identificador_Produto"].fillna(df.loc[mask_id_ruim, "Unidades"]).astype(str)
    )
    df["Descricao_Produto"] = df["Descricao_Produto"].fillna("").astype(str).str.strip()
    df["Nr_Person"] = df["Nr_Person"].fillna("").astype(str).str.strip()
    df["Qtde_Produto"] = df["Qtde_Produto"].fillna("").astype(str).str.strip()

    for col in ["Principal", "Correcao", "Juros_Atraso", "Multa_Atraso", "Vlr_Parcela"]:
        df[col] = df[col].apply(converter_valor)

    df["Parc_Num"] = df["Parcela"].apply(extrair_numero_parcela)
    df["Parc_Total"] = df["Parcela"].apply(extrair_total_parcela)
    df["Vencimento"] = pd.to_datetime(df["Vencimento"], format="%d/%m/%Y", errors="coerce")
    df["Venc. Pror."] = pd.to_datetime(df["Venc. Pror."], format="%d/%m/%Y", errors="coerce")

    df["Correcao_Atraso"] = (
        df["Vlr_Parcela"]
        - df["Multa_Atraso"]
        - df["Juros_Atraso"]
        - df["Correcao"]
        - df["Principal"]
    )

    return df


# =========================
# IDENTIFICADOR POR VENDA + CLIENTE_BASE
# =========================
def escolher_identificador_por_venda_cliente_base(df):
    if df is None or df.empty:
        return pd.DataFrame(columns=[
            "Venda", "Cliente_Base", "Identificador_Padrao",
            "Descricao_Padrao", "EmpObra_Padrao"
        ])

    base = df.copy()

    if "Cliente_Base" not in base.columns:
        base = adicionar_chave_cliente(base)

    if "Unidades" not in base.columns:
        base["Unidades"] = ""

    if "Identificador_Produto" not in base.columns:
        base["Identificador_Produto"] = ""

    if "Descricao_Produto" not in base.columns:
        base["Descricao_Produto"] = ""

    if "Emp/Obra" not in base.columns:
        base["Emp/Obra"] = ""

    base["Venda"] = base["Venda"].fillna("").astype(str).str.strip()
    base["Cliente_Base"] = base["Cliente_Base"].fillna("").astype(str).str.strip()
    base["Parcela"] = base["Parcela"].fillna("").astype(str).str.strip() if "Parcela" in base.columns else ""

    base["cand1"] = base["Unidades"].fillna("").astype(str).apply(normalizar_identificador)
    base["cand2"] = base["Identificador_Produto"].fillna("").astype(str).apply(normalizar_identificador)

    resultados = []

    for (venda, cliente_base), grupo in base.groupby(["Venda", "Cliente_Base"], dropna=False):
        confirmados = []
        parcelas_vals = grupo["Parcela"].astype(str).str.strip()

        # Pré-mapa por parcela: evita filtrar o grupo inteiro repetidas vezes.
        candidatos_por_parcela = {}
        for parcela_key, gp_ev in grupo.groupby(parcelas_vals, dropna=False, sort=False):
            pk = str(parcela_key).strip()
            if not pk:
                continue
            cand_set = set()
            for vv in gp_ev["cand1"].tolist():
                sv = str(vv).strip()
                if sv and (not identificador_truncado(sv)):
                    cand_set.add(sv)
            for vv in gp_ev["cand2"].tolist():
                sv = str(vv).strip()
                if sv and (not identificador_truncado(sv)):
                    cand_set.add(sv)
            candidatos_por_parcela[pk] = cand_set

        for parcela, c1, c2 in zip(parcelas_vals.tolist(), grupo["cand1"].tolist(), grupo["cand2"].tolist()):
            parcela = str(parcela).strip()

            for c in (c1, c2):
                c = str(c).strip()
                if not c:
                    continue
                if identificador_truncado(c):
                    continue
                if not identificador_tem_formato_endereco(c):
                    continue
                if parcela:
                    if c in candidatos_por_parcela.get(parcela, set()):
                        confirmados.append(c)
                else:
                    confirmados.append(c)

        identificador_padrao = escolher_identificador_melhor(confirmados)

        descricoes = [
            str(x).strip().upper()
            for x in grupo["Descricao_Produto"].tolist()
            if str(x).strip() != ""
        ]
        descricao_padrao = escolher_moda_texto(descricoes)

        empobras = [
            normalizar_emp_obra(x)
            for x in grupo["Emp/Obra"].tolist()
            if str(x).strip() != ""
        ]
        empobra_padrao = escolher_moda_texto(empobras)

        resultados.append({
            "Venda": venda,
            "Cliente_Base": cliente_base,
            "Identificador_Padrao": identificador_padrao,
            "Descricao_Padrao": descricao_padrao,
            "EmpObra_Padrao": empobra_padrao,
        })

    return pd.DataFrame(resultados)


def aplicar_padroes(df_receber, df_recebidos):
    if df_receber is None:
        df_receber = pd.DataFrame()

    if df_recebidos is None:
        df_recebidos = pd.DataFrame()

    if not df_receber.empty:
        df_receber = adicionar_chave_cliente(df_receber)

    if not df_recebidos.empty:
        df_recebidos = adicionar_chave_cliente(df_recebidos)

    mapa_receber = escolher_identificador_por_venda_cliente_base(df_receber)
    mapa_recebidos = escolher_identificador_por_venda_cliente_base(df_recebidos)

    mapa = pd.concat([mapa_receber, mapa_recebidos], ignore_index=True)

    if not mapa.empty:
        mapa = (
            mapa.sort_values(
                by=["Venda", "Cliente_Base", "Identificador_Padrao"],
                ascending=[True, True, True]
            )
            .drop_duplicates(subset=["Venda", "Cliente_Base"], keep="first")
        )

    if not df_receber.empty:
        df_receber = df_receber.merge(mapa, on=["Venda", "Cliente_Base"], how="left")
        unidades_norm_r = df_receber["Unidades"].fillna("").astype(str).str.strip()

        mask_vazio_ou_trunc = (
            unidades_norm_r.eq("")
            | pd.Series([identificador_truncado(v) for v in unidades_norm_r.tolist()], index=df_receber.index)
        )

        df_receber.loc[mask_vazio_ou_trunc, "Unidades"] = (
            df_receber.loc[mask_vazio_ou_trunc, "Identificador_Padrao"]
            .fillna(df_receber.loc[mask_vazio_ou_trunc, "Unidades"])
        )

        df_receber["Descricao_Produto"] = (
            df_receber["Descricao_Padrao"].fillna(df_receber["Descricao_Produto"])
        )
        df_receber["Emp/Obra"] = (
            df_receber["EmpObra_Padrao"].fillna(df_receber["Emp/Obra"])
        )

        df_receber.drop(
            columns=["Identificador_Padrao", "Descricao_Padrao", "EmpObra_Padrao"],
            inplace=True,
            errors="ignore",
        )

    if not df_recebidos.empty:
        df_recebidos = df_recebidos.merge(mapa, on=["Venda", "Cliente_Base"], how="left")
        unidades_norm_p = df_recebidos["Unidades"].fillna("").astype(str).str.strip()

        mask_vazio_ou_trunc = (
            unidades_norm_p.eq("")
            | pd.Series([identificador_truncado(v) for v in unidades_norm_p.tolist()], index=df_recebidos.index)
        )

        df_recebidos.loc[mask_vazio_ou_trunc, "Unidades"] = (
            df_recebidos.loc[mask_vazio_ou_trunc, "Identificador_Padrao"]
            .fillna(df_recebidos.loc[mask_vazio_ou_trunc, "Unidades"])
        )

        df_recebidos["Descricao_Produto"] = (
            df_recebidos["Descricao_Padrao"].fillna(df_recebidos["Descricao_Produto"])
        )
        df_recebidos["Emp/Obra"] = (
            df_recebidos["EmpObra_Padrao"].fillna(df_recebidos["Emp/Obra"])
        )

        df_recebidos.drop(
            columns=["Identificador_Padrao", "Descricao_Padrao", "EmpObra_Padrao"],
            inplace=True,
            errors="ignore",
        )

    return df_receber, df_recebidos


# =========================
# APORTES
# =========================
def separar_aportes_financeiros(df_recebidos):
    """
    Mantém critérios existentes (PJ/banco/valor alto via eh_aporte_financeiro).
    Não mistura aportes ao fluxo parcelado normal — apenas isola antes do dedup principal dos recebidos.
    """
    if df_recebidos is None or df_recebidos.empty:
        vazio = pd.DataFrame(columns=df_recebidos.columns if df_recebidos is not None else [])
        return vazio.copy(), vazio.copy()

    base = df_recebidos.copy()

    if "Cliente" not in base.columns:
        base["Cliente"] = ""

    if "Unidades" not in base.columns:
        base["Unidades"] = ""

    if "Descricao_Produto" not in base.columns:
        base["Descricao_Produto"] = ""

    if "Total_Dep" not in base.columns:
        base["Total_Dep"] = 0.0

    base["Eh_Aporte"] = [
        eh_aporte_financeiro(
            cliente=c,
            valor_total=v,
            identificador=i,
            descricao=d,
        )
        for c, v, i, d in zip(
            base["Cliente"].tolist(),
            base["Total_Dep"].tolist(),
            base["Unidades"].tolist(),
            base["Descricao_Produto"].tolist(),
        )
    ]

    df_aportes = base[base["Eh_Aporte"] == True].copy()
    df_recebidos_sem_aporte = base[base["Eh_Aporte"] == False].copy()

    return df_recebidos_sem_aporte, df_aportes


# =========================
# DEDUP
# =========================
def _moda_status_vencimento_agrupado(series):
    """Preserva Status_Vencimento no dedup (moda; empate: string mais longa)."""
    vals = []
    for v in series:
        s = str(v).strip()
        if s == "" or s.lower() == "nan":
            continue
        vals.append(s.upper())
    if not vals:
        return ""
    c = Counter(vals)
    max_f = max(c.values())
    cands = [k for k, v in c.items() if v == max_f]
    return sorted(cands, key=lambda x: (-len(x), x))[0]


def _escolher_unidades_pos_dedup_estrutural(series_unidades):
    """Reconcilia Unidades após dedup contratual (não-chave): melhor identificador ou texto mais completo."""
    valores = [x for x in series_unidades.tolist()]
    cands = []
    for v in valores:
        s = str(v).strip()
        if not s:
            continue
        ni = normalizar_identificador(s)
        if ni and not identificador_truncado(s):
            cands.append(ni)
    if cands:
        return escolher_identificador_melhor(cands)
    raw = [str(v).strip() for v in valores if str(v).strip()]
    if not raw:
        return ""
    return sorted(raw, key=lambda t: (-len(t), t))[0]


def _log_volume_dedup(tag, n_antes, n_depois):
    if DEBUG_DADOS and n_antes >= 0:
        pct = 100.0 * (1.0 - float(n_depois) / float(max(n_antes, 1)))
        print(f"[DEBUG][DEDUP_{tag}] linhas_antes={n_antes} depois={n_depois} reducao_pct={pct:.2f}")


_MIN_SCORE_ID_DEDUP_CONFLITO = 1000


def _id_chave_estrutural_para_linha(unidades, ident_prod):
    """Melhor identificador normalizado por linha com score mínimo (alta confiança)."""
    best = ""
    best_sc = -1
    for v in (unidades, ident_prod):
        sv = str(v or "").strip()
        if not sv:
            continue
        if identificador_truncado(sv):
            continue
        n = normalizar_identificador(sv)
        if not n:
            continue
        sc = score_identificador(n)
        if sc < _MIN_SCORE_ID_DEDUP_CONFLITO:
            continue
        if sc > best_sc:
            best_sc = sc
            best = n
    return best


def _aplicar_split_dedup_por_identificador(base, chave_cols):
    """Marca conflito e subchave quando a mesma chave contratual traz identificadores fortes distintos."""
    base = base.copy()
    uu = base["Unidades"].astype(str) if "Unidades" in base.columns else pd.Series("", index=base.index, dtype=str)
    ip = (
        base["Identificador_Produto"].astype(str)
        if "Identificador_Produto" in base.columns
        else pd.Series("", index=base.index, dtype=str)
    )
    keys = [_id_chave_estrutural_para_linha(u, p) for u, p in zip(uu.tolist(), ip.tolist())]
    base["_Id_Key_Dedup"] = keys
    base["_POSSIVEL_CONFLITO_DEDUP"] = False
    for _, g in base.groupby(chave_cols, sort=False):
        ids = {str(x).strip() for x in g["_Id_Key_Dedup"].tolist() if str(x).strip()}
        if len(ids) > 1:
            base.loc[g.index, "_POSSIVEL_CONFLITO_DEDUP"] = True
    base["__dedup_subkey"] = ""
    mask = base["_POSSIVEL_CONFLITO_DEDUP"].astype(bool)
    base.loc[mask, "__dedup_subkey"] = base.loc[mask, "_Id_Key_Dedup"].fillna("").astype(str)
    return base


def _debug_dedup_perda_por_venda(tag, base_antes, df_depois, col_venda="Venda"):
    if not DEBUG_DADOS or base_antes is None or df_depois is None:
        return
    if col_venda not in base_antes.columns or col_venda not in df_depois.columns:
        return
    ba = base_antes.groupby(base_antes[col_venda].astype(str).str.strip()).size()
    bd = df_depois.groupby(df_depois[col_venda].astype(str).str.strip()).size()
    delta = (ba - bd.reindex(ba.index).fillna(0)).astype(int)
    delta = delta[delta > 0].sort_values(ascending=False)
    if delta.empty:
        return
    print(f"[DEBUG][DEDUP_{tag}_VENDA] maiores_perdas_linhas(top12)={delta.head(12).to_dict()}")


def deduplicar_receber(df_receber, alertas_auditoria_out=None):
    if df_receber is None or df_receber.empty:
        return df_receber

    n_antes = len(df_receber)
    base = adicionar_chave_cliente(df_receber)

    if "Identificador_Produto" not in base.columns:
        base["Identificador_Produto"] = ""

    if "Status_Vencimento" not in base.columns:
        base["Status_Vencimento"] = ""

    if "Tipo" not in base.columns:
        base["Tipo"] = ""
    if "Unidades" not in base.columns:
        base["Unidades"] = ""

    base["Parcela_Norm"] = base["Parcela"].apply(normalizar_parcela)
    base["Vencimento_Key"] = pd.to_datetime(base["Vencimento"], errors="coerce").dt.strftime("%Y-%m-%d")
    base["Vencimento_Key"] = base["Vencimento_Key"].fillna("").astype(str)
    base["Tipo_Norm"] = base["Tipo"].astype(str).str.strip().str.upper()
    base["_Pr_R"] = pd.to_numeric(base["Principal"], errors="coerce").fillna(0).round(2)
    base["_Vlr_R"] = pd.to_numeric(base["Vlr_Parcela"], errors="coerce").fillna(0).round(2)

    # Chave contratual: Unidades não é eixo — duplicatas com unidade truncada/diferente colapsam aqui.
    chave = ["Venda", "Cliente_Base", "Parcela_Norm", "Vencimento_Key", "Tipo_Norm", "_Pr_R", "_Vlr_R"]
    base = _aplicar_split_dedup_por_identificador(base, chave)
    chave_g = chave + ["__dedup_subkey"]

    if alertas_auditoria_out is not None:
        alertas_auditoria_out.extend(coletar_alertas_grupos_deduplicacao(base, chave_g, "RECEBER"))

    df = (
        base.groupby(chave_g, as_index=False)
        .agg({
            "Cliente": escolher_cliente_exibicao,
            "Unidades": _escolher_unidades_pos_dedup_estrutural,
            "Emp/Obra": "first",
            "Descricao_Produto": "first",
            "Nr_Person": lambda x: ", ".join(
                sorted(set([str(v) for v in x if str(v).strip() != ""]))
            ),
            "Identificador_Produto": lambda x: escolher_identificador_melhor([
                normalizar_identificador(v)
                for v in x
                if str(v).strip() != "" and not identificador_truncado(v)
            ]),
            "Principal": "sum",
            "Correcao": "sum",
            "Juros_Atraso": "sum",
            "Multa_Atraso": "sum",
            "Vlr_Parcela": "max",
            "Parc_Num": "max",
            "Parc_Total": "max",
            "Correcao_Atraso": "sum",
            "Status_Vencimento": _moda_status_vencimento_agrupado,
            "Vencimento": "first",
            "_POSSIVEL_CONFLITO_DEDUP": "max",
        })
    )

    df.rename(
        columns={"_POSSIVEL_CONFLITO_DEDUP": "POSSIVEL_CONFLITO_DUPLICIDADE"},
        inplace=True,
    )
    df["POSSIVEL_CONFLITO_DUPLICIDADE"] = df["POSSIVEL_CONFLITO_DUPLICIDADE"].fillna(False).astype(bool)

    df.drop(
        columns=["Vencimento_Key", "Tipo_Norm", "_Pr_R", "_Vlr_R", "__dedup_subkey"],
        inplace=True,
        errors="ignore",
    )

    if "Parcela" in df.columns:
        df.drop(columns=["Parcela"], inplace=True, errors="ignore")
    df.rename(columns={"Parcela_Norm": "Parcela"}, inplace=True)

    _log_volume_dedup("RECEBER", n_antes, len(df))
    _debug_dedup_perda_por_venda("RECEBER", base, df)
    if DEBUG_DADOS and df["POSSIVEL_CONFLITO_DUPLICIDADE"].any():
        n_c = int(df["POSSIVEL_CONFLITO_DUPLICIDADE"].sum())
        print(f"[DEBUG][DEDUP_RECEBER] linhas_com_POSSIVEL_CONFLITO_DUPLICIDADE={n_c}")
    return df


def deduplicar_recebidos(df_recebidos, alertas_auditoria_out=None):
    if df_recebidos is None or df_recebidos.empty:
        return df_recebidos

    n_antes = len(df_recebidos)
    base = adicionar_chave_cliente(df_recebidos)

    if "Identificador_Produto" not in base.columns:
        base["Identificador_Produto"] = ""

    if "Tipo" not in base.columns:
        base["Tipo"] = ""

    base["Parcela_Norm"] = base["Parcela"].apply(normalizar_parcela)
    base["Data_Rec_Key"] = pd.to_datetime(base["Data_Rec"], errors="coerce").dt.strftime("%Y-%m-%d")
    base["Data_Rec_Key"] = base["Data_Rec_Key"].fillna("").astype(str)
    base["Tipo_Norm"] = base["Tipo"].astype(str).str.strip().str.upper()
    base["_Pr_R"] = pd.to_numeric(base["Principal"], errors="coerce").fillna(0).round(2)
    base["_TD_R"] = pd.to_numeric(base["Total_Dep"], errors="coerce").fillna(0).round(2)

    # Chave contratual: evita duplicidade silenciosa por variação de Unidades ou ruído numérico fino.
    chave_evento = [
        "Venda",
        "Cliente_Base",
        "Parcela_Norm",
        "Data_Rec_Key",
        "Tipo_Norm",
        "_Pr_R",
        "_TD_R",
    ]
    base = _aplicar_split_dedup_por_identificador(base, chave_evento)
    chave_g = chave_evento + ["__dedup_subkey"]

    if alertas_auditoria_out is not None:
        alertas_auditoria_out.extend(coletar_alertas_grupos_deduplicacao(base, chave_g, "RECEBIDOS"))

    df = (
        base.groupby(chave_g, as_index=False)
        .agg({
            "Cliente": escolher_cliente_exibicao,
            "Emp/Obra": "first",
            "Descricao_Produto": "first",
            "Nr_Person": lambda x: ", ".join(
                sorted(set([str(v) for v in x if str(v).strip() != ""]))
            ),
            "Unidades": lambda x: sorted(set([
                normalizar_identificador(v)
                for v in x
                if str(v).strip() != "" and not identificador_truncado(v)
            ])),
            "Identificador_Produto": lambda x: sorted(set([
                normalizar_identificador(v)
                for v in x
                if str(v).strip() != "" and not identificador_truncado(v)
            ])),
            "Vlr_Parcela": "max",
            "Correcao": "max",
            "Multa_Atraso": "max",
            "Juros_Atraso": "max",
            "Total_Dep": "max",
            "Total_Nao_Dep": "max",
            "Data_Rec": "first",
            "Parc_Num": "max",
            "Parc_Total": "max",
            "_POSSIVEL_CONFLITO_DEDUP": "max",
        })
    )

    def escolher_identificador_lista(lista_ids):
        todos = []
        for item in lista_ids:
            if isinstance(item, list):
                todos.extend(item)

        if not todos:
            return ""

        freq = Counter(todos)
        return sorted(
            freq.keys(),
            key=lambda x: (freq[x], score_identificador(x), len(x)),
            reverse=True
        )[0]

    df["Unidades"] = df["Unidades"].apply(escolher_identificador_lista)
    df["Identificador_Produto"] = df["Identificador_Produto"].apply(escolher_identificador_lista)
    df["is_paid"] = 1

    df.rename(
        columns={"_POSSIVEL_CONFLITO_DEDUP": "POSSIVEL_CONFLITO_DUPLICIDADE"},
        inplace=True,
    )
    if "POSSIVEL_CONFLITO_DUPLICIDADE" in df.columns:
        df["POSSIVEL_CONFLITO_DUPLICIDADE"] = df["POSSIVEL_CONFLITO_DUPLICIDADE"].fillna(False).astype(bool)

    df.drop(
        columns=["Data_Rec_Key", "Tipo_Norm", "_Pr_R", "_TD_R", "__dedup_subkey"],
        inplace=True,
        errors="ignore",
    )

    if "Parcela" in df.columns:
        df.drop(columns=["Parcela"], inplace=True, errors="ignore")
    df.rename(columns={"Parcela_Norm": "Parcela"}, inplace=True)

    _log_volume_dedup("RECEBIDOS", n_antes, len(df))
    _debug_dedup_perda_por_venda("RECEBIDOS", base, df)
    if DEBUG_DADOS and "POSSIVEL_CONFLITO_DUPLICIDADE" in df.columns and df["POSSIVEL_CONFLITO_DUPLICIDADE"].any():
        n_c = int(df["POSSIVEL_CONFLITO_DUPLICIDADE"].sum())
        print(f"[DEBUG][DEDUP_RECEBIDOS] linhas_com_POSSIVEL_CONFLITO_DUPLICIDADE={n_c}")
    return df


# =========================
# ABAS AUXILIARES
# =========================
def gerar_aba_sem_identificador(df_receber, df_recebidos):
    frames = []

    if df_receber is not None and not df_receber.empty:
        a = df_receber.copy()
        a["Origem"] = "Dados Receber"
        if "Total_Dep" not in a.columns:
            a["Total_Dep"] = 0.0
        if "Status_Vencimento" not in a.columns:
            a["Status_Vencimento"] = ""
        frames.append(
            a[[
                "Origem", "Emp/Obra", "Venda", "Cliente", "Nr_Person", "Unidades", "Parcela",
                "Status_Vencimento", "Principal", "Correcao", "Juros_Atraso", "Multa_Atraso",
                "Correcao_Atraso", "Vlr_Parcela", "Total_Dep"
            ]]
        )

    if df_recebidos is not None and not df_recebidos.empty:
        b = df_recebidos.copy()
        b["Origem"] = "Dados Recebidos"
        if "Correcao_Atraso" not in b.columns:
            b["Correcao_Atraso"] = 0.0
        if "Status_Vencimento" not in b.columns:
            b["Status_Vencimento"] = ""
        frames.append(
            b[[
                "Origem", "Emp/Obra", "Venda", "Cliente", "Nr_Person", "Unidades", "Parcela",
                "Status_Vencimento", "Principal", "Correcao", "Juros_Atraso", "Multa_Atraso",
                "Correcao_Atraso", "Vlr_Parcela", "Total_Dep"
            ]]
        )

    if not frames:
        return pd.DataFrame(
            columns=[
                "Origem", "Emp/Obra", "Venda", "Cliente", "Nr_Person", "Unidades", "Parcela",
                "Status_Vencimento", "Principal", "Correcao", "Juros_Atraso", "Multa_Atraso",
                "Correcao_Atraso", "Vlr_Parcela", "Total_Dep", "Motivo"
            ]
        )

    base = pd.concat(frames, ignore_index=True)
    base = adicionar_chave_cliente(base)
    if "Parcela" not in base.columns:
        base["Parcela"] = ""
    base["Parcela"] = base["Parcela"].fillna("").astype(str).str.strip()

    def motivo_identificador(x):
        x = normalizar_identificador(x)

        if x == "":
            return "Sem identificador preenchido"

        if not identificador_tem_formato_endereco(x):
            return "Identificador sem formato válido"

        if identificador_truncado(x):
            return "Identificador truncado/inválido"

        return ""

    base["Motivo"] = base["Unidades"].apply(motivo_identificador)

    # Reconcilia sem identificador por Venda + Cliente_Base + Parcela
    # antes de enviar para aba "Sem Identificador".
    chaves_com_id = set(
        (
            base.loc[base["Motivo"] == "", "Venda"].astype(str).str.strip()
            + "||"
            + base.loc[base["Motivo"] == "", "Cliente_Base"].astype(str).str.strip()
            + "||"
            + base.loc[base["Motivo"] == "", "Parcela"].astype(str).str.strip()
        ).tolist()
    )
    chave_linha = (
        base["Venda"].astype(str).str.strip()
        + "||"
        + base["Cliente_Base"].astype(str).str.strip()
        + "||"
        + base["Parcela"].astype(str).str.strip()
    )
    mask_sem_id = base["Motivo"] != ""
    mask_recuperavel = mask_sem_id & chave_linha.isin(chaves_com_id)
    base.loc[mask_recuperavel, "Motivo"] = ""

    # Regra forte adicional: se a mesma chave (Venda + Cliente_Base + Parcela)
    # aparece nas duas origens, considera vinculo recuperado mesmo sem identificador.
    origem_por_chave = (
        base.assign(_chave=chave_linha)
        .groupby("_chave")["Origem"]
        .nunique()
        .to_dict()
    )
    mask_chave_duas_origens = chave_linha.map(lambda x: origem_por_chave.get(x, 0) >= 2)
    base.loc[(base["Motivo"] != "") & mask_chave_duas_origens, "Motivo"] = ""

    # Venda + Parcela: se a parcela aparece nas duas origens, ha vinculo operacional.
    base["VP"] = (
        base["Venda"].astype(str).str.strip()
        + "||"
        + base["Parcela"].astype(str).str.strip()
    )
    vp_origens = base.groupby("VP")["Origem"].nunique()
    mask_vp_ambos = base["VP"].map(vp_origens) >= 2
    base.loc[(base["Motivo"] != "") & mask_vp_ambos, "Motivo"] = ""
    base.drop(columns=["VP"], inplace=True, errors="ignore")

    saida = base[base["Motivo"] != ""].copy()
    return saida[[
        "Origem", "Emp/Obra", "Venda", "Cliente", "Nr_Person", "Unidades", "Parcela",
        "Status_Vencimento", "Principal", "Correcao", "Juros_Atraso", "Multa_Atraso",
        "Correcao_Atraso", "Vlr_Parcela", "Total_Dep", "Motivo"
    ]]


def gerar_aba_auditoria_vendas(df_receber, df_recebidos):
    if (df_receber is None or df_receber.empty) and (df_recebidos is None or df_recebidos.empty):
        return pd.DataFrame(columns=[
            "Venda", "Cliente_Base", "Cliente_Exibicao",
            "Qtd Nr_Person Receber", "Qtd Nr_Person Recebidos",
            "Qtd Identificadores Receber", "Qtd Identificadores Recebidos",
            "Identificador Final", "Conferencia"
        ])

    if df_receber is None:
        df_receber = pd.DataFrame()

    if df_recebidos is None:
        df_recebidos = pd.DataFrame()

    if not df_receber.empty:
        df_receber = adicionar_chave_cliente(df_receber)

    if not df_recebidos.empty:
        df_recebidos = adicionar_chave_cliente(df_recebidos)

    base_r = pd.DataFrame()
    base_p = pd.DataFrame()

    if not df_receber.empty:
        base_r = (
            df_receber.groupby(["Venda", "Cliente_Base"], as_index=False)
            .agg({
                "Cliente": escolher_cliente_exibicao,
                "Nr_Person": lambda x: len(set([str(v) for v in x if str(v).strip() != ""])),
                "Unidades": lambda x: len(set([
                    str(v) for v in x
                    if str(v).strip() != "" and not identificador_truncado(v)
                ])),
            })
            .rename(columns={
                "Cliente": "Cliente_Exibicao",
                "Nr_Person": "Qtd Nr_Person Receber",
                "Unidades": "Qtd Identificadores Receber",
            })
        )

    if not df_recebidos.empty:
        base_p = (
            df_recebidos.groupby(["Venda", "Cliente_Base"], as_index=False)
            .agg({
                "Cliente": escolher_cliente_exibicao,
                "Nr_Person": lambda x: len(set([str(v) for v in x if str(v).strip() != ""])),
                "Unidades": lambda x: len(set([
                    str(v) for v in x
                    if str(v).strip() != "" and not identificador_truncado(v)
                ])),
            })
            .rename(columns={
                "Cliente": "Cliente_Exibicao_Rec",
                "Nr_Person": "Qtd Nr_Person Recebidos",
                "Unidades": "Qtd Identificadores Recebidos",
            })
        )

    aud = pd.merge(base_r, base_p, on=["Venda", "Cliente_Base"], how="outer").fillna(0)

    if "Cliente_Exibicao" not in aud.columns:
        aud["Cliente_Exibicao"] = ""

    if "Cliente_Exibicao_Rec" in aud.columns:
        aud["Cliente_Exibicao"] = aud["Cliente_Exibicao"].replace("", pd.NA).fillna(aud["Cliente_Exibicao_Rec"])
        aud.drop(columns=["Cliente_Exibicao_Rec"], inplace=True, errors="ignore")

    mapa = escolher_identificador_por_venda_cliente_base(
        pd.concat([df_receber, df_recebidos], ignore_index=True)
        if (not df_receber.empty or not df_recebidos.empty)
        else pd.DataFrame()
    )

    if not mapa.empty:
        aud = aud.merge(
            mapa[["Venda", "Cliente_Base", "Identificador_Padrao"]],
            on=["Venda", "Cliente_Base"],
            how="left"
        )
        aud = aud.rename(columns={"Identificador_Padrao": "Identificador Final"})
    else:
        aud["Identificador Final"] = ""

    def status_conf(row):
        conflitos = 0

        if row.get("Qtd Nr_Person Receber", 0) > 1:
            conflitos += 1
        if row.get("Qtd Nr_Person Recebidos", 0) > 1:
            conflitos += 1
        if row.get("Qtd Identificadores Receber", 0) > 1:
            conflitos += 1
        if row.get("Qtd Identificadores Recebidos", 0) > 1:
            conflitos += 1

        return "REVISAR" if conflitos > 0 else "OK"

    aud["Conferencia"] = aud.apply(status_conf, axis=1)

    return aud


def unificar_cliente_por_venda_parcela(df):
    """
    Para cada Venda + Parcela, unifica o nome de cliente (mais frequente).
    Garante que o mesmo contrato/parcela nao fique fragmentado em Cliente_Base
    distintos (caso tipico sem identificador / nomes truncados).
    Linhas sem numero de parcela nao sao fundidas por Parcela_Key vazia.
    """
    if df is None or df.empty:
        return df
    if "Venda" not in df.columns or "Parcela" not in df.columns or "Cliente" not in df.columns:
        return df

    out = df.copy()
    out = adicionar_chave_cliente(out)
    out["Venda"] = out["Venda"].fillna("").astype(str).str.strip()
    out["Parcela"] = out["Parcela"].fillna("").astype(str).str.strip()
    out["Parcela_Key"] = out["Venda"] + "||" + out["Parcela"]
    valid = out["Parcela"].astype(str).str.strip() != ""
    if valid.any():
        mapa_cl = out.loc[valid].groupby("Parcela_Key", dropna=False)["Cliente"].agg(escolher_cliente_exibicao)
        out.loc[valid, "Cliente"] = out.loc[valid, "Parcela_Key"].map(mapa_cl).fillna(out.loc[valid, "Cliente"])

    out = out.drop(columns=["Cliente_Norm", "Cliente_Base", "Chave_Cliente", "Parcela_Key"], errors="ignore")
    return adicionar_chave_cliente(out)


def sem_identificador_ainda_com_chave_na_base(aba_sem_id, df_receber):
    """
    Linhas que permanecem em Sem Identificador mas ja existem na base com a mesma
    chave Venda + Cliente_Base + Parcela (deveriam ter sido excluidas da aba).
    """
    if aba_sem_id is None or aba_sem_id.empty or df_receber is None or df_receber.empty:
        return pd.DataFrame()
    dr = adicionar_chave_cliente(df_receber.copy())
    if "Parcela" not in dr.columns:
        return pd.DataFrame()
    dr["__k"] = (
        dr["Venda"].astype(str).str.strip()
        + "||"
        + dr["Cliente_Base"].astype(str).str.strip()
        + "||"
        + dr["Parcela"].astype(str).str.strip()
    )
    chaves = set(dr["__k"].tolist())
    a = adicionar_chave_cliente(aba_sem_id.copy())
    a["__k"] = (
        a["Venda"].astype(str).str.strip()
        + "||"
        + a["Cliente_Base"].astype(str).str.strip()
        + "||"
        + a["Parcela"].astype(str).str.strip()
    )
    return a[a["__k"].isin(chaves)].drop(columns=["__k"], errors="ignore")


# =========================
# CONSOLIDADO
# =========================
def _validar_invariantes_exportacao_consolidado(consolidado_df):
    """
    Checagens explícitas (somente DEBUG_DADOS): quantidades e colunas financeiras presentes.
    Não altera dados; registra divergências no terminal para auditoria.
    """
    if consolidado_df is None or consolidado_df.empty or not DEBUG_DADOS:
        return
    qt = pd.to_numeric(consolidado_df.get("Qtd.Parc.Total", 0), errors="coerce").fillna(0).astype(int)
    qp = pd.to_numeric(consolidado_df.get("Qtd.Parc.Paga", 0), errors="coerce").fillna(0).astype(int)
    qa = pd.to_numeric(consolidado_df.get("Qtd.Parc.Atrasada", 0), errors="coerce").fillna(0).astype(int)
    qv = pd.to_numeric(consolidado_df.get("Qtd.Parc.A Vencer", 0), errors="coerce").fillna(0).astype(int)
    m = (qt < qp) | (qt < qa) | (qt < qv)
    if m.any():
        vendas = consolidado_df.loc[m, "Venda"].astype(str).str.strip().head(12).tolist()
        print(f"[DEBUG][INVARIANTE] Qtd.Parc.Total < componente operacional | vendas(amostra)={vendas}")
    for col in ("Vl.Pago", "Vl.Vencer", "Vl.Principal (Encargos)", "Vl.Carteira", "% Pago"):
        if col not in consolidado_df.columns:
            print(f"[DEBUG][INVARIANTE] coluna ausente no consolidado final: {col}")


def _nunique_venda_serie(serie_venda) -> int:
    s = serie_venda.fillna("").astype(str).str.strip()
    s = s[s != ""]
    return int(s.nunique()) if len(s) else 0


def _serie_soma_float(serie) -> float:
    return float(pd.to_numeric(serie, errors="coerce").fillna(0).astype(float).sum())


def _serie_soma_int(serie) -> int:
    return int(pd.to_numeric(serie, errors="coerce").fillna(0).astype(float).sum())


def _status_geral_pct_inad(pct_decimal) -> str:
    try:
        p = float(pct_decimal)
    except (TypeError, ValueError):
        return ""
    if p >= 0.15:
        return "ALTO"
    if p >= 0.05:
        return "MÉDIO"
    return "BAIXO"


def montar_dataframe_resumo_geral(df_consolidado: pd.DataFrame) -> pd.DataFrame:
    """
    Agrega o Consolidado já calculado por (Emp/Obra, Empreendimento).
    Soma quantidades e valores; recalcula Vl.Carteira e percentuais sobre os totais do grupo
    via _recalcular_vl_carteira_e_percentuais (mesmo contrato do motor).
    Ordenação: maior % Inadimplência; empate → maior valor de inadimplência (encargos).
    """
    if df_consolidado is None or getattr(df_consolidado, "empty", True):
        return pd.DataFrame()
    need = (
        "Emp/Obra",
        "Empreendimento",
        "Venda",
        "Qtd.Parc.Paga",
        "Vl.Pago",
        "Qtd.Parc.Atrasada",
        "Vl.Principal (Encargos)",
        "Qtd.Parc.A Vencer",
        "Vl.Vencer",
    )
    if any(c not in df_consolidado.columns for c in need):
        return pd.DataFrame()

    d = df_consolidado.loc[:, list(need)].copy()
    d["Emp/Obra"] = d["Emp/Obra"].fillna("").astype(str).str.strip().apply(normalizar_emp_obra)
    d["Empreendimento"] = (
        d["Empreendimento"]
        .fillna("")
        .astype(str)
        .apply(limpar_nome_empreendimento)
        .astype(str)
        .str.strip()
    )
    # Lote / linhas esparsas: se Empreendimento vier vazio só em parte das vendas da mesma obra,
    # usar a moda do nome legal já presente no consolidado para essa Emp/Obra (evita grupo com empreendimento em branco no resumo).
    preench = d[d["Empreendimento"].str.len() > 0]
    if not preench.empty:
        def _moda_grp(s: pd.Series) -> str:
            return escolher_moda_texto([str(x).strip() for x in s.tolist() if str(x).strip()])

        mapa_emp_por_obra = preench.groupby("Emp/Obra", dropna=False)["Empreendimento"].agg(_moda_grp)
        vazio = d["Empreendimento"].str.len() == 0
        d.loc[vazio, "Empreendimento"] = (
            d.loc[vazio, "Emp/Obra"].map(mapa_emp_por_obra).fillna("").astype(str).str.strip()
        )
    # Mapa oficial (sigla) quando ainda não há nome legal no consolidado.
    vazio_of = d["Empreendimento"].str.len() == 0
    if bool(vazio_of.any()):
        nomes_of = d.loc[vazio_of, "Emp/Obra"].map(empreendimento_oficial_para_emp_obra)
        ok_of = nomes_of.fillna("").astype(str).str.strip().ne("")
        d.loc[vazio_of & ok_of, "Empreendimento"] = nomes_of.loc[vazio_of & ok_of].astype(str).str.strip()
    # Padronização final por mapa oficial (sigla conhecida sempre prevalece na exibição).
    d["Empreendimento"] = _aplicar_nome_oficial_em_series(d["Emp/Obra"], d["Empreendimento"])
    # Apresentação executiva do resumo: evita rótulo vazio remanescente.
    d.loc[d["Empreendimento"].str.len() == 0, "Empreendimento"] = "NÃO INFORMADO"

    g = d.groupby(["Emp/Obra", "Empreendimento"], dropna=False, as_index=False)
    out = g.agg(
        _qtd_vendas=("Venda", _nunique_venda_serie),
        _qtd_parc_paga=("Qtd.Parc.Paga", _serie_soma_int),
        _vl_pago=("Vl.Pago", _serie_soma_float),
        _qtd_parc_inad=("Qtd.Parc.Atrasada", _serie_soma_int),
        _vl_enc=("Vl.Principal (Encargos)", _serie_soma_float),
        _qtd_parc_av=("Qtd.Parc.A Vencer", _serie_soma_int),
        _vl_vencer=("Vl.Vencer", _serie_soma_float),
    )
    out["Vl.Pago"] = out["_vl_pago"].round(2)
    out["Vl.Principal (Encargos)"] = out["_vl_enc"].round(2)
    out["Vl.Vencer"] = out["_vl_vencer"].round(2)
    out = _recalcular_vl_carteira_e_percentuais(out)
    out = out.sort_values(
        by=["% Inadimplência", "Vl.Principal (Encargos)"],
        ascending=[False, False],
    ).reset_index(drop=True)

    out["STATUS GERAL"] = out["% Inadimplência"].map(_status_geral_pct_inad)

    colunas_excel = [
        "Emp/Obra",
        "Empreendimento",
        "QTD VENDAS",
        "QTD PARC. PAGA",
        "VALOR PAGO",
        "QTD PARC. INADIMPLÊNCIA",
        "VALOR INADIMPLÊNCIA",
        "QTD PARC. A VENCER",
        "VALOR A VENCER",
        "VL.CARTEIRA",
        "% PAGO",
        "% INADIMPLÊNCIA",
        "% A VENCER",
        "STATUS GERAL",
    ]
    exibir = pd.DataFrame(
        {
            "Emp/Obra": out["Emp/Obra"],
            "Empreendimento": out["Empreendimento"],
            "QTD VENDAS": out["_qtd_vendas"].astype(int),
            "QTD PARC. PAGA": out["_qtd_parc_paga"].astype(int),
            "VALOR PAGO": out["Vl.Pago"],
            "QTD PARC. INADIMPLÊNCIA": out["_qtd_parc_inad"].astype(int),
            "VALOR INADIMPLÊNCIA": out["Vl.Principal (Encargos)"],
            "QTD PARC. A VENCER": out["_qtd_parc_av"].astype(int),
            "VALOR A VENCER": out["Vl.Vencer"],
            "VL.CARTEIRA": out["Vl.Carteira"],
            "% PAGO": out["% Pago"],
            "% INADIMPLÊNCIA": out["% Inadimplência"],
            "% A VENCER": out["% A Vencer"],
            "STATUS GERAL": out["STATUS GERAL"],
        }
    )
    return exibir[colunas_excel]


def ler_dataframe_consolidado_de_xlsx_motor(caminho_xlsx: str) -> pd.DataFrame:
    """
    Lê a aba de Consolidado gerada pelo motor (cabeçalho tabular na linha 8 do Excel).
    Ignora abas cujo nome sugira apenas critérios.
    """
    if not caminho_xlsx or not os.path.isfile(caminho_xlsx):
        return pd.DataFrame()
    try:
        xl = pd.ExcelFile(caminho_xlsx)
    except Exception:
        return pd.DataFrame()
    alvo = None
    for s in xl.sheet_names:
        su = str(s).upper().replace("Í", "I")
        if "CONSOLIDADO" in su and "CRIT" not in su:
            alvo = s
            break
    if alvo is None and xl.sheet_names:
        alvo = xl.sheet_names[0]
    if alvo is None:
        return pd.DataFrame()
    try:
        df = pd.read_excel(caminho_xlsx, sheet_name=alvo, header=7)
    except Exception:
        return pd.DataFrame()
    # Compatibilidade: cabeçalhos podem estar padronizados em maiúsculo no XLSX estilizado.
    ren = {
        "EMP/OBRA": "Emp/Obra",
        "EMPREENDIMENTO": "Empreendimento",
        "VENDA": "Venda",
        "CLIENTE": "Cliente",
        "IDENTIFICADOR": "Identificador",
        "STATUS VENDA": "Status venda",
        "VALOR DA PARCELA": "Valor Da Parcela",
        "QTD.PARC.TOTAL": "Qtd.Parc.Total",
        "QTD.PARC.PAGA": "Qtd.Parc.Paga",
        "QTD.PARC.ATRASADA": "Qtd.Parc.Atrasada",
        "QTD.PARC.PAGO": "Qtd.Parc.Paga",
        "VL.PAGO": "Vl.Pago",
        "QTD.PARC.VENCIDA": "Qtd.Parc.Atrasada",
        "VL.PRINCIPAL (ENCARGOS)": "Vl.Principal (Encargos)",
        "QTD.PARC.A VENCER": "Qtd.Parc.A Vencer",
        "VL.A VENCER": "Vl.Vencer",
        "VL.VENCER": "Vl.Vencer",
    }
    cols_up = {str(c).strip().upper(): c for c in df.columns}
    for src_up, dst in ren.items():
        src_real = cols_up.get(src_up)
        if src_real is not None and dst not in df.columns:
            df = df.rename(columns={src_real: dst})
    # Layout estilizado do consolidado usa alguns rótulos repetidos na linha 8.
    # Quando isso ocorrer, preservamos o contrato do motor pela posição A:AA.
    mapa_posicional = {
        0: "Emp/Obra",
        1: "Empreendimento",
        2: "Venda",
        3: "Cliente",
        4: "Identificador",
        5: "Status venda",
        6: "Valor Da Parcela",
        7: "Qtd.Parc.Total",
        8: "Qtd.Parc.Paga",
        9: "Vl.Pago",
        10: "Qtd.Parc.Atrasada",
        11: "Vl.Principal Atrasado",
        12: "Vl.Correção",
        13: "Vl.Juros",
        14: "Vl.Multas",
        15: "Vl.Correção Atraso",
        16: "Vl.Principal (Encargos)",
        17: "Qtd.Parc.A Vencer",
        18: "Vl.Vencer",
        19: "Vl.Carteira",
        20: "% Pago",
        21: "% Inadimplência",
        22: "% A Vencer",
        23: "DIA VENCIMENTO",
        24: "Status Construção",
        25: "Judicializado",
        26: "APORTE",
    }
    colunas_essenciais = {"Emp/Obra", "Venda", "Qtd.Parc.Paga", "Qtd.Parc.A Vencer", "Vl.Vencer"}
    if not colunas_essenciais.issubset(set(df.columns)) and len(df.columns) >= 27:
        novos = []
        for idx, col in enumerate(df.columns):
            novos.append(mapa_posicional.get(idx, col))
        df.columns = novos
    return df


def _recalcular_vl_carteira_e_percentuais(consolidado_df):
    """
    Vl.Carteira segue CARTEIRA_MODO_OFICIAL:
      - SALDO_ABERTO: Encargos + Vl.Vencer
      - POSICAO_TOTAL: Vl.Pago + Encargos + Vl.Vencer
    Percentuais por linha usam o mesmo denominador da posição (Pago + Encargos + Vencer), somando 100%
    após normalização (evita resíduo de arredondamento).
    """
    if consolidado_df is None or consolidado_df.empty:
        return consolidado_df
    c = consolidado_df
    vl_p = pd.to_numeric(c.get("Vl.Pago", 0), errors="coerce").fillna(0).astype(float)
    enc = pd.to_numeric(c.get("Vl.Principal (Encargos)", 0), errors="coerce").fillna(0).astype(float)
    vl_v = pd.to_numeric(c.get("Vl.Vencer", 0), errors="coerce").fillna(0).astype(float)
    c["Vl.Carteira"] = _calcular_vl_carteira_oficial(vl_p, enc, vl_v)
    tot = vl_p + enc + vl_v
    mask = tot.abs() > 1e-12
    raw_p = (vl_p / tot).where(mask, 0.0).fillna(0.0)
    raw_i = (enc / tot).where(mask, 0.0).fillna(0.0)
    raw_a = (vl_v / tot).where(mask, 0.0).fillna(0.0)
    s = raw_p + raw_i + raw_a
    factor = (1.0 / s).where(s > 1e-12, 1.0)
    c["% Pago"] = (raw_p * factor).clip(0, 1).round(6)
    c["% Inadimplência"] = (raw_i * factor).clip(0, 1).round(6)
    c["% A Vencer"] = (raw_a * factor).clip(0, 1).round(6)
    return c


def _calcular_vl_carteira_oficial(vl_p, enc, vl_v):
    """Calcula Vl.Carteira segundo o modo oficial ativo (ponto único de decisão funcional)."""
    if str(CARTEIRA_MODO_OFICIAL).strip().upper() == "POSICAO_TOTAL":
        return (vl_p + enc + vl_v).round(2)
    return (enc + vl_v).round(2)


def _descricao_vl_carteira_modo():
    modo = str(CARTEIRA_MODO_OFICIAL).strip().upper()
    if modo == "POSICAO_TOTAL":
        return "Vl.Carteira = Vl.Pago + Vl.Principal (Encargos) + Vl.Vencer (posição total da venda)"
    return "Vl.Carteira = Vl.Principal (Encargos) + Vl.Vencer (saldo em aberto)"


def _identificador_linha_bases_uau(row) -> str:
    for c in ("Identificador_Produto", "Unidades"):
        if c not in row.index:
            continue
        v = row.get(c)
        if v is None or (isinstance(v, float) and pd.isna(v)):
            continue
        s = str(v).strip()
        if s:
            return s
    return ""


def montar_dataframe_relatorio_analitico(df_receber, df_recebidos):
    """Uma linha por lançamento nas bases tratadas; colunas fixas do relatório analítico."""
    linhas = []
    if df_receber is not None and not df_receber.empty:
        for _, r in df_receber.iterrows():
            v = str(r.get("Venda", "") or "").strip()
            if not v:
                continue
            vp = float(pd.to_numeric(r.get("Vlr_Parcela", 0), errors="coerce") or 0)
            linhas.append({
                "VENDA": v,
                "CLIENTE": str(r.get("Cliente", "") or "").strip(),
                "IDENTIFICADOR": _identificador_linha_bases_uau(r),
                "VL.PARCELA": vp,
                "ORIGEM": "DADOS RECEBER",
            })
    if df_recebidos is not None and not df_recebidos.empty:
        for _, r in df_recebidos.iterrows():
            v = str(r.get("Venda", "") or "").strip()
            if not v:
                continue
            vp = float(pd.to_numeric(r.get("Vlr_Parcela", 0), errors="coerce") or 0)
            linhas.append({
                "VENDA": v,
                "CLIENTE": str(r.get("Cliente", "") or "").strip(),
                "IDENTIFICADOR": _identificador_linha_bases_uau(r),
                "VL.PARCELA": vp,
                "ORIGEM": "DADOS RECEBIDOS",
            })
    cols = ["VENDA", "CLIENTE", "IDENTIFICADOR", "VL.PARCELA", "ORIGEM"]
    if not linhas:
        return pd.DataFrame(columns=cols)
    return pd.DataFrame(linhas, columns=cols)


def _caixa_alta_exibicao_relatorio(df):
    """Padroniza textos exibidos em maiúsculas; preserva numéricos e datas."""
    if df is None or df.empty:
        return df
    out = df.copy()
    for c in out.columns:
        if pd.api.types.is_datetime64_any_dtype(out[c]):
            continue
        if pd.api.types.is_numeric_dtype(out[c]):
            continue

        def _up_cell(x):
            if x is None or (isinstance(x, float) and pd.isna(x)):
                return x
            if isinstance(x, (int, float)) and not isinstance(x, bool):
                return x
            return str(x).strip().upper()

        out[c] = out[c].map(_up_cell)
    return out


def _padronizar_rotulo_coluna_exibicao(col) -> str:
    c = str(col or "").strip().upper()
    if not c:
        return c
    mapa_exato = {
        "EMP/OBRA": "EMP/OBRA",
        "VENDA": "VENDA",
        "CLIENTE": "CLIENTE",
        "CLIENTE_BASE": "CLI.BASE",
        "IDENTIFICADOR_PRODUTO": "IDENTIFICADOR",
        "PARCELA": "PARC.(GERAL)",
        "PARC_NUM": "PARC.NUM",
        "PARC_TOTAL": "PARC.TOTAL",
        "VENCIMENTO": "VENC.DATA",
        "STATUS_VENCIMENTO": "STATUS",
        "DIA_VENCIMENTO_BOLETO": "DIA.VENC.",
        "MES_VENCIMENTO": "MES.VENC.",
        "ANO_VENCIMENTO": "ANO.VENC.",
        "CLASSIFICACAO_ADIMPLENCIA": "CLASSIFICAÇÃO",
        "PRINCIPAL": "PRINCIPAL",
        "CORRECAO": "CORREÇÃO",
        "JUROS_ATRASO": "JUROS ATRASO",
        "MULTA_ATRASO": "MULTA ATRASO",
        "CORRECAO_ATRASO": "CORREÇÃO ATRASO",
        "VLR_PARCELA": "VL.PARCELA",
        "DATA_REC": "DATA.REC.",
        "TIPO": "TIPO",
    }
    if c in mapa_exato:
        return mapa_exato[c]
    c = c.replace("PARCELAS RECEBIDAS", "QTD.PARC.PAGO")
    c = c.replace("PARCELAS INADIMPLENTES", "QTD.PARC.VENCIDA")
    c = c.replace("PARCELAS A VENCER", "QTD.PARC.A VENCER")
    c = re.sub(r"\bPARCELAS\b", "QTD.PARCELAS", c)
    c = c.replace("VALOR PAGO", "VL.PAGO")
    c = c.replace("VALOR VENCIDO", "VL.VENCIDO")
    c = c.replace("VALOR A VENCER", "VL.A VENCER")
    return c


def _padronizar_colunas_exibicao(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    out = df.copy()
    out.columns = [_padronizar_rotulo_coluna_exibicao(c) for c in out.columns]
    return out


def _remover_colunas_totalmente_vazias(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    keep = []
    for c in df.columns:
        s = df[c]
        if pd.api.types.is_numeric_dtype(s):
            keep.append(c)
            continue
        sx = s.fillna("").astype(str).str.strip()
        if sx.ne("").any():
            keep.append(c)
    return df[keep] if keep else df


def _validar_integridade_financeira(consolidado_df):
    """
    Confere identidades do motor: Vl.Carteira conforme CARTEIRA_MODO_OFICIAL e coerência de % Pago
    frente ao total de posição (somente diagnóstico; não altera dados).
    """
    if consolidado_df is None or consolidado_df.empty or not DEBUG_DADOS:
        return
    obr = (
        "Vl.Pago",
        "Vl.Vencer",
        "Vl.Principal Atrasado",
        "Vl.Principal (Encargos)",
        "Vl.Carteira",
        "% Pago",
    )
    for col in obr:
        if col not in consolidado_df.columns:
            print(f"[DEBUG][INTEGRIDADE_FIN] coluna ausente: {col}")
    c = consolidado_df
    vl_p = pd.to_numeric(c["Vl.Pago"], errors="coerce").fillna(0).round(2)
    vl_v = pd.to_numeric(c["Vl.Vencer"], errors="coerce").fillna(0).round(2)
    enc = pd.to_numeric(c["Vl.Principal (Encargos)"], errors="coerce").fillna(0).round(2)
    cart = pd.to_numeric(c["Vl.Carteira"], errors="coerce").fillna(0).round(2)
    esp_cart = _calcular_vl_carteira_oficial(vl_p, enc, vl_v)
    diff = (cart - esp_cart).abs()
    if (diff > 0.02).any():
        am = c.loc[diff > 0.02, "Venda"].astype(str).str.strip().head(15).tolist()
        print(
            "[DEBUG][INTEGRIDADE_FIN] Vl.Carteira divergente do modo oficial "
            f"({CARTEIRA_MODO_OFICIAL}) (tol 0.02) | amostra_vendas={am}"
        )
    tot_ref = (vl_p + enc + vl_v).round(8)
    pct = pd.to_numeric(c["% Pago"], errors="coerce").fillna(0).clip(0, 1)
    esp = pd.Series(0.0, index=c.index)
    m2 = tot_ref.abs() > 1e-9
    esp.loc[m2] = (vl_p / tot_ref).loc[m2]
    esp = esp.clip(lower=0, upper=1).fillna(0)
    d_pct = (pct - esp).abs()
    if (d_pct > 0.02).any():
        am = c.loc[d_pct > 0.02, "Venda"].astype(str).str.strip().head(12).tolist()
        print(f"[DEBUG][INTEGRIDADE_FIN] % Pago divergente de Vl.Pago/(Pago+Encargos+Vencer) | amostra_vendas={am}")


def ordenar_dataframe_uau_por_venda_parcela(df, col_data=None):
    """
    Ordenação determinística: Venda, Parc_Num, depois data relevante (Vencimento / Data_Rec) se existir.
    Não confiar na ordem do TXT.
    """
    if df is None or df.empty or "Venda" not in df.columns:
        return df
    d = df.copy()
    v_raw = d["Venda"].astype(str).str.strip()
    v_num = pd.to_numeric(v_raw, errors="coerce")
    if "Parc_Num" in d.columns:
        pn = pd.to_numeric(d["Parc_Num"], errors="coerce").fillna(0).astype(int)
    else:
        pn = pd.Series(0, index=d.index, dtype=int)
    col_d = col_data if col_data and col_data in d.columns else None
    if col_d:
        dt = pd.to_datetime(d[col_d], errors="coerce")
        d = d.assign(_ord_v=v_num.fillna(-1), _ord_vs=v_raw, _ord_pn=pn, _ord_dt=dt)
        d = d.sort_values(
            by=["_ord_v", "_ord_vs", "_ord_pn", "_ord_dt"],
            kind="mergesort",
            na_position="last",
        )
        return d.drop(columns=["_ord_v", "_ord_vs", "_ord_pn", "_ord_dt"]).reset_index(drop=True)
    d = d.assign(_ord_v=v_num.fillna(-1), _ord_vs=v_raw, _ord_pn=pn)
    d = d.sort_values(by=["_ord_v", "_ord_vs", "_ord_pn"], kind="mergesort")
    return d.drop(columns=["_ord_v", "_ord_vs", "_ord_pn"]).reset_index(drop=True)


def montar_consolidado(
    df_receber,
    df_recebidos,
    data_base,
    nome_empreendimento_arquivo="",
    vendas_aporte=None,
    registro_etapas_tempo=None,
    mapa_emp_obra_nome_legal=None,
):
    if df_receber is None:
        df_receber = pd.DataFrame()

    if df_recebidos is None:
        df_recebidos = pd.DataFrame()

    if df_receber.empty and df_recebidos.empty:
        return pd.DataFrame(), data_base, pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), []

    if data_base is None and not df_receber.empty and "Vencimento" in df_receber.columns:
        data_base = df_receber["Vencimento"].dropna().max()

    # Chave-base restaurada para evitar perda/zeramento por fragmentação.
    # A linha final única por venda continua sendo garantida no fechamento.
    chave = ["Venda", "Cliente_Base"]
    pendencias_qtd_total_reconciliacao = []

    def _dbg(msg):
        if DEBUG_DADOS:
            print(f"[DEBUG][montar_consolidado] {msg}")

    def _serie_num(df, col):
        if df is None or df.empty or col not in df.columns:
            return pd.Series(dtype="float64")
        return pd.to_numeric(df[col], errors="coerce").fillna(0)

    def _status_info(df, nome_df):
        if df is None or df.empty:
            _dbg(f"{nome_df}: vazio")
            return
        vendas = df["Venda"].fillna("").astype(str).str.strip().nunique() if "Venda" in df.columns else 0
        _dbg(f"{nome_df}: linhas={len(df)} | vendas_unicas={vendas} | colunas={list(df.columns)}")
        if "Status_Vencimento" in df.columns:
            st = df["Status_Vencimento"].fillna("").astype(str).str.strip().str.upper()
            v = int((st == "VENCIDO").sum())
            a = int((st == "A VENCER").sum())
            _dbg(f"{nome_df}: status VENCIDO={v} | A VENCER={a} | outros={int(len(df)-v-a)}")
            vp = _serie_num(df, "Vlr_Parcela")
            _dbg(
                f"{nome_df}: soma Vlr_Parcela VENCIDO={float(vp.loc[st=='VENCIDO'].sum()):.2f} | "
                f"A VENCER={float(vp.loc[st=='A VENCER'].sum()):.2f}"
            )
        if "Vlr_Parcela" in df.columns:
            _dbg(f"{nome_df}: Vlr_Parcela não nulos={int(pd.to_numeric(df['Vlr_Parcela'], errors='coerce').notna().sum())}")
        cols_ex = [c for c in ["Venda", "Cliente", "Parcela", "Vlr_Parcela", "Principal", "Status_Vencimento"] if c in df.columns]
        if cols_ex and DEBUG_DADOS:
            _dbg(f"{nome_df}: exemplos(5)={df[cols_ex].head(5).to_dict('records')}")

    def _garantir_cliente_base(df, nome_df):
        if df is None:
            return pd.DataFrame(columns=chave)
        if df.empty:
            if "Cliente_Base" not in df.columns:
                return pd.DataFrame(columns=list(dict.fromkeys(list(df.columns) + chave)))
            return df
        if "Cliente_Base" not in df.columns:
            _dbg(f"CHAVE_CHECK: {nome_df} sem Cliente_Base -> aplicando adicionar_chave_cliente()")
            df = adicionar_chave_cliente(df)
        return df

    # =====================================================
    # CONTRATO DO CONSOLIDADO UAU
    # =====================================================
    # 1. Chave de consolidação: somente número da Venda (identificador não define linha)
    # 2. Cliente: nome que mais aparece dentro da venda
    # 3. Identificador: informação auxiliar (moda entre fontes)
    # 4. Valor Da Parcela: moda de Principal por venda
    # 5. Qtd total: denominador dominante (frequência de Parc_Total na venda; empate → maior)
    # 6. Pagos: Contas Recebidas / Total_Dep
    # 7. Inadimplência: somente Status_Vencimento = VENCIDO (encargos / Qtd.Parc.Atrasada)
    # 8. A VENCER: somente Status_Vencimento = A VENCER (Vl.Vencer col. R, Qtd.Parc.A Vencer)
    # 9. Evento 1/1: tratar como contrato único
    # 10. Empreendimento: valor dominante do relatório

    # =====================================================
    # CÓPIAS BRUTAS PREPARADAS PARA FALLBACK FINAL
    # =====================================================
    df_receber_raw = adicionar_chave_cliente(df_receber.copy()) if not df_receber.empty else pd.DataFrame(columns=chave)
    df_recebidos_raw = adicionar_chave_cliente(df_recebidos.copy()) if not df_recebidos.empty else pd.DataFrame(columns=chave)
    df_receber_raw = _garantir_cliente_base(df_receber_raw, "df_receber_raw")
    df_recebidos_raw = _garantir_cliente_base(df_recebidos_raw, "df_recebidos_raw")
    _status_info(df_receber_raw, "df_receber_raw")
    _status_info(df_recebidos_raw, "df_recebidos_raw")

    # =====================================================
    # PREPARAÇÃO PRINCIPAL
    # =====================================================
    alertas_dedup_auditoria = []
    t_montar_inicio = time.perf_counter()
    tempo_dedup_receber = 0.0
    tempo_dedup_recebidos = 0.0
    _t_bloco_ref = time.perf_counter()
    _tempo_blocos_montar = {
        "preparacao_inicial": 0.0,
        "cliente_final": 0.0,
        "identificador_final": 0.0,
        "metadados_base": 0.0,
        "valor_parcela_qtd_financeiro": 0.0,
        "montagem_consolidado_base": 0.0,
        "reconciliacao_qtd_total": 0.0,
        "evento_1_1": 0.0,
        "sincronizacao_receber": 0.0,
        "fechamento_posprocessamento": 0.0,
    }

    def _tick_bloco(nome):
        nonlocal _t_bloco_ref
        _tempo_blocos_montar[nome] = _tempo_blocos_montar.get(nome, 0.0) + (time.perf_counter() - _t_bloco_ref)
        _t_bloco_ref = time.perf_counter()

    if not df_receber.empty:
        _linhas_before = len(df_receber)
        _vendas_before = df_receber["Venda"].fillna("").astype(str).str.strip().nunique() if "Venda" in df_receber.columns else 0
        df_receber = adicionar_chave_cliente(df_receber)
        _t0_dedup_r = time.perf_counter()
        df_receber = deduplicar_receber(df_receber, alertas_dedup_auditoria)
        tempo_dedup_receber = time.perf_counter() - _t0_dedup_r
        if registro_etapas_tempo is not None:
            registro_etapas_tempo.append(
                (
                    "deduplicar_receber",
                    tempo_dedup_receber,
                    len(df_receber),
                    len(df_recebidos) if not df_recebidos.empty else 0,
                    0,
                )
            )
        df_receber = ordenar_dataframe_uau_por_venda_parcela(df_receber, col_data="Vencimento")
        df_receber = _garantir_cliente_base(df_receber, "df_receber_deduplicado")
        # Reforço: classificar por Vencimento x data-base (mesma regra do processamento final).
        # Evita linhas sem status após dedup e alinha com o consolidado.
        if data_base is not None and "Vencimento" in df_receber.columns:
            if "Status_Vencimento" not in df_receber.columns:
                df_receber["Status_Vencimento"] = ""
            venc_ts = pd.to_datetime(df_receber["Vencimento"], errors="coerce")
            db_ts = pd.Timestamp(data_base)
            m_ok = venc_ts.notna()
            df_receber.loc[m_ok & (venc_ts < db_ts), "Status_Vencimento"] = "VENCIDO"
            df_receber.loc[m_ok & (venc_ts >= db_ts), "Status_Vencimento"] = "A VENCER"
        _linhas_after = len(df_receber)
        _vendas_after = df_receber["Venda"].fillna("").astype(str).str.strip().nunique() if "Venda" in df_receber.columns else 0
        st_col = df_receber["Status_Vencimento"].fillna("").astype(str).str.strip() if "Status_Vencimento" in df_receber.columns else pd.Series([""] * len(df_receber))
        _dbg(
            f"PREP_RECEBER: linhas { _linhas_before } -> { _linhas_after } | vendas { _vendas_before } -> { _vendas_after } | "
            f"status_vazio={int(st_col.eq('').sum())} | Vlr_Parcela_zero={int((_serie_num(df_receber,'Vlr_Parcela')==0).sum())} | "
            f"Principal_zero={int((_serie_num(df_receber,'Principal')==0).sum())}"
        )

    if not df_recebidos.empty:
        _linhas_before = len(df_recebidos)
        _vendas_before = df_recebidos["Venda"].fillna("").astype(str).str.strip().nunique() if "Venda" in df_recebidos.columns else 0
        df_recebidos = adicionar_chave_cliente(df_recebidos)
        _t0_dedup_p = time.perf_counter()
        df_recebidos = deduplicar_recebidos(df_recebidos, alertas_dedup_auditoria)
        tempo_dedup_recebidos = time.perf_counter() - _t0_dedup_p
        if registro_etapas_tempo is not None:
            registro_etapas_tempo.append(
                (
                    "deduplicar_recebidos",
                    tempo_dedup_recebidos,
                    len(df_receber) if not df_receber.empty else 0,
                    len(df_recebidos),
                    0,
                )
            )

        df_recebidos = ordenar_dataframe_uau_por_venda_parcela(df_recebidos, col_data="Data_Rec")
        df_recebidos = _garantir_cliente_base(df_recebidos, "df_recebidos_deduplicado")
        _linhas_after = len(df_recebidos)
        _vendas_after = df_recebidos["Venda"].fillna("").astype(str).str.strip().nunique() if "Venda" in df_recebidos.columns else 0
        _dbg(
            f"PREP_RECEBIDOS: linhas { _linhas_before } -> { _linhas_after } | vendas { _vendas_before } -> { _vendas_after } | "
            f"Vlr_Parcela_zero={int((_serie_num(df_recebidos,'Vlr_Parcela')==0).sum())} | "
            f"Principal_zero={int((_serie_num(df_recebidos,'Principal')==0).sum())}"
        )

    _tick_bloco("preparacao_inicial")

    if not df_receber.empty:
        alertas_dedup_auditoria.extend(coletar_alertas_conflito_duplicidade_flag(df_receber, "RECEBER"))
    if not df_recebidos.empty:
        alertas_dedup_auditoria.extend(coletar_alertas_conflito_duplicidade_flag(df_recebidos, "RECEBIDOS"))

    if not df_receber.empty:
        df_receber["Parcela_Key"] = (
            df_receber["Venda"].astype(str).str.strip()
            + "||"
            + df_receber["Parcela"].astype(str).str.strip()
        )

    if not df_recebidos.empty:
        df_recebidos["Parcela_Key"] = (
            df_recebidos["Venda"].astype(str).str.strip()
            + "||"
            + df_recebidos["Parcela"].astype(str).str.strip()
        )

    # =====================================================
    # FUNÇÕES INTERNAS
    # =====================================================

    def evento_unico_1_1(df_base, venda):
        if df_base is None or df_base.empty:
            return False

        grupo = df_base[df_base["Venda"].astype(str).str.strip().eq(str(venda).strip())].copy()

        if grupo.empty:
            return False

        parcelas = sorted(set([str(v).strip() for v in grupo["Parcela"].tolist() if str(v).strip() != ""]))
        return len(parcelas) == 1 and parcelas[0] == "1/1"

    def conferir_identificador_por_venda_parcela(df_base, venda, identificador):
        if df_base is None or df_base.empty:
            return False

        identificador = normalizar_identificador(identificador)
        if not identificador:
            return False

        grupo = df_base[df_base["Venda"].astype(str).str.strip().eq(str(venda).strip())].copy()
        if grupo.empty:
            return False

        candidatos = []

        if "Unidades" in grupo.columns:
            candidatos.extend([
                normalizar_identificador(v)
                for v in grupo["Unidades"].tolist()
                if str(v).strip() != "" and not identificador_truncado(v)
            ])

        if "Identificador_Produto" in grupo.columns:
            candidatos.extend([
                normalizar_identificador(v)
                for v in grupo["Identificador_Produto"].tolist()
                if str(v).strip() != "" and not identificador_truncado(v)
            ])

        candidatos = [c for c in candidatos if identificador_tem_formato_endereco(c)]
        return identificador in set(candidatos)

    # =====================================================
    # BASE DE CHAVES = UNIÃO ENTRE RECEBER E RECEBIDOS
    # =====================================================
    df_receber = _garantir_cliente_base(df_receber, "df_receber_base_chave")
    df_recebidos = _garantir_cliente_base(df_recebidos, "df_recebidos_base_chave")

    chaves_receber = (
        df_receber[chave].drop_duplicates()
        if not df_receber.empty
        else pd.DataFrame(columns=chave)
    )

    chaves_recebidos = (
        df_recebidos[chave].drop_duplicates()
        if not df_recebidos.empty
        else pd.DataFrame(columns=chave)
    )

    base_chaves = (
        pd.concat([chaves_receber, chaves_recebidos], ignore_index=True)
        .drop_duplicates()
        .reset_index(drop=True)
    )
    _dbg(f"MAP_BASE: base_chaves_linhas={len(base_chaves)}")

    # =====================================================
    # CLIENTE FINAL
    # =====================================================
    cliente_receber = (
        df_receber.groupby(chave, as_index=False)
        .agg({"Cliente": escolher_cliente_exibicao})
        if not df_receber.empty
        else pd.DataFrame(columns=chave + ["Cliente"])
    )

    cliente_recebidos = (
        df_recebidos.groupby(chave, as_index=False)
        .agg({"Cliente": escolher_cliente_exibicao})
        if not df_recebidos.empty
        else pd.DataFrame(columns=chave + ["Cliente"])
    )

    cliente_final = base_chaves.merge(
        cliente_receber.rename(columns={"Cliente": "Cliente_Receber"}),
        on=chave,
        how="left"
    )
    cliente_final = cliente_final.merge(
        cliente_recebidos.rename(columns={"Cliente": "Cliente_Recebidos"}),
        on=chave,
        how="left"
    )
    cliente_final["Cliente"] = cliente_final["Cliente_Receber"].fillna(cliente_final["Cliente_Recebidos"])
    cliente_final["Cliente"] = cliente_final["Cliente"].fillna("").astype(str).str.strip()
    cliente_final = cliente_final[chave + ["Cliente"]]
    _tick_bloco("cliente_final")

    # =====================================================
    # MAPA PADRÃO
    # =====================================================
    mapa_padrao = escolher_identificador_por_venda_cliente_base(
        pd.concat(
            [
                df_receber if not df_receber.empty else pd.DataFrame(),
                df_recebidos if not df_recebidos.empty else pd.DataFrame(),
            ],
            ignore_index=True
        )
    )
    mapa_padrao_venda = pd.DataFrame()
    if not mapa_padrao.empty:
        mapa_padrao_venda = (
            mapa_padrao.groupby("Venda", as_index=False)
            .agg({
                "Descricao_Padrao": lambda s: escolher_moda_texto([str(x) for x in s if str(x).strip()]),
                "EmpObra_Padrao": lambda s: escolher_moda_texto([str(x) for x in s if str(x).strip()]),
            })
        )

    # =====================================================
    # IDENTIFICADOR FINAL = moda de todas as ocorrencias (Receber + Recebidos) por Venda
    # =====================================================
    def _coletar_ocorrencias_identificadores(df):
        if df is None or df.empty:
            return pd.DataFrame(columns=["Venda", "id_norm"])
        linhas = []
        for _, row in df.iterrows():
            venda = str(row.get("Venda", "")).strip()
            for col in ["Unidades", "Identificador_Produto"]:
                if col not in row.index:
                    continue
                val = row.get(col)
                if val is None or str(val).strip() == "":
                    continue
                if identificador_truncado(val):
                    continue
                nid = normalizar_identificador(val)
                if nid:
                    linhas.append({"Venda": venda, "id_norm": nid})
        return pd.DataFrame(linhas)

    df_ids_oc = pd.concat(
        [
            _coletar_ocorrencias_identificadores(df_receber),
            _coletar_ocorrencias_identificadores(df_recebidos),
        ],
        ignore_index=True,
    )
    if df_ids_oc.empty:
        # Somente Venda + Identificador: se incluir Cliente_Base aqui, o merge(on="Venda")
        # duplica Cliente_Base → Cliente_Base_x / Cliente_Base_y e quebra chave + ["Identificador"].
        identificador_moda = pd.DataFrame(columns=["Venda", "Identificador"])
    else:
        identificador_moda = (
            df_ids_oc.groupby("Venda", as_index=False)
            .agg({"id_norm": moda_identificador_final_serie})
            .rename(columns={"id_norm": "Identificador"})
        )
    identificador_final = base_chaves.merge(identificador_moda, on=["Venda"], how="left")
    if "Cliente_Base" not in identificador_final.columns:
        if "Cliente_Base_x" in identificador_final.columns:
            identificador_final = identificador_final.rename(columns={"Cliente_Base_x": "Cliente_Base"})
            if "Cliente_Base_y" in identificador_final.columns:
                identificador_final = identificador_final.drop(columns=["Cliente_Base_y"])
        elif "Cliente_Base_y" in identificador_final.columns:
            identificador_final = identificador_final.rename(columns={"Cliente_Base_y": "Cliente_Base"})
        else:
            identificador_final["Cliente_Base"] = ""
    identificador_final["Identificador"] = identificador_final["Identificador"].fillna("").astype(str).str.strip()
    identificador_final = identificador_final[chave + ["Identificador"]]
    _tick_bloco("identificador_final")

    # =====================================================
    # METADADOS BASE
    # =====================================================
    meta_receber = (
        df_receber.groupby(chave, as_index=False)
        .agg({
            "Emp/Obra": lambda x: Counter(
                [normalizar_emp_obra(v) for v in x if str(v).strip() != ""]
            ).most_common(1)[0][0] if any(str(v).strip() != "" for v in x) else "",
            "Descricao_Produto": lambda x: Counter(
                [str(v).strip().upper() for v in x if str(v).strip() != ""]
            ).most_common(1)[0][0] if any(str(v).strip() != "" for v in x) else "",
            "Nr_Person": lambda x: ", ".join(
                sorted(set([str(v) for v in x if str(v).strip() != ""]))
            ),
        })
        .rename(columns={
            "Emp/Obra": "EmpObra_Receber",
            "Descricao_Produto": "Descricao_Receber",
            "Nr_Person": "NrPerson_Receber",
        })
        if not df_receber.empty
        else pd.DataFrame(columns=chave + ["EmpObra_Receber", "Descricao_Receber", "NrPerson_Receber"])
    )

    meta_recebidos = (
        df_recebidos.groupby(chave, as_index=False)
        .agg({
            "Emp/Obra": lambda x: Counter(
                [normalizar_emp_obra(v) for v in x if str(v).strip() != ""]
            ).most_common(1)[0][0] if any(str(v).strip() != "" for v in x) else "",
            "Descricao_Produto": lambda x: Counter(
                [str(v).strip().upper() for v in x if str(v).strip() != ""]
            ).most_common(1)[0][0] if any(str(v).strip() != "" for v in x) else "",
            "Nr_Person": lambda x: ", ".join(
                sorted(set([str(v) for v in x if str(v).strip() != ""]))
            ),
        })
        .rename(columns={
            "Emp/Obra": "EmpObra_Recebidos",
            "Descricao_Produto": "Descricao_Recebidos",
            "Nr_Person": "NrPerson_Recebidos",
        })
        if not df_recebidos.empty
        else pd.DataFrame(columns=chave + ["EmpObra_Recebidos", "Descricao_Recebidos", "NrPerson_Recebidos"])
    )

    base_principal = base_chaves.merge(meta_receber, on=chave, how="left")
    base_principal = base_principal.merge(meta_recebidos, on=chave, how="left")

    if not mapa_padrao_venda.empty:
        base_principal = base_principal.merge(
            mapa_padrao_venda[["Venda", "Descricao_Padrao", "EmpObra_Padrao"]],
            on="Venda",
            how="left"
        )
    else:
        base_principal["Descricao_Padrao"] = ""
        base_principal["EmpObra_Padrao"] = ""

    base_principal["Emp/Obra"] = base_principal["EmpObra_Receber"].fillna(base_principal["EmpObra_Recebidos"])
    base_principal["Emp/Obra"] = base_principal["Emp/Obra"].fillna(base_principal["EmpObra_Padrao"])
    base_principal["Descricao_Produto"] = base_principal["Descricao_Receber"].fillna(base_principal["Descricao_Recebidos"])
    base_principal["Descricao_Produto"] = base_principal["Descricao_Produto"].fillna(base_principal["Descricao_Padrao"])
    base_principal["Nr_Person"] = base_principal["NrPerson_Receber"].fillna(base_principal["NrPerson_Recebidos"])

    base_principal["Emp/Obra"] = base_principal["Emp/Obra"].fillna("").astype(str).str.strip()
    base_principal["Descricao_Produto"] = base_principal["Descricao_Produto"].fillna("").astype(str).str.strip()
    base_principal["Nr_Person"] = base_principal["Nr_Person"].fillna("").astype(str).str.strip()

    descricoes_validas = [
        limpar_nome_empreendimento(v)
        for v in base_principal["Descricao_Produto"].tolist()
        if str(v).strip() != "" and str(v).strip() != "-" and limpar_nome_empreendimento(v) != ""
    ]

    # Reforco: usa o descritivo real do empreendimento vindo do TXT como fallback,
    # sem cair para "Emp/Obra" (codigo/sigla).
    if not descricoes_validas and str(nome_empreendimento_arquivo or "").strip():
        descricoes_validas = [limpar_nome_empreendimento(nome_empreendimento_arquivo)]

    if descricoes_validas:
        empreendimento_dominante = escolher_moda_texto(descricoes_validas)
    else:
        empreendimento_dominante = ""

    if empreendimento_dominante:
        base_principal.loc[
            base_principal["Descricao_Produto"].astype(str).str.strip() == "",
            "Descricao_Produto"
        ] = empreendimento_dominante

    base_principal = base_principal[chave + ["Emp/Obra", "Descricao_Produto", "Nr_Person"]]
    _tick_bloco("metadados_base")

    # =====================================================
    # VALOR DA PARCELA: regra madura do projeto (moda de Principal)
    # =====================================================
    if not df_receber.empty:
        linhas_vp = []
        for key, g in df_receber.groupby(chave):
            if isinstance(key, tuple):
                row = {chave[i]: key[i] for i in range(len(chave))}
            else:
                row = {chave[0]: key}
            row["Valor Da Parcela"] = moda_valor_parcela_por_df_ou_grupo(g)
            linhas_vp.append(row)
        valor_parcela_receber = pd.DataFrame(linhas_vp)
    else:
        valor_parcela_receber = pd.DataFrame(columns=chave + ["Valor Da Parcela"])

    # =====================================================
    # QTD TOTAL DE PARCELAS
    # =====================================================
    def _qtd_total_por_venda_df(df, nome_col):
        if df is None or df.empty or "Parc_Total" not in df.columns or "Venda" not in df.columns:
            return pd.DataFrame(columns=["Venda", nome_col])
        d = df
        vk = d["Venda"].fillna("").astype(str).str.strip()
        linhas = []
        for v_key, g in d.groupby(vk, sort=False):
            vs = str(v_key).strip()
            if not vs:
                continue
            linhas.append({"Venda": vs, nome_col: calcular_qtd_parc_total(g)})
        return pd.DataFrame(linhas) if linhas else pd.DataFrame(columns=["Venda", nome_col])

    qtd_total_receber = _qtd_total_por_venda_df(df_receber, "QtdTotal_Receber")
    qtd_total_recebidos = _qtd_total_por_venda_df(df_recebidos, "QtdTotal_Recebidos")

    qtd_total = base_chaves.merge(qtd_total_receber, on="Venda", how="left")
    qtd_total = qtd_total.merge(qtd_total_recebidos, on="Venda", how="left")
    qtd_total["Qtd.Parc.Total"] = qtd_total["QtdTotal_Receber"].fillna(qtd_total["QtdTotal_Recebidos"])
    qtd_total["Qtd.Parc.Total"] = qtd_total["Qtd.Parc.Total"].fillna(0).astype(int)
    qtd_total = qtd_total[chave + ["Qtd.Parc.Total"]]

    mapa_moda_receber = _mapa_qtd_parc_total_por_venda(df_receber)
    mapa_moda_recebidos = _mapa_qtd_parc_total_por_venda(df_recebidos)

    # =====================================================
    # PAGOS / INADIMPLÊNCIA (VENCIDO) / A VENCER (status A VENCER)
    # =====================================================
    pagos = (
        df_recebidos.groupby(chave, as_index=False)
        .agg({
            "Parcela": contar_parcelas_distintas_padrao,
            "Total_Dep": "sum",
        })
        .rename(columns={"Parcela": "Qtd.Parc.Paga", "Total_Dep": "Vl.Pago"})
        if not df_recebidos.empty
        else pd.DataFrame(columns=chave + ["Qtd.Parc.Paga", "Vl.Pago"])
    )

    qtd_avencer_df = pd.DataFrame(columns=chave + ["Qtd.Parc.A Vencer"])

    if not df_receber.empty:
        if "Status_Vencimento" not in df_receber.columns:
            df_receber["Status_Vencimento"] = ""
        st = df_receber["Status_Vencimento"].astype(str).str.strip().str.upper()

        df_receber_vencido = df_receber[st == "VENCIDO"].copy()
        atrasadas = (
            df_receber_vencido.groupby(chave, as_index=False)
            .agg({
                "Parcela": contar_parcelas_distintas_padrao,
                "Principal": "sum",
                "Correcao": "sum",
                "Juros_Atraso": "sum",
                "Multa_Atraso": "sum",
                "Correcao_Atraso": "sum",
            })
            .rename(columns={
                "Parcela": "Qtd.Parc.Atrasada",
                "Principal": "Vl.Principal Atrasado",
                "Correcao": "Vl.Correção",
                "Juros_Atraso": "Vl.Juros",
                "Multa_Atraso": "Vl.Multas",
                "Correcao_Atraso": "Vl.Correção Atraso",
            })
        )

        mask_a_vencer = st == "A VENCER"
        vp_serie = (
            pd.to_numeric(df_receber["Vlr_Parcela"], errors="coerce").fillna(0)
            if "Vlr_Parcela" in df_receber.columns
            else pd.Series(0.0, index=df_receber.index)
        )
        df_vl_vencer = df_receber.loc[mask_a_vencer].copy()
        df_vl_vencer["_Vlr_Parcela_R"] = vp_serie.loc[mask_a_vencer]
        avencer_valor = (
            df_vl_vencer.groupby(chave, as_index=False)
            .agg({"_Vlr_Parcela_R": "sum"})
            .rename(columns={"_Vlr_Parcela_R": "Vl.Vencer"})
        )
        qtd_avencer_df = (
            df_receber.loc[mask_a_vencer].groupby(chave, as_index=False)
            .agg({"Parcela": contar_parcelas_distintas_padrao})
            .rename(columns={"Parcela": "Qtd.Parc.A Vencer"})
        )
    else:
        atrasadas = pd.DataFrame(columns=chave + [
            "Qtd.Parc.Atrasada", "Vl.Principal Atrasado", "Vl.Correção",
            "Vl.Juros", "Vl.Multas", "Vl.Correção Atraso",
        ])
        avencer_valor = pd.DataFrame(columns=chave + ["Vl.Vencer"])
    _tick_bloco("valor_parcela_qtd_financeiro")

    # =====================================================
    # MONTA CONSOLIDADO
    # =====================================================
    consolidado = base_principal.merge(cliente_final, on=chave, how="left")
    consolidado = consolidado.merge(identificador_final, on=chave, how="left")
    consolidado = consolidado.merge(valor_parcela_receber, on=chave, how="left")
    consolidado = consolidado.merge(qtd_total, on=chave, how="left")
    consolidado = consolidado.merge(pagos, on=chave, how="left")
    consolidado = consolidado.merge(atrasadas, on=chave, how="left")
    consolidado = consolidado.merge(avencer_valor, on=chave, how="left")
    consolidado = consolidado.merge(qtd_avencer_df, on=chave, how="left")
    _tick_bloco("montagem_consolidado_base")

    colunas_numericas = [
        "Valor Da Parcela", "Qtd.Parc.Total", "Qtd.Parc.Paga", "Qtd.Parc.A Vencer", "Vl.Pago",
        "Qtd.Parc.Atrasada", "Vl.Principal Atrasado", "Vl.Correção",
        "Vl.Juros", "Vl.Multas", "Vl.Correção Atraso", "Vl.Vencer"
    ]
    for col in colunas_numericas:
        if col not in consolidado.columns:
            consolidado[col] = 0
    consolidado[colunas_numericas] = (
        consolidado[colunas_numericas]
        .apply(pd.to_numeric, errors="coerce")
        .fillna(0)
    )
    colunas_qtd_int = ["Qtd.Parc.Total", "Qtd.Parc.Paga", "Qtd.Parc.Atrasada", "Qtd.Parc.A Vencer"]
    consolidado[colunas_qtd_int] = consolidado[colunas_qtd_int].round(0).astype(int)

    # Piso de auditoria: Qtd.Parc.Total >= max(Paga, Atrasada, A Vencer) após moda por venda.
    if not consolidado.empty:
        qp = pd.to_numeric(consolidado["Qtd.Parc.Paga"], errors="coerce").fillna(0).astype(int)
        qa = pd.to_numeric(consolidado["Qtd.Parc.Atrasada"], errors="coerce").fillna(0).astype(int)
        qv = pd.to_numeric(consolidado["Qtd.Parc.A Vencer"], errors="coerce").fillna(0).astype(int)
        piso_qtd = pd.concat([qp, qa, qv], axis=1).max(axis=1)
        qt = pd.to_numeric(consolidado["Qtd.Parc.Total"], errors="coerce").fillna(0).astype(int)
        consolidado["Qtd.Parc.Total"] = pd.concat([qt, piso_qtd], axis=1).max(axis=1).astype(int)

        # -------------------------------------------------------------
        # Reconciliação padrão Qtd.Parc.Total (Receber confiável vs fontes)
        # -------------------------------------------------------------
        mapa_uni_r = _mapa_universo_parcelas_distintas_por_venda(df_receber)
        mapa_uni_p = _mapa_universo_parcelas_distintas_por_venda(df_recebidos)
        def _int_qtd_total_cell(val):
            v = pd.to_numeric(val, errors="coerce")
            if pd.isna(v):
                return 0
            return int(v)

        map_qr = {}
        if not qtd_total_receber.empty and "Venda" in qtd_total_receber.columns:
            for _, rr in qtd_total_receber.iterrows():
                vk = str(rr["Venda"]).strip()
                if vk:
                    map_qr[vk] = _int_qtd_total_cell(rr["QtdTotal_Receber"])
        map_qp = {}
        if not qtd_total_recebidos.empty and "Venda" in qtd_total_recebidos.columns:
            for _, rr in qtd_total_recebidos.iterrows():
                vk = str(rr["Venda"]).strip()
                if vk:
                    map_qp[vk] = _int_qtd_total_cell(rr["QtdTotal_Recebidos"])

        vs_norm = consolidado["Venda"].fillna("").astype(str).str.strip()
        piso_max_por_venda = consolidado.assign(_piso_agg=piso_qtd).groupby(vs_norm)["_piso_agg"].max().to_dict()

        for venda_key in sorted(set(piso_max_por_venda.keys()), key=lambda x: (len(x), x)):
            if not str(venda_key).strip():
                continue
            piso_M = int(piso_max_por_venda.get(venda_key, 0) or 0)
            qr = int(map_qr.get(venda_key, 0) or 0)
            qp_moda = int(map_qp.get(venda_key, 0) or 0)
            ur = int(mapa_uni_r.get(venda_key, 0) or 0)
            up = int(mapa_uni_p.get(venda_key, 0) or 0)
            qt_fill = qr if qr > 0 else qp_moda

            confiavel, motivos = _receber_confiavel_qtd_parc_total(qr, qp_moda, ur, up, piso_M)
            qt_prev = int(max(qt_fill, piso_M))
            if confiavel:
                continue

            teto_san = max(qt_prev, 1) * 3 + 24
            qt_new = int(min(max(piso_M, qp_moda, qr), teto_san))

            mask_v = vs_norm.eq(venda_key)
            consolidado.loc[mask_v, "Qtd.Parc.Total"] = qt_new

            sub = consolidado.loc[mask_v]
            cli_rep = ""
            id_rep = ""
            if not sub.empty:
                if "Cliente" in sub.columns:
                    cli_rep = str(sub["Cliente"].fillna("").astype(str).iloc[0]).strip()
                if "Identificador" in sub.columns:
                    id_rep = str(sub["Identificador"].fillna("").astype(str).iloc[0]).strip()
            vp_rep = float(pd.to_numeric(sub["Valor Da Parcela"], errors="coerce").fillna(0).iloc[0]) if (not sub.empty and "Valor Da Parcela" in sub.columns) else 0.0
            qtd_pag_rep = int(pd.to_numeric(sub["Qtd.Parc.Paga"], errors="coerce").fillna(0).iloc[0]) if (not sub.empty and "Qtd.Parc.Paga" in sub.columns) else 0
            qtd_atr_rep = int(pd.to_numeric(sub["Qtd.Parc.Atrasada"], errors="coerce").fillna(0).iloc[0]) if (not sub.empty and "Qtd.Parc.Atrasada" in sub.columns) else 0
            qtd_av_rep = int(pd.to_numeric(sub["Qtd.Parc.A Vencer"], errors="coerce").fillna(0).iloc[0]) if (not sub.empty and "Qtd.Parc.A Vencer" in sub.columns) else 0
            qtd_saldo_rep = max(int(qt_new) - int(qtd_pag_rep + qtd_atr_rep + qtd_av_rep), 0)

            obs_parts = [
                f"Total_Receber(modas)={qr}",
                f"Total_Recebidos(modas)={qp_moda}",
                f"piso_operacional_max={piso_M}",
                f"total_adotado={qt_new}",
                f"uni_Receber={ur} uni_Recebidos={up}",
                f"antes_reconciliacao={qt_prev}",
            ]
            pendencias_qtd_total_reconciliacao.append({
                "Venda": venda_key,
                "Cliente": cli_rep,
                "Identificador": id_rep,
                "Parcela": "",
                "Tipo de Divergência": "QTD_TOTAL_RECONCILIACAO_FONTE",
                "Valor Parcela Divergente": "",
                "Origem da Divergência": "Receber subamostrado ou inconsistente frente a Recebidos / piso operacional",
                "Observação": " | ".join(obs_parts),
                "__VL_PARC_PAGAS": float(vp_rep * qtd_pag_rep),
                "__VL_PARC_INADIMPLENTES": float(vp_rep * qtd_atr_rep),
                "__VL_PARC_A_VENCER": float(vp_rep * qtd_av_rep),
                "__VL_SALDO": float(vp_rep * qtd_saldo_rep),
                "__QTD_PARC_PAGAS": int(qtd_pag_rep),
                "__QTD_PARC_INADIMPLENTES": int(qtd_atr_rep),
                "__QTD_PARC_A_VENCER": int(qtd_av_rep),
                "__QTD_SALDO": int(qtd_saldo_rep),
                "__QTD_TOTAL_RECEBER": qr,
                "__QTD_TOTAL_RECEBIDOS": qp_moda,
                "__QTD_TOTAL_PISO": piso_M,
                "__QTD_TOTAL_ADOTADO": qt_new,
                "__MOTIVO_RECONCILIACAO": "; ".join(motivos),
            })
    _tick_bloco("reconciliacao_qtd_total")

    consolidado["Identificador"] = consolidado["Identificador"].fillna("").astype(str).str.strip()
    consolidado["Descricao_Produto"] = consolidado["Descricao_Produto"].fillna("").astype(str).str.strip()
    consolidado["Cliente"] = consolidado["Cliente"].fillna("").astype(str).str.strip()

    if empreendimento_dominante:
        consolidado.loc[
            consolidado["Descricao_Produto"].astype(str).str.strip() == "",
            "Descricao_Produto"
        ] = empreendimento_dominante

    # =====================================================
    # EVENTO ÚNICO 1/1 SEM BASE NO RECEBER
    # =====================================================
    receber_keys = set()
    if not df_receber_raw.empty and "Venda" in df_receber_raw.columns:
        receber_keys = set(df_receber_raw["Venda"].astype(str).str.strip().tolist())

    venda_cons = consolidado["Venda"].fillna("").astype(str).str.strip()
    mask_sem_receber = ~venda_cons.isin(receber_keys)
    unico_1_1_set = set()
    if not df_recebidos_raw.empty and "Venda" in df_recebidos_raw.columns and "Parcela" in df_recebidos_raw.columns:
        tmp_e11 = pd.DataFrame({
            "Venda": df_recebidos_raw["Venda"].fillna("").astype(str).str.strip(),
            "Parcela": df_recebidos_raw["Parcela"].fillna("").astype(str).str.strip(),
        })
        tmp_e11 = tmp_e11.loc[tmp_e11["Venda"] != ""]
        if not tmp_e11.empty:
            pars_por_venda = tmp_e11.groupby("Venda")["Parcela"].agg(
                lambda s: sorted(set([p for p in s.tolist() if p != ""]))
            )
            unico_1_1_set = set([
                str(v) for v, pars in pars_por_venda.items()
                if len(pars) == 1 and pars[0] == "1/1"
            ])

    mask_e11 = mask_sem_receber & venda_cons.isin(unico_1_1_set)
    if mask_e11.any():
        consolidado.loc[mask_e11, "Qtd.Parc.Total"] = (
            pd.to_numeric(consolidado.loc[mask_e11, "Qtd.Parc.Total"], errors="coerce")
            .fillna(0)
            .clip(lower=1)
            .astype(int)
        )
        # Estrutural: há exatamente uma parcela 1/1 em Recebidos (unico_1_1); alinhar à
        # contagem usada na validação, sem depender de Vl.Pago (só Total_Dep).
        consolidado.loc[mask_e11, "Qtd.Parc.Paga"] = 1
        consolidado.loc[mask_e11, "Qtd.Parc.Atrasada"] = 0
        consolidado.loc[mask_e11, "Qtd.Parc.A Vencer"] = 0
        _vp_cur = pd.to_numeric(consolidado.loc[mask_e11, "Valor Da Parcela"], errors="coerce").fillna(0)
        _vlp = pd.to_numeric(consolidado.loc[mask_e11, "Vl.Pago"], errors="coerce").fillna(0)
        consolidado.loc[mask_e11, "Valor Da Parcela"] = _vp_cur.where(_vp_cur > 0, _vlp)
        consolidado.loc[mask_e11, "Vl.Principal Atrasado"] = 0
        consolidado.loc[mask_e11, "Vl.Correção"] = 0
        consolidado.loc[mask_e11, "Vl.Juros"] = 0
        consolidado.loc[mask_e11, "Vl.Multas"] = 0
        consolidado.loc[mask_e11, "Vl.Correção Atraso"] = 0
        # Vl.Vencer permanece da soma coluna R (fonte Dados Receber) — nao zerar aqui
    _tick_bloco("evento_1_1")

    # =====================================================
    # SINCRONIZAÇÃO OBRIGATÓRIA COM DADOS RECEBER (Vl.Vencer, Valor Da Parcela, Identificador)
    # =====================================================
    def _mapa_moda_parcela_m(df_receb):
        out = {}
        if df_receb is None or df_receb.empty:
            return out
        for venda, g in df_receb.groupby("Venda", dropna=False):
            out[str(venda).strip()] = moda_valor_parcela_por_df_ou_grupo(g)
        return out

    def _mapa_id_moda_final(df_r, df_p):
        dfo = pd.concat(
            [_coletar_ocorrencias_identificadores(df_r), _coletar_ocorrencias_identificadores(df_p)],
            ignore_index=True,
        )
        if dfo.empty:
            return {}
        dfo["Venda"] = dfo["Venda"].astype(str).str.strip()
        raw = (
            dfo.groupby("Venda")["id_norm"]
            .agg(moda_identificador_final_serie)
            .to_dict()
        )
        return {str(k).strip(): v for k, v in raw.items()}

    def _mapa_qtd_a_vencer_status(df_receb):
        if df_receb is None or df_receb.empty:
            return {}
        d = df_receb.copy()
        if "Status_Vencimento" not in d.columns:
            d["Status_Vencimento"] = ""
        st = d["Status_Vencimento"].astype(str).str.strip().str.upper()
        d = d.loc[st == "A VENCER"].copy()
        if d.empty:
            return {}
        raw = d.groupby("Venda")["Parcela"].agg(contar_parcelas_distintas_padrao).to_dict()
        return {str(k).strip(): int(v) for k, v in raw.items()}

    if not df_receber.empty:
        mv = mapa_vl_vencer_por_venda_receber_tratado(df_receber)
        mq = _mapa_qtd_a_vencer_status(df_receber)
        _dbg(f"MAP_INFO: consolidado_vendas_antes_map={consolidado['Venda'].fillna('').astype(str).str.strip().nunique()}")
        if not identificador_moda.empty and "Venda" in identificador_moda.columns:
            mi = (
                identificador_moda.assign(Venda=identificador_moda["Venda"].astype(str).str.strip())
                .set_index("Venda")["Identificador"]
                .to_dict()
            )
        else:
            mi = {}
        mm = _mapa_moda_parcela_m(df_receber) if DEBUG_DADOS else {}

        def _dbg_map(nome, mp):
            keys = list(mp.keys()) if isinstance(mp, dict) else []
            vals = list(mp.values()) if isinstance(mp, dict) else []
            _dbg(f"{nome}: chaves={len(keys)} | primeiras_chaves={keys[:10]} | primeiros_valores={vals[:10]}")

        _dbg_map("mv_Vl.Vencer", mv)
        _dbg_map("mq_Qtd.AVencer", mq)
        _dbg_map("mm_ValorParcela", mm)
        _dbg_map("mi_Identificador", mi)

        vk = consolidado["Venda"].astype(str).str.strip()
        set_v = set([v for v in vk.tolist() if v != ""])
        for nome, mp in [("mv", mv), ("mq", mq), ("mm", mm), ("mi", mi)]:
            ch = set([str(k).strip() for k in mp.keys()]) if isinstance(mp, dict) else set()
            aus = sorted(list(set_v - ch))
            _dbg(f"MAP_COMPAT[{nome}]: ausentes={len(aus)} | exemplos={aus[:20]}")

        _antes_vencer = consolidado[["Venda", "Vl.Vencer"]].copy()
        _antes_qtd = consolidado[["Venda", "Qtd.Parc.A Vencer"]].copy()
        _antes_parcela = consolidado[["Venda", "Valor Da Parcela"]].copy()
        _antes_id = consolidado[["Venda", "Identificador"]].copy()
        # Sem sobrescrita agressiva: só atualiza quando a chave existe no mapa.
        map_vl = vk.map(mv)
        mask_vl = map_vl.notna()
        consolidado.loc[mask_vl, "Vl.Vencer"] = pd.to_numeric(map_vl.loc[mask_vl], errors="coerce").fillna(
            pd.to_numeric(consolidado.loc[mask_vl, "Vl.Vencer"], errors="coerce").fillna(0)
        )

        map_qtd = vk.map(mq)
        mask_qtd = map_qtd.notna()
        consolidado.loc[mask_qtd, "Qtd.Parc.A Vencer"] = pd.to_numeric(
            map_qtd.loc[mask_qtd], errors="coerce"
        ).fillna(pd.to_numeric(consolidado.loc[mask_qtd, "Qtd.Parc.A Vencer"], errors="coerce").fillna(0)).round(0).astype(int)

        # Valor Da Parcela: manter o valor agregado-base; não sobrescrever no escuro por mapa.
        # (mapa mm permanece apenas para diagnóstico de compatibilidade)

        map_id = vk.map(mi)
        mask_id = map_id.notna() & map_id.astype(str).str.strip().ne("")
        consolidado.loc[mask_id, "Identificador"] = map_id.loc[mask_id].astype(str).str.strip()
        _dbg(
            f"MAP_APOS: Vl.Vencer nulos={int(consolidado['Vl.Vencer'].isna().sum())} zeros={int((pd.to_numeric(consolidado['Vl.Vencer'],errors='coerce').fillna(0)==0).sum())} "
            f"soma={float(pd.to_numeric(consolidado['Vl.Vencer'],errors='coerce').fillna(0).sum()):.2f}"
        )
        _dbg(
            f"MAP_APOS: Qtd.AVencer nulos={int(consolidado['Qtd.Parc.A Vencer'].isna().sum())} zeros={int((pd.to_numeric(consolidado['Qtd.Parc.A Vencer'],errors='coerce').fillna(0)==0).sum())} "
            f"soma={float(pd.to_numeric(consolidado['Qtd.Parc.A Vencer'],errors='coerce').fillna(0).sum()):.2f}"
        )
        _dbg(
            f"MAP_APOS: ValorParcela nulos={int(consolidado['Valor Da Parcela'].isna().sum())} zeros={int((pd.to_numeric(consolidado['Valor Da Parcela'],errors='coerce').fillna(0)==0).sum())} "
            f"soma={float(pd.to_numeric(consolidado['Valor Da Parcela'],errors='coerce').fillna(0).sum()):.2f}"
        )
        _dbg(
            f"MAP_APOS: Identificador vazios={int(consolidado['Identificador'].fillna('').astype(str).str.strip().eq('').sum())}"
        )
        if DEBUG_DADOS:
            _dbg(f"MAP_DIFF Vl.Vencer exemplos={_antes_vencer.merge(consolidado[['Venda','Vl.Vencer']], on='Venda', suffixes=('_antes','_depois')).head(10).to_dict('records')}")
            _dbg(f"MAP_DIFF Qtd.AVencer exemplos={_antes_qtd.merge(consolidado[['Venda','Qtd.Parc.A Vencer']], on='Venda', suffixes=('_antes','_depois')).head(10).to_dict('records')}")
            _dbg(f"MAP_DIFF ValorParcela exemplos={_antes_parcela.merge(consolidado[['Venda','Valor Da Parcela']], on='Venda', suffixes=('_antes','_depois')).head(10).to_dict('records')}")
            _dbg(f"MAP_DIFF Identificador exemplos={_antes_id.merge(consolidado[['Venda','Identificador']], on='Venda', suffixes=('_antes','_depois')).head(10).to_dict('records')}")
    _tick_bloco("sincronizacao_receber")

    # =====================================================
    # COLUNAS FINAIS
    # =====================================================
    colunas_finais = [
        "Emp/Obra", "Empreendimento", "Venda", "Cliente", "Identificador",
        "Status venda", "Valor Da Parcela", "Qtd.Parc.Total", "Qtd.Parc.Paga",
        "Vl.Pago", "Qtd.Parc.Atrasada", "Vl.Principal Atrasado", "Vl.Correção",
        "Vl.Juros", "Vl.Multas", "Vl.Correção Atraso", "Vl.Principal (Encargos)",
        "Qtd.Parc.A Vencer", "Vl.Vencer", "Vl.Carteira", "% Pago",
        "% Inadimplência", "% A Vencer", "DIA VENCIMENTO", "Status Construção", "Judicializado", "APORTE",
    ]

    # =====================================================
    # FECHAMENTO: transporte direto dos agregados (merge acima)
    # Nao recalcular parcelas/valores aqui — evita perda de massa e divergencia
    # em relacao a Dados Receber / Dados Recebidos.
    # =====================================================
    alertas_conferencia = []
    alertas_conferencia.extend(alertas_dedup_auditoria)
    alertas_conferencia.extend(coletar_alertas_cliente_base(df_receber_raw, df_recebidos_raw))
    alertas_conferencia.extend(montar_alertas_etl_de_metricas(obter_etl_stats_acumulado()))
    scores_qualidade_parcelas = (
        calcular_score_qualidade_parcelas_por_venda(df_receber)
        if not df_receber.empty
        else {}
    )
    alertas_conferencia.extend(auditoria_sequencia_parcelas_receber(df_receber))

    # =====================================================
    # VALIDAÇÕES / RECOMPOSIÇÃO FINAL
    # =====================================================
    if not df_receber.empty and "Status_Vencimento" in df_receber.columns:
        st_dbg = df_receber["Status_Vencimento"].fillna("").astype(str).str.strip().str.upper()
        base_v = df_receber.loc[st_dbg == "VENCIDO"].copy()
        _dbg(
            "INAD_BASE: "
            f"Principal={float(_serie_num(base_v,'Principal').sum()):.2f} | "
            f"Correcao={float(_serie_num(base_v,'Correcao').sum()):.2f} | "
            f"Juros={float(_serie_num(base_v,'Juros_Atraso').sum()):.2f} | "
            f"Multas={float(_serie_num(base_v,'Multa_Atraso').sum()):.2f} | "
            f"Correcao_Atraso={float(_serie_num(base_v,'Correcao_Atraso').sum()):.2f}"
        )

    consolidado["Vl.Principal (Encargos)"] = (
        consolidado["Vl.Principal Atrasado"]
        + consolidado["Vl.Correção"]
        + consolidado["Vl.Juros"]
        + consolidado["Vl.Multas"]
        + consolidado["Vl.Correção Atraso"]
    ).round(2)

    _dbg(
        "INAD_CONSOLIDADO: "
        f"PrincipalAtrasado={float(_serie_num(consolidado,'Vl.Principal Atrasado').sum()):.2f} | "
        f"Correcao={float(_serie_num(consolidado,'Vl.Correção').sum()):.2f} | "
        f"Juros={float(_serie_num(consolidado,'Vl.Juros').sum()):.2f} | "
        f"Multas={float(_serie_num(consolidado,'Vl.Multas').sum()):.2f} | "
        f"CorrecaoAtraso={float(_serie_num(consolidado,'Vl.Correção Atraso').sum()):.2f} | "
        f"Encargos={float(_serie_num(consolidado,'Vl.Principal (Encargos)').sum()):.2f}"
    )

    consolidado["Emp/Obra"] = consolidado["Emp/Obra"].fillna("").astype(str).apply(normalizar_emp_obra)

    consolidado["Empreendimento"] = (
        consolidado["Descricao_Produto"].fillna("").astype(str).apply(limpar_nome_empreendimento)
    )
    if empreendimento_dominante:
        consolidado.loc[
            consolidado["Empreendimento"].astype(str).str.strip() == "",
            "Empreendimento"
        ] = empreendimento_dominante

    # Lote unificado: nomes legais por bloco do TXT (cabeçalho antes das linhas de dados).
    if mapa_emp_obra_nome_legal:
        eo_norm = consolidado["Emp/Obra"].fillna("").astype(str).apply(normalizar_emp_obra)
        nom_por_linha = eo_norm.map(mapa_emp_obra_nome_legal)
        m_ok = nom_por_linha.notna() & nom_por_linha.astype(str).str.strip().ne("")
        if bool(m_ok.any()):
            consolidado.loc[m_ok, "Empreendimento"] = nom_por_linha.loc[m_ok].astype(str).str.strip()
    # Fallback universal por Emp/Obra: evita empreendimento vazio em linhas conhecidas.
    eo_norm = consolidado["Emp/Obra"].fillna("").astype(str).apply(normalizar_emp_obra)
    emp_txt = consolidado["Empreendimento"].fillna("").astype(str).apply(limpar_nome_empreendimento)
    mapa_empobra = (
        pd.DataFrame({"eo": eo_norm, "emp": emp_txt})
        .loc[lambda d: d["eo"].astype(str).str.strip().ne("") & d["emp"].astype(str).str.strip().ne("")]
        .groupby("eo", as_index=False)
        .agg({"emp": escolher_moda_texto})
    )
    if not mapa_empobra.empty:
        mapa_empobra = dict(zip(mapa_empobra["eo"].tolist(), mapa_empobra["emp"].tolist()))
        emp_fill = eo_norm.map(mapa_empobra)
        m_fill = emp_txt.astype(str).str.strip().eq("") & emp_fill.notna() & emp_fill.astype(str).str.strip().ne("")
        if bool(m_fill.any()):
            consolidado.loc[m_fill, "Empreendimento"] = emp_fill.loc[m_fill].astype(str).str.strip()

    _emp_of_serie = consolidado["Emp/Obra"].map(empreendimento_oficial_para_emp_obra)
    _m_emp_vazio = consolidado["Empreendimento"].fillna("").astype(str).str.strip().eq("")
    _m_of_ok = _emp_of_serie.fillna("").astype(str).str.strip().ne("")
    consolidado.loc[_m_emp_vazio & _m_of_ok, "Empreendimento"] = _emp_of_serie.loc[
        _m_emp_vazio & _m_of_ok
    ].astype(str).str.strip()
    # Padronização final: siglas oficiais devem exibir sempre o nome oficial.
    consolidado["Empreendimento"] = _aplicar_nome_oficial_em_series(
        consolidado["Emp/Obra"], consolidado["Empreendimento"]
    )

    consolidado["Status Construção"] = ""
    consolidado["Identificador"] = consolidado["Identificador"].fillna("").astype(str).str.strip()
    consolidado["DIA VENCIMENTO"] = ""
    consolidado["APORTE"] = "NÃO"

    # Garante 1 linha por venda no consolidado final.
    if not consolidado.empty:
        # Normaliza chave de venda antes do fechamento final para evitar fragmentação
        # por variações de formatação (ex.: espaços).
        consolidado["Venda"] = consolidado["Venda"].fillna("").astype(str).str.strip()

        mapa_id_vc_bruto, mapa_id_v_bruto = mapas_identificador_moda_de_bases_brutas(df_receber_raw, df_recebidos_raw)
        harmonizar_identificador_por_venda_cliente_dataframe(consolidado, alertas_conferencia)

        cols_soma = [
            "Qtd.Parc.Total", "Qtd.Parc.Paga", "Qtd.Parc.Atrasada", "Qtd.Parc.A Vencer",
            "Vl.Pago", "Vl.Principal Atrasado", "Vl.Correção", "Vl.Juros",
            "Vl.Multas", "Vl.Correção Atraso", "Vl.Vencer"
        ]
        for c in cols_soma + ["Valor Da Parcela"]:
            if c not in consolidado.columns:
                consolidado[c] = 0

        def _mais_frequente(series):
            vals = [str(v).strip() for v in series.tolist() if str(v).strip() != ""]
            return escolher_moda_texto(vals) if vals else ""

        def _moda_valor_da_parcela_coluna(series):
            return moda_valor_numerico_positivo(pd.to_numeric(series, errors="coerce").fillna(0))

        agg_map = {c: "sum" for c in cols_soma}
        # Total de parcelas por venda é denominador contratual, não soma de fragmentos.
        agg_map["Qtd.Parc.Total"] = "max"
        agg_map["Valor Da Parcela"] = _moda_valor_da_parcela_coluna
        agg_map.update({
            "Emp/Obra": _mais_frequente,
            "Empreendimento": _mais_frequente,
            "Cliente": _mais_frequente,
            "Identificador": _mais_frequente,
            "Status Construção": "first",
        })
        if "Descricao_Produto" in consolidado.columns:
            agg_map["Descricao_Produto"] = _mais_frequente

        consolidado = (
            consolidado
            .groupby("Venda", as_index=False)
            .agg(agg_map)
        )
        # Reforco: Valor Da Parcela = moda de Principal por venda na base bruta.
        if not df_receber_raw.empty:
            base_vp = df_receber_raw.copy()
            base_vp["Venda"] = base_vp["Venda"].fillna("").astype(str).str.strip()
            mapa_vlr_venda = {}
            for v, g in base_vp.groupby("Venda"):
                mapa_vlr_venda[v] = moda_valor_parcela_por_df_ou_grupo(g)
            consolidado["Valor Da Parcela"] = (
                consolidado["Venda"].astype(str).str.strip().map(mapa_vlr_venda).fillna(consolidado["Valor Da Parcela"])
            )

        # Reforço final do cliente/identificador dominantes por venda (não usar Cliente_Base como nome exibido).
        base_cli = pd.concat(
            [
                df_receber[["Venda", "Cliente"]] if not df_receber.empty and "Cliente" in df_receber.columns else pd.DataFrame(columns=["Venda", "Cliente"]),
                df_recebidos[["Venda", "Cliente"]] if not df_recebidos.empty and "Cliente" in df_recebidos.columns else pd.DataFrame(columns=["Venda", "Cliente"]),
            ],
            ignore_index=True,
        )
        if not base_cli.empty:
            base_cli["Venda"] = base_cli["Venda"].fillna("").astype(str).str.strip()
            base_cli["Cliente"] = base_cli["Cliente"].fillna("").astype(str).str.strip()
            base_cli = base_cli.loc[(base_cli["Venda"] != "") & (base_cli["Cliente"] != "")].copy()
            mapa_cli = (
                base_cli.groupby("Venda")["Cliente"]
                .agg(escolher_cliente_exibicao)
                .to_dict()
            )
            if DEBUG_DADOS:
                print(f"[DEBUG][CLIENTE_FINAL] vendas_processadas={len(mapa_cli)}")
            consolidado["Cliente"] = consolidado["Venda"].astype(str).str.strip().map(mapa_cli).fillna(consolidado["Cliente"]).astype(str).str.strip()

        base_ids = pd.concat(
            [_coletar_ocorrencias_identificadores(df_receber_raw), _coletar_ocorrencias_identificadores(df_recebidos_raw)],
            ignore_index=True,
        )
        if not base_ids.empty:
            base_ids["Venda"] = base_ids["Venda"].astype(str).str.strip()
            mapa_id_final = (
                base_ids.groupby("Venda")["id_norm"]
                .agg(moda_identificador_final_serie)
                .to_dict()
            )
            consolidado["Identificador"] = consolidado["Venda"].astype(str).str.strip().map(mapa_id_final).fillna(consolidado["Identificador"]).astype(str).str.strip()
        preencher_identificador_vazio_de_mapas_brutos(
            consolidado, mapa_id_vc_bruto, mapa_id_v_bruto, alertas_conferencia
        )
        consolidado["Vl.Principal (Encargos)"] = (
            consolidado["Vl.Principal Atrasado"]
            + consolidado["Vl.Correção"]
            + consolidado["Vl.Juros"]
            + consolidado["Vl.Multas"]
            + consolidado["Vl.Correção Atraso"]
        ).round(2)
        consolidado = _recalcular_vl_carteira_e_percentuais(consolidado)
        consolidado["Status venda"] = "ADIMPLENTE"
        consolidado.loc[consolidado["Qtd.Parc.Atrasada"] > 0, "Status venda"] = "INADIMPLENTE"
        consolidado.loc[
            (consolidado["Qtd.Parc.Atrasada"] == 0)
            & (consolidado["Vl.Vencer"] == 0)
            & (consolidado["Vl.Pago"] > 0),
            "Status venda"
        ] = "QUITADO"

        # DIA VENCIMENTO (por VENDA):
        # 1) moda do dia de Vencimento com Status_Vencimento = "A VENCER"
        # 2) fallback: moda do dia de Vencimento considerando todas as parcelas da venda
        # 3) empate: menor dia (critério estável)
        # 4) sem base confiável: vazio
        def _dia_moda_predominante(serie_dias):
            s = pd.to_numeric(serie_dias, errors="coerce").dropna().astype(int)
            s = s[(s >= 1) & (s <= 31)]
            if s.empty:
                return None
            vc = s.value_counts()
            max_freq = int(vc.max())
            cand = sorted([int(d) for d, q in vc.items() if int(q) == max_freq])
            return int(cand[0]) if cand else None

        mapa_dia_venc = {}
        if not df_receber.empty and "Venda" in df_receber.columns and "Vencimento" in df_receber.columns:
            base_dv = df_receber[["Venda", "Vencimento"]].copy()
            if "Status_Vencimento" in df_receber.columns:
                base_dv["Status_Vencimento"] = (
                    df_receber["Status_Vencimento"].fillna("").astype(str).str.strip().str.upper()
                )
            else:
                base_dv["Status_Vencimento"] = ""
            base_dv["Venda"] = base_dv["Venda"].fillna("").astype(str).str.strip()
            base_dv["_dia"] = pd.to_datetime(base_dv["Vencimento"], errors="coerce").dt.day
            base_dv = base_dv.loc[base_dv["Venda"].ne("")].copy()

            for venda, g in base_dv.groupby("Venda", sort=False):
                g_av = g.loc[g["Status_Vencimento"] == "A VENCER", "_dia"]
                d = _dia_moda_predominante(g_av)
                if d is None:
                    d = _dia_moda_predominante(g["_dia"])
                if d is not None:
                    mapa_dia_venc[str(venda).strip()] = int(d)

        if mapa_dia_venc:
            vv = consolidado["Venda"].fillna("").astype(str).str.strip()
            consolidado["DIA VENCIMENTO"] = vv.map(mapa_dia_venc)
            consolidado["DIA VENCIMENTO"] = pd.to_numeric(
                consolidado["DIA VENCIMENTO"], errors="coerce"
            ).astype("Int64")

        if vendas_aporte is None:
            vendas_aporte = set()
        vendas_aporte_norm = set([str(v).strip() for v in vendas_aporte if str(v).strip() != ""])
        if "Venda" in consolidado.columns:
            consolidado["APORTE"] = consolidado["Venda"].astype(str).str.strip().isin(vendas_aporte_norm).map(
                {True: "SIM", False: "NÃO"}
            )

    check_encargos = (
        consolidado["Vl.Principal (Encargos)"]
        - (
            consolidado["Vl.Principal Atrasado"]
            + consolidado["Vl.Correção"]
            + consolidado["Vl.Juros"]
            + consolidado["Vl.Multas"]
            + consolidado["Vl.Correção Atraso"]
        )
    ).round(2).sum()

    _vl_p_ck = pd.to_numeric(consolidado.get("Vl.Pago", 0), errors="coerce").fillna(0)
    _enc_ck = pd.to_numeric(consolidado.get("Vl.Principal (Encargos)", 0), errors="coerce").fillna(0)
    _vl_v_ck = pd.to_numeric(consolidado.get("Vl.Vencer", 0), errors="coerce").fillna(0)
    _cart_esp_ck = _calcular_vl_carteira_oficial(_vl_p_ck, _enc_ck, _vl_v_ck)
    check_carteira = (consolidado["Vl.Carteira"] - _cart_esp_ck).round(2).sum()

    if round(check_encargos, 2) != 0:
        print(f"[AVISO] Inconsistência na composição dos encargos: {check_encargos}")
        alertas_conferencia.append({
            "Venda": "GERAL",
            "Cliente_Base": "",
            "Tipo_Alerta": "ENCARGOS",
            "Mensagem": f"Composição dos encargos divergente: {check_encargos}",
            "Valor_Esperado": 0,
            "Valor_Encontrado": float(round(check_encargos, 2)),
            "Regra": "Vl.Principal (Encargos) = Principal + Correcao + Juros + Multas + Correcao Atraso",
            "Observacao": "Nao bloqueante",
        })

    if round(check_carteira, 2) != 0:
        print(f"[AVISO] Inconsistência na composição da carteira: {check_carteira}")
        alertas_conferencia.append({
            "Venda": "GERAL",
            "Cliente_Base": "",
            "Tipo_Alerta": "CARTEIRA",
            "Mensagem": f"Composição da carteira divergente: {check_carteira}",
            "Valor_Esperado": 0,
            "Valor_Encontrado": float(round(check_carteira, 2)),
            "Regra": _descricao_vl_carteira_modo(),
            "Observacao": "Nao bloqueante",
        })

    # Checagem automática interna de consistência de carteira.
    consolidado["_Erro_Interno_Carteira"] = (
        (
            _cart_esp_ck - consolidado["Vl.Carteira"]
        ).abs() > 0.01
    )
    qtd_erros_carteira = int(consolidado["_Erro_Interno_Carteira"].sum())
    if qtd_erros_carteira > 0:
        print(f"[AVISO] Inconsistência interna de carteira em {qtd_erros_carteira} venda(s).")
        alertas_conferencia.append({
            "Venda": "GERAL",
            "Cliente_Base": "",
            "Tipo_Alerta": "CARTEIRA",
            "Mensagem": f"Inconsistência interna de carteira em {qtd_erros_carteira} venda(s).",
            "Divergencia": qtd_erros_carteira,
        })

    qtd_quitados_sem_identificador = len(
        consolidado[
            (consolidado["Status venda"].astype(str).str.upper().str.strip() == "QUITADO")
            & (consolidado["Identificador"].astype(str).str.strip() == "")
        ]
    )
    print(f"Quitados sem identificador após ajuste: {qtd_quitados_sem_identificador}")

    for col in colunas_finais:
        if col not in consolidado.columns:
            consolidado[col] = 0 if col not in [
                "Emp/Obra", "Empreendimento", "Venda", "Cliente", "Identificador",
                "Status venda", "Status Construção", "Judicializado", "APORTE"
            ] else ""

    if not consolidado.empty and consolidado["Venda"].duplicated().any():
        alertas_conferencia.append({
            "Venda": "GERAL",
            "Cliente_Base": "",
            "Tipo_Alerta": "DUPLICIDADE",
            "Mensagem": "Existem linhas duplicadas por Venda no consolidado antes da exportacao.",
            "Divergencia": int(consolidado["Venda"].duplicated().sum()),
        })

    consolidado = consolidado.sort_values(
        by=["Qtd.Parc.Atrasada", "Venda"],
        ascending=[False, True]
    ).reset_index(drop=True)

    # Conferencia automatica por venda antes da exportacao (nao bloqueante).
    if not consolidado.empty:
        mapa_pago_base = {}
        mapa_principal_base = {}
        if not df_receber_raw.empty and "Principal" in df_receber_raw.columns:
            base_pri = adicionar_chave_cliente(df_receber_raw.copy())
            base_pri["Venda"] = base_pri["Venda"].fillna("").astype(str).str.strip()
            mapa_principal_base = base_pri.groupby("Venda")["Principal"].sum().to_dict()
        if not df_recebidos_raw.empty and "Total_Dep" in df_recebidos_raw.columns:
            base_pago = adicionar_chave_cliente(df_recebidos_raw.copy())
            base_pago["Venda"] = base_pago["Venda"].fillna("").astype(str).str.strip()
            mapa_pago_base = base_pago.groupby("Venda")["Total_Dep"].sum().to_dict()

        mapa_vencer_base = {}
        mapa_inad_base = {}
        if (
            not df_receber_raw.empty
            and "Vlr_Parcela" in df_receber_raw.columns
        ):
            base_vencer = adicionar_chave_cliente(df_receber_raw.copy())
            base_vencer["Venda"] = base_vencer["Venda"].fillna("").astype(str).str.strip()
            if "Status_Vencimento" not in base_vencer.columns:
                base_vencer["Status_Vencimento"] = ""
            mask_st = (
                base_vencer["Status_Vencimento"].astype(str).str.strip().str.upper() == "A VENCER"
            )
            base_vencer = base_vencer.loc[mask_st].copy()
            base_vencer["_vp"] = pd.to_numeric(base_vencer["Vlr_Parcela"], errors="coerce").fillna(0)
            mapa_vencer_base = base_vencer.groupby("Venda")["_vp"].sum().to_dict()

        if not df_receber_raw.empty and "Principal" in df_receber_raw.columns:
            base_inad = adicionar_chave_cliente(df_receber_raw.copy())
            base_inad["Venda"] = base_inad["Venda"].fillna("").astype(str).str.strip()
            if "Status_Vencimento" not in base_inad.columns:
                base_inad["Status_Vencimento"] = ""
            st_inad = base_inad["Status_Vencimento"].astype(str).str.strip().str.upper()
            base_inad = base_inad.loc[st_inad == "VENCIDO"].copy()
            mapa_inad_base = base_inad.groupby("Venda")["Principal"].sum().to_dict()
        else:
            mapa_inad_base = {}

        for _, r in consolidado.iterrows():
            venda_ref = str(r.get("Venda", "")).strip()
            vp = float(r.get("Valor Da Parcela", 0) or 0)
            if vp <= 0 and float(mapa_principal_base.get(venda_ref, 0) or 0) > 0:
                # venda possui base financeira e parcela nao pode ficar zerada.
                alertas_conferencia.append({
                    "Venda": venda_ref,
                    "Cliente_Base": "",
                    "Tipo_Alerta": "VALOR_PARCELA",
                    "Mensagem": "Valor Da Parcela zerado para venda com base financeira.",
                    "Valor_Esperado": "Moda(Principal) > 0",
                    "Valor_Encontrado": vp,
                    "Regra": "Valor Da Parcela = moda de Principal por venda",
                    "Observacao": "Nao bloqueante",
                })

            esperado_pago = float(mapa_pago_base.get(venda_ref, 0) or 0)
            encontrado_pago = float(r.get("Vl.Pago", 0) or 0)
            if abs(encontrado_pago - esperado_pago) > 0.01:
                alertas_conferencia.append({
                    "Venda": venda_ref,
                    "Cliente_Base": "",
                    "Tipo_Alerta": "PAGO",
                    "Mensagem": "Divergencia entre Vl.Pago e base de recebidos.",
                    "Valor_Esperado": esperado_pago,
                    "Valor_Encontrado": encontrado_pago,
                    "Regra": "Vl.Pago = soma(Total_Dep) em Dados Recebidos",
                    "Observacao": "Nao bloqueante",
                })

            esperado_vencer = float(mapa_vencer_base.get(venda_ref, 0) or 0)
            encontrado_vencer = float(r.get("Vl.Vencer", 0) or 0)
            if abs(encontrado_vencer - esperado_vencer) > 0.01:
                alertas_conferencia.append({
                    "Venda": venda_ref,
                    "Cliente_Base": "",
                    "Tipo_Alerta": "A_VENCER",
                    "Mensagem": "Divergencia entre Vl.Vencer e base de receber.",
                    "Valor_Esperado": esperado_vencer,
                    "Valor_Encontrado": encontrado_vencer,
                    "Regra": "Vl.Vencer = soma Vlr_Parcela (coluna R) com Status_Vencimento = A VENCER",
                    "Observacao": "Nao bloqueante",
                })

            esperado_inad = float(mapa_inad_base.get(venda_ref, 0) or 0)
            encontrado_inad = float(r.get("Vl.Principal Atrasado", 0) or 0)
            if abs(encontrado_inad - esperado_inad) > 0.01:
                alertas_conferencia.append({
                    "Venda": venda_ref,
                    "Cliente_Base": "",
                    "Tipo_Alerta": "INADIMPLENCIA",
                    "Mensagem": "Divergencia entre principal atrasado e base de receber.",
                    "Valor_Esperado": esperado_inad,
                    "Valor_Encontrado": encontrado_inad,
                    "Regra": "Vl.Principal Atrasado = soma Principal em Dados Receber com Status_Vencimento = VENCIDO",
                    "Observacao": "Nao bloqueante",
                })

    # Validação obrigatória pré-exportação: 1 linha por Venda + alinhamento estrito com Dados Receber
    if not consolidado.empty:
        consolidado["Venda"] = consolidado["Venda"].fillna("").astype(str).str.strip()
        if consolidado["Venda"].duplicated().any():
            dup_vendas = (
                consolidado.loc[consolidado["Venda"].duplicated(keep=False), "Venda"]
                .fillna("").astype(str).str.strip().tolist()
            )
            _dbg(f"FECHAMENTO_DUP: vendas duplicadas apos fechamento final: {sorted(set([v for v in dup_vendas if v]))[:20]}")

        if not df_receber.empty:
            dr = df_receber.copy()
            if "Status_Vencimento" not in dr.columns:
                dr["Status_Vencimento"] = ""
            dr["Venda"] = dr["Venda"].fillna("").astype(str).str.strip()
            st = dr["Status_Vencimento"].astype(str).str.strip().str.upper()

            # A VENCER (somente status A VENCER; coluna R = Vlr_Parcela)
            mapa_vencer = mapa_vl_vencer_por_venda_receber_tratado(dr)
            d_av = dr.loc[st == "A VENCER"].copy()
            mapa_qtd_av = {str(k).strip(): int(v) for k, v in d_av.groupby("Venda")["Parcela"].agg(contar_parcelas_distintas_padrao).to_dict().items()} if not d_av.empty else {}

            # INADIMPLÊNCIA (somente status VENCIDO)
            d_inad = dr.loc[st == "VENCIDO"].copy()
            if not d_inad.empty:
                g_inad = d_inad.groupby("Venda", as_index=False).agg({
                    "Parcela": contar_parcelas_distintas_padrao,
                    "Principal": "sum",
                    "Correcao": "sum",
                    "Juros_Atraso": "sum",
                    "Multa_Atraso": "sum",
                    "Correcao_Atraso": "sum",
                })
                mapa_qtd_inad = {str(k).strip(): int(v) for k, v in g_inad.set_index("Venda")["Parcela"].to_dict().items()}
                mapa_pri_inad = {str(k).strip(): float(v) for k, v in g_inad.set_index("Venda")["Principal"].to_dict().items()}
                mapa_cor_inad = {str(k).strip(): float(v) for k, v in g_inad.set_index("Venda")["Correcao"].to_dict().items()}
                mapa_jur_inad = {str(k).strip(): float(v) for k, v in g_inad.set_index("Venda")["Juros_Atraso"].to_dict().items()}
                mapa_mul_inad = {str(k).strip(): float(v) for k, v in g_inad.set_index("Venda")["Multa_Atraso"].to_dict().items()}
                mapa_cra_inad = {str(k).strip(): float(v) for k, v in g_inad.set_index("Venda")["Correcao_Atraso"].to_dict().items()}
            else:
                mapa_qtd_inad, mapa_pri_inad, mapa_cor_inad, mapa_jur_inad, mapa_mul_inad, mapa_cra_inad = {}, {}, {}, {}, {}, {}

            vk = consolidado["Venda"].astype(str).str.strip()
            consolidado["Vl.Vencer"] = vk.map(mapa_vencer).fillna(0).astype(float)
            consolidado["Qtd.Parc.A Vencer"] = vk.map(mapa_qtd_av).fillna(0).astype(int)
            consolidado["Qtd.Parc.Atrasada"] = vk.map(mapa_qtd_inad).fillna(0).astype(int)
            consolidado["Vl.Principal Atrasado"] = vk.map(mapa_pri_inad).fillna(0).astype(float)
            consolidado["Vl.Correção"] = vk.map(mapa_cor_inad).fillna(0).astype(float)
            consolidado["Vl.Juros"] = vk.map(mapa_jur_inad).fillna(0).astype(float)
            consolidado["Vl.Multas"] = vk.map(mapa_mul_inad).fillna(0).astype(float)
            consolidado["Vl.Correção Atraso"] = vk.map(mapa_cra_inad).fillna(0).astype(float)

            consolidado["Vl.Principal (Encargos)"] = (
                consolidado["Vl.Principal Atrasado"]
                + consolidado["Vl.Correção"]
                + consolidado["Vl.Juros"]
                + consolidado["Vl.Multas"]
                + consolidado["Vl.Correção Atraso"]
            ).round(2)
            consolidado = _recalcular_vl_carteira_e_percentuais(consolidado)

            # Universo distinto de parcelas na venda (Receber): não reduzir Qtd.Parc.Total abaixo disso
            # quando coerente com teto de sanidade (evita ruído explodindo o total).
            mapa_uni_v = {}
            for vv, gv in dr.groupby("Venda", sort=False):
                vsk = str(vv).strip()
                if not vsk or "Parcela" not in gv.columns:
                    continue
                mapa_uni_v[vsk] = int(contar_parcelas_distintas_padrao(gv["Parcela"]))
            for ix in consolidado.index:
                vsk = str(consolidado.at[ix, "Venda"]).strip()
                u = int(mapa_uni_v.get(vsk, 0) or 0)
                if u <= 0:
                    continue
                qt = int(consolidado.at[ix, "Qtd.Parc.Total"] or 0)
                qp = int(consolidado.at[ix, "Qtd.Parc.Paga"] or 0)
                qa = int(consolidado.at[ix, "Qtd.Parc.Atrasada"] or 0)
                qav = int(consolidado.at[ix, "Qtd.Parc.A Vencer"] or 0)
                piso = max(qp, qa, qav, 1)
                teto_sanidade = max(qt, piso, 1) * 3 + 24
                if u <= teto_sanidade:
                    consolidado.at[ix, "Qtd.Parc.Total"] = int(max(qt, u))

            alertas_conferencia.extend(
                auditoria_alertas_qtd_parcelas_consolidado(
                    consolidado,
                    mapa_uni_v,
                    mapa_moda_receber,
                    mapa_moda_recebidos,
                )
            )
        else:
            alertas_conferencia.extend(
                auditoria_alertas_qtd_parcelas_consolidado(
                    consolidado,
                    {},
                    mapa_moda_receber,
                    mapa_moda_recebidos,
                )
            )

    # Validação de fechamento da chave/técnica e escolha final por venda.
    _dbg("CHAVE_CHECK_FINAL: "
         f"df_receber_has_Cliente_Base={('Cliente_Base' in df_receber.columns) if isinstance(df_receber,pd.DataFrame) else False} | "
         f"df_recebidos_has_Cliente_Base={('Cliente_Base' in df_recebidos.columns) if isinstance(df_recebidos,pd.DataFrame) else False} | "
         f"base_chaves_has_Cliente_Base={('Cliente_Base' in base_chaves.columns) if isinstance(base_chaves,pd.DataFrame) else False}")
    _dbg(f"FECHAMENTO_VENDAS: entradas_unicas={df_receber['Venda'].fillna('').astype(str).str.strip().nunique() if not df_receber.empty and 'Venda' in df_receber.columns else 0} | "
         f"saida_unicas={consolidado['Venda'].fillna('').astype(str).str.strip().nunique() if not consolidado.empty and 'Venda' in consolidado.columns else 0} | "
         f"duplicadas_saida={int(consolidado['Venda'].fillna('').astype(str).str.strip().duplicated().sum()) if not consolidado.empty and 'Venda' in consolidado.columns else 0}")
    if not consolidado.empty and DEBUG_DADOS:
        _dbg(f"ESCOLHAS_FINAIS cliente(20)={consolidado[['Venda','Cliente']].head(20).to_dict('records')}")
        _dbg(f"ESCOLHAS_FINAIS identificador(20)={consolidado[['Venda','Identificador']].head(20).to_dict('records')}")
        _dbg(
            "AMOSTRA_FINAL(10)="
            + str(
                consolidado[["Venda", "Cliente", "Identificador", "Valor Da Parcela"]]
                .head(10)
                .to_dict("records")
            )
        )

    _tick_bloco("fechamento_posprocessamento")
    consolidado = consolidado[colunas_finais].copy()

    if registro_etapas_tempo is not None:
        _t_montar_total = time.perf_counter() - t_montar_inicio
        _t_nucleo = max(0.0, _t_montar_total - tempo_dedup_receber - tempo_dedup_recebidos)
        _lr_f = len(df_receber) if not df_receber.empty else 0
        _lp_f = len(df_recebidos) if not df_recebidos.empty else 0
        _lc_f = len(consolidado) if not consolidado.empty else 0
        registro_etapas_tempo.append(
            ("montar_consolidado_nucleo", _t_nucleo, _lr_f, _lp_f, _lc_f)
        )
        print(
            "[TEMPO] montar_consolidado_nucleo.blocos: "
            f"preparacao_inicial={_tempo_blocos_montar['preparacao_inicial']:.2f}s | "
            f"cliente_final={_tempo_blocos_montar['cliente_final']:.2f}s | "
            f"identificador_final={_tempo_blocos_montar['identificador_final']:.2f}s | "
            f"metadados_base={_tempo_blocos_montar['metadados_base']:.2f}s | "
            f"valor_parcela_qtd_financeiro={_tempo_blocos_montar['valor_parcela_qtd_financeiro']:.2f}s | "
            f"montagem_consolidado_base={_tempo_blocos_montar['montagem_consolidado_base']:.2f}s | "
            f"reconciliacao_qtd_total={_tempo_blocos_montar['reconciliacao_qtd_total']:.2f}s | "
            f"evento_1_1={_tempo_blocos_montar['evento_1_1']:.2f}s | "
            f"sincronizacao_receber={_tempo_blocos_montar['sincronizacao_receber']:.2f}s | "
            f"fechamento_posprocessamento={_tempo_blocos_montar['fechamento_posprocessamento']:.2f}s"
        )

    _validar_invariantes_exportacao_consolidado(consolidado)
    _validar_integridade_financeira(consolidado)

    alertas_conferencia.extend(auditoria_integridade_financeira_obrigatoria(consolidado))
    alertas_conferencia = classificar_alertas_confiabilidade(alertas_conferencia)

    df_alertas = pd.DataFrame(alertas_conferencia).drop_duplicates() if alertas_conferencia else pd.DataFrame()
    if DEBUG_DADOS and not consolidado.empty:
        sc_adj = ajustar_scores_com_alertas_tot(scores_qualidade_parcelas, alertas_conferencia)
        piores = sorted(sc_adj.items(), key=lambda x: x[1])[:12]
        print(f"[DEBUG][PARCELAS_SCORE] vendas_piores_score(amostra)={piores}")
        tot_alerts = [a for a in alertas_conferencia if str(a.get("Tipo_Alerta", "")).strip() == "PARCELAS_INCONSISTENTES_TOTAL"]
        if tot_alerts:
            print(f"[DEBUG][QTD_PARC] alertas_PARCELAS_INCONSISTENTES_TOTAL={len(tot_alerts)}")
        res_exec = gerar_resumo_auditoria_consolidado(consolidado, df_alertas, sc_adj)
        print(formatar_resumo_auditoria_para_log(res_exec))
        mapa_cf = calcular_confianca_final_por_venda(alertas_conferencia, scores_qualidade_parcelas)
        print(f"[DEBUG][CONFIANCA_VENDA] {resumo_confianca_executivo(mapa_cf)} | amostra_baixa={[(k,v) for k,v in sorted(mapa_cf.items(), key=lambda x: x[1]['Pontos'])[:8]]}")
    # Bases tratadas alinhadas ao motor (dedup/classificação) — usar na validação e exportação para reconciliação.
    return consolidado, data_base, df_alertas, df_receber, df_recebidos, pendencias_qtd_total_reconciliacao


# =========================
# ESTILO
# =========================
# Limites do modo turbo / autoajuste rápido na pós-formatação openpyxl.
# Rollback operacional: aumentar o valor desativa o modo agressivo sem reverter o PR inteiro.
LIMIAR_LINHAS_TURBO_CONSOLIDADO = 25000
LIMIAR_LINHAS_TURBO_RELATORIO_ANALITICO = 30000
LIMIAR_LINHAS_TURBO_PENDENCIAS = 12000


def _desfazer_merges_faixa_linhas(ws, linha_min: int, linha_max: int) -> None:
    """Remove merges que interceptam [linha_min, linha_max] (reaplicar blocos/cabeçalho sem conflito)."""
    for mcr in list(ws.merged_cells.ranges):
        if mcr.max_row < linha_min or mcr.min_row > linha_max:
            continue
        try:
            ws.unmerge_cells(str(mcr))
        except Exception:
            pass


def _autoajustar_colunas_e_linhas(
    ws,
    header_row: int = 8,
    data_start_row: int = 9,
    fixed_widths: dict | None = None,
    limite_coluna: int | None = None,
    max_scan_rows: int = 4000,
    modo_rapido: bool = False,
    ajustar_altura_linhas: bool = True,
) -> None:
    """
    Autoajuste de colunas/linhas por conteúdo, preservando larguras fixas informadas.
    Mantém cabeçalhos/merges estruturais e evita explosão de largura.
    """
    fixed = {str(k).upper(): float(v) for k, v in (fixed_widths or {}).items()}
    max_col = int(limite_coluna or ws.max_column or 1)
    max_row = int(ws.max_row or data_start_row)
    scan_until = min(max_row, max(1, int(max_scan_rows or 4000)))
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        if letter in fixed:
            ws.column_dimensions[letter].width = fixed[letter]
            continue
        if modo_rapido:
            v_header = ws.cell(row=header_row, column=col).value
            tam = len(str(v_header or "").strip())
            if tam <= 0:
                continue
            ws.column_dimensions[letter].width = min(max(tam + 3, 10), 28)
            continue
        best = 0
        for row in range(header_row, scan_until + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            txt = str(v)
            if not txt.strip():
                continue
            best = max(best, len(txt))
        if best <= 0:
            continue
        width = min(max(best + 2, 8), 42)
        cur = ws.column_dimensions[letter].width
        if cur is None:
            ws.column_dimensions[letter].width = width
        else:
            ws.column_dimensions[letter].width = max(float(cur), float(width))
    if modo_rapido or not ajustar_altura_linhas:
        return
    # Altura de linhas apenas para dados longos (mantém cabeçalho estável).
    for row in range(data_start_row, scan_until + 1):
        longest = 0
        for col in range(1, max_col + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            longest = max(longest, len(str(v)))
        if longest > 70:
            ws.row_dimensions[row].height = 30
        elif longest > 45:
            ws.row_dimensions[row].height = 22
        elif ws.row_dimensions[row].height is None:
            ws.row_dimensions[row].height = 15


def _nome_oficial_para_titulo_aba(ws, fallback: str) -> str:
    """Resolve nome oficial para título da aba usando primeira linha de dados (A9/B9)."""
    eo_a9 = normalizar_emp_obra(ws["A9"].value)
    of = empreendimento_oficial_para_emp_obra(eo_a9)
    if of:
        return of
    b9 = str(ws["B9"].value or "").strip().upper()
    if b9:
        return b9
    return str(fallback or "").strip().upper()


def _aplicar_estilo_aba_resumo_geral(wb, data_base, nome_empreendimento):
    """Mesma linguagem visual do Consolidado: painel (1–6), blocos (7), cabeçalho (8), dados (9+)."""
    if NOME_ABA_RESUMO_GERAL not in wb.sheetnames:
        return
    ws = wb[NOME_ABA_RESUMO_GERAL]
    _desfazer_merges_faixa_linhas(ws, 7, 8)
    azul_escuro = "10243F"
    verde = "92D050"
    vermelho = "F8696B"
    azul_claro = "00B0F0"
    amarelo = "FFFF00"
    branco = "FFFFFF"
    preto = "000000"
    cinza = "D9D9D9"
    borda_fina_branca = Side(style="thin", color="FFFFFF")
    borda_media_preta = Side(style="medium", color="000000")
    borda_fina_cinza = Side(style="thin", color="BFBFBF")

    ws["A1"] = "EMPREENDIMENTO"
    ws["B1"] = "CART.GERAL"
    try:
        ws.unmerge_cells("C1:L6")
    except Exception:
        pass
    ws.merge_cells("C1:L6")
    ws["C1"] = "RESUMO GERAL"
    ws["C1"].font = Font(name="Calibri", bold=True, color=branco, size=72)
    ws["C1"].fill = PatternFill("solid", fgColor=azul_escuro)
    ws["C1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["M1"] = "PAINEL ESTOQUE"
    ws["M2"] = "QTD.UNID.TOTAL"
    ws["M3"] = "QTD.UNID.INADIMPLENTES"
    ws["M4"] = "QTD.UNID.QUITADAS"
    ws["M5"] = "QTD.UNID.VENDIDAS"
    ws["M6"] = "QTD.UNID.DISPONIVEL"
    ws["A2"] = "DATA-BASE"
    ws["B2"] = data_base.strftime("%d/%m/%Y") if data_base else ""
    # Rótulos alinhados ao modelo visual (fórmulas / referências de coluna inalteradas).
    ws["A3"] = "QTD. VENDAS"
    ws["B3"] = "=SUBTOTAL(109,C9:C1048576)"
    ws["A4"] = "VL.CART.TOTAL"
    ws["B4"] = "=SUBTOTAL(109,J9:J1048576)"
    ws["A5"] = "VL.INADIMPLÊNCIA CART.TOTAL"
    ws["B5"] = "=SUBTOTAL(109,G9:G1048576)"
    ws["A6"] = "% INADIMPLÊNCIA CART.TOTAL"
    ws["B6"] = "=IFERROR(B5/B4,0)"
    for r in range(1, 7):
        ws.row_dimensions[r].height = 18
    ws.row_dimensions[7].height = 14.4
    ws.row_dimensions[8].height = 14.4

    for linha in range(1, 7):
        ws[f"A{linha}"].font = Font(name="Calibri", size=10, bold=True, color=branco)
        ws[f"A{linha}"].fill = PatternFill("solid", fgColor=azul_escuro)
        ws[f"A{linha}"].alignment = Alignment(horizontal="left", vertical="center")
        ws[f"A{linha}"].border = Border(
            left=borda_media_preta,
            right=borda_fina_branca,
            top=borda_media_preta if linha == 1 else borda_fina_branca,
            bottom=borda_fina_branca,
        )
        ws[f"B{linha}"].font = Font(name="Calibri", size=10, bold=True, color=preto)
        ws[f"B{linha}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"B{linha}"].border = Border(
            left=borda_fina_branca,
            right=borda_media_preta,
            top=borda_media_preta if linha == 1 else borda_fina_branca,
            bottom=borda_fina_branca,
        )
        ws[f"M{linha}"].font = Font(name="Calibri", size=10, bold=True, color=branco)
        ws[f"M{linha}"].fill = PatternFill("solid", fgColor=azul_escuro)
        ws[f"M{linha}"].alignment = Alignment(horizontal="left", vertical="center")
        ws[f"M{linha}"].border = Border(
            left=borda_media_preta,
            right=borda_fina_branca,
            top=borda_media_preta if linha == 1 else borda_fina_branca,
            bottom=borda_fina_branca,
        )

    # M1:M6 e N1:N6: contorno branco #FFFFFF + divisórias horizontais (painel estoque, anexo 5).
    _bd_ext_branco = Side(style="thin", color="FFFFFF")
    _bd_ext_preto = Side(style="medium", color="000000")
    for r in range(1, 7):
        ws[f"B{r}"].fill = PatternFill("solid", fgColor="BFBFBF")
        ws[f"M{r}"].border = Border(
            left=_bd_ext_branco,
            right=_bd_ext_branco,
            top=_bd_ext_branco if r == 1 else Side(style=None),
            bottom=_bd_ext_branco,
        )
        ws[f"N{r}"].fill = PatternFill("solid", fgColor="BFBFBF")
        ws[f"N{r}"].border = Border(
            left=_bd_ext_preto,
            right=_bd_ext_preto,
            top=_bd_ext_preto if r == 1 else Side(style=None),
            bottom=_bd_ext_preto if r == 6 else Side(style=None),
        )

    ws["B4"].number_format = "R$ #,##0.00"
    ws["B5"].number_format = "R$ #,##0.00"
    ws["B6"].number_format = "0.00%"
    ws["A5"] = "VL.INADIM.CART.TOTAL"
    ws["A6"] = "% INADIM.CART.TOTAL"

    blocos_rg = [
        ("A7:C7", "DADOS CADASTRO", azul_escuro, branco),
        ("D7:E7", "PAGO", verde, preto),
        ("F7:G7", "INADIMPLÊNCIA", vermelho, preto),
        ("H7:I7", "A VENCER", azul_claro, preto),
        ("J7:M7", "INDICADORES", amarelo, preto),
        ("N7:N7", "INFORMAÇÕES", "FFF2CC", preto),
    ]
    for faixa, titulo, cor, cor_fonte in blocos_rg:
        ws.merge_cells(faixa)
        celula = ws[faixa.split(":")[0]]
        celula.value = titulo
        celula.fill = PatternFill("solid", fgColor=cor)
        celula.font = Font(name="Calibri", size=10, bold=True, color=cor_fonte)
        celula.alignment = Alignment(horizontal="center", vertical="center")
        for row in ws[faixa]:
            for c in row:
                c.border = Border(
                    left=borda_fina_branca,
                    right=borda_fina_branca,
                    top=borda_media_preta,
                    bottom=borda_fina_branca,
                )

    for merged_range in list(ws.merged_cells.ranges):
        if str(merged_range) == "N7":
            ws.merged_cells.ranges.remove(merged_range)

    ws["F7"] = "INADIMPLENCIA"
    ws["N7"] = "INFORMAÇÕES"
    ws["N7"].fill = PatternFill("solid", fgColor="FFF2CC")
    ws["N7"].font = Font(name="Calibri", size=10, bold=True, color=preto)
    ws["N7"].alignment = Alignment(horizontal="center", vertical="center")
    ws["N7"].border = Border(
        left=borda_fina_branca,
        right=borda_fina_branca,
        top=borda_media_preta,
        bottom=borda_fina_branca,
    )

    hdr_seg_rg = [
        ("A", "C", azul_escuro, branco),
        ("D", "E", verde, preto),
        ("F", "G", vermelho, preto),
        ("H", "I", azul_claro, preto),
        ("J", "M", amarelo, preto),
        ("N", "N", "FFF2CC", preto),
    ]
    for c0, c1, fg, fc in hdr_seg_rg:
        for ci in range(column_index_from_string(c0), column_index_from_string(c1) + 1):
            cell = ws.cell(row=8, column=ci)
            if cell.value is not None:
                cell.value = _padronizar_rotulo_coluna_exibicao(str(cell.value))
            if ci == column_index_from_string("N"):
                cell.value = "CLASSIFICAÇÃO"
            cell.font = Font(name="Calibri", size=10, bold=True, color=fc)
            cell.fill = PatternFill("solid", fgColor=fg)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(
                left=borda_fina_branca,
                right=borda_fina_branca,
                top=borda_fina_branca,
                bottom=borda_media_preta,
            )

    headers_resumo = {
        "A": "EMP/OBRA",
        "B": "EMPREENDIMENTO",
        "C": "QTD VENDAS",
        "D": "QTD.PARC.",
        "E": "VALOR",
        "F": "QTD.PARC.",
        "G": "VALOR",
        "H": "QTD.PARC.",
        "I": "VALOR",
        "J": "VL.CARTEIRA",
        "K": "% PAGO",
        "L": "% INADIMPLENCIA",
        "M": "% A VENCER",
        "N": "CLASSIFICAÇÃO",
    }
    for col, titulo in headers_resumo.items():
        ws[f"{col}8"] = titulo

    ws["N8"].font = Font(name="Calibri", size=10, bold=True, color=preto)
    ws["N8"].fill = PatternFill("solid", fgColor="FFF2CC")
    ws["N8"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["N8"].border = Border(
        left=borda_fina_branca,
        right=borda_fina_branca,
        top=borda_fina_branca,
        bottom=borda_media_preta,
    )

    max_row_data = ws.max_row or 9
    max_col_data = ws.max_column or 14
    border_data = Border(
        left=borda_fina_cinza,
        right=borda_fina_cinza,
        top=borda_fina_cinza,
        bottom=borda_fina_cinza,
    )
    align_centro = Alignment(horizontal="center", vertical="center")
    colunas_moeda = frozenset({"E", "G", "I", "J"})
    colunas_inteiras = frozenset({"C", "D", "F", "H"})
    colunas_percentuais = frozenset({"K", "L", "M"})

    for linha in range(9, max_row_data + 1):
        for col in range(1, max_col_data + 1):
            cell = ws.cell(row=linha, column=col)
            letter = get_column_letter(col)
            cell.alignment = align_centro
            cell.border = border_data
            if linha % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F2F2F2")
            else:
                cell.fill = PatternFill("solid", fgColor="FFFFFF")
            if letter in colunas_moeda:
                cell.number_format = "R$ #,##0.00"
            elif letter in colunas_inteiras:
                cell.number_format = "0"
            elif letter in colunas_percentuais:
                cell.number_format = "0.00%"
        ws.row_dimensions[linha].height = 15.0

    for col in ["C", "E", "G", "I", "M", "N"]:
        for linha in range(7, max_row_data + 1):
            cell = ws[f"{col}{linha}"]
            cell.border = Border(
                left=cell.border.left,
                right=borda_media_preta,
                top=cell.border.top,
                bottom=cell.border.bottom,
            )

    # Destaque visual solicitado: coluna % INADIMPLÊNCIA (L) e CLASSIFICAÇÃO (N)
    # com a mesma paleta do status (alto/médio/baixo).
    if max_row_data >= 9:
        fill_alto = PatternFill("solid", fgColor="FF5E5E")
        fill_medio = PatternFill("solid", fgColor="FFFF00")
        fill_baixo = PatternFill("solid", fgColor="00B0F0")
        font_cond = Font(name="Calibri", size=10, bold=True, color=preto)

        ws.conditional_formatting.add(
            f"L9:L{max_row_data}",
            FormulaRule(formula=["$L9>=0.15"], stopIfTrue=True, fill=fill_alto, font=font_cond),
        )
        ws.conditional_formatting.add(
            f"L9:L{max_row_data}",
            FormulaRule(formula=["AND($L9>=0.06,$L9<0.15)"], stopIfTrue=True, fill=fill_medio, font=font_cond),
        )
        ws.conditional_formatting.add(
            f"L9:L{max_row_data}",
            FormulaRule(formula=['AND($L9<0.06,$L9<>"")'], stopIfTrue=True, fill=fill_baixo, font=font_cond),
        )

        ws.conditional_formatting.add(
            f"N9:N{max_row_data}",
            FormulaRule(formula=['OR(UPPER($N9)="ALTO",UPPER($N9)="ALTA")'], stopIfTrue=True, fill=fill_alto),
        )
        ws.conditional_formatting.add(
            f"N9:N{max_row_data}",
            FormulaRule(formula=['OR(UPPER($N9)="MÉDIO",UPPER($N9)="MEDIO")'], stopIfTrue=True, fill=fill_medio),
        )
        ws.conditional_formatting.add(
            f"N9:N{max_row_data}",
            FormulaRule(formula=['UPPER($N9)="BAIXO"'], stopIfTrue=True, fill=fill_baixo),
        )

        for linha in range(9, max_row_data + 1):
            cell_n = ws[f"N{linha}"]
            valor_n = str(cell_n.value or "").strip().upper()
            if valor_n in {"ALTO", "ALTA"}:
                cell_n.fill = fill_alto
                cell_n.font = font_cond
            elif valor_n in {"MÉDIO", "MEDIO"}:
                cell_n.fill = fill_medio
                cell_n.font = font_cond
            elif valor_n == "BAIXO":
                cell_n.fill = fill_baixo
                cell_n.font = font_cond

    def _norm_txt_rg(valor) -> str:
        s = str(valor or "").strip().upper()
        s = unicodedata.normalize("NFKD", s)
        s = "".join(ch for ch in s if not unicodedata.combining(ch))
        s = re.sub(r"\s+", " ", s)
        return s

    def _sigla_aba_consolidado(nome_aba: str) -> str:
        n = _norm_txt_rg(nome_aba)
        if (
            not n
            or n == _norm_txt_rg(NOME_ABA_RESUMO_GERAL)
            or n in {"DADOS RECEBER", "DADOS RECEBIDOS", "PEND.PARCELAS", "RELATORIO ANALITICO", "CRITERIOS ANALISES", "CONSOLIDADO ESTOQUE"}
        ):
            return ""
        m = re.match(r"^([A-Z0-9]{2,})", n)
        return m.group(1) if m else ""

    def _mapa_headers_aba(ws_aba):
        out = {}
        for c in ws_aba[8]:
            if c.value is None:
                continue
            out[_norm_txt_rg(c.value)] = c.column_letter
        return out

    def _achar_coluna(headers: dict, aliases) -> str:
        for a in aliases:
            na = _norm_txt_rg(a)
            if na in headers:
                return headers[na]
        for a in aliases:
            na = _norm_txt_rg(a)
            for h, col in headers.items():
                if na and na in h:
                    return col
        return ""

    mapa_sigla_para_aba = {}
    for nome_aba in wb.sheetnames:
        sig = _sigla_aba_consolidado(nome_aba)
        if sig and sig not in mapa_sigla_para_aba:
            mapa_sigla_para_aba[sig] = nome_aba

    # Lote por empreendimento: converte as linhas do RESUMO em fórmulas dinâmicas
    # para respeitar filtros aplicados em cada aba consolidada.
    for linha in range(9, max_row_data + 1):
        emp_obra = str(ws[f"A{linha}"].value or "").strip()
        if not emp_obra:
            continue
        sig = normalizar_emp_obra(emp_obra)
        nome_aba_cons = mapa_sigla_para_aba.get(sig)
        if not nome_aba_cons or nome_aba_cons not in wb.sheetnames:
            continue
        hs = _mapa_headers_aba(wb[nome_aba_cons])
        c_venda = _achar_coluna(hs, ("VENDA",))
        c_qpago = _achar_coluna(hs, ("QTD.PARC.PAGA", "QTD.PARC.PAGO", "QTD.PAGO"))
        c_vlpago = _achar_coluna(hs, ("VL.PAGO",))
        c_qinad = _achar_coluna(hs, ("QTD.PARC.VENCIDA", "QTD.PARC.ATRASADA", "QTD.PARC.INADIMPLENCIA"))
        c_vlinad = _achar_coluna(hs, ("VL.PRINCIPAL (ENCARGOS)", "VL.VENCIDO", "VALOR INADIMPLENCIA"))
        c_qav = _achar_coluna(hs, ("QTD.PARC.A VENCER",))
        c_vlav = _achar_coluna(hs, ("VL.A VENCER", "VL.VENCER"))
        c_vlcart = _achar_coluna(hs, ("VL.CARTEIRA",))
        if not all((c_venda, c_qpago, c_vlpago, c_qinad, c_vlinad, c_qav, c_vlav, c_vlcart)):
            continue
        aba_q = quote_sheetname(nome_aba_cons)
        ws[f"C{linha}"] = f"=SUBTOTAL(103,{aba_q}!${c_venda}$9:${c_venda}$1048576)"
        ws[f"D{linha}"] = f"=SUBTOTAL(109,{aba_q}!${c_qpago}$9:${c_qpago}$1048576)"
        ws[f"E{linha}"] = f"=SUBTOTAL(109,{aba_q}!${c_vlpago}$9:${c_vlpago}$1048576)"
        ws[f"F{linha}"] = f"=SUBTOTAL(109,{aba_q}!${c_qinad}$9:${c_qinad}$1048576)"
        ws[f"G{linha}"] = f"=SUBTOTAL(109,{aba_q}!${c_vlinad}$9:${c_vlinad}$1048576)"
        ws[f"H{linha}"] = f"=SUBTOTAL(109,{aba_q}!${c_qav}$9:${c_qav}$1048576)"
        ws[f"I{linha}"] = f"=SUBTOTAL(109,{aba_q}!${c_vlav}$9:${c_vlav}$1048576)"
        ws[f"J{linha}"] = f"=SUBTOTAL(109,{aba_q}!${c_vlcart}$9:${c_vlcart}$1048576)"
        ws[f"K{linha}"] = f"=IFERROR(E{linha}/J{linha},0)"
        ws[f"L{linha}"] = f"=IFERROR(G{linha}/J{linha},0)"
        ws[f"M{linha}"] = f"=IFERROR(I{linha}/J{linha},0)"
        ws[f"N{linha}"] = f"=IF(L{linha}>=15%,\"ALTO\",IF(L{linha}>=6%,\"MÉDIO\",\"BAIXO\"))"

    if max_row_data >= 9:
        ws.conditional_formatting.add(
            f"L9:L{max_row_data}",
            CellIsRule(operator="greaterThanOrEqual", formula=["0.15"], stopIfTrue=True, fill=fill_alto, font=font_cond),
        )
        ws.conditional_formatting.add(
            f"L9:L{max_row_data}",
            FormulaRule(formula=["AND($L9>=0.06,$L9<0.15)"], stopIfTrue=True, fill=fill_medio, font=font_cond),
        )
        ws.conditional_formatting.add(
            f"L9:L{max_row_data}",
            FormulaRule(formula=['AND($L9<0.06,$L9<>"")'], stopIfTrue=True, fill=fill_baixo, font=font_cond),
        )

        ws.conditional_formatting.add(
            f"N9:N{max_row_data}",
            FormulaRule(formula=['OR(UPPER($N9)="ALTO",UPPER($N9)="ALTA")'], stopIfTrue=True, fill=fill_alto, font=font_cond),
        )
        ws.conditional_formatting.add(
            f"N9:N{max_row_data}",
            FormulaRule(formula=['OR(UPPER($N9)="MÉDIO",UPPER($N9)="MEDIO")'], stopIfTrue=True, fill=fill_medio, font=font_cond),
        )
        ws.conditional_formatting.add(
            f"N9:N{max_row_data}",
            FormulaRule(formula=['UPPER($N9)="BAIXO"'], stopIfTrue=True, fill=fill_baixo, font=font_cond),
        )

    # Larguras do modelo de referência (Downloads\CARTEIRAS GERAL.xlsx), aba RESUMO GERAL.
    larguras_rg = {
        "A": 19.453125,
        "B": 27.0,
        "C": 13.62109375,
        "D": 11.68359375,
        "E": 17.6640625,
        "F": 11.68359375,
        "G": 15.5546875,
        "H": 11.68359375,
        "I": 16.5546875,
        "J": 17.6640625,
        "K": 9.62109375,
        "L": 17.8515625,
        "M": 23.0,
        "N": 18.0,
    }
    for col, largura in larguras_rg.items():
        ws.column_dimensions[col].width = largura

    _autoajustar_colunas_e_linhas(
        ws,
        header_row=8,
        data_start_row=9,
        fixed_widths=larguras_rg,
        ajustar_altura_linhas=False,
    )
    ws.freeze_panes = "A9"
    ws.auto_filter.ref = "A8:N8"

    # Mantém A6/B6 no mesmo padrão visual das linhas do painel (sem faixa cinza isolada).


def aplicar_estilo_arquivo_so_aba_resumo_geral(caminho_xlsx, data_base, nome_empreendimento):
    """Formata workbook que contém a aba RESUMO GERAL (uso no lote por empreendimento)."""
    wb = load_workbook(caminho_xlsx)
    if NOME_ABA_RESUMO_GERAL not in wb.sheetnames:
        wb.save(caminho_xlsx)
        return
    _aplicar_estilo_aba_resumo_geral(wb, data_base, nome_empreendimento)
    try:
        from openpyxl.workbook.properties import CalcProperties

        if wb.calculation is None:
            wb.calculation = CalcProperties()
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.calcCompleted = False
        wb.calculation.forceFullCalc = True
        if hasattr(wb.calculation, "calcMode"):
            wb.calculation.calcMode = "auto"
    except Exception:
        pass
    wb.save(caminho_xlsx)


def aplicar_estilo_arquivo_so_aba_consolidado_estoque(
    caminho_xlsx, data_base, nome_empreendimento, indicadores_painel=None
):
    """Formata workbook que contém apenas a aba CONSOLIDADO ESTOQUE (ex.: costura do lote por empreendimento)."""
    wb = load_workbook(caminho_xlsx)
    if "CONSOLIDADO ESTOQUE" not in wb.sheetnames:
        wb.save(caminho_xlsx)
        return
    _aplicar_estilo_aba_consolidado_estoque(wb, data_base, nome_empreendimento, indicadores_painel)
    try:
        from openpyxl.workbook.properties import CalcProperties

        if wb.calculation is None:
            wb.calculation = CalcProperties()
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.calcCompleted = False
        wb.calculation.forceFullCalc = True
        if hasattr(wb.calculation, "calcMode"):
            wb.calculation.calcMode = "auto"
    except Exception:
        pass
    wb.save(caminho_xlsx)


def _aplicar_estilo_aba_consolidado_estoque(wb, data_base, nome_empreendimento, indicadores_painel=None):
    """Painel gerencial (estoque + situação das vendidas), grade com filtro/congelamento e destaque de status."""
    from services.estoque_uau import CONSOLIDADO_ESTOQUE_PANDAS_STARTROW, calcular_indicadores_painel_consolidado_estoque

    nome_aba = "CONSOLIDADO ESTOQUE"
    if nome_aba not in wb.sheetnames:
        return
    ws = wb[nome_aba]
    ind = indicadores_painel if indicadores_painel is not None else calcular_indicadores_painel_consolidado_estoque(
        pd.DataFrame()
    )

    hdr_row = CONSOLIDADO_ESTOQUE_PANDAS_STARTROW + 1
    d0 = CONSOLIDADO_ESTOQUE_PANDAS_STARTROW + 2

    azul_escuro = "10243F"
    branco = "FFFFFF"
    preto = "000000"
    azul_claro_panel = "D9E1F2"
    borda_fina_branca = Side(style="thin", color="FFFFFF")
    borda_media_preta = Side(style="medium", color="000000")
    borda_fina_cinza = Side(style="thin", color="BFBFBF")

    def _borda_kpi(linha: int, topo_grosso: bool) -> None:
        top = borda_media_preta if topo_grosso else borda_fina_branca
        ws[f"A{linha}"].font = Font(bold=True, color=branco)
        ws[f"A{linha}"].fill = PatternFill("solid", fgColor=azul_escuro)
        ws[f"A{linha}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"A{linha}"].border = Border(
            left=borda_media_preta,
            right=borda_fina_branca,
            top=top,
            bottom=borda_fina_branca,
        )
        ws[f"B{linha}"].font = Font(bold=True, color=preto)
        ws[f"B{linha}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"B{linha}"].border = Border(
            left=borda_fina_branca,
            right=borda_media_preta,
            top=top,
            bottom=borda_fina_branca,
        )

    ws["A1"] = str(nome_empreendimento or "").strip().upper() or "EMPREENDIMENTOS"
    ws["B1"] = "EMPREENDIMENTOS -"
    ws["A2"] = data_base.strftime("%d/%m/%Y") if data_base else ""
    ws["B2"] = "DATA BASE -"
    for linha in (1, 2):
        _borda_kpi(linha, topo_grosso=(linha == 1))

    for rng in ("A3:P3", "A9:P9", "A18:P18"):
        ws.merge_cells(rng)
    for mcell, texto in (
        ("A3", "PAINEL CONSOLIDADO ESTOQUE"),
        ("A9", "INDICADORES DAS UNIDADES"),
        ("A18", "VISÃO ESTOQUE × FINANCEIRO (POR UNIDADE / IDENTIFICADOR)"),
    ):
        c = ws[mcell]
        c.value = texto
        c.fill = PatternFill("solid", fgColor=azul_escuro)
        c.font = Font(bold=True, color=branco)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    qt = max(int(ind.get("qtd_total", 0) or 0), 0)
    qi = int(ind.get("qtd_inadimplentes", 0) or 0)
    qq = int(ind.get("qtd_quitadas", 0) or 0)
    qv = int(ind.get("qtd_vendidas", 0) or 0)
    qd = int(ind.get("qtd_livres", 0) or 0)
    linhas = [
        (4, qt, "QTD.UNID.TOTAL -"),
        (5, qi, "QTD.UNID.INADIMPLENTES -"),
        (6, qq, "QTD.UNID.QUITADAS -"),
        (7, qv, "QTD.UNID.VENDIDAS -"),
        (8, qd, "QTD.UNID.DISPONIVEL -"),
    ]
    for linha, valor, rotulo in linhas:
        ws[f"A{linha}"] = valor
        ws[f"B{linha}"] = rotulo
        _borda_kpi(linha, topo_grosso=False)
        ws[f"A{linha}"].number_format = "0"
    ws["C5"] = f"=IFERROR(A5/A4,0)"
    ws["C6"] = f"=IFERROR(A6/A4,0)"
    ws["C7"] = f"=IFERROR(A7/A4,0)"
    ws["C8"] = f"=IFERROR(A8/A4,0)"
    for l in (5, 6, 7, 8):
        ws[f"C{l}"].number_format = '0.00"%"'
        ws[f"C{l}"].font = Font(bold=True, color=preto)
        ws[f"C{l}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"C{l}"].border = Border(left=borda_fina_branca, right=borda_media_preta, top=borda_fina_branca, bottom=borda_fina_branca)
    ws["A16"] = f"=SUBTOTAL(103,A{d0}:A1048576)"
    ws["B16"] = "QTD.UNID.TOTAL (FILTRO) -"
    ws["A17"] = f"=SUBTOTAL(109,M{d0}:M1048576)"
    ws["B17"] = "VL.CARTEIRA (FILTRO) -"
    for linha in (16, 17):
        ws[f"B{linha}"].font = Font(bold=True, color=preto)
        ws[f"B{linha}"].fill = PatternFill("solid", fgColor=azul_claro_panel)
        ws[f"B{linha}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"B{linha}"].border = Border(
            left=borda_media_preta,
            right=borda_fina_branca,
            top=borda_fina_branca,
            bottom=borda_fina_branca,
        )
        ws[f"A{linha}"].font = Font(bold=True, color=preto)
        ws[f"A{linha}"].fill = PatternFill("solid", fgColor=branco)
        ws[f"A{linha}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"A{linha}"].border = Border(
            left=borda_fina_branca,
            right=borda_media_preta,
            top=borda_fina_branca,
            bottom=borda_fina_branca,
        )
    ws["A17"].number_format = "R$ #,##0.00"

    for cell in ws[hdr_row]:
        cell.font = Font(bold=True, color=branco)
        cell.fill = PatternFill("solid", fgColor=azul_escuro)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=borda_fina_branca,
            right=borda_fina_branca,
            top=borda_fina_branca,
            bottom=borda_media_preta,
        )

    max_row_data = ws.max_row or d0
    max_col_data = min(ws.max_column or 18, 18)
    border_data = Border(
        left=borda_fina_cinza,
        right=borda_fina_cinza,
        top=borda_fina_cinza,
        bottom=borda_fina_cinza,
    )
    align_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)

    mapa_class = {
        "DISPONIVEL": "C6EFCE",
        "ADIMPLENTE": "E2EFDA",
        "INADIMPLENTE": "FFC7CE",
        "QUITADO": "D9E1F2",
    }

    for linha in range(d0, max_row_data + 1):
        clf = str(ws.cell(row=linha, column=6).value or "").strip().upper()
        fill_hex = mapa_class.get(clf)
        if not fill_hex:
            for k, cor in mapa_class.items():
                if clf.startswith(k):
                    fill_hex = cor
                    break
        fill_obj = PatternFill("solid", fgColor=fill_hex) if fill_hex else None

        for col in range(1, max_col_data + 1):
            cell = ws.cell(row=linha, column=col)
            cell.alignment = align_centro
            cell.border = border_data
            letter = get_column_letter(col)
            if letter in ("H", "J", "L", "M"):
                cell.number_format = "R$ #,##0.00"
            elif letter in ("G", "I", "K"):
                cell.number_format = "0"
            elif letter in ("N", "O", "P"):
                cell.number_format = '0.00"%"'
            if col == 6 and fill_obj:
                cell.fill = fill_obj
                cell.font = Font(bold=True, color=preto)

    for col in ("C", "F", "H", "J", "M", "P", "R"):
        for linha in range(18, max_row_data + 1):
            cell = ws[f"{col}{linha}"]
            cell.border = Border(
                left=cell.border.left,
                right=borda_media_preta,
                top=cell.border.top,
                bottom=cell.border.bottom,
            )

    larg = {
        "A": 14,
        "B": 28,
        "C": 25,
        "D": 12,
        "E": 28,
        "F": 16,
        "G": 10,
        "H": 14,
        "I": 10,
        "J": 14,
        "K": 12,
        "L": 14,
        "M": 14,
        "N": 10,
        "O": 10,
        "P": 10,
        "Q": 18,
        "R": 42,
    }
    for col, w in larg.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = f"A{d0}"
    if max_row_data >= hdr_row:
        ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(max_col_data)}{max_row_data}"


def aplicar_estilo_excel(
    caminho_saida,
    data_base,
    nome_empreendimento,
    nome_aba_principal,
    progress_cb=None,
    *,
    apenas_abas_apoio: bool = False,
    indicadores_estoque: dict | None = None,
):
    """
    progress_cb: função opcional(str) chamada em marcos da formatação (UI/logs).
    Não altera resultados financeiros — apenas observabilidade.
    apenas_abas_apoio: se True, formata só Dados Receber/Recebidos, relatório analítico e pendências
    (uso no lote por empreendimento com consolidados já estilizados em outras abas).
    """

    def _notify(msg):
        print(f"[ETAPA][excel_estilo] {msg}", flush=True)
        if progress_cb:
            try:
                progress_cb(msg)
            except Exception:
                pass

    _heartbeat_estado: dict[str, float] = {}

    def _notify_timed(chave: str, msg: str, intervalo_segundos: float = 1.5):
        agora = time.perf_counter()
        ultimo = _heartbeat_estado.get(chave, 0.0)
        if (agora - ultimo) >= intervalo_segundos:
            _heartbeat_estado[chave] = agora
            _notify(msg)

    def _norm_txt(s):
        t = str(s or "").strip().upper()
        repl = {
            "Á": "A",
            "À": "A",
            "Â": "A",
            "Ã": "A",
            "É": "E",
            "Ê": "E",
            "Í": "I",
            "Ó": "O",
            "Ô": "O",
            "Õ": "O",
            "Ú": "U",
            "Ç": "C",
        }
        for a, b in repl.items():
            t = t.replace(a, b)
        return " ".join(t.split())

    def _resolver_link_drive_empreendimento(ws_cons):
        if not MAPA_LINKS_DRIVE_EMPREENDIMENTO:
            return None
        mapa_norm = {
            _norm_txt(k).replace("–", "").replace("-", ""): str(v).strip()
            for k, v in MAPA_LINKS_DRIVE_EMPREENDIMENTO.items()
            if str(k).strip() != "" and str(v).strip() != ""
        }
        sigla = _norm_txt(str(ws_cons.title).split(" ")[0]).replace("–", "").replace("-", "")
        emp_nome = _norm_txt(nome_empreendimento)
        candidatos = []
        if sigla:
            candidatos.append(sigla)
        if emp_nome:
            candidatos.append(emp_nome.replace("–", "").replace("-", ""))
        for k in candidatos:
            url = mapa_norm.get(k)
            if isinstance(url, str) and url.strip():
                return url.strip()
        return None

    def _kpis_estoque_por_empreendimento(ws_cons):
        if "CONSOLIDADO ESTOQUE" not in wb.sheetnames:
            return None
        ws_es = wb["CONSOLIDADO ESTOQUE"]
        header_row = None
        for r in range(1, min(40, ws_es.max_row) + 1):
            vals = [_norm_txt(ws_es.cell(r, c).value) for c in range(1, min(35, ws_es.max_column) + 1)]
            if any(v == "EMPREENDIMENTO" for v in vals) and any(v.startswith("SITUA") for v in vals):
                header_row = r
                break
        if not header_row:
            return None
        mapa = {}
        for c in range(1, ws_es.max_column + 1):
            mapa[_norm_txt(ws_es.cell(header_row, c).value)] = c
        c_emp = mapa.get("EMPREENDIMENTO")
        c_eo = mapa.get("EMP/OBRA")
        c_sit = None
        for k, v in mapa.items():
            if k.startswith("SITUA"):
                c_sit = v
                break
        if not c_sit:
            return None
        emp_alvo = _norm_txt(nome_empreendimento)
        sigla = _norm_txt(str(ws_cons.title).split(" ")[0]).replace("–", "").replace("-", "")
        linhas = []
        for r in range(header_row + 1, ws_es.max_row + 1):
            emp = _norm_txt(ws_es.cell(r, c_emp).value) if c_emp else ""
            eo = _norm_txt(ws_es.cell(r, c_eo).value) if c_eo else ""
            ok_emp = bool(emp_alvo) and emp == emp_alvo
            ok_sig = bool(sigla) and (f"/{sigla}" in eo or eo.endswith(sigla))
            if ok_emp or ok_sig:
                linhas.append(r)
        if not linhas:
            return None
        total = len(linhas)
        q_inad = q_quit = q_disp = 0
        for r in linhas:
            s = _norm_txt(ws_es.cell(r, c_sit).value)
            if "INADIMPLENTE" in s:
                q_inad += 1
            elif "QUITADO" in s:
                q_quit += 1
            elif "DISPON" in s:
                q_disp += 1
        q_vend = max(total - q_disp, 0)
        return {
            "total": total,
            "inad": q_inad,
            "quit": q_quit,
            "vend": q_vend,
            "disp": q_disp,
        }

    wb = load_workbook(caminho_saida)

    azul_escuro = "10243F"
    verde = "92D050"
    vermelho = "F8696B"
    azul_claro = "00B0F0"
    amarelo = "FFFF00"
    branco = "FFFFFF"
    preto = "000000"
    cinza = "D9D9D9"

    borda_fina_branca = Side(style="thin", color="FFFFFF")
    borda_media_preta = Side(style="medium", color="000000")
    borda_fina_cinza = Side(style="thin", color="BFBFBF")

    aux_alignment_centro = Alignment(horizontal="center", vertical="center")
    aux_border_cinza = Border(
        left=borda_fina_cinza,
        right=borda_fina_cinza,
        top=borda_fina_cinza,
        bottom=borda_fina_cinza,
    )

    if not apenas_abas_apoio:
        ws = wb[nome_aba_principal]
        _desfazer_merges_faixa_linhas(ws, 7, 8)
        _notify("Consolidado: painel, blocos e cabeçalho")

        nome_oficial_titulo = _nome_oficial_para_titulo_aba(ws, nome_empreendimento)
        ws["A1"] = "EMPREENDIMENTO"
        ws["B1"] = nome_oficial_titulo
        try:
            ws.unmerge_cells("C1:U6")
        except Exception:
            pass
        ws.merge_cells("C1:U6")
        ws["C1"] = nome_oficial_titulo
        ws["C1"].font = Font(name="Calibri", bold=True, color=branco, size=72)
        ws["C1"].fill = PatternFill("solid", fgColor=azul_escuro)
        ws["C1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A2"] = "DATA-BASE"
        ws["B2"] = data_base.strftime("%d/%m/%Y") if data_base else ""
        ws["A3"] = "QTD. VENDAS"
        ws["B3"] = "=SUBTOTAL(103,C9:C1048576)"
        ws["A4"] = "VL.CART.TOTAL"
        ws["B4"] = "=SUBTOTAL(109,T9:T1048576)"
        ws["A5"] = "VL.INADIM.CART.TOTAL"
        ws["B5"] = "=SUBTOTAL(109,Q9:Q1048576)"
        ws["A6"] = "% INADIM.CART.TOTAL"
        ws["B6"] = "=IFERROR(B5/B4,0)"
        k_es = _kpis_estoque_por_empreendimento(ws)
        ws["Z1"] = "PAINEL ESTOQUE"
        ws["Z2"] = "QTD.UNID.TOTAL"
        ws["Z3"] = "QTD.UNID.INADIMPLENTES"
        ws["Z4"] = "QTD.UNID.QUITADAS"
        ws["Z5"] = "QTD.UNID.VENDIDAS"
        ws["Z6"] = "QTD.UNID.DISPONIVEL"
        link_drive = _resolver_link_drive_empreendimento(ws)
        ws["V6"] = "LINK DRIVE"
        if link_drive:
            ws["W6"] = "ACESSE CLICANDO AQUI"
            ws["W6"].hyperlink = link_drive
        else:
            ws["W6"] = "SEM LINK INFORMADO"
        try:
            ws.unmerge_cells("W6:Y6")
        except Exception:
            pass
        ws.merge_cells("W6:Y6")
        if k_es:
            tot = max(int(k_es["total"]), 0)
            ws["AA2"] = tot
            ws["AA3"] = int(k_es["inad"])
            ws["AA4"] = int(k_es["quit"])
            ws["AA5"] = int(k_es["vend"])
            ws["AA6"] = int(k_es["disp"])
        else:
            for c in ("AA2", "AA3", "AA4", "AA5", "AA6"):
                ws[c] = ""

        for linha in range(1, 7):
            ws[f"A{linha}"].font = Font(name="Calibri", size=10, bold=True, color=branco)
            ws[f"A{linha}"].fill = PatternFill("solid", fgColor=azul_escuro)
            ws[f"A{linha}"].alignment = Alignment(horizontal="left", vertical="center")
            ws[f"A{linha}"].border = Border(
                left=borda_media_preta, right=borda_fina_branca,
                top=borda_media_preta if linha == 1 else borda_fina_branca,
                bottom=borda_fina_branca
            )

            ws[f"B{linha}"].font = Font(name="Calibri", size=10, bold=True, color=preto)
            ws[f"B{linha}"].alignment = Alignment(horizontal="center", vertical="center")
            ws[f"B{linha}"].border = Border(
                left=borda_fina_branca, right=borda_media_preta,
                top=borda_media_preta if linha == 1 else borda_fina_branca,
                bottom=borda_fina_branca
            )
            ws[f"Z{linha}"].font = Font(name="Calibri", size=10, bold=True, color=branco)
            ws[f"Z{linha}"].fill = PatternFill("solid", fgColor=azul_escuro)
            ws[f"Z{linha}"].alignment = Alignment(horizontal="left", vertical="center")
            ws[f"Z{linha}"].border = Border(
                left=borda_media_preta, right=borda_fina_branca,
                top=borda_media_preta if linha == 1 else borda_fina_branca,
                bottom=borda_fina_branca
            )
            ws[f"AA{linha}"].font = Font(name="Calibri", size=10, bold=True, color=preto)
            ws[f"AA{linha}"].alignment = Alignment(horizontal="center", vertical="center")
            ws[f"AA{linha}"].border = Border(
                left=borda_fina_branca, right=borda_fina_branca,
                top=borda_media_preta if linha == 1 else borda_fina_branca,
                bottom=borda_fina_branca
            )
        ws["V6"].font = Font(name="Calibri", size=10, bold=True, color=branco)
        ws["V6"].fill = PatternFill("solid", fgColor=azul_escuro)
        ws["V6"].alignment = Alignment(horizontal="left", vertical="center")
        ws["V6"].border = Border(
            left=borda_media_preta, right=borda_fina_branca,
            top=borda_media_preta, bottom=borda_fina_branca
        )
        ws["W6"].font = Font(
            name="Calibri",
            size=10,
            bold=True,
            color="FFFFFF" if link_drive else "C00000",
            underline="single" if link_drive else None,
        )
        ws["W6"].fill = PatternFill("solid", fgColor="1D4ED8" if link_drive else "E5E7EB")
        ws["W6"].alignment = Alignment(horizontal="center", vertical="center")
        ws["W6"].border = Border(
            left=borda_fina_branca, right=borda_media_preta,
            top=borda_media_preta, bottom=borda_fina_branca
        )
        # Faixa V1:Y5: bloco azul contínuo, sem linhas internas (como anexo 2).
        for rr in range(1, 6):
            for cc in ("V", "W", "X", "Y"):
                c = ws[f"{cc}{rr}"]
                c.fill = PatternFill("solid", fgColor=azul_escuro)
                c.border = Border()

        # Z1:Z6 e AA1:AA6: bordas brancas iguais ao painel M/N do resumo (contorno + divisórias horizontais).
        _bd_w_panel = Side(style="thin", color="FFFFFF")
        _bd_ext_preto = Side(style="medium", color="000000")
        for rr in range(1, 7):
            ws[f"B{rr}"].fill = PatternFill("solid", fgColor="BFBFBF")
            ws[f"Z{rr}"].border = Border(
                left=borda_media_preta,
                right=_bd_w_panel,
                top=borda_media_preta if rr == 1 else _bd_w_panel,
                bottom=_bd_w_panel,
            )
            ws[f"AA{rr}"].fill = PatternFill("solid", fgColor="BFBFBF")
            ws[f"AA{rr}"].border = Border(
                left=_bd_ext_preto,
                right=_bd_ext_preto,
                top=_bd_ext_preto if rr == 1 else Side(style=None),
                bottom=_bd_ext_preto if rr == 6 else Side(style=None),
            )

        ws["B4"].number_format = 'R$ #,##0.00'
        ws["B5"].number_format = 'R$ #,##0.00'
        ws["B6"].number_format = '0.00%'
        for c in ("AA2", "AA3", "AA4", "AA5", "AA6"):
            ws[c].number_format = "0"

        # Blocos linha 7 conforme modelo (A:AA): mesmas cores; colunas de dados abaixo permanecem as do motor.
        blocos = [
            ("A7:F7", "DADOS CADASTRO", "10243F", branco),
            ("G7:H7", "DADOS FINANCEIRO", "C5D9F1", preto),
            ("I7:J7", "PAGO", "92D050", preto),
            ("K7:Q7", "INADIMPLENCIA", "FF5E5E", preto),
            ("R7:S7", "A VENCER", "00B0F0", preto),
            ("T7:W7", "INDICADORES", "FFFF00", preto),
            ("X7:AA7", "INFORMAÇÕES", "FFF2CC", preto),
        ]

        for faixa, titulo, cor, cor_fonte in blocos:
            ws.merge_cells(faixa)
            celula = ws[faixa.split(":")[0]]
            celula.value = titulo
            celula.fill = PatternFill("solid", fgColor=cor)
            celula.font = Font(name="Calibri", size=10, bold=True, color=cor_fonte)
            celula.alignment = Alignment(horizontal="center", vertical="center")

            for row in ws[faixa]:
                for c in row:
                    c.border = Border(
                        left=borda_fina_branca,
                        right=borda_fina_branca,
                        top=borda_media_preta,
                        bottom=borda_fina_branca
                    )

        titulos_linha_8 = {
            "A": "EMP/OBRA",
            "B": "EMPREENDIMENTO",
            "C": "VENDA",
            "D": "CLIENTE",
            "E": "IDENTIFICADOR",
            "F": "STATUS",
            "G": "VL.PARCELA",
            "H": "QTD.PARC.TOTAL",
            "I": "QTD.PARC.",
            "J": "VALOR",
            "K": "QTD.PARC.",
            "L": "VL.PRINCIPAL",
            "M": "VL.CORREÇÃO",
            "N": "VL.JUROS",
            "O": "VL.MULTAS",
            "P": "VL.CORREÇÃO ATRASO",
            "Q": "VL.PRINCIPAL (ENCARGOS)",
            "R": "QTD.PARC.",
            "S": "VALOR",
            "T": "VL.CARTEIRA",
            "U": "% PAGO",
            "V": "% INADIMPLENCIA",
            "W": "% A VENCER",
            "X": "DIA VENCIMENTO",
            "Y": "STATUS CONSTRUÇÃO",
            "Z": "JUDICIALIZADO",
            "AA": "APORTE",
        }
        hdr_seg_cons = [
            ("A", "F", "10243F", branco),
            ("G", "H", "C5D9F1", preto),
            ("I", "J", "92D050", preto),
            ("K", "Q", "FF5E5E", preto),
            ("R", "S", "00B0F0", preto),
            ("T", "W", "FFFF00", preto),
            ("X", "AA", "FFF2CC", preto),
        ]
        for c0, c1, fg, fc in hdr_seg_cons:
            for ci in range(column_index_from_string(c0), column_index_from_string(c1) + 1):
                cell = ws.cell(row=8, column=ci)
                cell.value = titulos_linha_8.get(get_column_letter(ci), cell.value)
                cell.font = Font(name="Calibri", size=10, bold=True, color=fc)
                cell.fill = PatternFill("solid", fgColor=fg)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = Border(
                    left=borda_fina_branca,
                    right=borda_fina_branca,
                    top=borda_fina_branca,
                    bottom=borda_media_preta
                )

        mapa_colunas_principal = {str(c.value).strip(): c.column_letter for c in ws[8] if c.value is not None}
        col_aporte = mapa_colunas_principal.get("APORTE", "")
        col_aporte_num = column_index_from_string(col_aporte) if col_aporte else None

        max_row_data = ws.max_row
        max_col_data = ws.max_column
        max_col_grade_consolidado = min(max_col_data, column_index_from_string("AA"))
        borda_fina_grade_consolidado = Side(style="thin", color="D8D8D8")
        border_data = Border(
            left=borda_fina_grade_consolidado,
            right=borda_fina_grade_consolidado,
            top=borda_fina_grade_consolidado,
            bottom=borda_fina_grade_consolidado,
        )
        align_centro_dados = Alignment(horizontal="center", vertical="center")
        fill_aporte_row = PatternFill("solid", fgColor="FFF2CC")
        fill_zebra_1 = PatternFill("solid", fgColor="FFFFFF")
        fill_zebra_2 = PatternFill("solid", fgColor="F2F2F2")
        fill_x_destaque = PatternFill("solid", fgColor="FFF2CC")
        fill_f_quit = PatternFill("solid", fgColor="C6EFCE")
        fill_f_inad = PatternFill("solid", fgColor="FFC7CE")
        fill_f_outros = PatternFill("solid", fgColor="BDD7EE")
        fill_k_gt3 = PatternFill("solid", fgColor="F8696B")
        fill_k_23 = PatternFill("solid", fgColor="FFFF00")
        fill_k_1 = PatternFill("solid", fgColor="D9D9D9")
        fill_k_0 = PatternFill("solid", fgColor="BDD7EE")
        font_f_bold_preto = Font(bold=True, color=preto)
        formato_contabil = '_-R$ * #,##0.00_-;[Red]-R$ * #,##0.00_-;_-R$ * "-"??_-;_-@_-'
        colunas_moeda = frozenset({"G", "J", "L", "M", "N", "O", "P", "Q", "S", "T"})
        colunas_inteiras = frozenset({"H", "I", "K", "R", "X"})
        colunas_percentuais = frozenset({"U", "V", "W"})

        col_letters_cache = [get_column_letter(col) for col in range(1, max_col_grade_consolidado + 1)]
        col_formato_cache = []
        for letter in col_letters_cache:
            if letter in colunas_moeda:
                col_formato_cache.append(formato_contabil)
            elif letter in colunas_inteiras:
                col_formato_cache.append('0')
            elif letter in colunas_percentuais:
                col_formato_cache.append('0.00%')
            else:
                col_formato_cache.append(None)
        col_f_idx = column_index_from_string("F")
        col_k_idx = column_index_from_string("K")
        col_x_idx = column_index_from_string("X")

        linhas_consolidado = max(max_row_data - 8, 0)
        modo_turbo_consolidado = linhas_consolidado >= LIMIAR_LINHAS_TURBO_CONSOLIDADO
        if modo_turbo_consolidado:
            _notify(
                "Consolidado: modo turbo ativado para base grande "
                f"({linhas_consolidado} linhas) — sem grade completa e zebra pesada."
            )
        else:
            _notify(f"Consolidado: formatação das linhas de dados (9..{max_row_data})")

        colunas_fechamento_idx = {
            column_index_from_string(col)
            for col in ("G", "K", "Q", "T", "X", "AA")
            if column_index_from_string(col) <= max_col_grade_consolidado
        }
        border_data_fechamento = border_data
        total_linhas_cons = max(max_row_data - 8, 1)
        for idx_linha, row_cells in enumerate(
            ws.iter_rows(min_row=9, max_row=max_row_data, min_col=1, max_col=max_col_grade_consolidado),
            start=1,
        ):
            if not row_cells:
                continue
            linha = row_cells[0].row
            ws.row_dimensions[linha].height = 15.0
            pct = int((idx_linha / total_linhas_cons) * 100)
            _notify_timed(
                "consolidado_linhas",
                f"Consolidado: linhas {linha}/{max_row_data} ({pct}%)",
                intervalo_segundos=1.2,
            )
            cell_f = row_cells[col_f_idx - 1] if len(row_cells) >= col_f_idx else None
            cell_k = row_cells[col_k_idx - 1] if len(row_cells) >= col_k_idx else None
            cell_aporte = row_cells[col_aporte_num - 1] if (col_aporte_num and len(row_cells) >= col_aporte_num) else None
            status = str((cell_f.value if cell_f is not None else "") or "").strip().upper()
            if status == "QUITADO":
                fill_f = fill_f_quit
            elif status == "INADIMPLENTE":
                fill_f = fill_f_inad
            else:
                fill_f = fill_f_outros
            valor = cell_k.value if cell_k is not None else 0
            try:
                qtd = int(valor or 0)
            except Exception:
                qtd = 0
            if qtd > 3:
                fill_k = fill_k_gt3
            elif qtd in (2, 3):
                fill_k = fill_k_23
            elif qtd == 1:
                fill_k = fill_k_1
            else:
                fill_k = fill_k_0
            aporte_sim = bool(
                cell_aporte is not None and str(cell_aporte.value or "").strip().upper() == "SIM"
            )
            for col, cell in enumerate(row_cells, start=1):
                fmt = col_formato_cache[col - 1]
                if fmt:
                    cell.number_format = fmt
                if not modo_turbo_consolidado:
                    cell.alignment = align_centro_dados
                cell.border = border_data_fechamento if col in colunas_fechamento_idx else border_data
                if col == col_f_idx:
                    cell.fill = fill_f
                elif col == col_x_idx:
                    cell.fill = fill_x_destaque
                elif modo_turbo_consolidado:
                    if aporte_sim and col <= min(max_col_data, 2):
                        cell.fill = fill_aporte_row
                else:
                    if aporte_sim:
                        cell.fill = fill_aporte_row
                        continue
                    cell.fill = fill_zebra_2 if (linha % 2 == 0) else fill_zebra_1

            if cell_f is not None:
                cell_f.font = font_f_bold_preto
            if cell_k is not None:
                cell_k.fill = fill_k
                cell_k.font = font_f_bold_preto

        # Larguras base do consolidado conforme referência; algumas colunas seguem autoajuste por conteúdo.
        larguras = {
            "A": 19.0,
            "B": 24.0,
            "C": 8.7,
            "D": 30.0,
            "E": 13.0,
            "F": 14.5,
            "G": 14.0,
            "H": 16.0,
            "I": 11.0,
            "J": 14.0,
            "K": 11.0,
            "L": 13.5,
            "M": 13.0,
            "N": 14.0,
            "O": 13.0,
            "P": 20.0,
            "Q": 24.0,
            "R": 11.0,
            "S": 13.0,
            "T": 14.0,
            "U": 9.0,
            "V": 17.2,
            "W": 12.2,
            "X": 24.0,
            "Y": 21.9,
            "Z": 22.0,
            "AA": 18.3,
        }

        for col, largura in larguras.items():
            if column_index_from_string(col) <= max_col_data or col == "AA":
                ws.column_dimensions[col].width = largura

        ajustes_largura_por_aba = {
            "CIDAN.CID.NOVA.ITP-PA": {
                "M": 14.203125,
                "N": 17.62109375,
                "O": 13.5,
                "Q": 24.0,
                "X": 24.0,
            },
            "NVLOT.NIL.VELOSO.RVD-GO": {
                "N": 14.7109375,
                "O": 13.5,
                "P": 23.421875,
                "Q": 24.0,
                "X": 24.0,
            },
        }
        for col, largura in ajustes_largura_por_aba.get(ws.title, {}).items():
            ws.column_dimensions[col].width = largura

        # Linhas superiores mais consistentes visualmente (painel/link/estoque).
        for r in range(1, 7):
            ws.row_dimensions[r].height = 15.0
        ws.row_dimensions[7].height = 14.4
        ws.row_dimensions[8].height = 14.4

        _scan_rows = 600 if max_row_data > 40000 else 1200 if max_row_data > 20000 else 4000
        _autoajustar_colunas_e_linhas(
            ws,
            header_row=8,
            data_start_row=9,
            fixed_widths=larguras,
            limite_coluna=column_index_from_string("AA"),
            max_scan_rows=_scan_rows,
            modo_rapido=max_row_data > LIMIAR_LINHAS_TURBO_CONSOLIDADO,
            ajustar_altura_linhas=False,
        )
        ws.freeze_panes = "A9"
        ult_col = get_column_letter(max(ws.max_column, column_index_from_string("AA")))
        ws.auto_filter.ref = f"A8:{ult_col}{ws.max_row}"
        # Planilha sem proteção: edição e exclusão de linhas liberadas.
    elif apenas_abas_apoio:
        _notify("Consolidado: ignorado (lote — abas técnicas unificadas)")

    NOME_ABA_REL_ANALITICO = "DADOS GERAL"
    if NOME_ABA_REL_ANALITICO in wb.sheetnames:
        _notify("Relatório analítico: painel e cabeçalho (sem borda célula a célula nas linhas de dados)")
        wa = wb[NOME_ABA_REL_ANALITICO]
        wa["A1"] = "EMPREENDIMENTO"
        wa["B1"] = str(nome_empreendimento or "").strip().upper()
        wa["A2"] = "DATA-BASE"
        wa["B2"] = data_base.strftime("%d/%m/%Y") if data_base else ""
        wa["A3"] = "QTD. REGISTROS (VISÍVEIS)"
        wa["B3"] = "=SUBTOTAL(103,A9:A1048576)"
        wa["A4"] = "QTD. CLIENTES (VISÍVEIS)"
        wa["B4"] = "=SUMPRODUCT(SUBTOTAL(103,OFFSET($B$9,ROW($B$9:$B$1048576)-ROW($B$9),0,1))/COUNTIFS($B$9:$B$1048576,$B$9:$B$1048576,$B$9:$B$1048576,\"<>\"))"
        wa["A5"] = "VL.PARCELAS (VISÍVEL)"
        wa["B5"] = "=SUBTOTAL(109,D9:D1048576)"
        wa["A6"] = "QTD. VENDAS (VISÍVEIS)"
        wa["B6"] = "=SUMPRODUCT(SUBTOTAL(103,OFFSET($C$9,ROW($C$9:$C$1048576)-ROW($C$9),0,1))/COUNTIFS($C$9:$C$1048576,$C$9:$C$1048576,$C$9:$C$1048576,\"<>\"))"
        for linha in range(1, 7):
            wa[f"A{linha}"].font = Font(name="Calibri", size=10, bold=True, color=branco)
            wa[f"A{linha}"].fill = PatternFill("solid", fgColor=azul_escuro)
            wa[f"A{linha}"].alignment = Alignment(horizontal="left", vertical="center")
            wa[f"A{linha}"].border = Border(
                left=borda_media_preta, right=borda_fina_branca,
                top=borda_media_preta if linha == 1 else borda_fina_branca,
                bottom=borda_fina_branca
            )
            wa[f"B{linha}"].font = Font(name="Calibri", size=10, bold=True, color=preto)
            wa[f"B{linha}"].alignment = Alignment(horizontal="center", vertical="center")
            wa[f"B{linha}"].fill = PatternFill("solid", fgColor=cinza)
            wa[f"B{linha}"].border = Border(
                left=borda_fina_branca, right=borda_media_preta,
                top=borda_media_preta if linha == 1 else borda_fina_branca,
                bottom=borda_fina_branca
            )
        wa["B5"].number_format = 'R$ #,##0.00'
        for col in ["A", "B"]:
            wa[f"{col}7"].fill = PatternFill("solid", fgColor="D9E1F2")
        hr_ra = 8
        for col_i in range(1, wa.max_column + 1):
            c = wa.cell(row=hr_ra, column=col_i)
            c.value = _padronizar_rotulo_coluna_exibicao(c.value)
        for cell in wa[hr_ra]:
            cell.font = Font(name="Calibri", size=10, bold=True, color=branco)
            cell.fill = PatternFill("solid", fgColor=azul_escuro)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(
                left=borda_fina_branca,
                right=borda_fina_branca,
                top=borda_fina_branca,
                bottom=borda_media_preta,
            )
        # Relatório com zebra cinza leve (legibilidade).
        # Em bases muito grandes, evitar borda/alinhamento célula a célula em toda a grade
        # para não bloquear a etapa final por vários minutos.
        data_start_ra = 9
        fill_rel_odd = PatternFill("solid", fgColor="FFFFFF")
        fill_rel_even = PatternFill("solid", fgColor="F2F2F2")
        linhas_ra = max(wa.max_row - data_start_ra + 1, 0)
        modo_turbo_rel_analitico = linhas_ra > LIMIAR_LINHAS_TURBO_RELATORIO_ANALITICO
        if modo_turbo_rel_analitico:
            _notify(
                "Relatório analítico: modo turbo ativado para base grande "
                f"({linhas_ra} linhas) — aplicando apenas formato essencial."
            )
            intervalo_nf = 10000
            for linha in range(data_start_ra, wa.max_row + 1):
                if (linha - data_start_ra) % intervalo_nf == 0:
                    pct_ra = int(((linha - data_start_ra + 1) / max(linhas_ra, 1)) * 100)
                    _notify_timed(
                        "relatorio_analitico_turbo",
                        f"Relatório analítico (turbo): linhas {linha}/{wa.max_row} ({pct_ra}%)",
                        intervalo_segundos=1.2,
                    )
                wa[f"D{linha}"].number_format = 'R$ #,##0.00'
        else:
            aplicar_borda_em_massa = wa.max_row <= 20000
            intervalo_heartbeat_ra = 2000
            _notify(f"Relatório analítico: zebra nas linhas ({data_start_ra}..{wa.max_row})")
            for linha in range(data_start_ra, wa.max_row + 1):
                if (linha - data_start_ra) % intervalo_heartbeat_ra == 0:
                    pct_ra = int(((linha - data_start_ra + 1) / max(linhas_ra, 1)) * 100)
                    _notify_timed(
                        "relatorio_analitico",
                        f"Relatório analítico: linhas {linha}/{wa.max_row} ({pct_ra}%)",
                        intervalo_segundos=1.2,
                    )
                wa[f"D{linha}"].number_format = 'R$ #,##0.00'
                row_fill = fill_rel_even if (linha % 2 == 0) else fill_rel_odd
                for col_i in range(1, wa.max_column + 1):
                    dcell = wa.cell(row=linha, column=col_i)
                    dcell.fill = row_fill
                    if aplicar_borda_em_massa:
                        dcell.border = aux_border_cinza
                        dcell.alignment = aux_alignment_centro
            if not aplicar_borda_em_massa:
                _notify("Relatório analítico: modo otimizado ativado (sem borda célula a célula em base grande)")
        wa.freeze_panes = "A9"
        wa.auto_filter.ref = f"A8:{get_column_letter(wa.max_column)}{wa.max_row}"
        for col_i in range(1, wa.max_column + 1):
            letter = get_column_letter(col_i)
            titulo = str(wa.cell(row=hr_ra, column=col_i).value or "")
            tu = titulo.upper()
            if "IDENTIFICADOR" in tu:
                wa.column_dimensions[letter].width = 25
            else:
                wa.column_dimensions[letter].width = min(max(len(titulo) + 4, 12), 48)

    # Padrao profissional das abas de apoio + resumo no topo.
    formatos_auxiliares = {
        "DADOS RECEBER": {
            "datas": ["VENC.DATA"],
            "moeda": ["PRINCIPAL", "CORREÇÃO", "JUROS ATRASO", "MULTA ATRASO", "VL.PARCELA", "CORREÇÃO ATRASO"],
            "inteiros": ["VENDA", "PARC.NUM", "PARC.TOTAL", "DIA.VENC.", "ANO.VENC."],
            "col_principal": "Principal",
            "col_pago": None,
            "col_avencer": "Principal",
            "col_inad": "Principal",
        },
        "DADOS RECEBIDOS": {
            "datas": ["DATA.REC."],
            "moeda": ["VL.PARCELA", "PRINCIPAL", "CORREÇÃO", "MULTA ATRASO", "JUROS ATRASO"],
            "inteiros": ["VENDA", "PARC.NUM", "PARC.TOTAL"],
            "col_principal": "Principal",
            "col_pago": "VL.PARCELA",
            "col_avencer": None,
            "col_inad": None,
        },
    }

    def _sum_coluna(ws_aux, col_letter, start_row):
        total = 0.0
        for linha in range(start_row, ws_aux.max_row + 1):
            val = ws_aux[f"{col_letter}{linha}"].value
            try:
                total += float(val or 0)
            except Exception:
                pass
        return total

    aux_alignment_centro = Alignment(horizontal="center", vertical="center")
    aux_border_cinza = Border(left=borda_fina_cinza, right=borda_fina_cinza, top=borda_fina_cinza, bottom=borda_fina_cinza)
    aux_header_font = Font(name="Calibri", size=10, bold=True, color=branco)
    aux_header_fill = PatternFill("solid", fgColor=azul_escuro)

    _notify("Abas DADOS RECEBER / DADOS RECEBIDOS: resumo, cabeçalho e formatos (sem borda em massa nos dados)")
    for nome_aba, cfg in formatos_auxiliares.items():
        if nome_aba not in wb.sheetnames:
            continue
        ws_aux = wb[nome_aba]
        ws_aux.insert_rows(1, amount=7)
        header_row = 8
        data_start = 9
        for col_i in range(1, ws_aux.max_column + 1):
            c = ws_aux.cell(row=header_row, column=col_i)
            c.value = _padronizar_rotulo_coluna_exibicao(c.value)
        mapa_colunas = {
            str(c.value).strip().upper(): c.column_letter
            for c in ws_aux[header_row]
            if c.value is not None
        }

        max_row_aux = ws_aux.max_row

        if nome_aba == "DADOS RECEBER":
            col_parcela = mapa_colunas.get("PARC.(GERAL)", "F")
            col_vlr = mapa_colunas.get("VL.PARCELA", "T")
            col_status = mapa_colunas.get("STATUS", "J")
            resumo = [
                ("QTD.PARCELAS", f"=SUBTOTAL(103,{col_parcela}{data_start}:{col_parcela}1048576)"),
                ("VL.PARCELAS", f"=SUBTOTAL(109,{col_vlr}{data_start}:{col_vlr}1048576)"),
                (
                    "QTD.PARC.VENCIDA",
                    f"=SUMPRODUCT(SUBTOTAL(103,OFFSET(${col_status}${data_start},ROW(${col_status}${data_start}:${col_status}1048576)-ROW(${col_status}${data_start}),0,1)),--(UPPER(TRIM(${col_status}${data_start}:${col_status}1048576))=\"VENCIDO\"))",
                ),
                (
                    "VL.VENCIDO",
                    f"=SUMPRODUCT(SUBTOTAL(109,OFFSET(${col_vlr}${data_start},ROW(${col_vlr}${data_start}:${col_vlr}1048576)-ROW(${col_vlr}${data_start}),0,1)),--(UPPER(TRIM(${col_status}${data_start}:${col_status}1048576))=\"VENCIDO\"),${col_vlr}${data_start}:${col_vlr}1048576)",
                ),
                (
                    "QTD.PARC.A VENCER",
                    f"=SUMPRODUCT(SUBTOTAL(103,OFFSET(${col_status}${data_start},ROW(${col_status}${data_start}:${col_status}1048576)-ROW(${col_status}${data_start}),0,1)),--(UPPER(TRIM(${col_status}${data_start}:${col_status}1048576))=\"A VENCER\"))",
                ),
                (
                    "VL.A VENCER",
                    f"=SUMPRODUCT(SUBTOTAL(109,OFFSET(${col_vlr}${data_start},ROW(${col_vlr}${data_start}:${col_vlr}1048576)-ROW(${col_vlr}${data_start}),0,1)),--(UPPER(TRIM(${col_status}${data_start}:${col_status}1048576))=\"A VENCER\"),${col_vlr}${data_start}:${col_vlr}1048576)",
                ),
            ]
        else:
            col_venda = mapa_colunas.get("VENDA", "B")
            col_cliente = mapa_colunas.get("CLIENTE", "C")
            col_parcela = mapa_colunas.get("PARC.(GERAL)", "F")
            col_pago = mapa_colunas.get("VL.PARCELA", "M")
            resumo = [
                ("QTD.VENDAS", f"=SUMPRODUCT(SUBTOTAL(103,OFFSET(${col_venda}${data_start},ROW(${col_venda}${data_start}:${col_venda}1048576)-ROW(${col_venda}${data_start}),0,1))/COUNTIFS(${col_venda}${data_start}:${col_venda}1048576,${col_venda}${data_start}:${col_venda}1048576,${col_venda}${data_start}:${col_venda}1048576,\"<>\"))"),
                ("QTD.CLIENTES", f"=SUMPRODUCT(SUBTOTAL(103,OFFSET(${col_cliente}${data_start},ROW(${col_cliente}${data_start}:${col_cliente}1048576)-ROW(${col_cliente}${data_start}),0,1))/COUNTIFS(${col_cliente}${data_start}:${col_cliente}1048576,${col_cliente}${data_start}:${col_cliente}1048576,${col_cliente}${data_start}:${col_cliente}1048576,\"<>\"))"),
                ("QTD.PARCELAS", f"=SUBTOTAL(103,{col_parcela}{data_start}:{col_parcela}1048576)"),
                ("VL.PAGO", f"=SUBTOTAL(109,{col_pago}{data_start}:{col_pago}1048576)"),
            ]

        ws_aux["A1"] = f"RESUMO - {nome_aba}"
        ws_aux["A1"].font = Font(bold=True, color=branco)
        ws_aux["A1"].fill = PatternFill("solid", fgColor=azul_escuro)
        ws_aux.merge_cells("A1:D1")
        for idx, (lbl, val) in enumerate(resumo, start=2):
            ws_aux[f"A{idx}"] = lbl
            ws_aux[f"B{idx}"] = val
            ws_aux[f"A{idx}"].font = Font(bold=True)
        for col in ["C", "D"]:
            for r in range(2, 7):
                ws_aux[f"{col}{r}"] = None

        # Linha de separacao visual entre resumo e tabela.
        for col in ["A", "B", "C", "D"]:
            ws_aux[f"{col}7"].fill = PatternFill("solid", fgColor="D9E1F2")

        if nome_aba == "DADOS RECEBER":
            for col in ["B2", "B3", "B4", "B5", "B6", "B7"]:
                r = col[1:]
                lbl = str(ws_aux[f"A{r}"].value or "")
                cel = ws_aux[col]
                v = cel.value
                if v is None:
                    continue
                if "QTD." in lbl.upper():
                    cel.number_format = "0"
                else:
                    cel.number_format = "R$ #,##0.00"
        else:
            for col in ["B2", "B3", "B4", "B5"]:
                r = col[1:]
                lbl = str(ws_aux[f"A{r}"].value or "")
                cel = ws_aux[col]
                v = cel.value
                if v is None:
                    continue
                if "QTD." in lbl.upper():
                    cel.number_format = "0"
                else:
                    cel.number_format = "R$ #,##0.00"

        # Cabecalho da tabela de dados.
        for cell in ws_aux[header_row]:
            cell.font = aux_header_font
            cell.fill = aux_header_fill
            cell.alignment = aux_alignment_centro
            cell.border = aux_border_cinza

        col_fmt_pairs = []
        for nome_col in cfg.get("datas", []):
            col = mapa_colunas.get(nome_col.upper())
            if col:
                col_fmt_pairs.append((col, "dd/mm/yyyy"))
        for nome_col in cfg.get("moeda", []):
            col = mapa_colunas.get(nome_col.upper())
            if col:
                col_fmt_pairs.append((col, 'R$ #,##0.00'))
        for nome_col in cfg.get("inteiros", []):
            col = mapa_colunas.get(nome_col.upper())
            if col:
                col_fmt_pairs.append((col, "0"))
        linhas_aux = max(max_row_aux - data_start + 1, 0)
        for col_letter, fmt in col_fmt_pairs:
            for idx_linha, linha in enumerate(range(data_start, max_row_aux + 1), start=1):
                if idx_linha % 12000 == 0:
                    pct_aux = int((idx_linha / max(linhas_aux, 1)) * 100)
                    _notify_timed(
                        f"{nome_aba}_fmt",
                        f"{nome_aba}: aplicando formato em {col_letter} ({linha}/{max_row_aux}, {pct_aux}%)",
                        intervalo_segundos=1.2,
                    )
                ws_aux[f"{col_letter}{linha}"].number_format = fmt

        max_col_aux = ws_aux.max_column
        # Ajuste simples de largura.
        for col_i in range(1, max_col_aux + 1):
            letter = ws_aux.cell(row=header_row, column=col_i).column_letter
            titulo = str(ws_aux.cell(row=header_row, column=col_i).value or "")
            tu = titulo.upper()
            if "IDENTIFICADOR" in tu:
                ws_aux.column_dimensions[letter].width = 25
            else:
                ws_aux.column_dimensions[letter].width = min(max(len(titulo) + 4, 12), 40)

        ws_aux.freeze_panes = f"A{data_start}"
        if max_row_aux >= header_row:
            ws_aux.auto_filter.ref = f"A{header_row}:{get_column_letter(max_col_aux)}{max_row_aux}"

    # Aba de pendências estruturais de parcelas (informativa).
    nome_aba_pen = (
        "PEND.PARCELAS" if "PEND.PARCELAS" in wb.sheetnames
        else ("PENDENCIAS_PARCELAS" if "PENDENCIAS_PARCELAS" in wb.sheetnames else ("Pendencias_Parcelas" if "Pendencias_Parcelas" in wb.sheetnames else ""))
    )
    if nome_aba_pen:
        _notify("Pendências: painel, cabeçalho e grade das linhas de divergência")
        ws_pen = wb[nome_aba_pen]
        ws_pen.insert_rows(1, amount=4)
        header_row = PENDENCIAS_PARCELAS_HEADER_ROW
        data_start = PENDENCIAS_PARCELAS_DATA_START_ROW
        max_row_pen = ws_pen.max_row
        max_col_pen = ws_pen.max_column
        mapa_pen = {str(c.value).strip().upper(): c.column_letter for c in ws_pen[header_row] if c.value is not None}

        col_vl_pag = mapa_pen.get("VL.PARC.PAGO", "")
        col_vl_inad = mapa_pen.get("VL.VENCIDO", "")
        col_vl_av = mapa_pen.get("VL.A VENCER", "")
        col_vl_saldo = mapa_pen.get("VL.SALDO", "")
        col_qt_pag = mapa_pen.get("QTD.PARC.PAGO", "")
        col_qt_inad = mapa_pen.get("QTD.PARC.VENCIDA", "")
        col_qt_av = mapa_pen.get("QTD.PARC.A VENCER", "")
        col_qt_saldo = mapa_pen.get("QTD.SALDO", "")

        ws_pen["A1"] = "VL.PARC.PAGAS"
        ws_pen["A2"] = "VL.PARC.INADIMPLENTES"
        ws_pen["A3"] = "VL.PARC.A VENCER"
        ws_pen["A4"] = "SALDO"
        ws_pen["E1"] = "QTD.PARC.PAGAS"
        ws_pen["E2"] = "QTD.PARC.INADIMPLENTES"
        ws_pen["E3"] = "QTD.PARC.A VENCER"
        ws_pen["E4"] = "SALDO"

        if col_vl_pag:
            ws_pen["B1"] = f"=SUBTOTAL(109,{col_vl_pag}{data_start}:{col_vl_pag}1048576)"
        if col_vl_inad:
            ws_pen["B2"] = f"=SUBTOTAL(109,{col_vl_inad}{data_start}:{col_vl_inad}1048576)"
        if col_vl_av:
            ws_pen["B3"] = f"=SUBTOTAL(109,{col_vl_av}{data_start}:{col_vl_av}1048576)"
        if col_vl_saldo:
            ws_pen["B4"] = f"=SUBTOTAL(109,{col_vl_saldo}{data_start}:{col_vl_saldo}1048576)"
        if col_qt_pag:
            ws_pen["F1"] = f"=SUBTOTAL(109,{col_qt_pag}{data_start}:{col_qt_pag}1048576)"
        if col_qt_inad:
            ws_pen["F2"] = f"=SUBTOTAL(109,{col_qt_inad}{data_start}:{col_qt_inad}1048576)"
        if col_qt_av:
            ws_pen["F3"] = f"=SUBTOTAL(109,{col_qt_av}{data_start}:{col_qt_av}1048576)"
        if col_qt_saldo:
            ws_pen["F4"] = f"=SUBTOTAL(109,{col_qt_saldo}{data_start}:{col_qt_saldo}1048576)"

        for linha in range(1, 5):
            for col_lbl, col_val in (("A", "B"), ("E", "F")):
                ws_pen[f"{col_lbl}{linha}"].font = Font(bold=True, color=branco)
                ws_pen[f"{col_lbl}{linha}"].fill = PatternFill("solid", fgColor=azul_escuro)
                ws_pen[f"{col_lbl}{linha}"].alignment = Alignment(horizontal="left", vertical="center")
                ws_pen[f"{col_val}{linha}"].font = Font(bold=True, color=preto)
                ws_pen[f"{col_val}{linha}"].alignment = Alignment(horizontal="center", vertical="center")
                ws_pen[f"{col_val}{linha}"].fill = PatternFill("solid", fgColor="EAF2FF")

        for c in ["B1", "B2", "B3", "B4"]:
            ws_pen[c].number_format = 'R$ #,##0.00'
        for c in ["F1", "F2", "F3", "F4"]:
            ws_pen[c].number_format = "0"

        for cell in ws_pen[header_row]:
            cell.font = aux_header_font
            cell.fill = aux_header_fill
            cell.alignment = aux_alignment_centro
            cell.border = aux_border_cinza

        fmt_moeda = 'R$ #,##0.00'
        fmt_inteiro = "0"
        for titulo_col, letra in mapa_pen.items():
            tc = str(titulo_col or "").strip().upper()
            if not tc:
                continue
            if tc.startswith("VL.") or tc == "VL.PARCELA DIVERGENTE":
                nf = fmt_moeda
            elif tc.startswith("QTD."):
                nf = fmt_inteiro
            else:
                continue
            for linha in range(data_start, max_row_pen + 1):
                ws_pen[f"{letra}{linha}"].number_format = nf

        linhas_pen = max(max_row_pen - data_start + 1, 0)
        modo_turbo_pendencias = linhas_pen >= LIMIAR_LINHAS_TURBO_PENDENCIAS
        if modo_turbo_pendencias:
            _notify(
                "Pendências: base grande detectada — mantendo cabeçalho e formatos essenciais sem grade total."
            )
        else:
            for idx_linha, linha in enumerate(range(data_start, max_row_pen + 1), start=1):
                if idx_linha % 4000 == 0:
                    pct_pen = int((idx_linha / max(linhas_pen, 1)) * 100)
                    _notify_timed(
                        "pendencias_grade",
                        f"Pendências: alinhamento/grade ({linha}/{max_row_pen}, {pct_pen}%)",
                        intervalo_segundos=1.2,
                    )
                for col_i in range(1, max_col_pen + 1):
                    c = ws_pen.cell(row=linha, column=col_i)
                    c.alignment = aux_alignment_centro
                    c.border = aux_border_cinza

        larg_pen = {
            "venda": 12,
            "cliente": 30,
            "identificador": 25,
            "parcela": 16,
            "tipo de divergência": 24,
            "origem da divergência": 28,
            "motivo reconciliação": 28,
            "observação": 48,
            "observacao": 48,
        }
        for cell in ws_pen[header_row]:
            tit = str(cell.value or "").strip()
            if not tit:
                continue
            w = larg_pen.get(tit.casefold())
            if w:
                ws_pen.column_dimensions[cell.column_letter].width = w
            elif "VL." in tit.upper():
                ws_pen.column_dimensions[cell.column_letter].width = 15
            elif "QTD." in tit.upper():
                ws_pen.column_dimensions[cell.column_letter].width = 14

        ws_pen.freeze_panes = f"A{data_start}"
        ws_pen.auto_filter.ref = f"A{header_row}:{get_column_letter(max_col_pen)}{max_row_pen}"

    # Regra global solicitada: linha 7 e 8 com altura fixa em todas as abas.
    for ws_any in wb.worksheets:
        ws_any.row_dimensions[7].height = 14.4
        ws_any.row_dimensions[8].height = 14.4

    if not apenas_abas_apoio:
        if NOME_ABA_RESUMO_GERAL in wb.sheetnames:
            _notify("Resumo Geral: painel, blocos e cabeçalho")
            _aplicar_estilo_aba_resumo_geral(wb, data_base, nome_empreendimento)

        if "CONSOLIDADO ESTOQUE" in wb.sheetnames:
            _notify("Consolidado Estoque: painel, cabeçalho e destaques")
            _aplicar_estilo_aba_consolidado_estoque(wb, data_base, nome_empreendimento, indicadores_estoque)

    # Aba explicativa executiva (CRITERIOS ANALISES): layout didático com hierarquia visual.
    nome_aba_crit = ""
    for cand in ("CRITERIOS ANALISES", "CRITERIOS", "Critérios"):
        if cand in wb.sheetnames:
            nome_aba_crit = cand
            break
    if nome_aba_crit:
        ws_cr = wb[nome_aba_crit]
        if ws_cr.title != "CRITERIOS ANALISES":
            ws_cr.title = "CRITERIOS ANALISES"
        ws_cr.insert_rows(1, amount=4)
        ws_cr.merge_cells("A1:C1")
        ws_cr["A1"] = "CRITERIOS ANALISES"
        ws_cr["A2"] = "EMPREENDIMENTO"
        ws_cr["B2"] = str(nome_empreendimento or "").strip().upper()
        ws_cr["A3"] = "DATA-BASE"
        ws_cr["B3"] = data_base.strftime("%d/%m/%Y") if data_base else ""
        ws_cr["A4"] = "GUIA DE LEITURA EXECUTIVA POR ABA"
        ws_cr.merge_cells("A4:C4")
        for c in ("A1", "A2", "A3", "A4"):
            ws_cr[c].font = Font(bold=True, color=branco)
            ws_cr[c].fill = PatternFill("solid", fgColor=azul_escuro)
            ws_cr[c].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        for c in ("B2", "B3"):
            ws_cr[c].font = Font(bold=True, color=preto)
            ws_cr[c].fill = PatternFill("solid", fgColor="EAF2FF")
            ws_cr[c].alignment = Alignment(horizontal="left", vertical="center")

        hr = 5
        for col_i in range(1, min(3, ws_cr.max_column) + 1):
            cc = ws_cr.cell(row=hr, column=col_i)
            cc.value = _padronizar_rotulo_coluna_exibicao(cc.value)
            cc.font = Font(bold=True, color=branco)
            cc.fill = PatternFill("solid", fgColor=azul_escuro)
            cc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cc.border = aux_border_cinza

        mapa_sec = {}
        for r in range(hr + 1, ws_cr.max_row + 1):
            sec = str(ws_cr.cell(r, 1).value or "").strip().upper()
            if sec and sec not in mapa_sec:
                mapa_sec[sec] = PatternFill("solid", fgColor="D9E1F2")
            if sec:
                ws_cr.cell(r, 1).fill = mapa_sec[sec]
                ws_cr.cell(r, 1).font = Font(bold=True, color=preto)
            ws_cr.cell(r, 2).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws_cr.cell(r, 3).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            for cidx in (1, 2, 3):
                ws_cr.cell(r, cidx).border = aux_border_cinza

        ws_cr.column_dimensions["A"].width = 28
        ws_cr.column_dimensions["B"].width = 34
        ws_cr.column_dimensions["C"].width = 120
        ws_cr.freeze_panes = "A6"
        ws_cr.auto_filter.ref = f"A5:C{ws_cr.max_row}"

    # Excel / OnlyOffice: recálculo completo ao abrir; modo automático quando suportado.
    try:
        from openpyxl.workbook.properties import CalcProperties

        if wb.calculation is None:
            wb.calculation = CalcProperties()
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.calcCompleted = False
        wb.calculation.forceFullCalc = True
        if hasattr(wb.calculation, "calcMode"):
            wb.calculation.calcMode = "auto"
    except Exception:
        pass

    _notify("Gravando workbook no disco (etapa final, pode levar alguns minutos em base grande)")
    if NOME_ABA_RESUMO_GERAL in wb.sheetnames and wb.sheetnames[0] != NOME_ABA_RESUMO_GERAL:
        try:
            idx_rg = wb.sheetnames.index(NOME_ABA_RESUMO_GERAL)
            wb.move_sheet(wb[NOME_ABA_RESUMO_GERAL], offset=-idx_rg)
        except Exception:
            pass
    if "CONSOLIDADO ESTOQUE" in wb.sheetnames and "CRITERIOS ANALISES" in wb.sheetnames:
        try:
            idx_es = wb.sheetnames.index("CONSOLIDADO ESTOQUE")
            idx_cr = wb.sheetnames.index("CRITERIOS ANALISES")
            if idx_es != max(0, idx_cr - 1):
                wb.move_sheet(wb["CONSOLIDADO ESTOQUE"], offset=(idx_cr - 1) - idx_es)
        except Exception:
            pass
    if "CRITERIOS ANALISES" in wb.sheetnames and wb.sheetnames[-1] != "CRITERIOS ANALISES":
        try:
            wb.move_sheet(wb["CRITERIOS ANALISES"], offset=len(wb.sheetnames))
        except Exception:
            pass
    wb.save(caminho_saida)


# =========================
# PROCESSO FINAL
# =========================
def processar_e_gerar_excel(
    caminho_receber,
    caminho_recebidos,
    caminho_saida,
    *,
    lote_unificado_empreendimentos_distintos=0,
    preservar_empreendimento_por_linha=False,
    nome_arquivo_xlsx_override=None,
    gerar_aba_resumo_geral=True,
    caminho_estoque=None,
    gerar_aba_consolidado_estoque=True,
    progresso_cb=None,
):
    """
    Parâmetros opcionais (somente orquestração em lote):
    - lote_unificado_empreendimentos_distintos: >0 ajusta rótulo B1 para consolidação geral de lote.
    - preservar_empreendimento_por_linha: não sobrescreve a coluna Empreendimento com um único canônico.
    - nome_arquivo_xlsx_override: nome fixo do .xlsx final (ex.: LOTE_UNIFICADO.xlsx).
    - gerar_aba_resumo_geral: False em workbooks temporários do lote por empreendimento (resumo único no arquivo final).
    - caminho_estoque: TXT opcional de relatório de estoque UAU (camada complementar; não altera o motor financeiro).
    - gerar_aba_consolidado_estoque: False em workbooks temporários do lote por empreendimento (aba única no workbook final).
    """
    inicio_execucao = time.time()
    reset_etl_stats_acumulado()
    t_perf0 = time.perf_counter()
    perf_etapas = []
    perf_alertas = []
    perf_extra_ligado = str(os.environ.get("PERF_EXTRA", "0")).strip() == "1"
    etapas_perf_extra = {
        "harmonizacao_aplicar_padroes",
        "montar_consolidado_nucleo",
        "montar_consolidado_total",
        "_validar_pre_exportacao",
        "excel_pos_formatacao_openpyxl",
    }
    print(f"[CONFIG] CARTEIRA_MODO_OFICIAL={CARTEIRA_MODO_OFICIAL}", flush=True)

    def _nlin_df(df):
        return 0 if df is None or getattr(df, "empty", True) else len(df)

    def _emit_perf_extra_noop(_nome, _dt, _lr, _lp, _lc):
        return None

    def _emit_perf_extra_real(nome, dt, lr, lp, lc):
        if nome not in etapas_perf_extra:
            return
        total_linhas = max(int(lr or 0) + int(lp or 0), 1)
        taxa = float(total_linhas) / max(float(dt), 1e-9)
        print(
            f"[TEMPO][EXTRA] etapa={nome} | linhas_total={total_linhas} | "
            f"taxa_linhas_s={taxa:.1f} | consolidado={int(lc or 0)}"
        )

    _emit_perf_extra = _emit_perf_extra_real if perf_extra_ligado else _emit_perf_extra_noop

    def _emit_perf(nome, dt, lr, lp, lc):
        if perf_extra_ligado:
            perf_etapas.append((nome, float(dt)))
        acum = time.perf_counter() - t_perf0
        print(
            f"[TEMPO] {nome}: {dt:.2f}s | acum={acum:.2f}s | "
            f"receber={lr} | recebidos={lp} | consolidado={lc}"
        )
        # Diagnóstico de gargalo: só com PERF_EXTRA=1 (evita prints extras no fluxo padrão).
        if float(dt) >= 90.0 and perf_extra_ligado:
            perf_alertas.append((str(nome), float(dt)))
            print(
                f"[ALERTA][GARGALO] etapa={nome} | tempo={dt:.2f}s | "
                "acao_sugerida=investigar_esta_etapa_primeiro"
            )
        _emit_perf_extra(nome, dt, lr, lp, lc)

    def _imprimir_ranking_perf():
        if not perf_etapas:
            return
        ordem = sorted(perf_etapas, key=lambda x: x[1], reverse=True)
        print("[TEMPO] Ranking (mais lenta -> mais rapida):")
        for i, (nome, seg) in enumerate(ordem, 1):
            print(f"  {i}. {nome}: {seg:.2f}s")
        if perf_alertas:
            print("[TEMPO] Etapas críticas (>= 90s):")
            for nome, seg in sorted(perf_alertas, key=lambda x: x[1], reverse=True):
                print(f"  - {nome}: {seg:.2f}s")

    def _dbg(msg):
        if DEBUG_VALIDACAO:
            print(f"[DEBUG][processar_e_gerar_excel] {msg}")

    def _emitir_progresso_motor(**payload):
        if callable(progresso_cb):
            try:
                progresso_cb(payload)
            except Exception:
                pass

    def _resumo_df(df, nome):
        if df is None or df.empty:
            _dbg(f"{nome}: vazio")
            return
        vendas = df["Venda"].fillna("").astype(str).str.strip().nunique() if "Venda" in df.columns else 0
        _dbg(f"{nome}: linhas={len(df)} | vendas_unicas={vendas} | colunas={list(df.columns)}")
        if "Status_Vencimento" in df.columns:
            st = df["Status_Vencimento"].fillna("").astype(str).str.strip().str.upper()
            _dbg(f"{nome}: status VENCIDO={int((st=='VENCIDO').sum())} | A VENCER={int((st=='A VENCER').sum())} | outros={int(len(df)-int((st=='VENCIDO').sum())-int((st=='A VENCER').sum()))}")
            if "Vlr_Parcela" in df.columns:
                vp = pd.to_numeric(df["Vlr_Parcela"], errors="coerce").fillna(0)
                _dbg(f"{nome}: soma Vlr_Parcela VENCIDO={float(vp.loc[st=='VENCIDO'].sum()):.2f} | A VENCER={float(vp.loc[st=='A VENCER'].sum()):.2f}")
        if "Vlr_Parcela" in df.columns:
            _dbg(f"{nome}: Vlr_Parcela não nulo={int(pd.to_numeric(df['Vlr_Parcela'], errors='coerce').notna().sum())}")
        cols_ex = [c for c in ["Venda", "Cliente", "Parcela", "Vlr_Parcela", "Principal", "Status_Vencimento"] if c in df.columns]
        if cols_ex and DEBUG_DADOS:
            _dbg(f"{nome}: exemplos(5)={df[cols_ex].head(5).to_dict('records')}")

    validacao_entrada_info = {}
    _t = time.perf_counter()
    try:
        validacao_entrada_info = validar_arquivos_entrada_uau(caminho_receber, caminho_recebidos)
    except ProcessamentoUAUErro:
        raise
    except Exception as e:
        raise ProcessamentoUAUErro(
            etapa="validação de entrada",
            funcao="validar_arquivos_entrada_uau",
            validacao="falha inesperada na validação inicial",
            mensagem=MSG_TXT_NAO_CONFIAVEL,
            campo_ou_aba="TXT",
            erro_tecnico=e,
        ) from e

    try:
        texto_receber_validado = validacao_entrada_info.get("texto_receber")
        texto_recebidos_validado = validacao_entrada_info.get("texto_recebidos")
        data_base_receber = extrair_data_base(caminho_receber, texto_pre_lido=texto_receber_validado)
        data_base_recebidos = extrair_data_base(caminho_recebidos, texto_pre_lido=texto_recebidos_validado)
        data_base = data_base_receber or data_base_recebidos
        nome_emp_receber = extrair_nome_empreendimento_txt(caminho_receber, texto_pre_lido=texto_receber_validado)
        nome_emp_recebidos = extrair_nome_empreendimento_txt(caminho_recebidos, texto_pre_lido=texto_recebidos_validado)
        nome_empreendimento_arquivo = escolher_moda_texto(
            [x for x in [nome_emp_receber, nome_emp_recebidos] if str(x).strip() != ""]
        ) if any(str(x).strip() != "" for x in [nome_emp_receber, nome_emp_recebidos]) else ""
        nome_empreendimento_arquivo = limpar_nome_empreendimento(nome_empreendimento_arquivo)
        if not nome_empreendimento_arquivo:
            nome_emp_receber_arquivo = extrair_nome_empreendimento_nome_arquivo(caminho_receber)
            nome_emp_recebidos_arquivo = extrair_nome_empreendimento_nome_arquivo(caminho_recebidos)
            nome_empreendimento_arquivo = escolher_moda_texto(
                [x for x in [nome_emp_receber_arquivo, nome_emp_recebidos_arquivo] if str(x).strip() != ""]
            ) if any(str(x).strip() != "" for x in [nome_emp_receber_arquivo, nome_emp_recebidos_arquivo]) else ""
            nome_empreendimento_arquivo = limpar_nome_empreendimento(nome_empreendimento_arquivo)
    except Exception as e:
        raise ProcessamentoUAUErro(
            etapa="leitura do arquivo",
            funcao="processar_e_gerar_excel",
            validacao="extração de metadados/data-base",
            mensagem="Falha ao ler metadados iniciais dos arquivos TXT.",
            campo_ou_aba="cabeçalho dos arquivos",
            erro_tecnico=e,
        ) from e
    _emit_perf("validacao_entrada", time.perf_counter() - _t, 0, 0, 0)

    _t = time.perf_counter()
    try:
        df_receber = carregar_receber_bruto(caminho_receber)
    except Exception as e:
        raise ProcessamentoUAUErro(
            etapa="leitura do arquivo Contas a Receber",
            funcao="carregar_receber_bruto",
            validacao="parse da aba Dados Receber",
            mensagem="Falha ao ler o arquivo de Contas a Receber.",
            campo_ou_aba="Dados Receber",
            erro_tecnico=e,
        ) from e
    _emit_perf("carregar_receber_bruto", time.perf_counter() - _t, _nlin_df(df_receber), 0, 0)

    _t = time.perf_counter()
    try:
        df_recebidos = carregar_recebidos_bruto(caminho_recebidos)
    except Exception as e:
        raise ProcessamentoUAUErro(
            etapa="leitura do arquivo Contas Recebidas",
            funcao="carregar_recebidos_bruto",
            validacao="parse da aba Dados Recebidos",
            mensagem="Falha ao ler o arquivo de Contas Recebidas.",
            campo_ou_aba="Dados Recebidos",
            erro_tecnico=e,
        ) from e
    _emit_perf(
        "carregar_recebidos_bruto",
        time.perf_counter() - _t,
        _nlin_df(df_receber),
        _nlin_df(df_recebidos),
        0,
    )

    _resumo_df(df_receber, "LEITURA df_receber bruto")
    _resumo_df(df_recebidos, "LEITURA df_recebidos bruto")

    # Reduz fragmentacao de cliente/base para mesma venda antes da consolidacao.
    _r0, _p0 = len(df_receber), len(df_recebidos)
    _rv0 = df_receber["Venda"].fillna("").astype(str).str.strip().nunique() if not df_receber.empty and "Venda" in df_receber.columns else 0
    _pv0 = df_recebidos["Venda"].fillna("").astype(str).str.strip().nunique() if not df_recebidos.empty and "Venda" in df_recebidos.columns else 0
    _t = time.perf_counter()
    df_receber = harmonizar_cliente_por_venda(df_receber)
    df_recebidos = harmonizar_cliente_por_venda(df_recebidos)
    _dbg(f"PREP harmonizar_cliente_por_venda: receber linhas {_r0}->{len(df_receber)} vendas {_rv0}->{(df_receber['Venda'].fillna('').astype(str).str.strip().nunique() if not df_receber.empty and 'Venda' in df_receber.columns else 0)}")
    _dbg(f"PREP harmonizar_cliente_por_venda: recebidos linhas {_p0}->{len(df_recebidos)} vendas {_pv0}->{(df_recebidos['Venda'].fillna('').astype(str).str.strip().nunique() if not df_recebidos.empty and 'Venda' in df_recebidos.columns else 0)}")

    try:
        _r1, _p1 = len(df_receber), len(df_recebidos)
        df_receber, df_recebidos = aplicar_padroes(df_receber, df_recebidos)
        _dbg(f"PREP aplicar_padroes: receber linhas {_r1}->{len(df_receber)} | recebidos {_p1}->{len(df_recebidos)}")
    except Exception as e:
        raise ProcessamentoUAUErro(
            etapa="normalização",
            funcao="aplicar_padroes",
            validacao="padronização de colunas e identificadores",
            mensagem="Falha durante a normalização dos dados.",
            campo_ou_aba="Dados Receber / Dados Recebidos",
            erro_tecnico=e,
        ) from e
    _emit_perf(
        "harmonizacao_aplicar_padroes",
        time.perf_counter() - _t,
        _nlin_df(df_receber),
        _nlin_df(df_recebidos),
        0,
    )

    # Higieniza texto e ordena abas auxiliares por data.
    def _limpar_texto_df(df):
        if df is None or df.empty:
            return df
        out = df.copy()
        cols_txt = out.select_dtypes(include=["object"]).columns
        for col in cols_txt:
            out[col] = out[col].fillna("").astype(str).str.replace(r"\s+", " ", regex=True).str.strip()
        return out

    _t = time.perf_counter()
    df_receber = _limpar_texto_df(df_receber)
    df_recebidos = _limpar_texto_df(df_recebidos)

    # Unifica cliente por Venda+Parcela para nao perder soma na consolidacao (sem identificador / nomes variados).
    _r2, _p2 = len(df_receber), len(df_recebidos)
    df_receber = unificar_cliente_por_venda_parcela(df_receber)
    df_recebidos = unificar_cliente_por_venda_parcela(df_recebidos)
    _dbg(f"PREP unificar_cliente_por_venda_parcela: receber linhas {_r2}->{len(df_receber)} | recebidos {_p2}->{len(df_recebidos)}")

    try:
        if not df_receber.empty and "Vencimento" in df_receber.columns:
            df_receber = df_receber.sort_values(by=["Vencimento", "Venda"], ascending=[True, True]).reset_index(drop=True)
            venc = pd.to_datetime(df_receber["Vencimento"], errors="coerce")
            db_ts = pd.Timestamp(data_base) if data_base is not None else None
            df_receber["Status_Vencimento"] = ""
            if db_ts is not None:
                m = venc.notna()
                df_receber.loc[m & (venc < db_ts), "Status_Vencimento"] = "VENCIDO"
                df_receber.loc[m & (venc >= db_ts), "Status_Vencimento"] = "A VENCER"
        elif not df_receber.empty:
            df_receber["Status_Vencimento"] = ""
    except Exception as e:
        raise ProcessamentoUAUErro(
            etapa="normalização",
            funcao="processar_e_gerar_excel",
            validacao="classificação Status_Vencimento",
            mensagem="Falha ao classificar VENCIDO/A VENCER em Dados Receber.",
            campo_ou_aba="Status_Vencimento",
            erro_tecnico=e,
        ) from e
    _resumo_df(df_receber, "LEITURA/PREP df_receber classificado")
    _resumo_df(df_recebidos, "LEITURA/PREP df_recebidos classificado")

    # Colunas críticas esperadas para consolidação
    for col in ["Status_Vencimento", "Vlr_Parcela", "Venda", "Parcela", "Principal"]:
        if col not in df_receber.columns:
            raise ProcessamentoUAUErro(
                etapa="consolidação",
                funcao="processar_e_gerar_excel",
                validacao="presença de colunas críticas",
                mensagem=f"Erro na etapa de consolidação: coluna {col} ausente em df_receber deduplicado/preparado.",
                campo_ou_aba=col,
            )
    if df_receber["Status_Vencimento"].fillna("").astype(str).str.strip().eq("").all():
        raise ProcessamentoUAUErro(
            etapa="consolidação",
            funcao="processar_e_gerar_excel",
            validacao="classificação de status",
            mensagem="Erro na etapa de consolidação: coluna Status_Vencimento vazia em df_receber deduplicado.",
            campo_ou_aba="Status_Vencimento",
        )
    if not df_recebidos.empty and "Data_Rec" in df_recebidos.columns:
        df_recebidos = df_recebidos.sort_values(by=["Data_Rec", "Venda"], ascending=[True, True]).reset_index(drop=True)

    _emit_perf(
        "preparacao_pos_padroes",
        time.perf_counter() - _t,
        _nlin_df(df_receber),
        _nlin_df(df_recebidos),
        0,
    )

    _t = time.perf_counter()
    df_recebidos_sem_aporte, df_aportes = separar_aportes_financeiros(df_recebidos)
    vendas_aporte = set()
    if df_aportes is not None and not df_aportes.empty and "Venda" in df_aportes.columns:
        vendas_aporte = set(df_aportes["Venda"].fillna("").astype(str).str.strip().tolist())
    _emit_perf(
        "separar_aportes_financeiros",
        time.perf_counter() - _t,
        _nlin_df(df_receber),
        _nlin_df(df_recebidos),
        0,
    )

    df_receber_entrada_motor = df_receber.copy()
    df_recebidos_entrada_motor = df_recebidos.copy()

    _mapa_eo_nome_legal_txt = None
    if int(lote_unificado_empreendimentos_distintos or 0) > 0 and texto_receber_validado:
        _m = mapa_emp_obra_nome_legal_de_texto_receber_multibloco(texto_receber_validado)
        if _m:
            _mapa_eo_nome_legal_txt = _m

    reg_montar = []
    _t = time.perf_counter()
    try:
        (
            df_consolidado,
            data_base,
            alertas_consolidado,
            df_receber_tratado,
            df_recebidos_tratado,
            pendencias_qtd_total_reconc_montar,
        ) = montar_consolidado(
            df_receber,
            df_recebidos,
            data_base,
            nome_empreendimento_arquivo=nome_empreendimento_arquivo,
            vendas_aporte=vendas_aporte,
            registro_etapas_tempo=reg_montar,
            mapa_emp_obra_nome_legal=_mapa_eo_nome_legal_txt,
        )
    except Exception as e:
        raise ProcessamentoUAUErro(
            etapa="consolidação final",
            funcao="montar_consolidado",
            validacao="agregação por venda / cálculo de indicadores",
            mensagem="Falha durante a montagem do Consolidado Venda.",
            campo_ou_aba="Consolidado Venda",
            erro_tecnico=e,
        ) from e
    _dt_montar_total = time.perf_counter() - _t
    for _nome_m, _dt_m, _lr_m, _lp_m, _lc_m in reg_montar:
        _emit_perf(_nome_m, _dt_m, _lr_m, _lp_m, _lc_m)
    _emit_perf(
        "montar_consolidado_total",
        _dt_montar_total,
        _nlin_df(df_receber_tratado),
        _nlin_df(df_recebidos_tratado),
        _nlin_df(df_consolidado),
    )

    def _validar_pre_exportacao(df_r, df_p, df_c, df_r_motor_entrada=None, df_p_motor_entrada=None):
        erros = []
        qtd_div_parcelas = 0
        contexto_primeira_falha = None
        pendencias_parcelas = []
        _t_pre = time.perf_counter()
        _tempo_blocos = {
            "preparacao_inicial": 0.0,
            "mapas_agregados": 0.0,
            "validacao_loop_vendas": 0.0,
            "auditoria_subgrupos": 0.0,
            "checagens_finais": 0.0,
        }

        def _pendencia_pre_export(tipo, regra_curta, observacao, venda_k="", cliente_k="", ident_k="", parcela_k="", valor_vp=0.0):
            nonlocal qtd_div_parcelas, contexto_primeira_falha
            vk = str(venda_k or "").strip()
            if contexto_primeira_falha is None and vk:
                contexto_primeira_falha = {
                    "Venda": vk,
                    "Cliente": str(cliente_k or "").strip(),
                    "Identificador": str(ident_k or "").strip(),
                    "Regra_Violada": str(regra_curta),
                    "Causa_Raiz": str(observacao),
                }
            pendencias_parcelas.append({
                "Venda": vk,
                "Cliente": str(cliente_k or "").strip(),
                "Identificador": str(ident_k or "").strip(),
                "Parcela": str(parcela_k or "").strip(),
                "Tipo de Divergência": str(tipo),
                "Valor Parcela Divergente": float(valor_vp),
                "Origem da Divergência": str(regra_curta)[:500],
                "Observação": str(observacao)[:2000],
                "__VL_PARC_PAGAS": 0.0,
                "__VL_PARC_INADIMPLENTES": 0.0,
                "__VL_PARC_A_VENCER": 0.0,
                "__VL_SALDO": 0.0,
                "__QTD_PARC_PAGAS": 0,
                "__QTD_PARC_INADIMPLENTES": 0,
                "__QTD_PARC_A_VENCER": 0,
                "__QTD_SALDO": 0,
            })
            qtd_div_parcelas += 1

        # 1) Leitura de dados obrigatórios
        if df_r is None or df_r.empty:
            erros.append("VAL-LEITURA: Dados Receber vazio.")
            return erros, None, pendencias_parcelas

        for col in ["Status_Vencimento", "Vlr_Parcela", "Venda", "Parcela"]:
            if col not in df_r.columns:
                erros.append(f"VAL-LEITURA: coluna obrigatoria ausente em Dados Receber: {col}.")

        if erros:
            return erros, None, pendencias_parcelas

        _tempo_blocos["preparacao_inicial"] += (time.perf_counter() - _t_pre)

        df_r = df_r.copy()
        df_r["Venda"] = df_r["Venda"].fillna("").astype(str).str.strip()
        if df_p is not None and not df_p.empty and "Venda" in df_p.columns:
            df_p = df_p.copy()
            df_p["Venda"] = df_p["Venda"].fillna("").astype(str).str.strip()

        # Entrada do motor (= df_receber_raw / df_recebidos_raw no montar): mesma base usada para moda de Principal e identificador final.
        df_r_vp_id = (df_r_motor_entrada if df_r_motor_entrada is not None else df_r).copy()
        df_r_vp_id["Venda"] = df_r_vp_id["Venda"].fillna("").astype(str).str.strip()
        df_p_vp_id = None
        if df_p_motor_entrada is not None and not df_p_motor_entrada.empty:
            df_p_vp_id = df_p_motor_entrada.copy()
            if "Venda" in df_p_vp_id.columns:
                df_p_vp_id["Venda"] = df_p_vp_id["Venda"].fillna("").astype(str).str.strip()
        elif df_p is not None and not df_p.empty:
            df_p_vp_id = df_p.copy()

        st = df_r["Status_Vencimento"].fillna("").astype(str).str.strip().str.upper()
        vp = pd.to_numeric(df_r["Vlr_Parcela"], errors="coerce")
        if st.eq("").all():
            erros.append("VAL-LEITURA: Status_Vencimento vazio em 100% das linhas.")
        if vp.notna().sum() == 0:
            erros.append("VAL-LEITURA: Vlr_Parcela sem valores numericos validos.")
        if (vp.fillna(0) <= 0).all():
            erros.append("VAL-LEITURA: Vlr_Parcela sem valores positivos.")

        # 2) Classificação estrita
        status_validos = {"VENCIDO", "A VENCER"}
        status_invalidos = sorted(set([x for x in st.unique().tolist() if x not in status_validos]))
        if status_invalidos:
            erros.append(f"VAL-CLASSIFICACAO: status invalidos encontrados em Dados Receber: {status_invalidos}.")
        if int((st == "VENCIDO").sum() + (st == "A VENCER").sum()) != int(len(df_r)):
            erros.append("VAL-CLASSIFICACAO: existem linhas fora de VENCIDO/A VENCER.")

        # 3..7) Consolidação / somas / identificador / parcela
        if df_c is None or df_c.empty:
            erros.append("VAL-CRITICO: Consolidado vazio.")
            return erros, None, pendencias_parcelas

        if df_c["Venda"].astype(str).str.strip().duplicated().any():
            erros.append("VAL-CONSOLIDACAO: venda repetida no consolidado.")

        vendas_base = set(df_r["Venda"].fillna("").astype(str).str.strip())
        vendas_base = {v for v in vendas_base if v != ""}
        vendas_cons = set(df_c["Venda"].fillna("").astype(str).str.strip())
        faltantes = sorted(vendas_base - vendas_cons)
        if faltantes:
            erros.append(f"VAL-CONSOLIDACAO: vendas da base nao consolidadas: {len(faltantes)}.")

        # Cliente final por venda: moda do nome real (não usar Cliente_Base como exibição)
        base_cli_valid = pd.concat(
            [
                df_r[["Venda", "Cliente"]] if "Cliente" in df_r.columns else pd.DataFrame(columns=["Venda", "Cliente"]),
                df_p[["Venda", "Cliente"]] if (df_p is not None and not df_p.empty and "Cliente" in df_p.columns) else pd.DataFrame(columns=["Venda", "Cliente"]),
            ],
            ignore_index=True,
        )
        if not base_cli_valid.empty and "Cliente" in df_c.columns:
            base_cli_valid["Venda"] = base_cli_valid["Venda"].fillna("").astype(str).str.strip()
            base_cli_valid["Cliente"] = base_cli_valid["Cliente"].fillna("").astype(str).str.strip()
            base_cli_valid = base_cli_valid.loc[
                (base_cli_valid["Venda"] != "") & (base_cli_valid["Cliente"] != "")
            ].copy()
            mapa_cli_esp = (
                base_cli_valid.groupby("Venda")["Cliente"]
                .agg(escolher_cliente_exibicao)
                .to_dict()
            )
            for row in df_c.itertuples(index=False):
                v = str(getattr(row, "Venda", "") or "").strip()
                esp = str(mapa_cli_esp.get(v, "")).strip()
                if esp == "":
                    continue
                enc = str(getattr(row, "Cliente", "") or "").strip()
                if enc != esp:
                    _pendencia_pre_export(
                        "VAL-CLIENTE",
                        "Cliente divergente da moda nas bases",
                        f"Consolidado exibe {enc!r}; moda nas bases {esp!r}.",
                        v,
                        enc,
                        str(getattr(row, "Identificador", "") or "").strip(),
                    )
                    break

        def _df_com_subgrupo_para_validacao(d):
            if d is None or d.empty:
                return d
            out = d.copy()
            if "Cliente_Base" not in out.columns:
                out = adicionar_chave_cliente(out)
            out["_K_SUB"] = out.apply(chave_subgrupo_contratual, axis=1)
            out["_CB_KEY"] = out["Cliente_Base"].fillna("").astype(str).str.strip()
            return out

        df_r = _df_com_subgrupo_para_validacao(df_r)
        if df_p is not None and not df_p.empty:
            df_p = _df_com_subgrupo_para_validacao(df_p)

        _t_mapas = time.perf_counter()
        # Somatórias por status na base
        df_venc = df_r.loc[st == "VENCIDO"].copy()
        df_av = df_r.loc[st == "A VENCER"].copy()

        def _preparar_cols_parcela(df_src):
            if df_src is None or df_src.empty:
                return df_src
            d = df_src
            if "Parcela" not in d.columns:
                d["Parcela"] = ""
            if "_PARC_NORM" not in d.columns:
                d["_PARC_NORM"] = d["Parcela"].map(normalizar_parcela)
            if "_DEN_AUD" not in d.columns:
                d["_DEN_AUD"] = d["_PARC_NORM"].map(_denominador_parcela_audit)
            if "_PRI_NUM" not in d.columns:
                if "Principal" in d.columns:
                    d["_PRI_NUM"] = pd.to_numeric(d["Principal"], errors="coerce").fillna(0.0)
                else:
                    d["_PRI_NUM"] = 0.0
            return d

        df_r = _preparar_cols_parcela(df_r)
        df_p = _preparar_cols_parcela(df_p)
        df_venc = _preparar_cols_parcela(df_venc)
        df_av = _preparar_cols_parcela(df_av)

        def _mapa_qtd_parc_distintas_por_venda_df(df):
            if df is None or df.empty or "Venda" not in df.columns or "Parcela" not in df.columns:
                return {}
            vk = df["Venda"].fillna("").astype(str).str.strip()
            parc_norm = (
                df["_PARC_NORM"]
                if "_PARC_NORM" in df.columns
                else df["Parcela"].map(normalizar_parcela)
            )
            out = {}
            for v_key, sub in pd.DataFrame({"_VK": vk, "_PN": parc_norm}).groupby("_VK", sort=False):
                vs = str(v_key).strip()
                if not vs:
                    continue
                vals = [p for p in sub["_PN"].tolist() if p]
                out[vs] = int(len(set(vals)))
            return out

        def _mapa_slice_por_venda(df):
            if df is None or df.empty or "Venda" not in df.columns:
                return {}
            vk = df["Venda"].fillna("").astype(str).str.strip()
            out = {}
            for v_key, sub in df.groupby(vk, sort=False):
                vs = str(v_key).strip()
                if not vs:
                    continue
                out[vs] = sub
            return out

        cache_qtd_por_venda_p = _mapa_qtd_parc_distintas_por_venda_df(df_p)
        cache_qtd_por_venda_venc = _mapa_qtd_parc_distintas_por_venda_df(df_venc)
        cache_qtd_por_venda_av = _mapa_qtd_parc_distintas_por_venda_df(df_av)
        cache_df_venda_r = _mapa_slice_por_venda(df_r)
        cache_df_venda_p = _mapa_slice_por_venda(df_p)
        cache_df_venda_venc = _mapa_slice_por_venda(df_venc)
        cache_df_venda_av = _mapa_slice_por_venda(df_av)
        cache_info_sub = {}

        # Auditoria universal de reconciliação A VENCER é diagnóstica e custosa;
        # executar somente em modo detalhado para não penalizar tempo padrão.
        if DEBUG_VALIDACAO:
            mapa_av_venda = mapa_vl_vencer_por_venda_receber_tratado(df_r)
            vl_cons_por_venda = {}
            for row in df_c.itertuples(index=False):
                vk = str(getattr(row, "Venda", "") or "").strip()
                if vk == "":
                    continue
                _vl = pd.to_numeric(getattr(row, "Vl_Vencer", 0), errors="coerce")
                vl_cons_por_venda[vk] = 0.0 if pd.isna(_vl) else float(_vl)
            audit_rows = []
            todas_v = set(mapa_av_venda.keys()) | set(vl_cons_por_venda.keys())
            for v in sorted(todas_v):
                soma_b = float(mapa_av_venda.get(v, 0) or 0)
                soma_c = float(vl_cons_por_venda.get(v, 0) or 0)
                diff = soma_b - soma_c
                if abs(soma_b) < 1e-9 and abs(soma_c) < 1e-9 and abs(diff) < 1e-9:
                    continue
                audit_rows.append((v, soma_b, soma_c, diff))
            audit_rows.sort(key=lambda x: -abs(x[3]))
            max_abs = max((abs(x[3]) for x in audit_rows), default=0.0)
            print(
                "[DEBUG][AUDITORIA_A_VENCER] resumo | "
                f"vendas_com_massa_av_ou_cons={len(audit_rows)} | max_abs_diff={max_abs:.2f}"
            )
            divergentes = [(v, sb, sc, d) for v, sb, sc, d in audit_rows if abs(d) > 0.05]
            if (not divergentes) and audit_rows:
                print("[DEBUG][AUDITORIA_A_VENCER] detalhe | todas_amostradas_ok_abaixo_0_05 (sem divergencia material)")
            limite_log = 30
            for venda, soma_av_base, vl_vencer_cons, diferenca in (divergentes[:limite_log] if divergentes else []):
                sub_av = cache_df_venda_av.get(venda) if cache_df_venda_av else None
                if sub_av is None:
                    sub_av = pd.DataFrame()
                qtd_linhas = int(len(sub_av))
                if "Parcela" in sub_av.columns and qtd_linhas:
                    qtd_parcelas_distintas = contar_parcelas_distintas_padrao(sub_av["Parcela"])
                    pars = [str(x).strip() for x in sub_av["Parcela"].tolist() if str(x).strip() != ""]
                else:
                    qtd_parcelas_distintas = 0
                    pars = []
                denoms = sorted({d for d in (_denominador_parcela_audit(p) for p in pars) if d > 0})
                padrao = "HETEROGENEO" if len(denoms) > 1 else "HOMOGENEO"
                ncb_av = int(sub_av["Cliente_Base"].nunique()) if "Cliente_Base" in sub_av.columns and qtd_linhas else 0
                sub_r_audit = cache_df_venda_r.get(venda) if cache_df_venda_r else None
                ncb_venda = (
                    int(sub_r_audit["Cliente_Base"].nunique())
                    if (
                        sub_r_audit is not None
                        and not sub_r_audit.empty
                        and "Cliente_Base" in sub_r_audit.columns
                        and venda
                    )
                    else 0
                )
                frag_cb = "SIM" if ncb_venda > 1 else "NAO"
                if abs(diferenca) <= 0.05:
                    causa = "OK"
                elif soma_av_base > vl_vencer_cons + 0.05:
                    causa = "SOMA_BASE_TRATADA_MAIOR_QUE_VL_VENCER_CONSOLIDADO"
                else:
                    causa = "VL_VENCER_CONSOLIDADO_MAIOR_QUE_BASE_TRATADA"
                print(
                    "[DEBUG][AUDITORIA_A_VENCER] "
                    f"venda={venda} | soma_av_base={soma_av_base:.2f} | vl_vencer_cons={vl_vencer_cons:.2f} | "
                    f"diferenca={diferenca:.2f} | qtd_linhas_av_base={qtd_linhas} | qtd_parcelas_av_base={qtd_parcelas_distintas} | "
                    f"denominadores={denoms} | padrao={padrao} | nunique_cliente_base_av={ncb_av} | "
                    f"fragmentacao_cliente_base_venda={frag_cb} | causa={causa}"
                )

        soma_inad_base = (
            pd.to_numeric(df_venc.get("Principal", 0), errors="coerce").fillna(0)
            + pd.to_numeric(df_venc.get("Correcao", 0), errors="coerce").fillna(0)
            + pd.to_numeric(df_venc.get("Juros_Atraso", 0), errors="coerce").fillna(0)
            + pd.to_numeric(df_venc.get("Multa_Atraso", 0), errors="coerce").fillna(0)
            + pd.to_numeric(df_venc.get("Correcao_Atraso", 0), errors="coerce").fillna(0)
        ).sum()
        soma_inad_cons = pd.to_numeric(df_c.get("Vl.Principal (Encargos)", 0), errors="coerce").fillna(0).sum()
        if abs(float(soma_inad_base) - float(soma_inad_cons)) > 0.05:
            _pendencia_pre_export(
                "VAL-SOMA_INAD",
                "Soma inadimplência base vs consolidado",
                f"Receber(VENCIDO)={float(soma_inad_base):.2f} | Consolidado={float(soma_inad_cons):.2f}.",
                "GERAL",
            )

        soma_av_base = pd.to_numeric(df_av["Vlr_Parcela"], errors="coerce").fillna(0).sum()
        soma_av_cons = pd.to_numeric(df_c.get("Vl.Vencer", 0), errors="coerce").fillna(0).sum()
        if abs(float(soma_av_base) - float(soma_av_cons)) > 0.05:
            _pendencia_pre_export(
                "VAL-SOMA_AVENCER",
                "Soma a vencer base vs consolidado",
                f"Receber(A VENCER)={float(soma_av_base):.2f} | Consolidado={float(soma_av_cons):.2f}.",
                "GERAL",
            )

        # Validação financeira adicional: Vl.Pago (Consolidado) vs Dados Recebidos (Total_Dep)
        soma_pago_base = 0.0
        if df_p is not None and not df_p.empty and "Total_Dep" in df_p.columns:
            soma_pago_base = float(pd.to_numeric(df_p["Total_Dep"], errors="coerce").fillna(0).sum())
        soma_pago_cons = float(pd.to_numeric(df_c.get("Vl.Pago", 0), errors="coerce").fillna(0).sum())
        if abs(soma_pago_base - soma_pago_cons) > 0.05:
            _pendencia_pre_export(
                "VAL-SOMA_PAGO",
                "Soma pago base vs consolidado",
                f"Recebidos(Total_Dep)={soma_pago_base:.2f} | Consolidado(Vl.Pago)={soma_pago_cons:.2f}.",
                "GERAL",
            )

        # Composição financeira interna de encargos no consolidado
        soma_enc_composta = float((
            pd.to_numeric(df_c.get("Vl.Principal Atrasado", 0), errors="coerce").fillna(0)
            + pd.to_numeric(df_c.get("Vl.Correção", 0), errors="coerce").fillna(0)
            + pd.to_numeric(df_c.get("Vl.Juros", 0), errors="coerce").fillna(0)
            + pd.to_numeric(df_c.get("Vl.Multas", 0), errors="coerce").fillna(0)
            + pd.to_numeric(df_c.get("Vl.Correção Atraso", 0), errors="coerce").fillna(0)
        ).sum())
        if abs(soma_enc_composta - float(soma_inad_cons)) > 0.05:
            _pendencia_pre_export(
                "VAL-SOMA_ENCARGOS_COMP",
                "Composição interna de encargos no consolidado",
                f"Soma componentes={soma_enc_composta:.2f} | Vl.Principal (Encargos)={float(soma_inad_cons):.2f}.",
                "GERAL",
            )

        # Validação estrutural universal de parcelas por venda
        # (sem exceções locais): definir total confiável + validar fechamento.
        # Contagens por venda (df_p / df_venc / df_av): mapas cache_qtd_por_venda_* (pré-calculados acima).

        def _mapa_info_parcelas(df_a, df_b):
            frames = []
            for src in [df_a, df_b]:
                if src is None or src.empty or "Venda" not in src.columns:
                    continue
                _parc_src = (
                    src["Parcela"]
                    if "Parcela" in src.columns
                    else pd.Series("", index=src.index, dtype=str)
                )
                _pri_src = (
                    src["Principal"]
                    if "Principal" in src.columns
                    else pd.Series(0.0, index=src.index, dtype=float)
                )
                t = pd.DataFrame({
                    "Venda": src["Venda"].fillna("").astype(str).str.strip(),
                    "_PARC_NORM": src["_PARC_NORM"] if "_PARC_NORM" in src.columns else _parc_src.map(normalizar_parcela),
                    "_DEN_AUD": src["_DEN_AUD"] if "_DEN_AUD" in src.columns else _parc_src.map(normalizar_parcela).map(_denominador_parcela_audit),
                    "_PRI_NUM": src["_PRI_NUM"] if "_PRI_NUM" in src.columns else pd.to_numeric(_pri_src, errors="coerce").fillna(0.0),
                })
                t = t.loc[t["Venda"] != "", ["Venda", "_PARC_NORM", "_DEN_AUD", "_PRI_NUM"]]
                if not t.empty:
                    frames.append(t)
            if not frames:
                return {}

            base = pd.concat(frames, ignore_index=True)
            if base.empty:
                return {}

            out = {}
            for venda, g in base.groupby("Venda", dropna=False):
                pars = [p for p in g["_PARC_NORM"].tolist() if p != ""]
                pars_set = set(pars)
                denoms = set([int(x) for x in g["_DEN_AUD"].tolist() if int(x) > 0])
                qtd_dist = len(pars_set)
                max_den = max(denoms) if denoms else 0
                possui_1_1 = "1/1" in pars_set
                denom_stats = {}
                for den_val, g_den in g.groupby("_DEN_AUD", sort=False, dropna=False):
                    try:
                        den_int = int(den_val)
                    except (TypeError, ValueError):
                        continue
                    if den_int <= 0:
                        continue
                    parcelas_den = set(
                        [x for x in g_den["_PARC_NORM"].tolist() if str(x).strip() != ""]
                    )
                    soma_pri_den = float(
                        pd.to_numeric(g_den["_PRI_NUM"], errors="coerce").fillna(0).sum()
                    )
                    denom_stats[den_int] = {
                        "qtd_parcelas": int(len(parcelas_den)),
                        "soma_principal": float(soma_pri_den),
                    }

                denom_dominante = 0
                criterio_dominante = "sem_dominante"
                if denom_stats:
                    max_freq = max([v["qtd_parcelas"] for v in denom_stats.values()]) or 1
                    max_pri = max([v["soma_principal"] for v in denom_stats.values()]) or 1.0
                    best_score = -1.0
                    for den, st_den in denom_stats.items():
                        freq_norm = float(st_den["qtd_parcelas"]) / float(max_freq)
                        pri_norm = float(st_den["soma_principal"]) / float(max_pri) if max_pri > 0 else 0.0
                        cobertura = float(st_den["qtd_parcelas"]) / float(max(qtd_dist, 1))
                        # Universo contratual dominante: frequência + massa financeira + cobertura.
                        score = (freq_norm * 0.55) + (pri_norm * 0.30) + (cobertura * 0.15)
                        if score > best_score:
                            best_score = score
                            denom_dominante = int(den)
                    criterio_dominante = "score_freq_principal_cobertura"

                # Hierarquia de total confiável:
                # 1) maior denominador válido
                # 2) contagem distinta de parcelas reais
                # 3) fallback contrato único 1/1
                if max_den > 0 and len(denoms) > 1:
                    # Em padrão heterogêneo, usa universo contratual dominante.
                    qtd_total = int(denom_dominante or max_den)
                    criterio = "heterogeneo_universo_contratual_dominante"
                elif max_den > 0:
                    qtd_total = max(max_den, qtd_dist)
                    criterio = "max_denominador_vs_qtd_distinta"
                elif qtd_dist > 0:
                    qtd_total = qtd_dist
                    criterio = "qtd_distinta_parcelas"
                elif possui_1_1:
                    qtd_total = 1
                    criterio = "fallback_unico_1_1"
                else:
                    qtd_total = 0
                    criterio = "sem_base_confiavel"

                out[str(venda).strip()] = {
                    "qtd_total": int(qtd_total),
                    "criterio_total": criterio,
                    "qtd_distinta": int(qtd_dist),
                    "denominadores": sorted([int(x) for x in denoms]),
                    "denominador_dominante": int(denom_dominante or 0),
                    "criterio_dominante": criterio_dominante,
                    "soma_principal_por_denominador": {int(k): float(v["soma_principal"]) for k, v in denom_stats.items()},
                    "frequencia_parcelas_por_denominador": {int(k): int(v["qtd_parcelas"]) for k, v in denom_stats.items()},
                    "heterogeneo": len(denoms) > 1,
                    "possui_1_1": bool(possui_1_1),
                }
            return out

        def _info_parcelas_subgrupo(tmp_r_sub, tmp_p_sub):
            frames = []
            for src in [tmp_r_sub, tmp_p_sub]:
                if src is None or src.empty:
                    continue
                _parc_src = (
                    src["Parcela"]
                    if "Parcela" in src.columns
                    else pd.Series("", index=src.index, dtype=str)
                )
                _pri_src = (
                    src["Principal"]
                    if "Principal" in src.columns
                    else pd.Series(0.0, index=src.index, dtype=float)
                )
                t = pd.DataFrame({
                    "_PARC_NORM": src["_PARC_NORM"] if "_PARC_NORM" in src.columns else _parc_src.map(normalizar_parcela),
                    "_DEN_AUD": src["_DEN_AUD"] if "_DEN_AUD" in src.columns else _parc_src.map(normalizar_parcela).map(_denominador_parcela_audit),
                    "_PRI_NUM": src["_PRI_NUM"] if "_PRI_NUM" in src.columns else pd.to_numeric(_pri_src, errors="coerce").fillna(0.0),
                })
                t = t.loc[:, ["_PARC_NORM", "_DEN_AUD", "_PRI_NUM"]]
                if not t.empty:
                    frames.append(t)
            if not frames:
                return {}
            g2 = pd.concat(frames, ignore_index=True)
            if g2.empty:
                return {}
            pars = [p for p in g2["_PARC_NORM"].tolist() if p != ""]
            pars_set = set(pars)
            denoms = set([int(x) for x in g2["_DEN_AUD"].tolist() if int(x) > 0])
            qtd_dist = len(pars_set)
            max_den = max(denoms) if denoms else 0
            possui_1_1 = "1/1" in pars_set
            denom_stats = {}
            for den_val, g_den in g2.groupby("_DEN_AUD", sort=False, dropna=False):
                try:
                    den_int = int(den_val)
                except (TypeError, ValueError):
                    continue
                if den_int <= 0:
                    continue
                parcelas_den = set(
                    [x for x in g_den["_PARC_NORM"].tolist() if str(x).strip() != ""]
                )
                soma_pri_den = float(
                    pd.to_numeric(g_den["_PRI_NUM"], errors="coerce").fillna(0).sum()
                )
                denom_stats[den_int] = {
                    "qtd_parcelas": int(len(parcelas_den)),
                    "soma_principal": float(soma_pri_den),
                }
            denom_dominante = 0
            criterio_dominante = "sem_dominante"
            if denom_stats:
                max_freq = max([v["qtd_parcelas"] for v in denom_stats.values()]) or 1
                max_pri = max([v["soma_principal"] for v in denom_stats.values()]) or 1.0
                best_score = -1.0
                for den, st_den in denom_stats.items():
                    freq_norm = float(st_den["qtd_parcelas"]) / float(max_freq)
                    pri_norm = float(st_den["soma_principal"]) / float(max_pri) if max_pri > 0 else 0.0
                    cobertura = float(st_den["qtd_parcelas"]) / float(max(qtd_dist, 1))
                    score = (freq_norm * 0.55) + (pri_norm * 0.30) + (cobertura * 0.15)
                    if score > best_score:
                        best_score = score
                        denom_dominante = int(den)
                criterio_dominante = "score_freq_principal_cobertura"
            if max_den > 0 and len(denoms) > 1:
                qtd_total = int(denom_dominante or max_den)
                criterio = "heterogeneo_universo_contratual_dominante"
            elif max_den > 0:
                qtd_total = max(max_den, qtd_dist)
                criterio = "max_denominador_vs_qtd_distinta"
            elif qtd_dist > 0:
                qtd_total = qtd_dist
                criterio = "qtd_distinta_parcelas"
            elif possui_1_1:
                qtd_total = 1
                criterio = "fallback_unico_1_1"
            else:
                qtd_total = 0
                criterio = "sem_base_confiavel"
            return {
                "qtd_total": int(qtd_total),
                "criterio_total": criterio,
                "qtd_distinta": int(qtd_dist),
                "denominadores": sorted([int(x) for x in denoms]),
                "denominador_dominante": int(denom_dominante or 0),
                "criterio_dominante": criterio_dominante,
                "soma_principal_por_denominador": {int(k): float(v["soma_principal"]) for k, v in denom_stats.items()},
                "frequencia_parcelas_por_denominador": {int(k): int(v["qtd_parcelas"]) for k, v in denom_stats.items()},
                "heterogeneo": len(denoms) > 1,
                "possui_1_1": bool(possui_1_1),
            }

        mapa_info_parcelas = _mapa_info_parcelas(df_r, df_p)
        mapa_total_confiavel = {k: int(v.get("qtd_total", 0) or 0) for k, v in mapa_info_parcelas.items()}
        mapa_set_pag = mapa_conjunto_parcelas_por_venda(df_p)
        mapa_set_venc = mapa_conjunto_parcelas_por_venda(df_venc)
        mapa_set_av = mapa_conjunto_parcelas_por_venda(df_av)

        def _amostra_conjunto_parcelas(s, lim=5):
            if not s:
                return "[]"
            lst = sorted(s)[:lim]
            return str(lst) + ("..." if len(s) > lim else "")

        def _montar_mapa_chaves_subgrupo_venda():
            partes = []
            for d in (df_r, df_p, df_venc, df_av):
                if d is None or d.empty:
                    continue
                if "Venda" not in d.columns or "Cliente_Base" not in d.columns or "_K_SUB" not in d.columns:
                    continue
                p = d[["Venda", "Cliente_Base", "_K_SUB"]].copy()
                p["Venda"] = p["Venda"].fillna("").astype(str).str.strip()
                p["Cliente_Base"] = p["Cliente_Base"].fillna("").astype(str).str.strip()
                p["_K_SUB"] = p["_K_SUB"].fillna("").astype(str).str.strip()
                p = p.loc[p["Venda"] != ""]
                if not p.empty:
                    partes.append(p)
            if not partes:
                return {}
            base = pd.concat(partes, ignore_index=True).drop_duplicates()
            out = {}
            for venda_key, g in base.groupby("Venda", sort=False):
                out[str(venda_key)] = sorted(
                    {(str(cb).strip(), str(ks).strip()) for cb, ks in zip(g["Cliente_Base"], g["_K_SUB"])}
                )
            return out

        mapa_chaves_subgrupo_venda = _montar_mapa_chaves_subgrupo_venda()
        _tempo_blocos["mapas_agregados"] += (time.perf_counter() - _t_mapas)

        def _chaves_subgrupo_para_venda(vsk):
            return mapa_chaves_subgrupo_venda.get(str(vsk).strip(), [])

        _cache_diag_parcelas_venda = {}

        def _diag_parcelas_lazy(vsk):
            """Diagnóstico pesado só sob demanda + memoização por venda (validação pré-exportação)."""
            vs = str(vsk).strip()
            if vs not in _cache_diag_parcelas_venda:
                _cache_diag_parcelas_venda[vs] = _diagnostico_parcelas_venda(vs, df_r, df_p, df_c)
            return _cache_diag_parcelas_venda[vs]

        _t_loop = time.perf_counter()
        total_vendas_validacao = int(len(df_c))
        t_valid_loop_ini = time.perf_counter()
        tempo_subgrupos_total = 0.0
        tempo_subgrupos_filtragem = 0.0
        tempo_subgrupos_norm_agreg = 0.0
        qtd_subgrupos_avaliados = 0
        _empty_df = pd.DataFrame()

        def _metricas_parcelas_df(df_part):
            if df_part is None or df_part.empty or "Parcela" not in df_part.columns:
                return 0, set()
            if "_PARC_NORM" in df_part.columns:
                pars_norm = df_part["_PARC_NORM"]
            else:
                pars_norm = df_part["Parcela"].map(normalizar_parcela)
            vals = [p for p in pars_norm.tolist() if p]
            if not vals:
                return 0, set()
            s = set(vals)
            return int(len(s)), s

        def _mapa_metricas_venda_total(df_src):
            """Métricas de parcelas por venda inteira (evita recomputar a cada iteração do consolidado)."""
            out = {}
            if df_src is None or df_src.empty or "Venda" not in df_src.columns:
                return out
            vk = df_src["Venda"].fillna("").astype(str).str.strip()
            base = df_src.loc[vk.ne("")]
            if base.empty:
                return out
            for v_key, g in base.groupby("Venda", sort=False, dropna=False):
                vs = str(v_key).strip()
                if not vs:
                    continue
                out[vs] = _metricas_parcelas_df(g)
            return out

        def _mapa_subgrupos_somente_por_venda(df_src):
            """Slices por subgrupo (sem métricas; ex.: df_r)."""
            out = {}
            if df_src is None or df_src.empty:
                return out
            if "Venda" not in df_src.columns or "_CB_KEY" not in df_src.columns or "_K_SUB" not in df_src.columns:
                return out
            base = df_src.loc[df_src["Venda"].fillna("").astype(str).str.strip().ne("")]
            if base.empty:
                return out
            for venda_key, df_venda in base.groupby("Venda", sort=False, dropna=False):
                vs = str(venda_key).strip()
                if not vs:
                    continue
                mapa_sub = {}
                for (cbk, ksub), g in df_venda.groupby(["_CB_KEY", "_K_SUB"], sort=False, dropna=False):
                    mapa_sub[(str(cbk).strip(), str(ksub).strip())] = g
                out[vs] = mapa_sub
            return out

        def _mapa_subgrupos_e_metricas_por_venda(df_src):
            """Uma passagem: DataFrames por subgrupo + métricas de parcelas (evita 2x groupby no mesmo df)."""
            out_sub = {}
            out_met = {}
            if df_src is None or df_src.empty:
                return out_sub, out_met
            if "Venda" not in df_src.columns or "_CB_KEY" not in df_src.columns or "_K_SUB" not in df_src.columns:
                return out_sub, out_met
            base = df_src.loc[df_src["Venda"].fillna("").astype(str).str.strip().ne("")]
            if base.empty:
                return out_sub, out_met
            for venda_key, df_venda in base.groupby("Venda", sort=False, dropna=False):
                vs = str(venda_key).strip()
                if not vs:
                    continue
                m_sub = {}
                m_met = {}
                for (cbk, ksub), g in df_venda.groupby(["_CB_KEY", "_K_SUB"], sort=False, dropna=False):
                    key = (str(cbk).strip(), str(ksub).strip())
                    m_sub[key] = g
                    m_met[key] = _metricas_parcelas_df(g)
                out_sub[vs] = m_sub
                out_met[vs] = m_met
            return out_sub, out_met

        mapa_sub_p_por_venda, mapa_metricas_sub_p_por_venda = _mapa_subgrupos_e_metricas_por_venda(df_p)
        mapa_sub_v_por_venda, mapa_metricas_sub_v_por_venda = _mapa_subgrupos_e_metricas_por_venda(df_venc)
        mapa_sub_a_por_venda, mapa_metricas_sub_a_por_venda = _mapa_subgrupos_e_metricas_por_venda(df_av)
        mapa_sub_r_por_venda = _mapa_subgrupos_somente_por_venda(df_r)

        metrics_venda_total_p = _mapa_metricas_venda_total(df_p)
        metrics_venda_total_venc = _mapa_metricas_venda_total(df_venc)
        metrics_venda_total_av = _mapa_metricas_venda_total(df_av)

        _n_df_c = len(df_c)

        def _col_list_str(df, col, n):
            if col not in df.columns:
                return [""] * n
            return df[col].fillna("").astype(str).str.strip().tolist()

        def _col_list_float(df, col, n):
            if col not in df.columns:
                return [0.0] * n
            return pd.to_numeric(df[col], errors="coerce").fillna(0.0).tolist()

        _c_venda = _col_list_str(df_c, "Venda", _n_df_c)
        _c_cli = _col_list_str(df_c, "Cliente", _n_df_c)
        _c_id = _col_list_str(df_c, "Identificador", _n_df_c)
        _c_qtd_paga = _col_list_float(df_c, "Qtd.Parc.Paga", _n_df_c)
        _c_qtd_atr = _col_list_float(df_c, "Qtd.Parc.Atrasada", _n_df_c)
        _c_qtd_av = _col_list_float(df_c, "Qtd.Parc.A Vencer", _n_df_c)
        _c_qtd_tot = _col_list_float(df_c, "Qtd.Parc.Total", _n_df_c)
        _c_valparc = _col_list_float(df_c, "Valor Da Parcela", _n_df_c)
        _c_vl_enc = _col_list_float(df_c, "Vl.Principal (Encargos)", _n_df_c)
        _c_vl_vencer = _col_list_float(df_c, "Vl.Vencer", _n_df_c)

        _cols_p_empty = (
            list(df_p.columns) if df_p is not None and not df_p.empty else ["Venda", "Parcela", "Principal"]
        )
        _cols_venc_empty = list(df_venc.columns) if not df_venc.empty else []
        _cols_av_empty = list(df_av.columns) if not df_av.empty else []
        _cols_r_empty = list(df_r.columns) if not df_r.empty else []

        for i in range(_n_df_c):
            idx_valid = i + 1
            venda = str(_c_venda[i] or "").strip()
            if venda == "":
                continue
            row = {
                "Venda": venda,
                "Cliente": str(_c_cli[i] or "").strip(),
                "Identificador": str(_c_id[i] or "").strip(),
                "Qtd.Parc.Paga": float(_c_qtd_paga[i]),
                "Qtd.Parc.Atrasada": float(_c_qtd_atr[i]),
                "Qtd.Parc.A Vencer": float(_c_qtd_av[i]),
                "Qtd.Parc.Total": float(_c_qtd_tot[i]),
                "Valor Da Parcela": float(_c_valparc[i]),
            }
            if perf_extra_ligado and (idx_valid % 200) == 0:
                print(
                    f"[TEMPO][VALIDACAO_PROGRESSO] "
                    f"processadas={idx_valid}/{total_vendas_validacao} "
                    f"({(100.0 * idx_valid / max(total_vendas_validacao, 1)):.1f}%) | "
                    f"elapsed={(time.perf_counter() - t_valid_loop_ini):.2f}s"
                )

            pago_cons = int(float(row.get("Qtd.Parc.Paga", 0) or 0))
            atr_cons = int(float(row.get("Qtd.Parc.Atrasada", 0) or 0))
            av_cons = int(float(row.get("Qtd.Parc.A Vencer", 0) or 0))
            total_cons = int(float(row.get("Qtd.Parc.Total", 0) or 0))

            pago_base = int(cache_qtd_por_venda_p.get(venda, 0))
            atr_base = int(cache_qtd_por_venda_venc.get(venda, 0))
            av_base = int(cache_qtd_por_venda_av.get(venda, 0))
            total_base = int(mapa_total_confiavel.get(venda, 0) or 0)
            soma_fechamento = int(pago_cons + atr_cons + av_cons)
            info = mapa_info_parcelas.get(venda, {})
            criterio_total = str(info.get("criterio_total", "na"))
            heterogeneo = bool(info.get("heterogeneo", False))
            denoms = info.get("denominadores", [])
            den_dom = int(info.get("denominador_dominante", 0) or 0)
            crit_dom = str(info.get("criterio_dominante", "na"))
            soma_pri_den = info.get("soma_principal_por_denominador", {})
            freq_den = info.get("frequencia_parcelas_por_denominador", {})
            qtd_dist = int(info.get("qtd_distinta", 0) or 0)
            padrao = "HETEROGENEO" if heterogeneo else "HOMOGENEO"

            set_pag = mapa_set_pag.get(venda, set())
            set_venc = mapa_set_venc.get(venda, set())
            set_av = mapa_set_av.get(venda, set())
            muni = metricas_universo_parcelas_operacionais(set_pag, set_venc, set_av)
            U_dist = int(muni["universo_total_distinto"])
            ipv = muni["intersec_pag_venc"]
            ipa = muni["intersec_pag_av"]
            iva = muni["intersec_venc_av"]
            tem_sobreposicao = bool(muni["tem_sobreposicao"])

            def _registrar_pendencia_parcelas(tipo_divergencia, regra_violada, observacao, parcela_ref=""):
                nonlocal contexto_primeira_falha, qtd_div_parcelas
                if contexto_primeira_falha is None:
                    contexto_primeira_falha = {
                        "Venda": venda,
                        "Cliente": str(row.get("Cliente", "") or "").strip(),
                        "Identificador": str(row.get("Identificador", "") or "").strip(),
                        "Regra_Violada": regra_violada,
                        "Causa_Raiz": observacao,
                    }
                valor_parcela = float(pd.to_numeric(row.get("Valor Da Parcela", 0), errors="coerce") or 0)
                qtd_pag = int(pago_cons)
                qtd_atr = int(atr_cons)
                qtd_avc = int(av_cons)
                qtd_solta = max(int(U_dist) - int(total_cons), 0)
                pendencias_parcelas.append({
                    "Venda": venda,
                    "Cliente": str(row.get("Cliente", "") or "").strip(),
                    "Identificador": str(row.get("Identificador", "") or "").strip(),
                    "Parcela": str(parcela_ref or "").strip(),
                    "Tipo de Divergência": str(tipo_divergencia),
                    "Valor Parcela Divergente": float(valor_parcela),
                    "Origem da Divergência": str(regra_violada),
                    "Observação": str(observacao),
                    "__VL_PARC_PAGAS": float(valor_parcela * qtd_pag),
                    "__VL_PARC_INADIMPLENTES": float(valor_parcela * qtd_atr),
                    "__VL_PARC_A_VENCER": float(valor_parcela * qtd_avc),
                    "__VL_SALDO": float(valor_parcela * int(qtd_solta)),
                    "__QTD_PARC_PAGAS": int(qtd_pag),
                    "__QTD_PARC_INADIMPLENTES": int(qtd_atr),
                    "__QTD_PARC_A_VENCER": int(qtd_avc),
                    "__QTD_SALDO": int(qtd_solta),
                })
                qtd_div_parcelas += 1

            def _log_diag(causa, escopo="GLOBAL_VENDA", ref_subgrupo=""):
                if not DEBUG_VALIDACAO:
                    return
                ref_sg = f" | ref_subgrupo={ref_subgrupo!r}" if str(ref_subgrupo or "").strip() else ""
                print(
                    "[DEBUG][AUDITORIA_PARCELAS] "
                    f"escopo={escopo}{ref_sg} | "
                    f"venda={venda} | padrao={padrao} | criterio_total={criterio_total} | "
                    f"denominadores={denoms} | denominador_dominante={den_dom} | criterio_dominante={crit_dom} | "
                    f"freq_por_den={freq_den} | soma_principal_por_den={soma_pri_den} | qtd_distinta={qtd_dist} | "
                    f"parcelas_pag_distintas={muni['n_pag']} | parcelas_venc_distintas={muni['n_venc']} | parcelas_av_distintas={muni['n_av']} | "
                    f"universo_total_distinto={U_dist} | "
                    f"intersec_pag_venc_n={len(ipv)} | intersec_pag_venc={_amostra_conjunto_parcelas(ipv)} | "
                    f"intersec_pag_av_n={len(ipa)} | intersec_pag_av={_amostra_conjunto_parcelas(ipa)} | "
                    f"intersec_venc_av_n={len(iva)} | intersec_venc_av={_amostra_conjunto_parcelas(iva)} | "
                    f"tem_sobreposicao={tem_sobreposicao} | "
                    f"qtd_paga_base={pago_base} | qtd_atrasada_base={atr_base} | qtd_avencer_base={av_base} | "
                    f"qtd_total_base={total_base} | qtd_paga_cons={pago_cons} | qtd_atrasada_cons={atr_cons} | "
                    f"qtd_avencer_cons={av_cons} | qtd_total_cons={total_cons} | soma_fechamento={soma_fechamento} | "
                    f"causa={causa}"
                )

            def _resumo_divergencia_global(causa_str):
                if not DEBUG_VALIDACAO:
                    return
                print(
                    "[DEBUG][AUDITORIA_DIVERGENCIA_RESUMO] "
                    f"venda={venda} | escopo=GLOBAL_VENDA | "
                    f"camada=QUANTIDADES_E_UNIVERSO_AGREGADO_VENDA | causa={causa_str!r} | "
                    f"subgrupos_contratuais_avaliados_neste_momento=NAO | "
                    f"obs=validacao_por_subuniverso_so_ocorre_apos_passar_neste_escopo"
                )

            if pago_cons != pago_base:
                _log_diag("DIVERGENCIA_QTD_PAGA")
                _resumo_divergencia_global("DIVERGENCIA_QTD_PAGA")
                msg_e, ctx_e = _montar_erro_val_parcelas_qtd(
                    "QTD_PAGA",
                    venda,
                    "Qtd.Parc.Paga",
                    pago_cons,
                    pago_base,
                    "contagem de rótulos de parcela distintos (normalizar_parcela) em Dados Recebidos",
                    row,
                    _diag_parcelas_lazy(venda),
                    "Qtd.Parc.Paga no consolidado deve coincidir com a quantidade de parcelas canônicas distintas em Dados Recebidos para a mesma venda (base tratada na validação).",
                    "Possível dessincronia entre bases tratadas, duplicidade residual em Recebidos, ou divergência de texto de parcela (ex.: 1/120 vs 01/120) após normalização.",
                )
                _registrar_pendencia_parcelas(
                    "QTD_PAGA",
                    "Qtd.Parc.Paga no consolidado deve coincidir com a quantidade de parcelas canônicas distintas em Dados Recebidos para a mesma venda (base tratada na validação).",
                    "Possível dessincronia entre bases tratadas, duplicidade residual em Recebidos, ou divergência de texto de parcela (ex.: 1/120 vs 01/120) após normalização.",
                )
                continue
            if atr_cons != atr_base:
                _log_diag("DIVERGENCIA_QTD_ATRASADA")
                _resumo_divergencia_global("DIVERGENCIA_QTD_ATRASADA")
                msg_e, ctx_e = _montar_erro_val_parcelas_qtd(
                    "QTD_ATRASADA",
                    venda,
                    "Qtd.Parc.Atrasada",
                    atr_cons,
                    atr_base,
                    "contagem de parcelas distintas em Dados Receber com Status_Vencimento=VENCIDO",
                    row,
                    _diag_parcelas_lazy(venda),
                    "Qtd.Parc.Atrasada no consolidado deve coincidir com parcelas canônicas distintas em linhas VENCIDO na base Receber tratada.",
                    "Status ou datas de vencimento podem divergir do consolidado; ou parcelas duplicadas / rótulos heterogêneos na base Receber.",
                )
                _registrar_pendencia_parcelas(
                    "QTD_ATRASADA",
                    "Qtd.Parc.Atrasada no consolidado deve coincidir com parcelas canônicas distintas em linhas VENCIDO na base Receber tratada.",
                    "Status ou datas de vencimento podem divergir do consolidado; ou parcelas duplicadas / rótulos heterogêneos na base Receber.",
                )
                continue
            if av_cons != av_base:
                _log_diag("DIVERGENCIA_QTD_A_VENCER")
                _resumo_divergencia_global("DIVERGENCIA_QTD_A_VENCER")
                msg_e, ctx_e = _montar_erro_val_parcelas_qtd(
                    "QTD_A_VENCER",
                    venda,
                    "Qtd.Parc.A Vencer",
                    av_cons,
                    av_base,
                    "contagem de parcelas distintas em Dados Receber com Status_Vencimento=A VENCER",
                    row,
                    _diag_parcelas_lazy(venda),
                    "Qtd.Parc.A Vencer no consolidado deve coincidir com parcelas canônicas distintas em linhas A VENCER na base Receber tratada.",
                    "Classificação A VENCER vs VENCIDO na base Receber pode não bater com a data-base usada no motor; ou duplicidade de linhas.",
                )
                _registrar_pendencia_parcelas(
                    "QTD_A_VENCER",
                    "Qtd.Parc.A Vencer no consolidado deve coincidir com parcelas canônicas distintas em linhas A VENCER na base Receber tratada.",
                    "Classificação A VENCER vs VENCIDO na base Receber pode não bater com a data-base usada no motor; ou duplicidade de linhas.",
                )
                continue
            if (not heterogeneo) and total_base > 0:
                ref_min_g, ref_max_g = ajustar_total_confiavel_global(
                    total_base, pago_base, atr_base, av_base, tem_sobreposicao
                )
                if total_cons < ref_min_g or total_cons > ref_max_g:
                    _log_diag("DIVERGENCIA_QTD_TOTAL_UNIVERSO_HOMOGENEO")
                    _resumo_divergencia_global("DIVERGENCIA_QTD_TOTAL_UNIVERSO_HOMOGENEO")
                    extras_tot = [
                        f"  Referência homogênea: ref_min={ref_min_g} ref_max={ref_max_g} (ajustar_total_confiavel_global)",
                        f"  Total confiável (mapa Receber+Recebidos): {total_base}",
                    ]
                    msg_e, ctx_e = _montar_erro_val_parcelas_qtd(
                        "QTD_TOTAL_HOMOGENEO",
                        venda,
                        "Qtd.Parc.Total",
                        total_cons,
                        total_base,
                        "total confiável derivado de denominadores / universo (mapa_info_parcelas) em universo homogêneo",
                        row,
                        _diag_parcelas_lazy(venda),
                        "Com universo homogêneo, Qtd.Parc.Total deve situar-se entre ref_min e ref_max definidos pela função ajustar_total_confiavel_global.",
                        "Denominador dominante ou universo distinto pode não refletir o total gravado no consolidado; verificar lift de total vs piso.",
                        extras_linhas=extras_tot,
                    )
                    _registrar_pendencia_parcelas(
                        "QTD_TOTAL_HOMOGENEO",
                        "Com universo homogêneo, Qtd.Parc.Total deve situar-se entre ref_min e ref_max definidos pela função ajustar_total_confiavel_global.",
                        "Denominador dominante ou universo distinto pode não refletir o total gravado no consolidado; verificar lift de total vs piso.",
                    )
                    continue
            if heterogeneo and total_base > 0 and total_cons != total_base:
                _log_diag("DIVERGENCIA_QTD_TOTAL_UNIVERSO_HETEROGENEO")

            if (not heterogeneo) and total_cons < max(pago_cons, atr_cons, av_cons):
                _log_diag("QTD_TOTAL_MENOR_QUE_COMPONENTE")
                _resumo_divergencia_global("QTD_TOTAL_MENOR_QUE_COMPONENTE")
                piso_c = max(pago_cons, atr_cons, av_cons)
                extras_p = [
                    f"  Piso operacional max(Paga,Atrasada,AVencer): {piso_c}",
                ]
                msg_e, ctx_e = _montar_erro_val_parcelas_qtd(
                    "QTD_TOTAL_INCONSISTENTE",
                    venda,
                    "Qtd.Parc.Total",
                    total_cons,
                    piso_c,
                    "piso=max(Qtd.Parc.Paga, Qtd.Parc.Atrasada, Qtd.Parc.A Vencer) no consolidado",
                    row,
                    _diag_parcelas_lazy(venda),
                    "Em universo homogêneo, Qtd.Parc.Total não pode ser inferior ao maior componente operacional (pago, atrasada, a vencer).",
                    "Total contratual no consolidado ficou abaixo de um dos componentes já agregados — revisar regra de piso no montar_consolidado.",
                    extras_linhas=extras_p,
                )
                _registrar_pendencia_parcelas(
                    "QTD_TOTAL_INCONSISTENTE",
                    "Em universo homogêneo, Qtd.Parc.Total não pode ser inferior ao maior componente operacional (pago, atrasada, a vencer).",
                    "Total contratual no consolidado ficou abaixo de um dos componentes já agregados — revisar regra de piso no montar_consolidado.",
                )
                continue
            if heterogeneo and total_cons < max(pago_cons, atr_cons, av_cons):
                _log_diag("QTD_TOTAL_MENOR_QUE_COMPONENTE_UNIVERSO_HETEROGENEO")

            # Conflito lógico na base Receber: mesma parcela em VENCIDO e A VENCER.
            if iva:
                _log_diag("PARCELA_EM_VENCIDO_E_A_VENCER_MESMA_REFERENCIA")
                _resumo_divergencia_global("PARCELA_EM_VENCIDO_E_A_VENCER_MESMA_REFERENCIA")
                am_iv = _amostra_lista_parcelas_norm(iva, lim=30)
                msg_e, ctx_e = _montar_erro_val_parcelas_qtd(
                    "PARCELA_DUAL_STATUS",
                    venda,
                    "Parcela (Receber)",
                    len(iva),
                    0,
                    "interseção VENCIDO ∩ A VENCER (não deveria existir)",
                    row,
                    _diag_parcelas_lazy(venda),
                    "A mesma parcela canônica não pode aparecer simultaneamente em linhas VENCIDO e A VENCER na base Receber.",
                    "Classificação por data-base gerou sobreposição; ou exportação UAU com linhas duplicadas em status distintos.",
                    extras_linhas=[f"  Parcelas na interseção (amostra): {am_iv}"],
                )
                obs_dual = (
                    "Classificação por data-base gerou sobreposição; ou exportação UAU com linhas duplicadas em status distintos."
                )
                for parc_ref in sorted(iva, key=lambda z: str(z)):
                    _registrar_pendencia_parcelas(
                        "PARCELA_DUAL_STATUS",
                        "A mesma parcela canônica não pode aparecer simultaneamente em linhas VENCIDO e A VENCER na base Receber.",
                        obs_dual,
                        parcela_ref=str(parc_ref),
                    )
                continue

            # Contrato por subuniverso (Tipo + denominador + identificador): teto e fechamento não misturam naturezas.
            chaves_sub = _chaves_subgrupo_para_venda(venda)
            if not chaves_sub:
                chaves_sub = [("", "__TODO_VENDA__")]

            def _rotulos_brutos_parcela_df(df_part):
                if df_part is None or df_part.empty or "Parcela" not in df_part.columns:
                    return []
                return sorted({
                    str(x).strip()
                    for x in df_part["Parcela"].tolist()
                    if str(x).strip() != ""
                })

            falha_sub = False
            subgrupos_passaram = []
            svp = cache_df_venda_p.get(venda) if cache_df_venda_p else None
            svv = cache_df_venda_venc.get(venda) if cache_df_venda_venc else None
            sva = cache_df_venda_av.get(venda) if cache_df_venda_av else None
            svr = cache_df_venda_r.get(venda) if cache_df_venda_r else None
            _t_sg_filt = time.perf_counter()
            mapa_sub_p = mapa_sub_p_por_venda.get(venda, {})
            mapa_sub_v = mapa_sub_v_por_venda.get(venda, {})
            mapa_sub_a = mapa_sub_a_por_venda.get(venda, {})
            mapa_sub_r = mapa_sub_r_por_venda.get(venda, {})
            mapa_metricas_sub_p = mapa_metricas_sub_p_por_venda.get(venda, {})
            mapa_metricas_sub_v = mapa_metricas_sub_v_por_venda.get(venda, {})
            mapa_metricas_sub_a = mapa_metricas_sub_a_por_venda.get(venda, {})
            venda_metricas_p = metrics_venda_total_p.get(venda, (0, set()))
            venda_metricas_v = metrics_venda_total_venc.get(venda, (0, set()))
            venda_metricas_a = metrics_venda_total_av.get(venda, (0, set()))
            tempo_subgrupos_filtragem += (time.perf_counter() - _t_sg_filt)

            def _resumo_divergencia_sub(causa_str, idx_sg, cbf_sg, ks_sg):
                if not DEBUG_VALIDACAO:
                    return
                nao_aval = [f"{a}|{b}" for a, b in chaves_sub[idx_sg + 1 :]]
                print(
                    "[DEBUG][AUDITORIA_DIVERGENCIA_RESUMO] "
                    f"venda={venda} | escopo=SUBGRUPO_LOCAL | "
                    f"camada=VALIDACAO_CONTRATUAL_SUBUNIVERSO | causa={causa_str!r} | "
                    f"subgrupo_falhou_cliente_base={cbf_sg!r} | subgrupo_falhou_chave={ks_sg!r} | "
                    f"indice_ordem_execucao={idx_sg} | total_subgrupos_planejados={len(chaves_sub)} | "
                    f"subgrupos_ok_sem_erro_ate_agora={subgrupos_passaram!r} | "
                    f"subgrupos_nao_avaliados={nao_aval!r}"
                )

            t_sub_ini = time.perf_counter()
            for idx, (cb_fin, k_sub) in enumerate(chaves_sub):
                qtd_subgrupos_avaliados += 1
                cbf = str(cb_fin or "").strip()
                ks = str(k_sub or "").strip()
                ref_sg = f"{cbf}|{ks}"
                if ks == "__TODO_VENDA__":
                    sp_df = (
                        cache_df_venda_p.get(venda)
                        if (
                            df_p is not None
                            and not df_p.empty
                            and "Venda" in df_p.columns
                            and cache_df_venda_p
                        )
                        else None
                    )
                    if sp_df is None:
                        sp_df = _empty_df
                    sv_df = (
                        cache_df_venda_venc.get(venda)
                        if (not df_venc.empty and "Venda" in df_venc.columns and cache_df_venda_venc)
                        else None
                    )
                    if sv_df is None:
                        sv_df = _empty_df
                    sa_df = (
                        cache_df_venda_av.get(venda)
                        if (not df_av.empty and "Venda" in df_av.columns and cache_df_venda_av)
                        else None
                    )
                    if sa_df is None:
                        sa_df = _empty_df
                    sr_vc = (
                        cache_df_venda_r.get(venda)
                        if (not df_r.empty and cache_df_venda_r)
                        else None
                    )
                    if sr_vc is None:
                        sr_vc = _empty_df
                    tipo_base_log = "__AGREGADO_VENDA__"
                    denom_log = ""
                    id_log = ""
                    qtd_pago_sub, set_pag_sub = venda_metricas_p
                    qtd_venc_sub, set_venc_sub = venda_metricas_v
                    qtd_av_sub, set_av_sub = venda_metricas_a
                else:
                    k_sub_key = (cbf, ks)
                    sp_df = mapa_sub_p.get(k_sub_key, _empty_df)
                    sv_df = mapa_sub_v.get(k_sub_key, _empty_df)
                    sa_df = mapa_sub_a.get(k_sub_key, _empty_df)
                    sr_vc = mapa_sub_r.get(k_sub_key, _empty_df)
                    partes_sg = ks.split("|||", 2)
                    tipo_base_log = partes_sg[0] if len(partes_sg) > 0 else ""
                    denom_log = partes_sg[1] if len(partes_sg) > 1 else ""
                    id_log = partes_sg[2] if len(partes_sg) > 2 else ""
                    qtd_pago_sub, set_pag_sub = mapa_metricas_sub_p.get(k_sub_key, (0, set()))
                    qtd_venc_sub, set_venc_sub = mapa_metricas_sub_v.get(k_sub_key, (0, set()))
                    qtd_av_sub, set_av_sub = mapa_metricas_sub_a.get(k_sub_key, (0, set()))

                tmp_p = sp_df if not sp_df.empty else pd.DataFrame(columns=_cols_p_empty)
                if not df_venc.empty:
                    tmp_venc = sv_df if not sv_df.empty else pd.DataFrame(columns=_cols_venc_empty)
                else:
                    tmp_venc = pd.DataFrame()
                if not df_av.empty:
                    tmp_av = sa_df if not sa_df.empty else pd.DataFrame(columns=_cols_av_empty)
                else:
                    tmp_av = pd.DataFrame()
                tmp_r = sr_vc if not sr_vc.empty else pd.DataFrame(columns=_cols_r_empty)

                soma_fechamento_sub = int(qtd_pago_sub + qtd_venc_sub + qtd_av_sub)
                _t_sg_norm = time.perf_counter()
                k_cache_sub = (venda, cbf, ks)
                if k_cache_sub not in cache_info_sub:
                    cache_info_sub[k_cache_sub] = _info_parcelas_subgrupo(tmp_r, tmp_p)
                info_sub = cache_info_sub[k_cache_sub]
                total_base_sub = int(info_sub.get("qtd_total", 0) or 0)
                heterogeneo_sub = bool(info_sub.get("heterogeneo", False))
                total_cons_sub = 0

                muni_sg = metricas_universo_parcelas_operacionais(set_pag_sub, set_venc_sub, set_av_sub)
                tempo_subgrupos_norm_agreg += (time.perf_counter() - _t_sg_norm)
                U_dist_sg = int(muni_sg["universo_total_distinto"])
                tem_sobreposicao_sub = bool(muni_sg["tem_sobreposicao"])

                teto_contratual_validacao = max(
                    int(total_cons_sub or 0),
                    int(total_base_sub or 0),
                )

                chave_fin_log = f"{venda}||{cbf}"
                if DEBUG_VALIDACAO:
                    print(
                        "[DEBUG][AUDITORIA_SUBGRUPO_CONTRATUAL] "
                        f"venda={venda} | chave_financeira={chave_fin_log} | cliente_equiparado={cbf!r} | "
                        f"tipo_base={tipo_base_log!r} | denominador={denom_log!r} | identificador_base={id_log!r} | "
                        f"qtd_pago_sub={qtd_pago_sub} | qtd_venc_sub={qtd_venc_sub} | qtd_av_sub={qtd_av_sub} | "
                        f"set_pag_sub={sorted(set_pag_sub)} | set_venc_sub={sorted(set_venc_sub)} | set_av_sub={sorted(set_av_sub)} | "
                        f"universo_total_distinto={U_dist_sg} | soma_fechamento={soma_fechamento_sub} | "
                        f"total_base={total_base_sub} | total_cons={total_cons_sub} | "
                        f"teto_contratual_validacao={teto_contratual_validacao} | "
                        f"heterogeneo_sub={heterogeneo_sub} | tem_sobreposicao_sub={tem_sobreposicao_sub}"
                    )

                if teto_contratual_validacao > 0 and U_dist_sg > teto_contratual_validacao:
                    _log_diag(
                        "UNIVERSO_DISTINTO_EXCEDE_TETO_CONTRATUAL_VALIDACAO",
                        escopo="SUBGRUPO_LOCAL",
                        ref_subgrupo=ref_sg,
                    )

                    brut_p = _rotulos_brutos_parcela_df(tmp_p)
                    brut_v = _rotulos_brutos_parcela_df(tmp_venc)
                    brut_a = _rotulos_brutos_parcela_df(tmp_av)
                    uniq_raw_union_n = len(set(brut_p) | set(brut_v) | set(brut_a))
                    alias_amostra = []
                    for b in sorted(set(brut_p) | set(brut_v) | set(brut_a))[:50]:
                        cn = normalizar_parcela(b)
                        if cn and b.strip() != cn:
                            alias_amostra.append(f"{b!r}→{cn!r}")
                    if DEBUG_VALIDACAO:
                        print(
                            "[DEBUG][AUDITORIA_PARCELAS_UNIVERSO] "
                            f"venda={venda} | chave_financeira={chave_fin_log} | subgrupo={ks!r} | "
                            f"total_cons={total_cons_sub} | total_base={total_base_sub} | "
                            f"universo_total_distinto={U_dist_sg} | "
                            f"set_pag_canon={sorted(set_pag_sub)} | set_venc_canon={sorted(set_venc_sub)} | "
                            f"set_av_canon={sorted(set_av_sub)} | "
                            f"rotulos_brutos_distintos_union_n={uniq_raw_union_n} | "
                            f"indicio_inflacao_por_rotulo_inconsistente={bool(uniq_raw_union_n > U_dist_sg)} | "
                            f"mapa_bruto_para_canon_amostra={alias_amostra[:20]}"
                        )

                        print(
                            "[DEBUG][AUDITORIA_TETO_CONTRATUAL] "
                            f"venda={venda} | chave_financeira={chave_fin_log} | subgrupo={ks!r} | "
                            f"heterogeneo={heterogeneo_sub} | "
                            f"qtd_total_cons={total_cons_sub} | qtd_total_base={total_base_sub} | "
                            f"teto_contratual_validacao={teto_contratual_validacao} | "
                            f"universo_total_distinto={U_dist_sg}"
                        )

                    _resumo_divergencia_sub(
                        "UNIVERSO_DISTINTO_EXCEDE_TETO_CONTRATUAL_VALIDACAO", idx, cbf, ks
                    )
                    extras_u = [
                        f"  Subgrupo: Cliente_Base={cbf!r} chave={ks!r}",
                        f"  Universo distinto sub={U_dist_sg} teto_contratual_validacao={teto_contratual_validacao} total_cons_sub={total_cons_sub} total_base_sub={total_base_sub}",
                    ]
                    msg_e, ctx_e = _montar_erro_val_parcelas_qtd(
                        "UNIVERSO_EXCEDE_TETO_SUB",
                        venda,
                        "Universo parcelas (subgrupo)",
                        U_dist_sg,
                        teto_contratual_validacao,
                        "teto_contratual_validacao = max(total_cons_sub, total_base_sub) no subgrupo contratual",
                        row,
                        _diag_parcelas_lazy(venda),
                        "No subuniverso contratual, a quantidade de parcelas canônicas distintas não pode exceder o teto contratual de validação.",
                        "Mistura de contratos no mesmo subgrupo, inflação de rótulos de parcela, ou total consolidado abaixo do universo real.",
                        extras_linhas=extras_u,
                    )
                    _registrar_pendencia_parcelas(
                        "UNIVERSO_EXCEDE_TETO_SUB",
                        "No subuniverso contratual, a quantidade de parcelas canônicas distintas não pode exceder o teto contratual de validação.",
                        "Mistura de contratos no mesmo subgrupo, inflação de rótulos de parcela, ou total consolidado abaixo do universo real.",
                    )
                    falha_sub = True
                    continue

                if soma_fechamento_sub > teto_contratual_validacao and (not heterogeneo_sub) and (not tem_sobreposicao_sub):
                    _log_diag(
                        "SOMA_FECHAMENTO_MAIOR_QUE_TOTAL_UNIVERSO_HOMOGENEO_DISJUNTO",
                        escopo="SUBGRUPO_LOCAL",
                        ref_subgrupo=ref_sg,
                    )
                    _resumo_divergencia_sub(
                        "SOMA_FECHAMENTO_MAIOR_QUE_TOTAL_UNIVERSO_HOMOGENEO_DISJUNTO", idx, cbf, ks
                    )
                    extras_f = [
                        f"  Subgrupo: Cliente_Base={cbf!r} chave={ks!r}",
                        f"  soma_fechamento_sub={soma_fechamento_sub} teto={teto_contratual_validacao} heterogeneo_sub={heterogeneo_sub} sobreposicao_sub={tem_sobreposicao_sub}",
                    ]
                    msg_e, ctx_e = _montar_erro_val_parcelas_qtd(
                        "FECHAMENTO_SUB_HOMOGENEO",
                        venda,
                        "Soma Paga+Atrasada+AVencer (subgrupo)",
                        soma_fechamento_sub,
                        teto_contratual_validacao,
                        "teto_contratual_validacao no subgrupo (universo homogêneo disjunto)",
                        row,
                        _diag_parcelas_lazy(venda),
                        "Com universos disjuntos e homogêneos no subgrupo, a soma das quantidades não pode exceder o teto contratual.",
                        "Parcelas contadas em mais de um papel sem sobreposição declarada, ou teto subestimado frente ao contrato.",
                        extras_linhas=extras_f,
                    )
                    _registrar_pendencia_parcelas(
                        "FECHAMENTO_SUB_HOMOGENEO",
                        "Com universos disjuntos e homogêneos no subgrupo, a soma das quantidades não pode exceder o teto contratual.",
                        "Parcelas contadas em mais de um papel sem sobreposição declarada, ou teto subestimado frente ao contrato.",
                    )
                    falha_sub = True
                    continue
                if soma_fechamento_sub > teto_contratual_validacao and heterogeneo_sub and (not tem_sobreposicao_sub):
                    _log_diag(
                        "SOMA_FECHAMENTO_MAIOR_QUE_TOTAL_UNIVERSO_HETEROGENEO_DISJUNTO",
                        escopo="SUBGRUPO_LOCAL",
                        ref_subgrupo=ref_sg,
                    )
                    _resumo_divergencia_sub(
                        "SOMA_FECHAMENTO_MAIOR_QUE_TOTAL_UNIVERSO_HETEROGENEO_DISJUNTO", idx, cbf, ks
                    )
                    extras_f2 = [
                        f"  Subgrupo: Cliente_Base={cbf!r} chave={ks!r}",
                        f"  soma_fechamento_sub={soma_fechamento_sub} teto={teto_contratual_validacao} heterogeneo_sub={heterogeneo_sub}",
                    ]
                    msg_e, ctx_e = _montar_erro_val_parcelas_qtd(
                        "FECHAMENTO_SUB_HETEROGENEO",
                        venda,
                        "Soma Paga+Atrasada+AVencer (subgrupo)",
                        soma_fechamento_sub,
                        teto_contratual_validacao,
                        "teto_contratual_validacao no subgrupo (universo heterogêneo disjunto)",
                        row,
                        _diag_parcelas_lazy(venda),
                        "Com universos disjuntos e heterogêneos no subgrupo, a soma das quantidades não pode exceder o teto contratual.",
                        "Contratos com denominadores mistos no mesmo subgrupo; revisar chave_subgrupo_contratual e totais.",
                        extras_linhas=extras_f2,
                    )
                    _registrar_pendencia_parcelas(
                        "FECHAMENTO_SUB_HETEROGENEO",
                        "Com universos disjuntos e heterogêneos no subgrupo, a soma das quantidades não pode exceder o teto contratual.",
                        "Contratos com denominadores mistos no mesmo subgrupo; revisar chave_subgrupo_contratual e totais.",
                    )
                    falha_sub = True
                    continue
                if soma_fechamento_sub > teto_contratual_validacao and tem_sobreposicao_sub:
                    _log_diag(
                        "SOMA_CRUA_MAIOR_QUE_TOTAL_EXPLICADA_POR_SOBREPOSICAO_UNIVERSOS",
                        escopo="SUBGRUPO_LOCAL",
                        ref_subgrupo=ref_sg,
                    )

                subgrupos_passaram.append(ref_sg)

            if falha_sub:
                tempo_subgrupos_total += (time.perf_counter() - t_sub_ini)
                continue
            tempo_subgrupos_total += (time.perf_counter() - t_sub_ini)

            if DEBUG_VALIDACAO and chaves_sub:
                print(
                    "[DEBUG][AUDITORIA_SUBGRUPO_RESUMO] "
                    f"venda={venda} | escopo=CONTRATO_SUBUNIVERSO | resultado=TODOS_SUBGRUPOS_OK | "
                    f"total_subgrupos_validados={len(chaves_sub)} | "
                    f"subgrupos_sem_erro={subgrupos_passaram!r}"
                )

            # Lacuna de fechamento: mesma referência dual do teto na venda (consolidado vs total confiável),
            # alinhada à arquitetura por subgrupo sem reabrir o loop de subgrupos.
            teto_referencia_lacuna = max(int(total_cons or 0), int(total_base or 0))
            if teto_referencia_lacuna > 0 and soma_fechamento < teto_referencia_lacuna:
                if heterogeneo:
                    causa_het = "PADRAO_HETEROGENEO_COM_LACUNA_DE_FECHAMENTO"
                    _log_diag(causa_het)
                else:
                    _log_diag("LACUNA_FECHAMENTO_UNIVERSO_HOMOGENEO")

        # Valor da parcela (moda de Principal por venda) e não zerado indevido
        _tempo_blocos["validacao_loop_vendas"] += (time.perf_counter() - _t_loop)
        _tempo_blocos["auditoria_subgrupos"] += tempo_subgrupos_total
        tempo_subgrupos_comp_alert = max(
            float(tempo_subgrupos_total) - float(tempo_subgrupos_filtragem) - float(tempo_subgrupos_norm_agreg),
            0.0,
        )
        _t_finais = time.perf_counter()
        mapa_parcela = {}
        for venda, g in df_r_vp_id.groupby(df_r_vp_id["Venda"].fillna("").astype(str).str.strip(), dropna=False):
            if str(venda).strip() == "":
                continue
            mapa_parcela[str(venda).strip()] = moda_valor_parcela_por_df_ou_grupo(g)
        for i in range(_n_df_c):
            venda = str(_c_venda[i] or "").strip()
            esperado = float(mapa_parcela.get(venda, 0) or 0)
            encontrado = float(_c_valparc[i] or 0)
            if esperado > 0 and encontrado <= 0:
                _pendencia_pre_export(
                    "VAL-PARCELA",
                    "Valor Da Parcela zerado indevido",
                    f"Moda nas bases={esperado:.2f}; consolidado={encontrado:.2f}.",
                    venda,
                    str(_c_cli[i] or "").strip(),
                    str(_c_id[i] or "").strip(),
                    valor_vp=float(encontrado),
                )
                break
            if esperado > 0 and abs(encontrado - esperado) > 0.05:
                _pendencia_pre_export(
                    "VAL-PARCELA",
                    "Valor Da Parcela divergente da moda",
                    f"Moda nas bases={esperado:.2f}; consolidado={encontrado:.2f}.",
                    venda,
                    str(_c_cli[i] or "").strip(),
                    str(_c_id[i] or "").strip(),
                    valor_vp=float(encontrado),
                )
                break

        # Identificador mais frequente por venda (desempate: mais completo)
        def _coletar_ids(df):
            itens = []
            if df is None or df.empty:
                return itens
            for _, r in df.iterrows():
                venda = str(r.get("Venda", "")).strip()
                if not venda:
                    continue
                for c in ["Unidades", "Identificador_Produto"]:
                    if c not in r.index:
                        continue
                    v = r.get(c)
                    if str(v).strip() == "":
                        continue
                    if identificador_truncado(v):
                        continue
                    n = normalizar_identificador(v)
                    if n:
                        itens.append((venda, n))
            return itens

        ids_all = _coletar_ids(df_r_vp_id) + _coletar_ids(df_p_vp_id)
        mapa_id = {}
        if ids_all:
            tmp = pd.DataFrame(ids_all, columns=["Venda", "id"])
            for venda, g in tmp.groupby("Venda"):
                freq = Counter(g["id"].tolist())
                max_f = max(freq.values())
                cands = [k for k, v in freq.items() if v == max_f]
                mapa_id[str(venda).strip()] = sorted(cands, key=lambda x: (-score_identificador(x), -len(str(x)), str(x)))[0]

        if mapa_id:
            for i in range(_n_df_c):
                venda = str(_c_venda[i] or "").strip()
                esperado = str(mapa_id.get(venda, "")).strip()
                if esperado == "":
                    continue
                encontrado = str(_c_id[i] or "").strip()
                if encontrado != esperado:
                    _pendencia_pre_export(
                        "VAL-IDENTIFICADOR",
                        "Identificador divergente da moda nas bases",
                        f"Esperado={esperado!r} | Consolidado={encontrado!r}.",
                        venda,
                        str(_c_cli[i] or "").strip(),
                        encontrado,
                    )
                    break

        # Zeros críticos indevidos quando base possui valor
        for i in range(_n_df_c):
            venda = str(_c_venda[i] or "").strip()
            if int(cache_qtd_por_venda_venc.get(venda, 0)) > 0 and float(_c_vl_enc[i] or 0) <= 0:
                _pendencia_pre_export(
                    "VAL-CRITICO_INAD",
                    "Inadimplência zerada com parcelas VENCIDO na base",
                    f"Qtd parcelas VENCIDO na base tratada={int(cache_qtd_por_venda_venc.get(venda, 0))}.",
                    venda,
                    str(_c_cli[i] or "").strip(),
                    str(_c_id[i] or "").strip(),
                )
                break
            if int(cache_qtd_por_venda_av.get(venda, 0)) > 0 and float(_c_vl_vencer[i] or 0) <= 0:
                _pendencia_pre_export(
                    "VAL-CRITICO_AVENCER",
                    "A vencer zerado com parcelas A VENCER na base",
                    f"Qtd parcelas A VENCER na base tratada={int(cache_qtd_por_venda_av.get(venda, 0))}.",
                    venda,
                    str(_c_cli[i] or "").strip(),
                    str(_c_id[i] or "").strip(),
                )
                break

        # Sem identificador não pode fragmentar venda em múltiplas linhas
        if "Identificador" in df_c.columns:
            dtmp = df_c.copy()
            dtmp["Venda"] = dtmp["Venda"].fillna("").astype(str).str.strip()
            dtmp["Identificador"] = dtmp["Identificador"].fillna("").astype(str).str.strip()
            dtmp = dtmp.loc[dtmp["Venda"] != ""]
            sem_id_multilinha = (
                dtmp.loc[dtmp["Identificador"] == ""]
                .groupby("Venda")
                .size()
            )
            sem_id_multilinha = sem_id_multilinha[sem_id_multilinha > 1]
            if not sem_id_multilinha.empty:
                for v_frag, nlin in sem_id_multilinha.items():
                    _pendencia_pre_export(
                        "VAL-IDENTIFICADOR_FRAGMENTACAO",
                        "Venda sem identificador em múltiplas linhas no consolidado",
                        f"Linhas sem identificador para a mesma venda: {int(nlin)}.",
                        str(v_frag),
                    )

        # Log de auditoria consolidado
        total_vendas = int(df_c["Venda"].fillna("").astype(str).str.strip().ne("").sum())
        total_pago = float(pd.to_numeric(df_c.get("Vl.Pago", 0), errors="coerce").fillna(0).sum())
        total_vencido = float(pd.to_numeric(df_c.get("Vl.Principal (Encargos)", 0), errors="coerce").fillna(0).sum())
        total_avencer = float(pd.to_numeric(df_c.get("Vl.Vencer", 0), errors="coerce").fillna(0).sum())
        qtd_sem_identificador = int(df_c.get("Identificador", "").fillna("").astype(str).str.strip().eq("").sum()) if "Identificador" in df_c.columns else 0
        if DEBUG_VALIDACAO:
            print(
                "[DEBUG][AUDITORIA_CONSOLIDADO] "
                f"vendas={total_vendas} | total_pago={total_pago:.2f} | total_vencido={total_vencido:.2f} | "
                f"total_a_vencer={total_avencer:.2f} | vendas_div_parcelas={qtd_div_parcelas} | "
                f"vendas_sem_identificador={qtd_sem_identificador}"
            )
        if perf_extra_ligado:
            print(
                f"[TEMPO] _validar_pre_exportacao.subgrupos: {tempo_subgrupos_total:.2f}s | "
                f"subgrupos_avaliados={qtd_subgrupos_avaliados}"
            )
            print(
                f"[TEMPO] _validar_pre_exportacao.subgrupos_partes: "
                f"filtragem={tempo_subgrupos_filtragem:.2f}s | "
                f"normalizacao_agregacao={tempo_subgrupos_norm_agreg:.2f}s | "
                f"comparacao_alertas={tempo_subgrupos_comp_alert:.2f}s"
            )
        _tempo_blocos["checagens_finais"] += (time.perf_counter() - _t_finais)
        if perf_extra_ligado:
            print(
                f"[TEMPO] _validar_pre_exportacao.preparacao_inicial: {_tempo_blocos['preparacao_inicial']:.2f}s | "
                f"mapas_agregados: {_tempo_blocos['mapas_agregados']:.2f}s | "
                f"loop_vendas: {_tempo_blocos['validacao_loop_vendas']:.2f}s | "
                f"subgrupos: {_tempo_blocos['auditoria_subgrupos']:.2f}s | "
                f"checagens_finais: {_tempo_blocos['checagens_finais']:.2f}s"
            )

        return erros, contexto_primeira_falha, pendencias_parcelas

    _t = time.perf_counter()
    try:
        if df_consolidado is not None and not df_consolidado.empty:
            _antes_cols = [
                "Vl.Vencer", "Qtd.Parc.A Vencer", "Valor Da Parcela", "Identificador",
                "Qtd.Parc.Atrasada", "Vl.Principal Atrasado", "Vl.Correção",
                "Vl.Juros", "Vl.Multas", "Vl.Correção Atraso", "Vl.Principal (Encargos)"
            ]
            _antes_val = {
                c: (
                    float(pd.to_numeric(df_consolidado[c], errors="coerce").fillna(0).sum())
                    if c in df_consolidado.columns and c != "Identificador"
                    else int(df_consolidado[c].fillna("").astype(str).str.strip().eq("").sum()) if c in df_consolidado.columns else None
                )
                for c in _antes_cols
            }
            _dbg(f"VALID_PRE antes: {_antes_val}")
        erros_validacao, contexto_validacao_falha, pendencias_parcelas = _validar_pre_exportacao(
            df_receber_tratado,
            df_recebidos_tratado,
            df_consolidado,
            df_r_motor_entrada=df_receber_entrada_motor,
            df_p_motor_entrada=df_recebidos_entrada_motor,
        )
        pendencias_parcelas = list(pendencias_qtd_total_reconc_montar or []) + list(pendencias_parcelas or [])
        if df_consolidado is not None and not df_consolidado.empty:
            _depois_val = {
                c: (
                    float(pd.to_numeric(df_consolidado[c], errors="coerce").fillna(0).sum())
                    if c in df_consolidado.columns and c != "Identificador"
                    else int(df_consolidado[c].fillna("").astype(str).str.strip().eq("").sum()) if c in df_consolidado.columns else None
                )
                for c in _antes_cols
            }
            _dbg(f"VALID_PRE depois: {_depois_val}")
        if erros_validacao:
            print("[ERRO] Validacao pre-exportacao falhou. Excel NAO sera gerado.")
            for e in erros_validacao:
                for linha in str(e).split("\n"):
                    print(f"[ERRO] {linha}")
            exc_val = ProcessamentoUAUErro(
                etapa="validação",
                funcao="_validar_pre_exportacao",
                validacao="regras críticas pré-exportação",
                mensagem=erros_validacao[0],
                campo_ou_aba="Dados Receber / Dados Recebidos / Consolidado Venda",
                contexto=contexto_validacao_falha or {},
            )
            print(exc_val.formatar_relatorio_completo())
            _emit_perf(
                "_validar_pre_exportacao",
                time.perf_counter() - _t,
                _nlin_df(df_receber_tratado),
                _nlin_df(df_recebidos_tratado),
                _nlin_df(df_consolidado),
            )
            _imprimir_ranking_perf()
            raise exc_val
    except ProcessamentoUAUErro:
        raise
    except Exception as e:
        raise ProcessamentoUAUErro(
            etapa="validação",
            funcao="_validar_pre_exportacao",
            validacao="regras críticas pré-exportação",
            mensagem="Falha inesperada durante validação final do consolidado.",
            campo_ou_aba="Dados Receber / Consolidado Venda",
            erro_tecnico=e,
        ) from e
    _emit_perf(
        "_validar_pre_exportacao",
        time.perf_counter() - _t,
        _nlin_df(df_receber_tratado),
        _nlin_df(df_recebidos_tratado),
        _nlin_df(df_consolidado),
    )

    if alertas_consolidado is None or alertas_consolidado.empty:
        alertas_consolidado = pd.DataFrame(
            columns=[
                "Venda", "Cliente_Base", "Tipo_Alerta", "Mensagem", "Divergencia",
                "Valor_Esperado", "Valor_Encontrado", "Regra", "Observacao"
            ]
        )
    else:
        for c in ["Valor_Esperado", "Valor_Encontrado", "Regra", "Observacao"]:
            if c not in alertas_consolidado.columns:
                alertas_consolidado[c] = ""

    nome_empreendimento = ""
    if not df_consolidado.empty and "Empreendimento" in df_consolidado.columns:
        descricoes = [
            str(v).strip()
            for v in df_consolidado["Empreendimento"].tolist()
            if str(v).strip() != ""
        ]
        nome_empreendimento = escolher_moda_texto(descricoes) if descricoes else ""

    if not nome_empreendimento:
        nome_empreendimento = "EMPREENDIMENTO"

    _emp_arquivo_st = str(nome_empreendimento_arquivo or "").strip()
    if _emp_arquivo_st:
        nome_empreendimento = _emp_arquivo_st

    _n_lote_uni = int(lote_unificado_empreendimentos_distintos or 0)
    if _n_lote_uni > 0:
        if _n_lote_uni == 1:
            nome_empreendimento = "CONSOLIDAÇÃO GERAL — 1 EMPREENDIMENTO (LOTE)"
        else:
            nome_empreendimento = f"CONSOLIDAÇÃO GERAL — {_n_lote_uni} EMPREENDIMENTOS UNIFICADOS"

    # Alinha a coluna exportada ao mesmo nome canônico de B1 / arquivo (sem alterar montar_consolidado).
    if (
        not preservar_empreendimento_por_linha
        and not df_consolidado.empty
        and "Empreendimento" in df_consolidado.columns
    ):
        _canon_emp_exp = str(nome_empreendimento or "").strip()
        if _canon_emp_exp and _canon_emp_exp != "EMPREENDIMENTO":
            df_consolidado["Empreendimento"] = _canon_emp_exp

    if str(nome_arquivo_xlsx_override or "").strip():
        nome_arquivo_final = str(nome_arquivo_xlsx_override).strip()
        if not nome_arquivo_final.lower().endswith(".xlsx"):
            nome_arquivo_final = nome_arquivo_final + ".xlsx"
    else:
        nome_arquivo_final = montar_nome_arquivo_empreendimento(
            df_consolidado, nome_empreendimento_canonico=nome_empreendimento
        )

    pasta_saida = os.path.dirname(caminho_saida)
    if not pasta_saida:
        pasta_saida = "."

    caminho_saida_final = os.path.join(pasta_saida, nome_arquivo_final)

    limpar_pasta_saida_excel_antigos(pasta_saida)

    nome_aba_principal = "Consolidado Venda"
    if not df_consolidado.empty and "Emp/Obra" in df_consolidado.columns:
        _eo_nonnull = df_consolidado["Emp/Obra"].dropna()
        if not _eo_nonnull.empty:
            _emp = str(_eo_nonnull.astype(str).iloc[0]).strip()
            if _emp and _emp.lower() != "nan":
                _sig = sanitizar_nome_arquivo(extrair_sigla_empreendimento(_emp))
                if _sig:
                    nome_aba_principal = f"{_sig} – Consolidado"
    nome_aba_pendencias = "PEND.PARCELAS"
    nome_aba_criterios = "CRITERIOS ANALISES"

    # Ordem logica das abas de apoio para facilitar conferencia.
    def _ordenar_colunas(df, ordem):
        if df is None or df.empty:
            return df
        cols = [c for c in ordem if c in df.columns]
        resto = [c for c in df.columns if c not in cols]
        return df[cols + resto]

    def _sem_colunas_internas_export(df):
        if df is None or df.empty:
            return df
        drop = [
            c for c in (
                "Parcela_Key",
                "POSSIVEL_CONFLITO_DUPLICIDADE",
                "__dedup_subkey",
                "_Id_Key_Dedup",
                "_POSSIVEL_CONFLITO_DEDUP",
            )
            if c in df.columns
        ]
        return df.drop(columns=drop) if drop else df

    def _mapa_identificador_estrito(*dfs):
        base = {}
        for df in dfs:
            if df is None or df.empty:
                continue
            for row in df.itertuples(index=False):
                venda = str(getattr(row, "Venda", "") or "").strip()
                cli_b = str(getattr(row, "Cliente_Base", "") or "").strip()
                if not venda or not cli_b:
                    continue
                cand = []
                for col in ("Identificador_Produto", "Unidades"):
                    raw = getattr(row, col, "")
                    if raw is None or str(raw).strip() == "":
                        continue
                    if identificador_truncado(raw):
                        continue
                    nid = normalizar_identificador(raw)
                    if nid:
                        cand.append(nid)
                if not cand:
                    continue
                chave = (venda, cli_b)
                bucket = base.setdefault(chave, [])
                bucket.extend(cand)
        return {
            k: moda_identificador_final_serie(pd.Series(v))
            for k, v in base.items()
            if v
        }

    def _resolver_identificador_export(row, mapa_id):
        for col in ("Identificador_Produto", "Unidades"):
            raw = row.get(col, "")
            if raw is None or str(raw).strip() == "":
                continue
            if identificador_truncado(raw):
                continue
            nid = normalizar_identificador(raw)
            if nid:
                return nid
        venda = str(row.get("Venda", "") or "").strip()
        cli_b = str(row.get("Cliente_Base", "") or "").strip()
        if venda and cli_b:
            from_map = str(mapa_id.get((venda, cli_b), "") or "").strip()
            if from_map:
                return from_map
        return "NLOC"

    def _normalizar_identificador_export_serie(serie: pd.Series) -> pd.Series:
        if serie is None:
            return pd.Series(dtype="object")

        def _resolver_item(raw):
            if raw is None:
                return ""
            txt = str(raw).strip()
            if not txt or identificador_truncado(txt):
                return ""
            return normalizar_identificador(txt) or ""

        return serie.fillna("").map(_resolver_item).astype(str)

    def _aplicar_identificador_export(df_src: pd.DataFrame, mapa_id: dict) -> None:
        if df_src is None or df_src.empty:
            return
        idx = df_src.index
        serie_id = (
            _normalizar_identificador_export_serie(df_src["Identificador_Produto"])
            if "Identificador_Produto" in df_src.columns
            else pd.Series("", index=idx, dtype="object")
        )
        serie_un = (
            _normalizar_identificador_export_serie(df_src["Unidades"])
            if "Unidades" in df_src.columns
            else pd.Series("", index=idx, dtype="object")
        )
        venda = (
            df_src["Venda"].fillna("").astype(str).str.strip()
            if "Venda" in df_src.columns
            else pd.Series("", index=idx, dtype="object")
        )
        cli_b = (
            df_src["Cliente_Base"].fillna("").astype(str).str.strip()
            if "Cliente_Base" in df_src.columns
            else pd.Series("", index=idx, dtype="object")
        )
        chaves = pd.Series(list(zip(venda.tolist(), cli_b.tolist())), index=idx, dtype="object")
        serie_mapa = chaves.map(mapa_id).fillna("").astype(str).str.strip()
        serie_final = serie_id.where(serie_id.ne(""), serie_un)
        serie_final = serie_final.where(serie_final.ne(""), serie_mapa)
        df_src["Identificador_Produto"] = serie_final.where(serie_final.ne(""), "NLOC")

    _t_xlsx = time.perf_counter()
    # Colunas auxiliares solicitadas para DADOS RECEBER:
    # - DIA_VENCIMENTO_BOLETO: dia do mês extraído do vencimento
    # - MES_VENCIMENTO / ANO_VENCIMENTO: referência temporal tabular
    # - CLASSIFICACAO_ADIMPLENCIA: adimplente/inadimplente por status de vencimento
    if df_receber_tratado is not None and not df_receber_tratado.empty:
        _venc_dt = pd.to_datetime(df_receber_tratado.get("Vencimento"), errors="coerce")
        df_receber_tratado["DIA_VENCIMENTO_BOLETO"] = _venc_dt.dt.day.fillna(0).astype(int)
        _meses_pt = {
            1: "JANEIRO",
            2: "FEVEREIRO",
            3: "MARÇO",
            4: "ABRIL",
            5: "MAIO",
            6: "JUNHO",
            7: "JULHO",
            8: "AGOSTO",
            9: "SETEMBRO",
            10: "OUTUBRO",
            11: "NOVEMBRO",
            12: "DEZEMBRO",
        }
        _mes_num = _venc_dt.dt.month
        df_receber_tratado["MES_VENCIMENTO"] = _mes_num.map(_meses_pt).fillna("")
        df_receber_tratado["ANO_VENCIMENTO"] = _venc_dt.dt.year.fillna(0).astype(int)
        _st_v = (
            df_receber_tratado.get("Status_Vencimento", "")
            .fillna("")
            .astype(str)
            .str.strip()
            .str.upper()
        )
        df_receber_tratado["CLASSIFICACAO_ADIMPLENCIA"] = _st_v.map(
            {
                "VENCIDO": "INADIMPLENTE",
                "A VENCER": "ADIMPLENTE",
            }
        ).fillna("")
    mapa_id_estrito = _mapa_identificador_estrito(df_receber_tratado, df_recebidos_tratado)
    _aplicar_identificador_export(df_receber_tratado, mapa_id_estrito)
    _aplicar_identificador_export(df_recebidos_tratado, mapa_id_estrito)

    df_receber = _ordenar_colunas(_sem_colunas_internas_export(df_receber_tratado), [
        "Emp/Obra", "Venda", "Cliente", "Cliente_Base", "Identificador_Produto",
        "Parcela", "Parc_Num", "Parc_Total", "Vencimento", "Status_Vencimento",
        "DIA_VENCIMENTO_BOLETO", "MES_VENCIMENTO", "ANO_VENCIMENTO", "CLASSIFICACAO_ADIMPLENCIA",
        "Principal", "Correcao", "Juros_Atraso", "Multa_Atraso", "Correcao_Atraso", "Vlr_Parcela",
    ])
    df_recebidos = _ordenar_colunas(_sem_colunas_internas_export(df_recebidos_tratado), [
        "Emp/Obra", "Venda", "Cliente", "Cliente_Base", "Identificador_Produto",
        "Parcela", "Parc_Num", "Parc_Total", "Data_Rec", "Tipo",
        "Principal", "Correcao", "Juros_Atraso", "Multa_Atraso", "Vlr_Parcela"
    ])
    nome_aba_analitico = "DADOS GERAL"
    df_relatorio_analitico = montar_dataframe_relatorio_analitico(df_receber, df_recebidos)
    df_relatorio_analitico = _caixa_alta_exibicao_relatorio(df_relatorio_analitico)

    df_consolidado = _caixa_alta_exibicao_relatorio(df_consolidado)
    df_receber = _caixa_alta_exibicao_relatorio(df_receber)
    df_recebidos = _caixa_alta_exibicao_relatorio(df_recebidos)
    df_receber = _padronizar_colunas_exibicao(df_receber)
    df_recebidos = _padronizar_colunas_exibicao(df_recebidos)
    df_receber = _remover_colunas_totalmente_vazias(df_receber)
    df_recebidos = _remover_colunas_totalmente_vazias(df_recebidos)

    cols_pend = [
        "VENDA",
        "CLIENTE",
        "IDENTIFICADOR",
        "PARCELA",
        "TIPO DE DIVERGÊNCIA",
        "QTD.PARC.PAGO",
        "QTD.PARC.VENCIDA",
        "QTD.PARC.A VENCER",
        "QTD.SALDO",
        "QTD.TOTAL RECEBER",
        "QTD.TOTAL RECEBIDOS",
        "QTD.TOTAL ADOTADO",
        "VL.PARCELA DIVERGENTE",
        "VL.PARC.PAGO",
        "VL.VENCIDO",
        "VL.A VENCER",
        "VL.SALDO",
        "ORIGEM DA DIVERGÊNCIA",
        "MOTIVO RECONCILIAÇÃO",
        "OBSERVAÇÃO",
    ]
    mapa_renome_pend = {
        "Venda": "VENDA",
        "Cliente": "CLIENTE",
        "Identificador": "IDENTIFICADOR",
        "Parcela": "PARCELA",
        "Tipo de Divergência": "TIPO DE DIVERGÊNCIA",
        "Valor Parcela Divergente": "VL.PARCELA DIVERGENTE",
        "Origem da Divergência": "ORIGEM DA DIVERGÊNCIA",
        "Observação": "OBSERVAÇÃO",
        "__VL_PARC_PAGAS": "VL.PARC.PAGO",
        "__VL_PARC_INADIMPLENTES": "VL.VENCIDO",
        "__VL_PARC_A_VENCER": "VL.A VENCER",
        "__VL_SALDO": "VL.SALDO",
        "__QTD_PARC_PAGAS": "QTD.PARC.PAGO",
        "__QTD_PARC_INADIMPLENTES": "QTD.PARC.VENCIDA",
        "__QTD_PARC_A_VENCER": "QTD.PARC.A VENCER",
        "__QTD_SALDO": "QTD.SALDO",
        "__QTD_TOTAL_RECEBER": "QTD.TOTAL RECEBER",
        "__QTD_TOTAL_RECEBIDOS": "QTD.TOTAL RECEBIDOS",
        "__QTD_TOTAL_ADOTADO": "QTD.TOTAL ADOTADO",
        "__MOTIVO_RECONCILIACAO": "MOTIVO RECONCILIAÇÃO",
    }
    if pendencias_parcelas:
        df_pendencias_parcelas = pd.DataFrame(pendencias_parcelas)
        aux_defaults = {
            "__VL_PARC_PAGAS": 0.0,
            "__VL_PARC_INADIMPLENTES": 0.0,
            "__VL_PARC_A_VENCER": 0.0,
            "__VL_SALDO": 0.0,
            "__QTD_PARC_PAGAS": 0,
            "__QTD_PARC_INADIMPLENTES": 0,
            "__QTD_PARC_A_VENCER": 0,
            "__QTD_SALDO": 0,
        }
        for c, v0 in aux_defaults.items():
            if c in df_pendencias_parcelas.columns:
                df_pendencias_parcelas[c] = pd.to_numeric(df_pendencias_parcelas[c], errors="coerce").fillna(v0)
        df_pendencias_parcelas = df_pendencias_parcelas.rename(columns=mapa_renome_pend)
        for c in cols_pend:
            if c not in df_pendencias_parcelas.columns:
                df_pendencias_parcelas[c] = ""
        df_pendencias_parcelas = df_pendencias_parcelas[cols_pend]
    else:
        df_pendencias_parcelas = pd.DataFrame(columns=cols_pend)
    df_pendencias_parcelas = _caixa_alta_exibicao_relatorio(df_pendencias_parcelas)
    df_criterios = pd.DataFrame({
        "SEÇÃO": [
            "A. CONSOLIDADOS",
            "A. CONSOLIDADOS",
            "A. CONSOLIDADOS",
            "B. DADOS RECEBER",
            "B. DADOS RECEBER",
            "B. DADOS RECEBER",
            "C. DADOS RECEBIDOS",
            "C. DADOS RECEBIDOS",
            "C. DADOS RECEBIDOS",
            "D. DADOS GERAL",
            "D. DADOS GERAL",
            "D. DADOS GERAL",
            "E. PEND.PARCELAS",
            "E. PEND.PARCELAS",
            "E. PEND.PARCELAS",
            "F. RESUMO GERAL",
            "F. RESUMO GERAL",
            "F. RESUMO GERAL",
            "G. CONSOLIDADO ESTOQUE",
            "G. CONSOLIDADO ESTOQUE",
            "G. CONSOLIDADO ESTOQUE",
            "G. CONSOLIDADO ESTOQUE",
        ],
        "TEMA": [
            "COMO ANALISAR",
            "CRITÉRIOS UTILIZADOS",
            "LÓGICAS DEFINIDAS",
            "COMO ANALISAR",
            "O QUE REPRESENTA",
            "CRITÉRIOS UTILIZADOS",
            "COMO ANALISAR",
            "O QUE REPRESENTA",
            "CRITÉRIOS UTILIZADOS",
            "COMO ANALISAR",
            "O QUE REPRESENTA",
            "CRITÉRIOS UTILIZADOS",
            "COMO ANALISAR",
            "O QUE REPRESENTA",
            "COMO INTERPRETAR AS DIVERGÊNCIAS",
            "COMO ANALISAR",
            "O QUE REPRESENTA",
            "COMO USAR GERENCIALMENTE",
            "COMO ANALISAR",
            "O QUE REPRESENTA",
            "COMO INTERPRETAR A SITUAÇÃO DAS UNIDADES",
            "CRITÉRIOS COMPLEMENTARES",
        ],
        "DESCRIÇÃO": [
            "CONFRONTE QTD.PARCELAS E VALORES POR VENDA E IDENTIFICADOR PARA VALIDAR A POSIÇÃO FINANCEIRA E A CLASSIFICAÇÃO OPERACIONAL.",
            f"VL.CARTEIRA SEGUE CONTRATO OFICIAL: {_descricao_vl_carteira_modo()}. AS COMPOSIÇÕES DE VL.PAGO, VL.VENCIDO E VL.A VENCER PERMANECEM HOMOLOGADAS.",
            "PERCENTUAIS SÃO NORMALIZADOS PARA FECHAMENTO CONSISTENTE QUANDO HÁ POSIÇÃO; LINHAS SEM POSIÇÃO MANTÊM 0/0 COMO PADRÃO TÉCNICO.",
            "USE OS FILTROS PARA SEGMENTAR EMP/OBRA, VENDA, CLIENTE E STATUS_VENCIMENTO; O PAINEL SUPERIOR RESPONDE AO FILTRO.",
            "BASE OPERACIONAL DE CONTAS A RECEBER, COM VISÃO DE PARCELAS, VENCIMENTO/PRORROGAÇÃO E COMPONENTES DE VALOR DA PARCELA.",
            "QTD.PARCELAS, QTD.PARC.VENCIDA, QTD.PARC.A VENCER, VL.VENCIDO E VL.A VENCER SÃO APRESENTADOS NO PAINEL COM FÓRMULAS DINÂMICAS.",
            "APLIQUE FILTRO POR VENDA/CLIENTE PARA CONFERIR QTD.PARC.PAGO E VL.PAGO COM RASTREABILIDADE DE RECEBIMENTO.",
            "BASE DE CONTAS RECEBIDAS, COM PARCELAS PAGAS E VALORES EFETIVAMENTE DEPOSITADOS.",
            "PAINEL DINÂMICO EXIBE QTD.VENDAS, QTD.CLIENTES, QTD.PARCELAS E VL.PAGO.",
            "UTILIZE COMO VISÃO SINTÉTICA PARA CRUZAMENTO RÁPIDO ENTRE CLIENTE, IDENTIFICADOR E VL.PARCELA.",
            "CONSOLIDA REGISTROS DE APOIO EM FORMATO ENXUTO PARA LEITURA EXECUTIVA.",
            "CONTÉM APENAS CAMPOS ESSENCIAIS PARA LEITURA E TRIAGEM; NÃO ALTERA REGRAS DO CONSOLIDADO.",
            "ANALISE PRIMEIRO TIPO DE DIVERGÊNCIA, DEPOIS QTD. E VL. PARA PRIORIZAR CASOS CRÍTICOS.",
            "MOSTRA DIFERENÇAS ESTRUTURAIS ENTRE UNIVERSOS DE PARCELAS (PAGAS, VENCIDAS, A VENCER) COM CONTEXTO DE VENDA/CLIENTE/IDENTIFICADOR.",
            "QTD. E VL. SÃO ORGANIZADOS PARA RECONCILIAÇÃO; MANTÉM UTILIDADE DE AUDITORIA SEM RUÍDO TÉCNICO EXCESSIVO.",
            "INICIE PELO TOTAL E PELOS BLOCOS DE STATUS PARA ENTENDER A SAÚDE GLOBAL DA CARTEIRA.",
            "SÍNTESE GERENCIAL DAS VENDAS, PARCELAS E VALORES POR EMPREENDIMENTO.",
            "USE PARA PRIORIZAR AÇÕES DE COBRANÇA, PERFORMANCE COMERCIAL E ACOMPANHAMENTO EXECUTIVO.",
            "CRUZE SITUAÇÃO DAS UNIDADES COM DADOS FINANCEIROS POR IDENTIFICADOR.",
            "REPRESENTA A VISÃO DE ESTOQUE COM CLASSIFICAÇÃO DE UNIDADES (DISPONÍVEL, ADIMPLENTE, INADIMPLENTE, QUITADO).",
            "O PAINEL SUPERIOR SEGREGA ESTOQUE E SITUAÇÃO DAS VENDIDAS; A GRADE PERMITE ANÁLISE DETALHADA POR UNIDADE.",
            "CLASSIFICAÇÃO DE ESTOQUE É COMPLEMENTAR E NÃO MODIFICA VALORES FINANCEIROS HOMOLOGADOS.",
        ],
    })
    df_criterios = _caixa_alta_exibicao_relatorio(df_criterios)
    df_relatorio_analitico = _padronizar_colunas_exibicao(df_relatorio_analitico)
    df_relatorio_analitico = _remover_colunas_totalmente_vazias(df_relatorio_analitico)

    df_resumo_geral = pd.DataFrame()
    if gerar_aba_resumo_geral:
        df_resumo_geral = montar_dataframe_resumo_geral(df_consolidado)

    from services.estoque_uau import (
        COLUNAS_SAIDA_CONSOLIDADO_ESTOQUE,
        CONSOLIDADO_ESTOQUE_PANDAS_STARTROW,
        NOME_ABA_CONSOLIDADO_ESTOQUE,
        calcular_indicadores_painel_consolidado_estoque,
        carregar_estoque_bruto,
        montar_dataframe_consolidado_estoque,
    )

    if gerar_aba_consolidado_estoque:
        df_estoque_in = carregar_estoque_bruto(caminho_estoque or "")
        df_consolidado_estoque = montar_dataframe_consolidado_estoque(df_consolidado, df_estoque_in)
        if df_consolidado_estoque.empty:
            df_consolidado_estoque = pd.DataFrame(columns=COLUNAS_SAIDA_CONSOLIDADO_ESTOQUE)
    else:
        df_consolidado_estoque = pd.DataFrame()

    try:
        with pd.ExcelWriter(caminho_saida_final, engine="openpyxl") as writer:
            if gerar_aba_resumo_geral and not df_resumo_geral.empty:
                df_resumo_geral.to_excel(
                    writer, sheet_name=NOME_ABA_RESUMO_GERAL, index=False, startrow=7
                )
            df_consolidado.to_excel(writer, sheet_name=nome_aba_principal, index=False, startrow=7)
            if gerar_aba_consolidado_estoque:
                df_consolidado_estoque.to_excel(
                    writer,
                    sheet_name=NOME_ABA_CONSOLIDADO_ESTOQUE,
                    index=False,
                    startrow=CONSOLIDADO_ESTOQUE_PANDAS_STARTROW,
                )
            df_receber.to_excel(writer, sheet_name="DADOS RECEBER", index=False)
            df_recebidos.to_excel(writer, sheet_name="DADOS RECEBIDOS", index=False)
            df_relatorio_analitico.to_excel(writer, sheet_name=nome_aba_analitico, index=False, startrow=7)
            df_pendencias_parcelas.to_excel(writer, sheet_name=nome_aba_pendencias, index=False)
            df_criterios.to_excel(writer, sheet_name=nome_aba_criterios, index=False)
        _emit_perf(
            "excel_escrever_openpyxl",
            time.perf_counter() - _t_xlsx,
            _nlin_df(df_receber),
            _nlin_df(df_recebidos),
            _nlin_df(df_consolidado),
        )
        _t = time.perf_counter()
        print(
            "[ETAPA] Formatação final do Excel em andamento "
            "(em bases grandes esta etapa costuma ser a mais longa).",
            flush=True,
        )
        _emitir_progresso_motor(
            status="processando",
            mensagem="Formatação final do Excel em andamento.",
            item_atual_abas="FORMATAÇÃO FINAL",
            abas_item=["FORMATAÇÃO FINAL"],
            tempo_decorrido_segundos=max(0.0, time.perf_counter() - t_perf0),
        )
        _ind_est = (
            calcular_indicadores_painel_consolidado_estoque(df_consolidado_estoque)
            if gerar_aba_consolidado_estoque
            else None
        )
        def _cb_estilo(msg: str):
            txt = str(msg or "").strip()
            if not txt:
                return
            _emitir_progresso_motor(
                status="processando",
                mensagem=f"Formatação final: {txt}",
                item_atual_abas=txt[:120],
                abas_item=[txt[:80]],
                tempo_decorrido_segundos=max(0.0, time.perf_counter() - t_perf0),
            )
        aplicar_estilo_excel(
            caminho_saida=caminho_saida_final,
            data_base=data_base,
            nome_empreendimento=nome_empreendimento,
            nome_aba_principal=nome_aba_principal,
            indicadores_estoque=_ind_est,
            progress_cb=_cb_estilo,
        )
        _emit_perf(
            "excel_pos_formatacao_openpyxl",
            time.perf_counter() - _t,
            _nlin_df(df_receber),
            _nlin_df(df_recebidos),
            _nlin_df(df_consolidado),
        )
    except Exception as e:
        raise ProcessamentoUAUErro(
            etapa="exportação do Excel",
            funcao="processar_e_gerar_excel/aplicar_estilo_excel",
            validacao="escrita do arquivo de saída",
            mensagem="Falha ao gerar ou estilizar o arquivo Excel final.",
            campo_ou_aba="Consolidado Venda / Dados Receber / Dados Recebidos",
            erro_tecnico=e,
        ) from e

    fim_execucao = time.time()
    tempo_total = fim_execucao - inicio_execucao
    print(f"[INFO] Tempo de execução: {tempo_total:.2f} segundos")
    if perf_extra_ligado:
        _imprimir_ranking_perf()
    return caminho_saida_final, tempo_total
